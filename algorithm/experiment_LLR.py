## Get traffic measures via SUMO, real travel times of vehicles and accuracy of prediction when vehicles are
## rerouted every routing period using travel time information from the Local Level Routing algorithm for areas modelled
## by the Local Level Routing algorithm and time-dependant travel times inputed by option --path_weight_files for areas
## outside. It also uses an algorithm that chooses which artificial edge a vehicle uses based on queue length and/or
## number of vehicles, instead of travel times.
## Otherwise all vehicles would choose only one of the artificial edges that has lowest travel time.

## IMPORT MODULES

from __future__ import absolute_import
from __future__ import print_function

import os
import sys
import subprocess
import optparse
import math
import numpy as np
import scipy.stats as st
from xml.etree import ElementTree
import xml.dom.minidom
import pickle
import time
import openpyxl
import logging
from threading import Thread
import warnings
from sumolib.net import readNet  # noqa
import itertools
import csv
import logging

timestamp = time.strftime("%Y%m%dT%H%M%S")
log_file_name = "llr_zizkov_" + timestamp + ".log"

# Create a logger object
LOGGER = logging.getLogger('llr_zizkov')
LOGGER.setLevel(logging.DEBUG)
# Create file handler which logs even debug messages
fh = logging.FileHandler(log_file_name, mode='w')
fh.setLevel(logging.DEBUG)
# Create console handler with a higher log level
ch = logging.StreamHandler()
ch.setLevel(logging.INFO)
# create formatter and add it to the handlers
formatter = logging.Formatter('%(levelname)-8s %(message)s')
ch.setFormatter(formatter)
formatter = logging.Formatter('%(asctime)s - %(levelname)-8s %(message)s')
fh.setFormatter(formatter)
# add the handlers to logger
LOGGER.addHandler(ch)
LOGGER.addHandler(fh)

warnings.filterwarnings("ignore")
np.seterr(all='raise')

DO_PRINT = False
DO_DATA_COLLECTION = False
DO_NOTTS_TRACKING = True
DO_LLREX_TRACKING = True
DO_ROUTE_POLYGONS = True
HAVE_NOTTS_POLY = False
HAVE_LLREX_POLY = False

TRACKED_VEHID = None
fcd_data = []
rerouted_vehicles = []
entrance_time = {}

TRACKED_NOTTS = None
NOTTS_TIMESTAMP_SET = set()
NOTTS_VEHICLE_SET = set()
NOTTS_DATA = {}
NOTTS_COLOR = (255,255,0,255)
NOTTS_LAYER = 12
NOTTS_RADIUS = 42.0
NOTTS_POLY_ID = 'notts_route'
NOTTS_POLY_LAYER = 1

TRACKED_LLREX = None
LLREX_TIMESTAMP_SET = set()
LLREX_VEHICLE_SET = set()
LLREX_DATA = {}
LLREX_COLOR = (255,0,0,255)
LLREX_LAYER = 10
LLREX_RADIUS = 50.0
LLREX_POLY_ID = 'llrex_route'
LLREX_POLY_LAYER = 0

ROUTE_TIMESTAMP_SET = set()
ROUTE_DATA = {}

# Import python modules from the $SUMO_HOME/tools directory
try:
    # sys.path.append(os.path.join(os.path.dirname(__file__), '..', '..', '..', '..', "tools"))
    sys.path.append(
        os.path.join(os.environ.get("SUMO_HOME", os.path.join(os.path.dirname(__file__), "..", "..", "..")), "tools"))
    from sumolib import checkBinary  # noqa
except ImportError:
    sys.exit(
        "missing declare environment variable 'SUMO_HOME' as the root directory of SUMO installation "
        "(it should contain folders 'bin', 'tools' and 'docs')")


## INPUTS THAT CAN BE CHANGED BY OPTION PARSER

def get_options():
    """Initialing Script from Command Line"""
    optParser = optparse.OptionParser()

    # Experiment Scenario Name
    optParser.add_option("--scenario_name",
                         default="mid.LLR.100PR",
                         action="store", type="string",
                         help="Name of the simulation scenario")

    # Python Interface Connection
    optParser.add_option("--interface", action="store", type="string", default="traci",
                         help="Define which connection interface to choose, if SUMO either traci or libsumo")

    # SUMO GUI Option
    optParser.add_option("--nogui", action="store_true", default=False, help="run the commandline version of sumo")

    # SUMO Simulation Paths Inputs and Outputs
    optParser.add_option("--simul_filename",
                         default="zizkov.sumocfg",
                         action="store", type="string",
                         help="SUMO simulation file")
    optParser.add_option("--net_filename",
                         default="zizkov_flora_closed.net.xml",
                         action="store", type="string",
                         help="SUMO network file")
    optParser.add_option("--fixed_route_files",
                         default="routes.mid.rou.xml,routes.probes.rou.xml",
                         action="store", type="string",
                         help="SUMO fixed (same for all scenarios) vehicle route files, use comma as separator")
    optParser.add_option("--dyn_route_files",
                         default="",
                         action="store", type="string",
                         help="SUMO dynamic (appending sce. name) vehicle route files, use comma as separator")
    optParser.add_option("--fixed_additional_files",
                         default="tll_static_closed.xml",
                         action="store", type="string",
                         help="SUMO fixed (same for all scenarios) additional files, use comma as separator")
    optParser.add_option("--dyn_additional_files",
                         default="additional.add.xml",
                         action="store", type="string",
                         help="SUMO dynamic (appending sce. name) additional files, use comma as separator")
    optParser.add_option("--path_inputs",
                         default="./Inputs_Outputs/",
                         action="store", type="string",
                         help="path of SUMO input files")
    optParser.add_option("--path_outputs",
                         default="./Inputs_Outputs/Necessary_Outputs/",
                         action="store", type="string",
                         help="path of SUMO output files")

    # SUMO Simulation Time Interval
    optParser.add_option("--begin", default="43200", action="store", type="string",
                         help="Begin step simulation")
    optParser.add_option("--end", default="55800", action="store", type="string",
                         help="End step simulation")
    optParser.add_option("--step_length", default="1", action="store", type="string",
                         help="Simulation step length in seconds")
    optParser.add_option("--warm_up_t", default=1800, action="store", type="int",
                         help="Warm up time for the travel time accuracy report")

    # SUMO Routing Parameters
    optParser.add_option("--routing_probability", default="1", action="store", type="string",
                         help="Probability of a vehicle with routing device")
    optParser.add_option("--routing_period", default="120", action="store", type="string",
                         help="Interval between routing queries")
    optParser.add_option("--routing_pre_period", default="1", action="store", type="string",
                         help="Interval between routing queries before trip")
    optParser.add_option("--routing_init_loaded_weights", default="true", action="store", type="string",
                         help="True if vehicles have onboard unit with pre-loaded average travel times,"
                              "and seting travel times per vehicle and not per edge at once")
    optParser.add_option("--tts_per_veh", action="store_true", default=False,
                         help="travel times are set to detected CAVs instead of globaly per modelled edges."
                         "Useful only for SUMO (Traci/Libsumo)")

    # SUMO Context Subscription Parameters
    optParser.add_option("--std_comm_range", default=1000, action="store", type="float",
                         help="communication range in meters. Useful only for SUMO (Traci/Libsumo)")

    # SUMO Lane Closure Input for Local Level Routing
    optParser.add_option("--interval_lane_closure", nargs=2, default=(43201,55801), action="store", type="float",
                         help="begin and end time of lane closure")
    optParser.add_option("--lane_to_close", default="dz_50178032#1.132_1,gneE46_0", action="store",
                         type="string", help="network id of lane to close. Use comma as separator")
    optParser.add_option("--closed_art_edge", default="gneE46", action="store",
                         type="string", help="network id of artificial edge that will close. Use comma as separator")
    optParser.add_option("--dist_to_close", default=-1, action="store", type="float",
                         help="distance to stop line the lane should be closed, if -1 it uses lane length")

    # Path Weight Files (Travel Times)
    optParser.add_option("--path_weight_files",
                         default="./Inputs_Outputs/Scenario_Dependent_Travel_Times/",
                         action="store", type="string",
                         help="path of file with Travel Times")

    # Data for Automatic Collection of Network Information
    # For this case, using SUMO outputs from previous simulations of travel time collection.
    optParser.add_option("--use_detectors_data", action="store_true", default=False,
                         help="Use this option use SUMO instant detectors data instead of flow and turning rates")
    optParser.add_option("--centralized_sys", action="store_true", default=False,
                         help="Use this option in case the system is expected to be centralized"
                              "ie. no LLR messages exchanges between TCs, all realtime data managed by one TC")
    optParser.add_option("--tlights_filename",
                         default="tll_static_closed.xml",
                         action="store", type="string",
                         help="SUMO signalized junctions traffic light plans file")
    optParser.add_option("--turndefs_filename",
                         default="zizkov_closed.turndefs.xml",
                         action="store", type="string",
                         help="SUMO edge turn definitions file")
    optParser.add_option("--flows_filename",
                         default="flows.mid.xml",
                         action="store", type="string",
                         help="SUMO edge flows file")
    optParser.add_option("--vtypes_filename",
                         default="",
                         action="store", type="string",
                         help="SUMO vehicle types file")
    optParser.add_option("--path_det_data",
                         default="./Inputs_Outputs/Scenario_Detectors_Data/",
                         action="store", type="string",
                         help="path of SUMO lane detectors' and outputs files")

    # Local Level Routing Parameters
    optParser.add_option("--no_probes", action="store_false", default=False,
                         help="Use this option to not use real neither artificial vehicles as probe vehicles "
                              "to estimate travel times. Only the deterministic queueing model is used then.")
    optParser.add_option("--std_t_max", default=120, action="store", type="float",
                         help="cycle prediction in seconds. Must be multiple of std_len_range")
    optParser.add_option("--std_len_range", default=20, action="store", type="float",
                         help="time interval of arrival ranges in seconds. Must be divisor of std_t_max")
    optParser.add_option("--queue_pred_resolution", default=5, action="store", type="float",
                         help="resolution of the queue prediction for TC modelled lanes in seconds,"
                              "notice that lanes not modelled by TC, the TC receives using std_len_range resolution")
    optParser.add_option("--std_reaction_t", default=-1, action="store", type="float",
                         help="standard reaction time (it can be different per vehicle type)."
                              "If -1, it uses step length")
    optParser.add_option("--std_accel_adjust", default=1, action="store", type="float",
                         help="standard adjustment of vehicle acceleration capability.")
    optParser.add_option("--std_decel_adjust", default=1, action="store", type="float",
                         help="standard adjustment of vehicle perceived deceleration")
    optParser.add_option("--delay_arrivals", default=0, action="store", type="int",
                         help="Delay time of arrivals (in sec) on the algorithm "
                              "(when take some time to vehicles reach modelled area of the network)")
    optParser.add_option("--max_tc_interactions", default=2, action="store", type="int",
                         help="Number of changes of traffic controllers order to be explored per algorithm update")
    optParser.add_option("--route_share_prob", default=1, action="store", type="float",
                         help="Probability that CAVs will share their route to RSUs."
                         "Useful only for SUMO (Traci/Libsumo)")

    options, args = optParser.parse_args()
    return options


## MAIN ENTRY POINT FOR SUMO

if __name__ == "__main__":
    options = get_options()

    net = readNet(options.path_inputs + options.net_filename)

    # this script has been called from the command line. It will start sumo as a server, then connect and run
    try:
        if options.interface == "libsumo":
            import libsumo as traci
            # As GUI is not working yet with libsumo, use only without GUI
            sumoBinary = checkBinary('sumo')
            INTERFACE = "libsumo"
        elif options.interface == "traci":
            raise ImportError
        else:
            # Import here module that connects the Python script with CAVs data interface
            pass
    except ImportError:
        import traci
        INTERFACE = "traci"
        if options.nogui:
            sumoBinary = checkBinary('sumo')
        else:
            sumoBinary = checkBinary('sumo-gui')

    # For simulation purposes, it may be desired to test a scenario where the system would be centralized
    # into one traffic controller
    if options.centralized_sys:
        # If centralized system scenario, use the same CAVs of the decentralized system scenario
        routing_prob = 0
    else:
        routing_prob = options.routing_probability

    # Define SUMO additional files
    add_filenames = options.fixed_additional_files.split(",")
    if options.dyn_additional_files != "":
        add_filenames.extend([dynf.split(".")[0] + "." + options.scenario_name + "." + ".".join(dynf.split(".")[1:])
                              for dynf in options.dyn_additional_files.split(",")])
    additional_files = options.path_inputs + options.path_inputs.join([afile + "," for afile in add_filenames][:-1]
                                                                      + [add_filenames[-1]])

    # Define SUMO route files
    route_filenames = options.fixed_route_files.split(",")
    if options.dyn_route_files != "":
        route_filenames.extend([dynf.split(".")[0] + "." + options.scenario_name + "." + ".".join(dynf.split(".")[1:])
             for dynf in options.dyn_route_files.split(",")])
    route_files = options.path_inputs + options.path_inputs.join([rfile + "," for rfile in route_filenames][:-1]
                                                                 + [route_filenames[-1]])

    # this is the normal way of using traci. sumo is started as a
    # subprocess and then the python script connects and runs
    params = [sumoBinary, "-c", "zizkov_alt.sumocfg",
                 "--net", options.path_inputs + options.net_filename,
                 "--route-files", route_files,
                 "--additional-files", additional_files,
                 # "--log", options.path_outputs.replace("Necessary_Outputs","Unecessary_Outputs") + "SUMOlog."
                 # + options.scenario_name + ".txt",
                 "--error-log", options.path_outputs.replace("Necessary_Outputs","Unecessary_Outputs") + "SUMOlog."
                 + options.scenario_name + ".txt",
                 "--vehroute-output",options.path_outputs + "vehroute_" + timestamp + "." + options.scenario_name + ".xml",
                 "--tripinfo-output",options.path_outputs + "tripinfo_" + timestamp + "." + options.scenario_name + ".xml",
                 "--begin",options.begin,
                 "--end",options.end,
                 "--emission-output", "llr_emission_" + timestamp + ".out.xml",
                 "--gui-settings-file", "zizkov_alt_1600x1000.view.xml",
                 "--window-size", "1600,1000",
                 "--window-pos", "0,0",
                 "--step-length", options.step_length,
                 "--device.rerouting.probability",routing_prob,
                 "--device.emissions.probability", "1",
                 "--device.rerouting.period", "100000",
                 "--device.rerouting.adaptation-interval", "0",
                 "--device.rerouting.pre-period", "0" ] \
                 + (lambda reac_t: ["--default.action-step-length", reac_t]
                 if reac_t != -1 else [])(options.std_reaction_t) \
                 + (lambda bool: ["--device.rerouting.init-with-loaded-weights",options.routing_init_loaded_weights]
                 if bool == True else [])(options.tts_per_veh) \
                 + (lambda bool: ["--weight-files",options.path_weight_files + "travel_times."
                                                   + options.scenario_name.replace("LLR.","")
                                                   + "." + options.routing_period + "P" + ".xml"]
                 if bool == True else [])(options.tts_per_veh)
                 
    traci.start(params)
    LOGGER.info("Started SUMO as: " + " ".join(params))
    
    if DO_NOTTS_TRACKING:
        with open('fcd_notts_samples.csv', 'rb') as cf:
            fr = csv.reader(cf)
            skip_row = True
            for row in fr:
                if skip_row:
                    skip_row = False
                    continue
                ts = float(row[0])
                veh_id = row[1]
                veh_x = float(row[18])
                veh_y = float(row[19])
                NOTTS_DATA[ts] = (veh_x, veh_y)
                NOTTS_TIMESTAMP_SET.add(ts)
                NOTTS_VEHICLE_SET.add(veh_id)

        LOGGER.info('NOTTS vehicles to track: {:s}'.format(str(NOTTS_VEHICLE_SET)))
    else:
        LOGGER.info('NOTTS vehicle tracking is not active')

    if DO_LLREX_TRACKING:
        with open('fcd_llr_samples.csv', 'rb') as cf:
            fr = csv.reader(cf)
            skip_row = True
            for row in fr:
                if skip_row:
                    skip_row = False
                    continue
                ts = float(row[0])
                veh_id = row[1]
                veh_x = float(row[18])
                veh_y = float(row[19])
                LLREX_DATA[ts] = (veh_x, veh_y)
                LLREX_TIMESTAMP_SET.add(ts)
                LLREX_VEHICLE_SET.add(veh_id)

        LOGGER.info('LLR vehicles to track: {:s}'.format(str(LLREX_VEHICLE_SET)))
    else:
        LOGGER.info('LLR vehicle tracking is not active')
        
    if DO_ROUTE_POLYGONS:
        with open('fcd_routes.csv', 'rb') as cf:
            fr = csv.reader(cf)
            skip_row = True
            for row in fr:
                if skip_row:
                    skip_row = False
                    continue
                ts = float(row[0])
                veh_id = row[1]
                command = row[2]
                gain_dist = float(row[3])
                gain_time = float(row[4])
                edge_list = row[5:]
                ROUTE_DATA[ts] = (command, veh_id, gain_dist, gain_time, edge_list)
                ROUTE_TIMESTAMP_SET.add(ts)

        LOGGER.info('Routes will be tracked')
    else:
        LOGGER.info('Route tracking is not active')
        
## GLOBAL CONSTANTS

# For simulation purposes, define TraCI Constants when using SUMO
LANE_VEH_NUM = traci.constants.LAST_STEP_VEHICLE_NUMBER
LANE_QUEUE_NUM = traci.constants.LAST_STEP_VEHICLE_HALTING_NUMBER
LANE_OCCUPANCY = traci.constants.LAST_STEP_OCCUPANCY
VEHIDS = traci.constants.LAST_STEP_VEHICLE_ID_LIST
DOMAIN = traci.constants.CMD_GET_VEHICLE_VARIABLE
TYPE = traci.constants.VAR_TYPE
LANE = traci.constants.VAR_LANE_ID
ROUTE = traci.constants.VAR_EDGES
POSITION = traci.constants.VAR_LANEPOSITION
SPEED = traci.constants.VAR_SPEED
ACC = traci.constants.VAR_ACCEL
DEC = traci.constants.VAR_APPARENT_DECEL
MINGAPT = traci.constants.VAR_TAU
SPEEDFACTOR = traci.constants.VAR_SPEED_FACTOR
LENGTH = traci.constants.VAR_LENGTH
MINGAPD = traci.constants.VAR_MINGAP
VEH_INFO = [TYPE, LANE, ROUTE, POSITION, SPEED, ACC, DEC, MINGAPT, SPEEDFACTOR, LENGTH, MINGAPD]
LANE_METRICS = [LANE_QUEUE_NUM, LANE_VEH_NUM]

# SUMO Version
try:
    SUMOversion = traci.getVersion()[1]
    if SUMOversion == "SUMO UNKNOWN":
        raise AttributeError
    else:
        if "_" in SUMOversion:
            SUMO_RELEASE = int(SUMOversion.split("_")[0][-1])
            SUMO_PATCH = int(SUMOversion.split("_")[1])
        else:
            SUMO_RELEASE = int(SUMOversion.split(".")[0][-1])
            SUMO_PATCH = int(SUMOversion.split(".")[1])
except (KeyError, AttributeError):
    # If module has no getVersion() or version is unkown, assumes that version is not earlier than 1.1.0
    SUMO_RELEASE = 1
    SUMO_PATCH = 1

# The prior edges are edges before artificial edges,
# artificial edges are edges representing one lane of a multi-lane incoming edge at certain junction.
# The index of artificial edge is the lane number of the prior edge
PRIOR_ARTIFICIAL_EDGES = {"gneE247": ("gneE248","gneE249"),
                          "fl_156262933#4": ("gneE246","gneE245"),
                          "fl_330512996#6.19": ("gneE240.50","gneE241.49"),
                          "gneE57": ("gneE60","gneE59"),
                          "fl_80779632#4": ("gneE148","gneE149"),
                          "dz_142147141": ("gneE143","gneE144"),
                          "dz_50178032#1.132": ("gneE47","gneE45")}

PRIOR_COMMON_AFTER_ART_EDGES = {"gneE247": "gneE293",
                                "fl_156262933#4": "fl_483444577#0",
                                "fl_330512996#6.19": "fl_448565742.11",
                                "gneE57": "fl_4641298#3",
                                "fl_80779632#4": "dz_330512997",
                                "dz_142147141": "fl_330512996#6.19",
                                "dz_50178032#1.132": "dz_142147149#0"}

# List of junctions with RSU
JCTS_WH_RSU = ("dz_gneJ16","cluster_dz_gneJ10","kOlsanske","cluster_gneJ114","cluster_gneJ14","cluster_gneJ291",
               "cluster_fl_gneJ25","cluster_gneJ272","cluster_gneJ240", "cluster_or_1140095002", "dz_21673405")

# List of junctions that are signalized but must be set as non-connected in the algorithm
# the default setting is that signalized junctions are connected
SIGNALIZED_JCTS_NON_C = ()
# SIGNALIZED_JCTS_NON_C = ("cluster_gneJ294", "cluster_or_1140095002", "cluster_gneJ289")

# List of junctions that are unsignalized but must be set as connected in the algorithm
UNSIGNALIZED_JCTS_C = ("dz_gneJ19","gneJ13","dz_5328335609","cluster_dz_1138616737","dz_21673419",
                       "dz_1131753614","dz_gneJ2","dz_306268757","dz_1131736467","gneJ97","fl_1134588976",
                       "fl_gneJ13","fl_4641299#11_2","fl_gneJ44","gneJ6","fl_gneJ50","cluster_gneJ161","fl_gneJ52",
                       "or_1826281371","fl_gneJ32","or_21673293","or_1826281337","or_1140094980","or_1140088627",
                       "or_1140088592","fl_1134613358","fl_cluster_2","fl_2014792260","gneJ313",
                       "fl_3450114685","fl_371075952", "cluster_dz_21311889", "cluster_dz_21311892")

# Definition or creating a reference for the groups of junctions to be modelled by each TC
TCS_REF_JCTS = ( # Leave empty tuple for a traffic controller for each traffic light controlled junctions
                 # (if TC list is reference <"ref"> or defined already order <"def">,
                 # TC list of junctions <"jctID"> where if "def" if is already the order
                 # and "ref" if the order will be based on the dist to the TC reference point,
                 # TC reference point <x,y>)
                 # TC list reference is junctions to be on the same TC but may include other junctions
                 # based on the junction distance to the TC reference point
                 # Leave empty the list of junction and empty ref. coordinates cases are not applicable
                (("def.",("gneJ13","dz_gneJ16","dz_gneJ19")),()),
                (("def.",("dz_21673405","dz_1131753603","dz_5328335609","cluster_dz_gneJ10",
                          "cluster_dz_1138616737","gneJ6")),()),
                (("def.",("fl_gneJ44","fl_gneJ50","kOlsanske","fl_4641299#11_2","cluster_gneJ161","fl_gneJ52")),()),
                (("def.",("or_1140088592","or_1140088627","cluster_gneJ240","gneJ98")),()),
                (("def.",("or_1140094980","cluster_or_1140095002")),()),
                (("def.",("or_1826281337","or_21673293","cluster_gneJ272","or_1826281371")),()),
                (("def.",("fl_gneJ32","cluster_gneJ294","cluster_fl_gneJ25",
                          "fl_3450114685","fl_371075952","fl_244011065","cluster_fl_21311916")),()),
                (("def.",("cluster_fl_244012009","fl_1134613358","fl_cluster_2","fl_2014792260",
                          "gneJ313","cluster_gneJ291", "fl_244011064", "cluster_gneJ306")),()),
                (("def.",("fl_gneJ13","fl_1134588976","cluster_gneJ114","gneJ97","cluster_dz_21311892",
                          "dz_25973094","cluster_dz_21311889")),()),
                (("def.",("cluster_gneJ289","dz_1131736467","dz_1131753614","dz_21673419","cluster_gneJ14",
                          "dz_306268750","dz_306268757","dz_gneJ2")),()))

# Origin coordinates <x,y> in case not given TCS_REF_JCTS.
ORIGIN_COORD = () # Leave empty if TCS_REF_JCTS is given

# Coordinates of the Network Limits (for plotting the map with information)
NET_XLIM = "1510,3700"
NET_YLIM = "435,2815"
NET_XLIM_CLOSED = ["2860,3120"] # one index for each desirable zoom of certain area. It will create one fig. per index
NET_YLIM_CLOSED = ["1950,2230"] # one index for each desirable zoom of certain area. It will create one fig. per index

# Maximum Values of Traffic Flow Volumes
MIN_FLOW = 0.1  # Min Flow to calculate connection capacity. In veh/s = 1 veh/10 s. This is 360 veh/h
LOW_FLOW = 0.25 # Low Flow. In veh/s = 1 veh/4 s. This is 900 veh/h
MID_FLOW = 0.5  # Medium Flow. In veh/s = 1 veh/2 s. This is 1800 veh/h.
MIN_HDWY = 1    # Minimum headway between vehicles. In seconds.
                # This would be a maximum capacity, MAX_FLOW = 1 veh/s. This is 3600 veh/h.

# Standard Deviation of the Error for the Queue Length Estimation (Assuming Error Following Normal Distribution)
# In percentage (of real queue length in vehicle numbers)
QUEUE_EST_LOW_DEV = 0.2
QUEUE_EST_MID_DEV = 0.25
QUEUE_EST_HIGH_DEV = 0.35

# GNSS Typical Error, in meters
GNSS_ERROR = 7 # at best location

# Physics Values
FRICTION_COEF = 0.7 # Coeficient of friction, dimensionless
GRAVITY_ACC = 10 # Gravitational acceleration, in m/s^2
STD_ACC_T = 2 # Standard acceleration lost time, in seconds

# Scenario and Path Names
SCE_NAME = options.scenario_name
PATH_OUTPUT = options.path_outputs

# Net File
NET_FILE = options.path_inputs + options.net_filename

# SUMO GUI Option and Connection
NOGUI = options.nogui

# SUMO Simulation Time Interval
BEGIN_T = round(float(options.begin),1)
END_T = round(float(options.end),1)
STEP_LENGTH = round(float(options.step_length),1)
WARM_UP_T = round(float(options.warm_up_t), 1)

# SUMO Routing Parameters
ROUTING_PROB = float(options.routing_probability) # Penetration rate of CAVs
ROUTING_PERIOD = float(options.routing_period) # Interval between routing queries before trip
TTS_PER_VEH = options.tts_per_veh # Boolean if setting travel times to each detected CAV or globally per edge

# SUMO Context Subscription Parameters
STD_COMM_RANGE = options.std_comm_range # communication range in meters

# Road Closure
BEGIN_LANE_CLOSURE = options.interval_lane_closure[0]
END_LANE_CLOSURE = options.interval_lane_closure[1]
LANES2CLOSE = options.lane_to_close.split(",")
EDGE2CLOSE = options.closed_art_edge.split(",")
DIST2CLOSE = options.dist_to_close

# Data for Automatic Collection of Network Information
CENTRALIZED_SYS_BOOL = options.centralized_sys
if options.use_detectors_data:
    TURN_DEFS_FILE = None
    FLOWS_FILE = None
    # Number of intervals to have the edge attributes specific for each interval.
    # Default every hour
    NUM_INTERVS = max(int((END_T - BEGIN_T) / 3600), 1)
    END_DETECTORS_FILE = options.path_det_data + "det_end_output." + SCE_NAME.replace("LLR.","") + ".xml"
    BEGIN_DETECTORS_FILE = options.path_det_data + "det_begin_output." + SCE_NAME.replace("LLR.","") + ".xml"
    INF_DETECTORS_FILE = options.path_det_data + "detectors." + SCE_NAME.replace("LLR.","") + ".add.xml"
else:
    INF_DETECTORS_FILE = None
    END_DETECTORS_FILE = None
    BEGIN_DETECTORS_FILE = None
    TURN_DEFS_FILE = options.path_inputs + options.turndefs_filename
    FLOWS_FILE = options.path_inputs + options.flows_filename

if options.vtypes_filename != "":
    VTYPES_FILE = options.path_inputs + options.vtypes_filename
else:
    VTYPES_FILE = FLOWS_FILE
if options.tlights_filename != "":
    MOV_GROUPS_FILE = options.path_inputs + options.tlights_filename
else:
    MOV_GROUPS_FILE = NET_FILE

# For simulation purposes, get list of vehicles to be set as CAVs if a centralized system scenario
# to use same vehicles of the decentralized system scenario, to ensure same conditions between scenarios
if CENTRALIZED_SYS_BOOL:
    rem_cavs_list = set([])
    rem_noncavs_list = set([])
    VEH_LIST_FILE = options.path_outputs + "tripinfo." + options.scenario_name.replace(".C","") + ".xml"
    tree_all_vehs = ElementTree.parse(VEH_LIST_FILE)
    trips = list(tree_all_vehs.iter('tripinfo'))
    for t in trips:
        if "routing" in t.get('devices'):
            rem_cavs_list.add(t.get('id'))
        else:
            rem_noncavs_list.add(t.get('id'))

# Local Level Routing Database for the Collected Network Information
DB_FILE = options.path_inputs + "database." + SCE_NAME + ".pkl"

# Local Level Routing Parameters
USE_PROBES = options.no_probes # bool if using artificial and real vehicles as probes
MAX_T_MAX_CYCLE = 3 # Maximum Value For Extrapolation of Vehicle Movements Over T_MAX
                    # If vehicle don't cross up this value, it will not try to cross it anymore, ie. [0,1,2]
STD_T_MAX = options.std_t_max # cycle prediction in seconds
LEN_RANGE = round(float(options.std_len_range), 1)# time interval of ranges dividing t_max in seconds
QUEUE_PRED_RES = options.queue_pred_resolution # time resolution of queue prediction for a TC modelled lanes
STD_REACTION_T = (lambda reac_t: STEP_LENGTH if reac_t == -1 # standard reaction time
                  else reac_t)(options.std_reaction_t) # (but it can be different for each vehicle type)
STD_ACCEL_ADJUST = options.std_accel_adjust # standard adjustment of vehicle acceleration capability
STD_DECEL_ADJUST = options.std_decel_adjust # standard adjustment of vehicle perceived deceleration
BEGIN_ARRIVALS = BEGIN_T + options.delay_arrivals # delay time of arrivals (in sec) on the algorithm
TC_IT_MAX = options.max_tc_interactions # Number of changes of traffic controllers order to be explored
ROUTE_SHARE_PROB = options.route_share_prob # Probability that CAVs will share their route to RSUs

# CAVs that entered LLR area for Aggregating Results
CAVS_AREA_FILE = options.path_outputs + "CAVs_modelled_area." + SCE_NAME + ".xlsx"

# nonCAVs that entered LLR area for Aggregating Results
NONCAVS_AREA_FILE = options.path_outputs + "nonCAVs_modelled_area." + SCE_NAME + ".xlsx"

# All vehicles that entered LLR area for Aggregating Results
ALL_VEHS_AREA_FILE = options.path_outputs + "allVehs_modelled_area." + SCE_NAME + ".xlsx"

# File with Measured, Estimated and Std. of Travel Times
TTS_ACCUR_FULL = options.path_outputs + "TTs_accur." + SCE_NAME + ".xlsx"
TTS_ACCUR_AGGREGATED = options.path_outputs + "python_edge_traffic." + SCE_NAME + ".xml"

# File with Measured, Estimated and Std. of departures
DEPS_ACCUR_FULL = options.path_outputs + "deps_accur." + SCE_NAME + ".xlsx"

# File with all Connection Edge IDs for Setting Travel Time Zero on Them and Check if Vehicle is on an Internal Lane
ALL_CONN_AREA_FILE = options.path_outputs + "allConn_modelled_area." + SCE_NAME.split(".")[0] + ".xlsx"

# Get Enviroment Variables
MY_ENV = os.environ.copy()
MY_ENV["PYTHONPATH"] = "/usr/sbin:/sbin:" + MY_ENV["PYTHONPATH"]



## LOGGING
logging.basicConfig(filename=options.path_outputs.replace("Necessary_Outputs/","Unecessary_Outputs/") + 'PYTHONlog.' +
                             SCE_NAME + ".log",level=logging.INFO,format='%(asctime)s %(message)s')
logger = logging.getLogger(__name__)
logger.info('Start Simulation')



## MONITOR SCRIPT RUNTIME

# Hierarchy of Functions (Top ones sum all its bottom ones)
f_names = ["1 rcvLdmDataMsgs",                                     # 0
           "2 sendLdmDataMsgs",                                    # 1
           "3 getCAVinRange",                                      # 2
           "1.1/6.1 updateTravelTimeDB",                           # 3
           "4 BroadcastAreaTravelTime",                            # 4
           "5 getSignalizedJctMovGroupPlans",                      # 5
           "6.1.1/8.2.1.3.1 updateDynConnCapacity",                # 6
           "9 getVehAlreadyOnLane",                                # 7
           "12.1 getInLaneQueueFromSUMO",                          # 8
           "2.1 updateDynLaneInflowProfile",                       # 9
           "7 estArrivals",                                        # 10
           "8 estDepartures",                                      # 11
           "6 estEdgeTravelTime",                                  # 12
           "6.1 estTravelTimeNonProbeVehs",                        # 13
           "2.2 updateQueueLength",                                # 14
           "10 ResetVehsOnLane",                                   # 15
           "10.1 lanes2reset,vehs_to_del",                         # 16
           "10.2 map_old2new_num",                                 # 17
           "10.3/11.2.1 delVehAttrs",                              # 18
           "10.4 reset vehs_on_lane and others",                   # 19
           "11 genCAVOnLane",                                      # 20
           "7.2.4/8.2.5/8.2.2.4/11.5/12.5 defVehAttrOnLane",       # 21
           "12 genStartingVehOnLane",                              # 22
           "11.2/9.1 delAttrPrevRouteEdgs",                        # 23
           "7.2.1/8.2.2.1/11.1/12.2 setVehParams",                 # 24
           "8.1 orderVehLane",                                     # 25
           "8.2.1.1.1.2.1 getTimeSpeedDistOfThreshDist",           # 26
           "8.2.1.(2/3/1.1.3) getTimeSpeedDistOfAccConstMov",      # 27
           "8.2.1.(2/3/1.2) getTimeSpeedDistOfDecelMov",           # 28
           "8.2 defVehMovement",                                   # 29
           "8.2.1.(1/2/3) estRefNextArrRange",                     # 30
           "8.2.1 estVehConnDepartureTime",                        # 31
           "8.2.1.1 Braking and Correcting",                       # 32
           "8.2.1.2 Accelerating and Delaying",                    # 33
           "8.2.1.3 Waiting Opposite",                             # 34
           "7.2.2/8.2.3/8.2.2.2/11.3/12.3 defVehNextEdgeJct",      # 35
           "7.2.3/8.2.4/8.2.2.3/11.4/12.4 defVehConnMovgNextLane", # 36
           "14 setEdgesAttrbInterval",                             # 37
           "4.1 setAdaptedTraveltime",                             # 38
           "4.2 rerouteTraveltime",                                # 39
           "7.1 Def Arriving Headways",                            # 40
           "7.2 Insert Arriving Vehs",                             # 41
           "8.3 Insert Probes",                                    # 42
           "7.3 Reseting not lane order vehs",                     # 43
           "8.2.1.1.2.1 Storing Inc. Lane Queue Length",           # 44
           "8.2.1.2.1 Retrieving Out. Lane Queue Length",          # 45
           "8.2.1.1.1 estvehconn init + threshdist",               # 46
           "8.2.1.1.2 Braking",                                    # 47
           "8.2.1.1.3 Correcting",                                 # 48
           "8.2.1.1.1.1 estvehconn init",                          # 49
           "8.2.1.1.1.2 Only Thresdist",                           # 50
           "Vacant",                                               # 51
           "8.2.1 def arrivals as probes",                         # 52
           "8.2.2 gen missing probes",                             # 53
           "8.2.1.3.1/6.1.1 estDelayConnCap",                      # 54
           "15 adjQueuePred",                                      # 55
           "Total"]
timings = [[f_name,0,0.0] for f_name in f_names] # function name, sum time, % total



## MAIN SCRIPT FUNCTIONS

def frange(start, end=None, inc=None):
    """A range function, that does accept float increments"""

    if end == None:
        end = start + 0.0
        start = 0.0

    if inc == None:
        inc = 1.0

    L = []
    while 1:
        next = start + len(L) * inc
        if inc > 0 and next >= end:
            break
        elif inc < 0 and next <= end:
            break
        L.append(next)

    return L

def getExcelTargetList(filename):
    """Get Values of Each Row of the First Collum of an Excel file"""

    book = openpyxl.load_workbook(filename, data_only=True)
    sheet = book.active
    rows = sheet.rows
    file_input_vals = []
    for row in rows:
        for cell in row:
            file_input_vals.append(cell.value)

    return file_input_vals
    
    
def log_uncaught_exceptions(*exc_info): 
    LOGGER.critical('Unhandled exception:', exc_info=exc_info)
    

def get_marker_shape(vx, vy, r=50.0, np=16):
    circle_shape = list()
    for i in range(np):
        phi = i * 2 * math.pi / np
        p = (vx + r*math.sin(phi), vy + r*math.cos(phi))
        circle_shape.append(p)
    return circle_shape


def add_marker(vehID, polygonID):
    vx, vy = traci.vehicle.getPosition(vehID)
    circle_shape = get_marker_shape(vx, vy)
    traci.polygon.add(polygonID, shape=circle_shape, color=(255,0,0,255), fill=True, polygonType="circle", layer=10)
    # traci.polygon.setFilled(polygonID, True)
    traci.polygon.addDynamics(polygonID, vehID, rotate=False)


def add_marker_xy(x, y, polygonID, redraw=False, remove=False, marker_color=(255,255,0,255), marker_layer=11, marker_radius=42.0):
    p_inner = polygonID + "_inner"
    p_outer = polygonID + "_outer"
    if remove:
        traci.polygon.remove(p_inner)
        traci.polygon.remove(p_outer)
    else:
        marker_shape_outer = get_marker_shape(x, y, r=marker_radius+4.0)
        marker_shape_inner = get_marker_shape(x, y, r=marker_radius)
        if redraw:
            traci.polygon.setShape(p_inner, marker_shape_inner)
            traci.polygon.setShape(p_outer, marker_shape_outer)
        else:
            LOGGER.info('drawing NEW polygon {:s} at ({:f},{:f})'.format(polygonID, x, y))
            # LOGGER.info('new shape is {:s}'.format(str(marker_shape)))
            traci.polygon.add(p_outer, shape=marker_shape_outer, color=(0,0,0,255), fill=True, polygonType="circle", layer=marker_layer)
            traci.polygon.add(p_inner, shape=marker_shape_inner, color=marker_color, fill=True, polygonType="circle", layer=marker_layer+1)


def draw_vehicle_route(routeID="route", route=None, redraw=False, route_color=None, route_layer=0):
    global net
    if redraw:
        traci.polygon.remove(routeID)
        if route_color is None:
            route_color = (255,255,255,255)
    elif route_color is None:
        route_color = (0,0,255,255)    
    if route:
        lanes = []
        for e in route:
            lanes.append(net.getEdge(e).getLane(0))
        route_shape = list(itertools.chain(*list(l.getShape() for l in lanes)))
        LOGGER.debug(route_shape)
        traci.polygon.add(routeID, shape=route_shape, color=route_color, lineWidth=30, polygonType="route", layer=route_layer)


def get_route_params(route):
    global net
    route_tt = 0.0
    route_len = 0.0
    e_tt_max = 0.0
    e_tt_max_id = None
    for e in route:
        e_tt = traci.edge.getTraveltime(e)
        if e_tt > e_tt_max:
            e_tt_max = e_tt
            e_tt_max_id = e
        route_tt += e_tt
        route_len += net.getEdge(e).getLength()
    # if e_tt_max > 100.0:
    #    LOGGER.warn('edge `{:s}` traveltime: {:f}'.format(e_tt_max_id, e_tt_max))
    return route_tt, route_len
    

def get_fcd_row(vehID):
    accel = traci.vehicle.getAcceleration(vehID)
    wt = traci.vehicle.getWaitingTime(vehID)
    fuel = traci.vehicle.getFuelConsumption(vehID)
    e_co = traci.vehicle.getCOEmission(vehID)
    e_co2 = traci.vehicle.getCO2Emission(vehID)
    e_nox = traci.vehicle.getNOxEmission(vehID)
    return ACTUAL_T, vehID, wt, accel, fuel, e_co, e_co2, e_nox


def handleSUMOsimulationData(time2updte):
    """Get traffic data data and use by calling other functions.
    This function is special for SUMO"""

    global lastStep_area_CAVs
    global TRACKED_VEHID
    global TRACKED_NOTTS
    global TRACKED_LLREX
    global NOTTS_TIMESTAMP_SET
    global NOTTS_DATA
    global HAVE_LLREX_POLY
    global HAVE_NOTTS_POLY

    if DO_DATA_COLLECTION and ACTUAL_T % 900 == 0:
        LOGGER.debug("saving rerouted data at time {:f}".format(ACTUAL_T))
        with open('llr_rerouted_{:s}_{:.0f}.csv'.format(timestamp, ACTUAL_T), 'wb') as csvfile:
            r_writer = csv.writer(csvfile, delimiter=',')
            for row in rerouted_vehicles:
                r_writer.writerow(row)

    # NOTTS vehicle tracking
    if ACTUAL_T in NOTTS_TIMESTAMP_SET:
        veh_x, veh_y = NOTTS_DATA[ACTUAL_T]
        if TRACKED_NOTTS:
            add_marker_xy(veh_x, veh_y, "notts", redraw=True, marker_radius=NOTTS_RADIUS)
        else:
            # New vehicle entered the simulation
            LOGGER.info('Adding NOTTS track at timestamp {:f}'.format(ACTUAL_T))
            add_marker_xy(veh_x, veh_y, "notts", marker_color=NOTTS_COLOR, marker_radius=NOTTS_RADIUS, marker_layer=NOTTS_LAYER)
            TRACKED_NOTTS = True
    else:
        if TRACKED_NOTTS:
            LOGGER.info('Removing NOTTS vehicle track at timestamp {:f}'.format(ACTUAL_T))
            add_marker_xy(None, None, "notts", remove=True)
            if HAVE_NOTTS_POLY:
                LOGGER.info('Removing NOTTS route track at timestamp {:f}'.format(ACTUAL_T))
                draw_vehicle_route(NOTTS_POLY_ID, redraw=True)
                HAVE_NOTTS_POLY = False
        TRACKED_NOTTS = False
    
    # LLREX vehicle tracking
    if ACTUAL_T in LLREX_TIMESTAMP_SET:
        veh_x, veh_y = LLREX_DATA[ACTUAL_T]
        if TRACKED_LLREX:
            add_marker_xy(veh_x, veh_y, "llrex", redraw=True, marker_radius=LLREX_RADIUS)
        else:
            LOGGER.info('Adding LLREX vehicle track at timestamp {:f}'.format(ACTUAL_T))
            add_marker_xy(veh_x, veh_y, "llrex", marker_color=LLREX_COLOR, marker_radius=LLREX_RADIUS, marker_layer=LLREX_LAYER)
            TRACKED_LLREX = True
    else:
        if TRACKED_LLREX:
            LOGGER.info('Removing LLREX vehicle track at timestamp {:f}'.format(ACTUAL_T))
            add_marker_xy(None, None, "llrex", remove=True)
            if HAVE_LLREX_POLY:
                LOGGER.info('Removing LLREX route track at timestamp {:f}'.format(ACTUAL_T))
                draw_vehicle_route(LLREX_POLY_ID, redraw=True)
                HAVE_LLREX_POLY = False
        TRACKED_LLREX = False
    
    # LLREX route drawing
    if ACTUAL_T in ROUTE_TIMESTAMP_SET:
        action, veh_id, gain_dist, gain_time, edge_list = ROUTE_DATA[ACTUAL_T]
        if action == 'entered_simulation':
            LOGGER.info('New NOTTS/LLREX track for vehicle `{:s} at timestamp {:f}'.format(veh_id, ACTUAL_T))
            draw_vehicle_route(NOTTS_POLY_ID, route=edge_list, route_color=NOTTS_COLOR, route_layer=NOTTS_POLY_LAYER)
            HAVE_NOTTS_POLY = True
        elif action == 'rerouted':
            LOGGER.info('Rerouting LLREX track for vehicle `{:s} at timestamp {:f}'.format(veh_id, ACTUAL_T))
            LOGGER.info('New track has distance gain {:f} m and travel time gain {:f}'.format(gain_dist, gain_time))
            if HAVE_LLREX_POLY:
                draw_vehicle_route(LLREX_POLY_ID, route=edge_list, redraw=True)
            else:
                draw_vehicle_route(LLREX_POLY_ID, route=edge_list, route_color=LLREX_COLOR, route_layer=LLREX_POLY_LAYER)
                HAVE_LLREX_POLY = True
           
            
    
    if CENTRALIZED_SYS_BOOL:
        # If centralized system (scenario), use list of CAVs from decentralized scenario (only if needed)
        setCAVsBasedOnList()
    else:
        # Delete vehicle from the dictionary of time to reroute as it arrived as its destination
        polygonID = "track"
        departedCAVs = set(vehs_time2route.keys()).intersection(traci.simulation.getArrivedIDList())
        for vehID in departedCAVs:
            del vehs_time2route[vehID]
            if DO_DATA_COLLECTION:
                rerouted_vehicles.append([ACTUAL_T, vehID, "left_simulation", ACTUAL_T - entrance_time[vehID]])
                if vehID == TRACKED_VEHID:
                    # delete the route polygon
                    draw_vehicle_route(redraw=True)
                    # polygon has been removed automatically
                    # traci.polygon.remove(polygonID)
                    # dump the fcd data
                    with open('fcd_llr_{:s}_{:s}.csv'.format(timestamp, vehID), 'wb') as csvfile:
                        fcd_writer = csv.writer(csvfile, delimiter=',')
                        for row in fcd_data:
                            fcd_writer.writerow(row)
                    # forget the vehicle
                    TRACKED_VEHID = None
        # Reroute vehicle is time to reroute is zero, otherwise reduce step length
        for vehID in vehs_time2route:
            if DO_DATA_COLLECTION and vehID == TRACKED_VEHID:
                row = get_fcd_row(vehID)
                fcd_data.append(row)
            if vehs_time2route[vehID] == 0:
                # Reset routing period
                vehs_time2route[vehID] = ROUTING_PERIOD
                #
                if DO_DATA_COLLECTION:
                    old_route = traci.vehicle.getRoute(vehID)
                # Reroute the vehicle
                traci.vehicle.rerouteTraveltime(vehID, currentTravelTimes=False)  # default is currentTravelTimes=True
                #
                if DO_DATA_COLLECTION:
                    new_route = traci.vehicle.getRoute(vehID)
                    have_different_route = (old_route != new_route)
                    if have_different_route:
                        old_tt, old_len = get_route_params(old_route)
                        new_tt, new_len = get_route_params(new_route)
                        route_difference = set(old_route).symmetric_difference(set(new_route))
                        if len(route_difference) > 2:
                            # Substantial route change
                            if vehID == TRACKED_VEHID:
                                # Tracked vehicle was rerouted
                                draw_vehicle_route(route=new_route, redraw=True)
                                LOGGER.debug("highighted vehicle {:s} rerouted at time {:f}\n - from `{:s}`\n - to `{:s}`".format(vehID, ACTUAL_T, str(old_route), str(new_route)))
                            else:
                                # Untracked vehicle was rerouted
                                LOGGER.debug("vehicle {:s} rerouted".format(vehID))
                            str_difference = str(route_difference)
                            rerouted_vehicles.append([ACTUAL_T, vehID, "rerouted"] + list(new_route) + 
                                ["old_tt", old_tt, "new_tt", new_tt, "old_len", old_len, "new_len", new_len, "route_diff"] + list(route_difference))
                            LOGGER.debug("  difference of routes: " + str_difference)
                            LOGGER.debug("  difference of lengths: {:f}-{:f} = {:f}".format(new_len, old_len, new_len - old_len))
                            LOGGER.debug("  difference of traveltimes: {:f}-{:f} = {:f}".format(new_tt, old_tt, new_tt - old_tt))
                            
                if False: # 'probe' in vehID: 
                    if (TRACKED_VEHID is None) and have_different_route:
                        # rerouted vehicle
                        LOGGER.debug("vehicle {:s} rerouted at time {:f}\n - from `{:s}`\n - to `{:s}`".format(vehID, ACTUAL_T, str(old_route), str(new_route)))
                        # traci.vehicle.highlight(vehID, size=50)
                        add_marker(vehID, polygonID)
                        TRACKED_VEHID = vehID
                        draw_vehicle_route(route=new_route)
                    # else:
                    #    LOGGER.debug('attempt to reroute {:s} at time {:f} without route change'.format(vehID, ACTUAL_T))
            else:
                vehs_time2route[vehID] -= STEP_LENGTH

        # Add vehicle to the dictionary of time to reroute if equipped with rerouteing device
        for vehID in traci.simulation.getDepartedIDList():
            if traci.vehicle.getParameter(vehID, "has.rerouting.device") == "true":
                vehs_time2route[vehID] = 20
                if DO_DATA_COLLECTION:
                    route = traci.vehicle.getRoute(vehID) 
                    rerouted_vehicles.append([ACTUAL_T, vehID, "entered_simulation"] + list(route))
                    entrance_time[vehID] = ACTUAL_T
                # Do not reroute immediately at the beginning
                # traci.vehicle.rerouteTraveltime(vehID, currentTravelTimes=False)  # default is currentTravelTimes=True
            if DO_DATA_COLLECTION and vehID in NOTTS_VEHICLE_SET:
                if TRACKED_VEHID is None:
                    LOGGER.info("LLR vehicle {:s} highlighted at time {:f}".format(vehID, ACTUAL_T))
                    add_marker(vehID, polygonID)
                    TRACKED_VEHID = vehID
                    new_route = traci.vehicle.getRoute(vehID)
                    LOGGER.info("LLR vehicle {:s} inserted with route {:s}".format(vehID, str(new_route)))
                    draw_vehicle_route(route=new_route)
                    row = get_fcd_row(vehID)
                    fcd_data.append(row)
                else:
                    raise ValueError('got new vehicle while the old is still running')

    for edgeID in all_InOut_edges:
        # For all modelled lanes (incoming and outgoing lanes of junctions)
        tc_ind = all_InOut_edges[edgeID]
        laneIDs = edg_objs[tc_ind].edgeID_to_laneIDs[edgeID]
        lanes_vehs = dict()
        edge_vehs = set([])
        for laneID in laneIDs:
            # Get list of vehicles running on the modelled lanes
            lanes_vehs.update({laneID: traci.lane.getSubscriptionResults(laneID)[VEHIDS]})
            edge_vehs.update(lanes_vehs[laneID])

        if len(edge_vehs) > 0:
            try:
                departed_vehs = last_step_edge_vehs[edgeID].difference(edge_vehs)
                arrived_vehs = edge_vehs.difference(last_step_edge_vehs[edgeID])
                last_step_edge_vehs.update({edgeID: edge_vehs.copy()})
                if time2updte == 0:
                    beginTime = ACTUAL_T
                else:
                    beginTime = ACTUAL_T - (LEN_RANGE - time2updte)
                for _ in departed_vehs:
                    # For measuring accuracy of departures modelling
                    if beginTime >= BEGIN_T + WARM_UP_T + STEP_LENGTH and beginTime <= END_T + STEP_LENGTH:
                        try:
                            deps_accur_vals[edgeID]["measured"][beginTime] += 1
                        except KeyError:
                            deps_accur_vals[edgeID]["measured"].update({beginTime: 1})

                # For counting number o vehicles
                if NOGUI == False:
                    SUMO_lasttStep_area_updDeps.append(len(departed_vehs))

                # If edge is a generating edge of the whole network, the else of the for loop will happen
                for laneID in laneIDs:
                    try:
                        if edg_objs[tc_ind].lane_inflow_profile[laneID][0] != ():
                            jctID = edg_objs[tc_ind].inLaneID_to_jctID[laneID]
                            for neight_jct in jct_objs[tc_ind].commonFromNeigh_laneIDs[jctID]:
                                if laneID in jct_objs[tc_ind].commonFromNeigh_laneIDs[jctID][neight_jct]:
                                    raise
                        else:
                            raise
                    except:
                        break
                else:
                    # For counting number o vehicles
                    SUMO_lasttStep_area_updGen.append(len(arrived_vehs))

                # Add to the set of vehicles on the lane within update
                vehs_last_range[edgeID].update(edge_vehs)

            except KeyError:
                Not_InLaneID = 1

            for vehID in edge_vehs:
                # all_vehs_area.add(vehID) # only when travel time accuracy not estimated
                # Measure vehicle edge travel time for measuring accuracy of travel time prediction
                measureVehicleRealTravelTime(vehID, edgeID, time2updte)

                # For counting number of vehicles
                if traci.vehicle.getParameter(vehID, "has.rerouting.device") == "true":
                    # Add vehicles to the list of detected CAVs,
                    # in order to subscribe their data if they are new CAVs
                    tStep_area_CAVs.add(vehID)
                    # Add vehicle to the set of CAVs detected in the modelled area
                    cavs_area.add(vehID)
                else:
                    # Add vehicle to the set of nonCAVs detected in the modelled area
                    tStep_area_nonCAVs.add(vehID)
                    noncavs_area.add(vehID)

            # Algorithm to spread vehicles evenly to next edge (only artificial ones) based on queue length and/or number
            if PRIOR_ARTIFICIAL_EDGES.has_key(edgeID):
                defPriorArtEdgeVehiclesNextEdge(edgeID, edge_vehs, lanes_vehs)

    for edgeID in allConn_modelled_area:
        # For all connections between modelled edges
        try:
            edge_vehs = traci.edge.getSubscriptionResults(edgeID)[VEHIDS]
            for vehID in edge_vehs:
                # all_vehs_area.add(vehID) # only when travel time accuracy not estimated
                if traci.vehicle.getParameter(vehID, "has.rerouting.device") == "true":
                    # Add vehicles to the list of detected CAVs,
                    # in order to subscribe their data if they are new CAVs
                    tStep_area_CAVs.add(vehID)
                    # Add vehicle to the set of CAVs detected in the modelled area
                    cavs_area.add(vehID)
                else:
                    # Add vehicle to the set of nonCAVs detected in the modelled area
                    tStep_area_nonCAVs.add(vehID)
                    noncavs_area.add(vehID)
        except (AttributeError, TypeError):
            noVehicles = 1

    if INTERFACE == "traci" or (CENTRALIZED_SYS_BOOL == False and NOGUI == False):
        # If using Traci, it is needed to subscribe vehicles for data as Libsumo doesn't support subscriptions yet
        # or if using GUI it is easier to see CAVs by colouring them
        # Check new CAVs of the time step
        try:
            timestep_new_CAVs = tStep_area_CAVs.difference(lastStep_area_CAVs)
        except NameError:
            lastStep_area_CAVs = set([])
            timestep_new_CAVs = tStep_area_CAVs

        if CENTRALIZED_SYS_BOOL == False and NOGUI == False:
            # Colour new CAVs
            for vehID in timestep_new_CAVs:
                if "probe" not in vehID:
                    # Change colour of CAV to blue (default is white)
                    traci.vehicle.setColor(vehID, (0, 0, 255))

        if INTERFACE == "traci":
            for vehID in timestep_new_CAVs:
                # Subscribe vehicle to get data from new CAV
                traci.vehicle.subscribe(vehID, VEH_INFO)

            # Define CAVs that left the modelled area
            timestep_out_CAVs = lastStep_area_CAVs.difference(tStep_area_CAVs)
            timestep_out_CAVs.difference_update(set(traci.simulation.getArrivedIDList()))
            for vehID in timestep_out_CAVs:
                # Unsubscribe vehicle to stop getting data from CAV that left modelled area
                traci.vehicle.unsubscribe(vehID)

        # Store the CAVs of current time step
        lastStep_area_CAVs = tStep_area_CAVs.copy()


def getCAVdataViaLibsumo(TC_seq):
    """Get data from CAV vehicles using Libsumo functions"""

    all_CAVs_data = dict()
    for vehID in tStep_area_CAVs:
        # Get data directly for all vehicles within modelled area
        all_CAVs_data.update({vehID: dict()})
        all_CAVs_data[vehID].update({TYPE: traci.vehicle.getTypeID(vehID)})
        all_CAVs_data[vehID].update({LANE: traci.vehicle.getLaneID(vehID)})
        all_CAVs_data[vehID].update({ROUTE: traci.vehicle.getRoute(vehID)})
        all_CAVs_data[vehID].update({POSITION: traci.vehicle.getLanePosition(vehID)})
        all_CAVs_data[vehID].update({SPEED: traci.vehicle.getSpeed(vehID)})
        all_CAVs_data[vehID].update({ACC: traci.vehicle.getAccel(vehID)})
        all_CAVs_data[vehID].update({DEC: traci.vehicle.getDecel(vehID)})
        all_CAVs_data[vehID].update({MINGAPT: traci.vehicle.getTau(vehID)})
        all_CAVs_data[vehID].update({SPEEDFACTOR: traci.vehicle.getSpeedFactor(vehID)})
        all_CAVs_data[vehID].update({LENGTH: traci.vehicle.getLength(vehID)})
        all_CAVs_data[vehID].update({MINGAPD: traci.vehicle.getMinGap(vehID)})

    for tc_ind in TC_seq:
        for RSU_id in jct_objs[tc_ind].ids:
            # Initialize RSU_CAVs_data, define that all junctions have RSU, as libsumo doesn't support context
            # subscriptions yet
            jct_objs[tc_ind].RSU_CAVs_data.update({RSU_id: dict()})

    for vehID in tStep_area_CAVs:
        in_laneID = all_CAVs_data[vehID][LANE]
        for tc_ind in TC_seq:
            try:
                # As context subscriptions are not working with libsumo define vehicle's detected RSU as the junction
                # of the incoming lane vehicle is running on and not the real RSU detecting it
                jctID = edg_objs[tc_ind].inLaneID_to_jctID[in_laneID]
                # If found jctID modelled by traffic controller tc_ind
                jct_objs[tc_ind].RSU_CAVs_data[jctID].update({vehID: all_CAVs_data[vehID]})
                veh_speed = jct_objs[tc_ind].RSU_CAVs_data[jctID][vehID][SPEED]
                if veh_speed == 0:
                    # SUMO has a bug that when it inserts vehicles for the first time it generates them
                    # with speed 0, change the speed to lane max speed.
                    in_laneID = jct_objs[tc_ind].RSU_CAVs_data[jctID][vehID][LANE]
                    vehID_dist = round(edg_objs[tc_ind].lane_length[in_laneID] - \
                                       jct_objs[tc_ind].RSU_CAVs_data[jctID][vehID][POSITION], 1)
                    vehID_len = jct_objs[tc_ind].RSU_CAVs_data[jctID][vehID][LENGTH]
                    if vehID_dist >= edg_objs[tc_ind].lane_length[in_laneID] - vehID_len - 0.2:
                        jct_objs[tc_ind].RSU_CAVs_data[jctID][vehID][SPEED] = \
                            edg_objs[tc_ind].lane_max_speed[in_laneID] \
                            * jct_objs[tc_ind].RSU_CAVs_data[jctID][vehID][SPEEDFACTOR]
                break
            except KeyError:
                # If vehicle's jctID is not modelled by the traffic controller tc_ind, try next tc_ind
                skip = 1
        else:
            # If in_laneID is not incoming or outgoing lane, check which junction vehicle based on its laneID name
            # In SUMO connection IDs contain the jctID within its name
            for tc_ind in TC_seq:
                found_jct = 0
                for jctID in jct_objs[tc_ind].ids:
                    # Old way to check if connection is within jctID
                    # ("_").join(in_laneID.split("_")[:-2])[1:] == jctID
                    # New way to check if connection is within jctID
                    if jctID in in_laneID:
                        jct_objs[tc_ind].RSU_CAVs_data[jctID].update({vehID: all_CAVs_data[vehID]})
                        found_jct = 1
                        break

                if found_jct == 1:
                    break
    # Delete dictionary with all CAVs data as it was assigned already to suitable RSU_CAVs_data
    del all_CAVs_data


def getCAVdataViaTraci(TC_seq):
    """Get data from CAV vehicles using Traci functions"""

    all_CAVs_data = dict()
    for vehID in tStep_area_CAVs:
        # Get data of all subscribed vehicles
        all_CAVs_data.update({vehID: traci.vehicle.getSubscriptionResults(vehID)})

    for tc_ind in TC_seq:
        RSU_ids = [jctID for jctID in jct_objs[tc_ind].ids if "Coop." in jct_objs[tc_ind].jctType[jctID]]
        for RSU_id in RSU_ids:
            # Initialize RSU_CAVs_data only for cooperative junctions
            jct_objs[tc_ind].RSU_CAVs_data.update({RSU_id: dict()})
            # Get vehicles that fulfil the context subscription of RSU_id
            RSU_vehs = traci.junction.getContextSubscriptionResults(RSU_id)
            try:
                for vehID in RSU_vehs:
                    # From the subscriptions get only data from CAVs fulfilling the context subscription
                    try:
                        jct_objs[tc_ind].RSU_CAVs_data[RSU_id].update({vehID: all_CAVs_data[vehID]})
                        veh_speed = jct_objs[tc_ind].RSU_CAVs_data[RSU_id][vehID][SPEED]
                        if veh_speed == 0:
                            # SUMO has a bug that when it inserts vehicles for the first time it generates them
                            # with speed 0, change the speed to lane max speed.
                            in_laneID = jct_objs[tc_ind].RSU_CAVs_data[RSU_id][vehID][LANE]
                            vehID_dist = round(edg_objs[tc_ind].lane_length[in_laneID] - \
                                               jct_objs[tc_ind].RSU_CAVs_data[RSU_id][vehID][POSITION], 1)
                            vehID_len = jct_objs[tc_ind].RSU_CAVs_data[RSU_id][vehID][LENGTH]
                            if vehID_dist >= edg_objs[tc_ind].lane_length[in_laneID] - vehID_len - 0.2:
                                jct_objs[tc_ind].RSU_CAVs_data[RSU_id][vehID][SPEED] = \
                                    edg_objs[tc_ind].lane_max_speed[in_laneID] \
                                    * jct_objs[tc_ind].RSU_CAVs_data[RSU_id][vehID][SPEEDFACTOR]

                    except KeyError:
                        # When vehicle within RSU comm. range but not running on modelled edges by the junction's TC
                        # or vehicle is a nonCAV
                        skip = 1
            except (AttributeError, TypeError):
                noVehicles = 1

    # Delete dictionary with all CAVs data as it was assigned already to suitable RSU_CAVs_data
    del all_CAVs_data


def setCAVsBasedOnList():
    """Set in SUMO if a vehicle is CAV or not based on a list of CAV and nonCAV vehicles"""

    # Get list of vehicles from SUMO
    net_departed_vehs = set(traci.simulation.getDepartedIDList())
    net_new_CAVs = rem_cavs_list.intersection(net_departed_vehs)
    net_new_nonCAVs = rem_noncavs_list.intersection(net_departed_vehs)

    # Vehicles not simulated in the previous scenario
    net_to_define = net_departed_vehs.difference(net_new_CAVs.union(net_new_nonCAVs))
    for veh2define in net_to_define:
        # If vehicle was not simulated in the previous scenario, define it CAV or not based on penetration rate
        if np.random.binomial(1, ROUTING_PROB):
            # Define as CAV
            net_new_CAVs.add(veh2define)
        else:
            # Define as nonCAV
            net_new_nonCAVs.add(veh2define)

    # Delete vehicle from the dictionary of time to reroute as it arrived as its destination
    departedCAVs = set(vehs_time2route.keys()).intersection(traci.simulation.getArrivedIDList())
    for vehID in departedCAVs:
        del vehs_time2route[vehID]

    # Reroute vehicle is time to reroute is zero, otherwise reduce step length
    for vehID in vehs_time2route:
        if vehs_time2route[vehID] == 0:
            vehs_time2route.update({vehID: ROUTING_PERIOD})
            traci.vehicle.rerouteTraveltime(vehID,
                                            currentTravelTimes=False)  # default is currentTravelTimes=True
        else:
            vehs_time2route[vehID] -= STEP_LENGTH

    for cavID in net_new_CAVs:
        try:
            # Update list of remaining CAVs to be simulated
            rem_cavs_list.remove(cavID)
        except KeyError:
            skip = 1

        # Set vehicle as CAV
        traci.vehicle.setParameter(cavID, "has.rerouting.device", "true")
        if NOGUI == False and "probe" not in cavID:
            # Change colour of CAV to blue (default is white)
            traci.vehicle.setColor(cavID, (0, 0, 255))

        # Add vehicle to the dictionary of time to reroute if equipped with rerouteing device
        vehs_time2route.update({cavID: ROUTING_PERIOD})
        traci.vehicle.rerouteTraveltime(cavID, currentTravelTimes=False)  # default is currentTravelTimes=True

    # Update list of remaining nonCAVs to be simulated
    for nonCavID in net_new_nonCAVs:
        try:
            rem_noncavs_list.remove(nonCavID)
        except KeyError:
            skip = 1


def measureVehicleRealTravelTime(vehID, edgeID, time2updte):
    """Store which edge vehicle is at each time step and measure travel time once changes edge.
    Only for simulation purposes"""

    # This configuration makes that vehicles detected on a internal lanes will not
    # be acounted as left the edge, only when arrived at the next one.
    try:
        if allVehs_Edge_ArrTime[vehID][0] != edgeID:
            # If vehicle is in another edge than the last one detected
            time_after_updte = LEN_RANGE - allVehs_Edge_ArrTime[vehID][2]
            # range is between [ACTUAL_T, ACTUAL_T + LEN_RANGE],
            # when time_after_updte = 0, means a new range
            # and time_after_updte is always lower and equal than LEN_RANGE
            # used to define which range the travel time is suitable
            if time_after_updte == LEN_RANGE:
                time_after_updte = 0

            beginTime = allVehs_Edge_ArrTime[vehID][1] - time_after_updte
            if beginTime >= BEGIN_T + WARM_UP_T + STEP_LENGTH and beginTime <= END_T + STEP_LENGTH:
                try:
                    # Store measured travel time for the range
                    tts_accur_vals[allVehs_Edge_ArrTime[vehID][0]]["measured"][beginTime].append(
                        ACTUAL_T - allVehs_Edge_ArrTime[vehID][1])
                except KeyError:
                    # If not found values for the range, create a key for it and store measured
                    # travel time
                    tts_accur_vals[allVehs_Edge_ArrTime[vehID][0]]["measured"].update(
                        {beginTime: [ACTUAL_T - allVehs_Edge_ArrTime[vehID][1]]})

            # Update new edge to monitor
            allVehs_Edge_ArrTime.update({vehID: (edgeID, ACTUAL_T, time2updte)})

    except KeyError:
        # If vehicle not detected before, start monitoring vehicle's edge
        # time2updte is necessary to define which range the measurement is suitable
        # as this function runs at every simulation time step and not every algorithm update
        allVehs_Edge_ArrTime.update({vehID: (edgeID, ACTUAL_T, time2updte)})


def measureCAVRealTravelTime(edg_obj, veh_obj, vehID, edgeID, in_laneID):
    """Store which edge a CAV is at each time step and measure travel time once changes edge"""

    try:
        if veh_obj.cavs_metrics_samples["travel_times"][vehID][0] != edgeID:
            # If vehicle is in another edge than the last one detected
            former_inLaneID = veh_obj.cavs_metrics_samples["travel_times"][vehID][2]
            if len(veh_obj.cavs_metrics_samples["travel_times"][vehID][3].keys()) > 1:
                # Not able to store movgID due too short lane length or movgID of a different incoming lane
                possible_movgIDs = set([])
                parallel_lanes = [former_inLaneID] \
                                 + [parallel_lane for parallel_lane
                                    in edg_obj.edgeID_to_laneIDs[edg_obj.laneID_to_edgeID[former_inLaneID]]
                                    if parallel_lane != former_inLaneID]

                for parallel_laneID in parallel_lanes:
                    for conn in edg_obj.inLaneID_to_connIDs[parallel_laneID]:
                        if edg_obj.connID_to_outLaneID[conn] in edg_obj.edgeID_to_laneIDs[edgeID]:
                            possible_movgIDs.add(edg_obj.connID_to_movgID[conn])
                    if len(possible_movgIDs) > 0:
                        break

                former_inLaneID = parallel_laneID
                try:
                    movgID = np.random.choice(list(possible_movgIDs))
                except ValueError:
                    # If not found possibly movgID (when CAV detected beyond next possible lane
                    # i.e. at least two too short lanes and CAV detected on internal lanes - within the junction)
                    # use a random movgID
                    movgID = np.random.choice(list(edg_obj.inLaneID_to_movgIDs[former_inLaneID]))
            else:
                movgID = veh_obj.cavs_metrics_samples["travel_times"][vehID][3].keys()[0]

            # Store measured travel time
            CAVarrTime = veh_obj.cavs_metrics_samples["travel_times"][vehID][1]
            prevTravelTime = veh_obj.cavs_metrics_samples["travel_times"][vehID][3][movgID]
            CAVtravelTime = ACTUAL_T - veh_obj.cavs_metrics_samples["travel_times"][vehID][1]
            edg_obj.lane_cavs_metrics[former_inLaneID]["travel_times"][movgID] = (CAVarrTime,
                                                                                  prevTravelTime,
                                                                                  CAVtravelTime)
            # Update new edge to monitor
            prevTravelTime = edg_obj.tc_edge_movgs_tt[edgeID]
            veh_obj.cavs_metrics_samples["travel_times"].update({vehID: [edgeID, ACTUAL_T, in_laneID, prevTravelTime]})

    except KeyError:
        # If vehicle not detected before, start monitoring vehicle's edge
        prevTravelTime = edg_obj.tc_edge_movgs_tt[edgeID]
        veh_obj.cavs_metrics_samples["travel_times"].update({vehID: [edgeID, ACTUAL_T, in_laneID, prevTravelTime]})


def MonitorQueueLenUsingCAVs(jct_obj, edg_obj):
    """Monitor the furthest CAV to stop line once it queues to define queue length"""

    for jctID in jct_obj.ids:
        if "Mod. Non-C." in jct_obj.jctType[jctID]:
            # If non-connected junction
            for in_laneID in jct_obj.in_laneIDs[jctID]:
                # Get vehicles running on the lane
                if INTERFACE in ("libsumo", "traci"):
                    lane_CAVS = traci.lane.getSubscriptionResults(in_laneID)[VEHIDS]
                else:
                    LOGGER.debug("Get the ID of all CAVs on lane via other interface")

                try:
                    for vehID in lane_CAVS:
                        # Assuming lane_CAVS ordered by the furthest from the stop line
                        try:
                            # Get CAV speed
                            # (in case o SUMO it also gets nonCAV IDs but the next line does not find RSU_id, so it just
                            # skip the vehicle)
                            RSU_id = jct_obj.cavs_in_range[jctID][vehID][0]
                            veh_speed = jct_obj.RSU_CAVs_data[RSU_id][vehID][SPEED]
                            if veh_speed == 0:
                                # If last CAV stopped, define the queue length in meters
                                # (assuming the position has the begin of the lane, i.e. furthest from stop line)
                                vehID_dist = round(edg_obj.lane_length[in_laneID] - \
                                                   jct_obj.RSU_CAVs_data[RSU_id][vehID][POSITION], 1)
                                vehID_len = jct_obj.RSU_CAVs_data[RSU_id][vehID][LENGTH]
                                # Define queue length with stopped vehicle
                                edg_obj.lane_cavs_metrics[in_laneID]["queue_len"]["by_CAV"] = int(np.ceil(vehID_dist
                                                                                                          + vehID_len))
                                queue_pred_ref_t = edg_obj.lane_queue_pred[in_laneID][0]
                                for pred_range in range(0, len(edg_obj.lane_queue_pred[in_laneID][1])):
                                    begTime = queue_pred_ref_t + pred_range * LEN_RANGE
                                    if begTime >= ACTUAL_T:
                                        edg_obj.lane_cavs_metrics[in_laneID]["queue_len"]["by_LLR_pred"] = \
                                            edg_obj.lane_queue_pred[in_laneID][1][pred_range]
                                        break
                                break
                        except KeyError:
                            continue
                except TypeError:
                    noVehicles = 1


def defPriorArtEdgeVehiclesNextEdge(edgeID, edge_vehs, lanes_vehs):
    """Define the next artificial edge of vehicles on prior artificial edges"""

    global lowest_art_lanes_measure, lowest_prior_lanes_measure

    num_lanes = len(priorAndArtificial_edgesLanes[edgeID])
    CAVs_route = dict()
    # Select artificial edges only that are not closed
    aval_art_edges = [art_edge for art_edge in PRIOR_ARTIFICIAL_EDGES[edgeID]
                      if art_edge not in EDGE2CLOSE]
    # Update vehicles that already had defined their next edges (artificial one)
    edge_endLane_vehs[edgeID].intersection_update(edge_vehs)
    edge_defined_vehs[edgeID].intersection_update(edge_vehs)

    # For the closest vehicle to stop line on each lane, check if going to next edge on the next time step.
    closest_stopline_cavs = set([])
    for laneID in lanes_vehs:
        try:
            # Assuming that only 1 vehicle per lane can decide next edge per time step
            # Assuming edge_vehs_list are ordered by their distance to stop line
            last_vehID = lanes_vehs[laneID][-1]
            if traci.vehicle.getParameter(last_vehID, "has.rerouting.device") == "true":
                try:
                    # Get vehicle next edges
                    veh_route_index = traci.vehicle.getRouteIndex(last_vehID)
                    veh_route = list(traci.vehicle.getRoute(last_vehID)[veh_route_index:])
                    if veh_route[2] == PRIOR_COMMON_AFTER_ART_EDGES[edgeID]:
                        veh_speed = traci.vehicle.getSpeed(last_vehID)
                        veh_displacement = traci.vehicle.getLanePosition(last_vehID)
                        # Define distance to stop line after next time step
                        dist_nextSteps = veh_displacement + veh_speed
                        if dist_nextSteps >= prior_edges_len[edgeID]:
                            closest_stopline_cavs.add(last_vehID)
                            CAVs_route.update({last_vehID: veh_route[:]})
                    else:
                        edge_defined_vehs[edgeID].add(last_vehID)
                except IndexError:
                    # Vehicle finishing trip
                    skip = 1
            else:
                edge_defined_vehs[edgeID].add(last_vehID)
        except IndexError:
            NoVehicle = 1
    if len(closest_stopline_cavs) > 0:
        # If so, get artificial edge metric and define next edge
        # Among all lanes of artificial edges, choose edge with following priority:
        # 1) shortest queue
        # 2) lowest number of vehicles
        sum_lane = 0
        for lane_metric in LANE_METRICS:
            lowest_art_lanes_measure = (None, float("inf"))
            for art_edgeID in PRIOR_ARTIFICIAL_EDGES[edgeID]:
                if art_edgeID not in EDGE2CLOSE or (ACTUAL_T < BEGIN_LANE_CLOSURE or ACTUAL_T > END_LANE_CLOSURE):
                    # If artificial edge is not closed
                    for lane_num, laneID in enumerate(priorAndArtificial_edgesLanes[art_edgeID]):
                        # For each lane of artificial edge,
                        # choose lane with either shortest queue or number of vehicles
                        # and take its respective artificial edge
                        lane_measure = traci.lane.getSubscriptionResults(laneID)[lane_metric]
                        sum_lane += lane_measure
                        if lane_measure < lowest_art_lanes_measure[1]:
                            # Update lowest metric
                            lowest_art_lanes_measure = (art_edgeID, lane_measure)
                        elif lane_num == (num_lanes - 1) and lane_measure == lowest_art_lanes_measure[1]:
                            # All queue lanes are the same, then maintain same prior lane or change
                            lowest_art_lanes_measure = None
            if sum_lane > 0:
                # If sum of lane metric is zero, use next measure
                break

    # For other vehicles, define current lane only for those not defined before
    edge_vehs.difference_update(edge_defined_vehs[edgeID])
    edge_vehs_to_monitor = set([])
    # but if lanes of current edge are too much different, make for all vehicles aiming to balance the load
    if num_lanes > 1:
        # If only more than 1 possible next artificial edge choose lane with following priority:
        # 1) shortest queue
        # 2) lowest number of vehicles
        sum_lane = 0
        for lane_metric in LANE_METRICS:
            lowest_prior_lanes_measure = (None, float("inf"))
            for lane_num, laneID in enumerate(priorAndArtificial_edgesLanes[edgeID]):
                # For each lane of prior edge, choose lane with lowest metric and its respective artificial edge
                try:
                    art_edgeID = PRIOR_ARTIFICIAL_EDGES[edgeID][lane_num]
                except IndexError:
                    art_edgeID = aval_art_edges[-1]
                if PRIOR_ARTIFICIAL_EDGES[edgeID][lane_num] not in EDGE2CLOSE \
                or (ACTUAL_T < BEGIN_LANE_CLOSURE or ACTUAL_T > END_LANE_CLOSURE):
                    lane_measure = traci.lane.getSubscriptionResults(laneID)[lane_metric]
                    if lane_metric == LANE_QUEUE_NUM:
                        # Due SUMO bug of new vehicle being inserted with speed zero, decrease it from the lane
                        # queue number
                        try:
                            first_vehID = lanes_vehs[laneID][0]
                            if traci.vehicle.getSpeed(first_vehID) <= 0.1:
                                lane_measure -= 1
                        except IndexError:
                            NoVehs = 1
                    sum_lane += lane_measure
                    if lane_measure < lowest_prior_lanes_measure[1]:
                        if lane_metric == LANE_QUEUE_NUM \
                        and 5 < lowest_prior_lanes_measure[1] < float("inf") \
                        and lane_measure < 0.6 * lowest_prior_lanes_measure[1]:
                            # If shortest lane is half of the longest queue and bigger than 5 vehicles,
                            # reset all vehicles to distribute better
                            edge_defined_vehs[edgeID] = set([])
                        # Update lowest queue
                        lowest_prior_lanes_measure = (art_edgeID, lane_measure)
                    elif lane_num == (num_lanes - 1) and lane_measure == lowest_prior_lanes_measure[1]:
                        # All queue lanes are the same, then maintain same prior lane
                        lowest_prior_lanes_measure = None

            if sum_lane > 0:
                # If sum of lane metric is zero, use next measure
                break
    else:
        # If only 1 possible next artificial edge (only one lane)
        lowest_prior_lanes_measure = None

    for vehID in edge_vehs:
        if lowest_prior_lanes_measure != None \
        and traci.vehicle.getParameter(vehID, "has.rerouting.device") == "true":
            try:
                # Get vehicle next edges
                veh_displacement = traci.vehicle.getLanePosition(vehID)
                veh_len = traci.vehicle.getLength(vehID)
                if veh_displacement >= veh_len + 0.2:
                    # Due SUMO bug that inserts vehicle with speed zero, skip vehicles being inserted right now,
                    # consider vehicle only on next time step
                    veh_route_index = traci.vehicle.getRouteIndex(vehID)
                    veh_route = list(traci.vehicle.getRoute(vehID)[veh_route_index:])
                    if veh_route[2] == PRIOR_COMMON_AFTER_ART_EDGES[edgeID]:
                        edge_vehs_to_monitor.add(vehID)
                        CAVs_route.update({vehID: veh_route[:]})
                    else:
                        edge_defined_vehs[edgeID].add(vehID)
            except IndexError:
                # Vehicle finishing trip
                skip = 1
        else:
            edge_defined_vehs[edgeID].add(vehID)

    # For vehicles that hasn't defined yet current lanes or next edges running on the prior artificial edges
    decide_current_lane = set([])
    decide_next_lane = set([])
    maintain_next_lane = set([])
    for vehID in closest_stopline_cavs:
        if vehID not in edge_endLane_vehs[edgeID]:
            # Set vehicle as already defined next artificial edge
            edge_endLane_vehs[edgeID].add(vehID)
            if lowest_art_lanes_measure != None:
                # Decide next lane (respective artificial edge),
                # based on the shortest queue or lowest number of vehicles of next lane
                decide_next_lane.add(vehID)
            else:
                # Maintain same current lane, to define respective artificial edge
                # of current lane of prior edge
                maintain_next_lane.add(vehID)
        else:
            # Maintain same current lane, to define respective artificial edge of current lane of prior edge
            maintain_next_lane.add(vehID)

    for vehID in edge_vehs_to_monitor:
        edge_defined_vehs[edgeID].add(vehID)
        veh_speed = traci.vehicle.getSpeed(vehID)
        if veh_speed > 0.1 and lowest_prior_lanes_measure != None:
            # Decide current lane, to define respective artificial edge
            # of the shortest queue of current prior edge
            decide_current_lane.add(vehID)
        # elif veh_speed == 0:
        #     # Maintain same current lane, to define respective artificial edge
        #     # of current lane of prior edge
        #     maintain_next_lane.add(vehID)

    # Input into SUMO suitable next artificial edge
    for vehID in maintain_next_lane:
        veh_route = CAVs_route[vehID][:]
        lane_num = traci.vehicle.getLaneIndex(vehID)
        if PRIOR_ARTIFICIAL_EDGES[edgeID][lane_num] not in EDGE2CLOSE \
        or (ACTUAL_T < BEGIN_LANE_CLOSURE or ACTUAL_T > END_LANE_CLOSURE):
            # The next artificial edge is based on the current lane of the prior edge
            veh_route[1] = PRIOR_ARTIFICIAL_EDGES[edgeID][lane_num]
        else:
            # If more prior lanes than artificial edges
            veh_route[1] = aval_art_edges[-1]
        traci.vehicle.setRoute(vehID, veh_route)
    for vehID in decide_current_lane:
        veh_route = CAVs_route[vehID][:]
        if veh_route[1] != lowest_prior_lanes_measure[0]:
            veh_route = CAVs_route[vehID][:]
            # The artificial edge is based on the shortest queue of current lane of the prior edge
            veh_route[1] = lowest_prior_lanes_measure[0]
            traci.vehicle.setRoute(vehID, veh_route)
    for vehID in decide_next_lane:
        veh_route = CAVs_route[vehID][:]
        if lowest_art_lanes_measure != None and veh_route[1] != lowest_art_lanes_measure[0]:
            # The next artificial edge is based on the lowest metric of all artificial edges
            veh_route[1] = lowest_art_lanes_measure[0]
            traci.vehicle.setRoute(vehID, veh_route)


def getInLaneQueueFromSUMO(jct_obj, # TC junctions object
                           edg_obj, # TC edges object
                           veh_obj, # TC vehicle object
                           jctID, # network junction ID
                           in_laneID): # network lane ID
    """Get the number of vehicles queuing on the incoming lane and add an error of the measurement.
    This function requires SUMO and it is for simulation purposes only."""

    start8 = time.time()

    edgeID = edg_obj.laneID_to_edgeID[in_laneID]
    lane_subscription = traci.lane.getSubscriptionResults(in_laneID)
    # Get queue length (in number of vehicles)
    real_queue_num = lane_subscription[LANE_QUEUE_NUM]  # SUMO

    if real_queue_num > 0:
        # Add expected error for the queue length estimation, but only if lane is not full
        # This is to avoid adding errors in different sections of same real lane,
        # in other words, it just adds errors to the tail of the real lane queue
        real_queue_occupancy = lane_subscription[LANE_OCCUPANCY]  # SUMO
        if real_queue_occupancy < 0.85:
            # The error is based on the flow of vehicles and the position of the last CAV in the queue and proportion of CAVs
            lane_vehs = lane_subscription[VEHIDS]
            num_CAVs_queueing = 0
            num_nonCAVs_queueing = 0
            distance_last_veh_queue = 0
            found_dist = 0
            for veh_ind,vehID in enumerate(lane_vehs):
                # Assuming lane_vehs are ordered by their distance from the beginning of the lane
                # (last vehicle on the queue)
                try:
                    RSU_id = jct_obj.cavs_in_range[jctID][vehID][0]
                    if traci.vehicle.getSpeed(vehID) <= veh_obj.thresh_stopped:
                        # SUMO has a bug that when it inserts vehicles for the first time it generates them
                        # with speed 0, decrease such vehicles from the queue count of SUMO
                        vehID_dist = round(edg_obj.lane_length[in_laneID] - \
                                           jct_obj.RSU_CAVs_data[RSU_id][vehID][POSITION], 1)
                        vehID_len = jct_obj.RSU_CAVs_data[RSU_id][vehID][LENGTH]
                        if vehID_dist < edg_obj.lane_length[in_laneID] - vehID_len - 0.2:
                            num_CAVs_queueing += 1
                            found_dist = 1
                        else:
                            real_queue_num -= 1
                except KeyError:
                    # Not a CAV
                    if traci.vehicle.getSpeed(vehID) <= veh_obj.thresh_stopped:
                        # SUMO has a bug that when it inserts vehicles for the first time it generates them
                        # with speed 0, decrease such vehicles from the queue count of SUMO
                        vehID_dist = round(edg_obj.lane_length[in_laneID] - traci.vehicle.getLanePosition(vehID), 1)
                        vehID_len = traci.vehicle.getLength(vehID)
                        if vehID_dist < edg_obj.lane_length[in_laneID] - vehID_len - 0.2:
                            num_nonCAVs_queueing += 1
                            if found_dist == 0:
                                distance_last_veh_queue += 1
                        else:
                            real_queue_num -= 1

            if real_queue_num > 0:
                # Define the queue error factor based on the location of the CAV within the queue
                # (if CAV at end, it is known the ground truth)
                queue_error_factor = distance_last_veh_queue / float(real_queue_num)
                if len(vehs_last_range[edgeID]) / LEN_RANGE < LOW_FLOW:
                    error_perc = np.random.normal(0, QUEUE_EST_LOW_DEV) * queue_error_factor
                elif len(vehs_last_range[edgeID]) / LEN_RANGE < MID_FLOW:
                    error_perc = np.random.normal(0, QUEUE_EST_MID_DEV) * queue_error_factor
                else:
                    error_perc = np.random.normal(0, QUEUE_EST_HIGH_DEV) * queue_error_factor
                # Define the percentage of queueing CAVs of the real queue
                try:
                    perc_CAVs_queueing = num_CAVs_queueing / float(num_CAVs_queueing + num_nonCAVs_queueing)
                except ZeroDivisionError:
                    perc_CAVs_queueing = 0
            else:
                error_perc = 0
                perc_CAVs_queueing = 0
        else:
            error_perc = 0
            perc_CAVs_queueing = 0

        # Change the number of queueing vehicles based on queue length estimation error
        queue_num = int(round(real_queue_num + real_queue_num * (error_perc * (1 - perc_CAVs_queueing)), 0))
    else:
        queue_num = 0

    # For Counting Number of All Vehicles
    if NOGUI == False:
        tStep_area_queues.append(real_queue_num)
        tStep_area_queuesError.append(abs(queue_num - real_queue_num))

    end8 = time.time()
    timings[8][1] += end8 - start8

    return queue_num


def getTCorder(objs, # Lits of all TC objects
               tc_it, # algorithm order interaction
               time2updte, # time to next algorithm update
               jct_order): # Actual order of junctions to be explored
    """Define the order of Traffic Controllers to be explored"""

    TC_seq = range(0, len(objs))
    if time2updte <= 0:
        # If it is the first iteraction, then repeat order of last algorithm update
        # If last iteraction, then invert order of current algorithm update
        if (TC_IT_MAX == 2 and ((jct_order == -1 and tc_it == 2) or (jct_order == 1 and tc_it == 1))) \
        or (TC_IT_MAX == 1 and (jct_order == -1)):
            # Make forward order
            jct_order = 1
        else:
            # Make reverse order
            TC_seq.reverse()
            jct_order = -1

    return TC_seq, jct_order


def BroadcastAreaTravelTime(TC_seq): # TC sequence
    """Prepare the information to be sent to vehicles around RSUs.
     When using SUMO it simulations such CAVs getting the traffic network edge travel times."""

    global beginTime, endTime

    start4 = time.time()

    if BEGIN_LANE_CLOSURE > BEGIN_T + STEP_LENGTH \
    and ACTUAL_T >= BEGIN_LANE_CLOSURE \
    and ACTUAL_T <= BEGIN_LANE_CLOSURE + ROUTING_PERIOD:
        # Reroute CAVs if lane closure happened after beginning of simulation and if actual time is within the
        # routing period, so vehicles don't wait for reroute due routing period to react to lane closures.
        reroute_due_lane_closure = 1
    else:
        reroute_due_lane_closure = 0
    broadcastTTs = 0
    if INTERFACE in ("libsumo", "traci"):
        if TTS_PER_VEH or reroute_due_lane_closure:
            # If setting travel time per each vehicle
            cavs2broadcast = dict()
            for tc_ind in TC_seq:
                for jctID in jct_objs[tc_ind].ids:
                    # Broadcast only to CAVs that were detected on an incoming lane modelled by the TC
                    # and also didn't receive the travel time information since the last update
                    cavs2broadcast.update({tc_ind: {cav for cav in jct_objs[tc_ind].cavs_in_range[jctID].keys()
                                                    if cav not in CAVsAlreadyBroadcastedTo}})
            if cavs2broadcast != set([]):
                # Prepare broadcast information only if there is vehicles that need to receive TTs
                broadcastTTs = 1
        else:
            # If setting travel time globally per edge
            broadcastTTs = 1
    else:
        LOGGER.debug("Use here another condition for defining when transmitting TTs to CAVs")

    if broadcastTTs == 1:
        tc_ind = 0 # all traffic controller have same information when simulation, when not simulation, there is just
                   # one traffic controller which is tc_ind = 0
        for edgeID in edg_objs[tc_ind].traffic_net_tt:
            # The ACTUAL_T when the travel times were collected
            begin_arr_range = edg_objs[tc_ind].traffic_net_tt[edgeID][0]
            # Define the time intervals fixed time length LEN_RANGE that are valid each travel time
            ranges = range(0, len(edg_objs[tc_ind].traffic_net_tt[edgeID][1]))
            ranges_tt = []
            # Provision of the travel times the junction calculated for its own incoming edges and
            # also received from other TCs in the same traffic network through LDM messages
            beginTime = begin_arr_range
            last_end_t = begin_arr_range
            provided_period = max(len(edg_objs[tc_ind].traffic_net_tt[edgeID][1]) * MAX_T_MAX_CYCLE,
                                  np.ceil(ROUTING_PERIOD/LEN_RANGE))
            while beginTime <= begin_arr_range + (provided_period * LEN_RANGE):
                # also beginTime == begin_arr_range + (provided_period * LEN_RANGE) to ensure that
                # a vehicle receiving travel times between algorithm updates will receive  a full provided_period
                # update for its interval between reroutes
                for arr_range in ranges:
                    beginTime = last_end_t + arr_range * LEN_RANGE
                    if beginTime > begin_arr_range + (provided_period * LEN_RANGE):
                        break
                    else:

                        start38 = time.time()

                        endTime = beginTime + LEN_RANGE
                        traveltime = edg_objs[tc_ind].traffic_net_tt[edgeID][1][arr_range]
                        if last_end_t == begin_arr_range:
                            # Account the travel time estimated during the t_max
                            ranges_tt.append(traveltime)

                        if INTERFACE in ("libsumo", "traci") and TTS_PER_VEH:
                            for cavID in cavs2broadcast[tc_ind]:
                                # Set traffic network travel times to each vehicle detected on incoming
                                # lanes modelled by the TC
                                if INTERFACE in ("libsumo", "traci"):
                                    traci.vehicle.setAdaptedTraveltime(cavID,            # SUMO vehID
                                                                       edgeID,           # SUMO edgeID
                                                                       traveltime,       # traveltime
                                                                       beginTime,        # beginTime
                                                                       endTime)          # endTime
                                else:
                                    LOGGER.debug("Use here another interface to transmit TTs specifically to each CAV")

                                # Add vehicle to the list of vehicles already routed during current range
                                CAVsAlreadyBroadcastedTo.add(cavID)
                            LOGGER.debug("broadcasted traveltime of `{:s}` as {:f} for [{:f},{:f}]".format(edgeID, traveltime, beginTime, endTime))
                        else:
                            # Set traffic network travel times globally to for every vehicle
                            if INTERFACE in ("libsumo", "traci"):
                                traci.edge.adaptTraveltime(edgeID,           # SUMO edgeID
                                                           traveltime,       # traveltime
                                                           beginTime,        # beginTime
                                                           endTime)          # endTime
                                LOGGER.debug("traveltime of `{:s}` adapted to {:f} for [{:f},{:f}]".format(edgeID, traveltime, beginTime, endTime))
                            else:
                                LOGGER.debug("Use here another interface to transmit TTs to all CAVs at once")

                        end38 = time.time()
                        timings[38][1] += end38 - start38

                last_end_t = endTime

            # Add an additional range for anytime after the provided_period time (to assure FIFO rule in SUMO)
            # Make the end time of the range as the end time of the simulation
            # Set the travel time of this last range as the average travel time estimated during the provided_period
            # If this is not implemented, vehicles would use OBUs travel time for arrivals at an edge
            # after the provided_period that could be shorter than the last range travel time + LEN_RANGE
            # This would mean that a vehicle arriving later would depart earlier, which may create
            # irrational loops in the route
            traveltime = round(np.mean(ranges_tt), 1)
            endTime = END_T + 30000
            if INTERFACE in ("libsumo", "traci") and TTS_PER_VEH:
                for cavID in cavs2broadcast[tc_ind]:
                    if INTERFACE in ("libsumo", "traci"):
                        traci.vehicle.setAdaptedTraveltime(cavID,  # SUMO vehID
                                                           edgeID,  # SUMO edgeID
                                                           traveltime,  # traveltime
                                                           beginTime,  # beginTime
                                                           endTime)  # endTime
                    else:
                        LOGGER.debug("Use here another interface to transmit TTs specifically to each CAV")

            else:
                if INTERFACE in ("libsumo", "traci"):
                    traci.edge.adaptTraveltime(edgeID,  # SUMO edgeID
                                               traveltime,  # traveltime
                                               beginTime,  # beginTime
                                               endTime)  # endTime
                else:
                    LOGGER.debug("Use here another interface to transmit TTs to all CAVs at once")

        start39 = time.time()

        if reroute_due_lane_closure:
            for cavID in cavs2broadcast[tc_ind]:
                if cavID not in routedCAVsLaneClosure:
                    # Vehicle route itself with the provided area travel times
                    # Use only when lane closure and if closed after the beginning of simulation
                    if INTERFACE in ("libsumo", "traci"):
                        traci.vehicle.rerouteTraveltime(cavID, currentTravelTimes=False)
                    else:
                        LOGGER.debug("Use here another interface to advice CAVs to reroute")

                    routedCAVsLaneClosure.add(cavID)

        end39 = time.time()
        timings[39][1] += end39 - start39

    end4 = time.time()
    timings[4][1] += end4 - start4


## CLASSES AND THEIR FUNCTIONS

class LdmDataMsgs:
    """Messages to be exchanged between Traffic Controllers (TCs) in the same traffic network"""

    def __init__(self, name):
        """Instances of the message object that represents an area of connected TCs"""
        self.name = name # object name
        self.msgbox = dict() # Format of exchanged messages: {destination: {origin: [inflow msg.,
                                                                                   # travel time msg.,
                                                                                   # queue msg.]}}
        self.num_tcs = None # num TCs in the traffic network
        self.tc_border_jcts = dict() # each TC's modelled junctions that are at the border with other TCs
        self.tc_border_jcts.update({"IDs": []}) # junction IDs
        self.tc_border_jcts.update({"TC": []}) # junctions' respective TCs IDs


    def setMsgsDicts(self):
        """Initialize message object dictionaries"""

        for dest_jct_ind,dest_jctID in enumerate([jct for jct in self.tc_border_jcts["IDs"]]):
            # Define msgbox's possible destination junctions for each origin junction
            self.msgbox.update({dest_jctID: {origin_jctID: [] for origin_jct_ind,origin_jctID
                                             in enumerate(self.tc_border_jcts["IDs"])
                                             if origin_jctID != dest_jctID
                                             and self.tc_border_jcts["TC"][dest_jct_ind]
                                             != self.tc_border_jcts["TC"][origin_jct_ind]}})

        self.num_tcs = len(set(self.tc_border_jcts["TC"]))


    def rcvLdmDataMsgs(self,
                       jct_obj, # TC junction object
                       edg_obj, # TC edge object
                       time2updte): # Remaining time for the algorithm to update
        """Define messages to be received from neighbouring junctions modelled by other TCs
        but in the same traffic network"""

        start0 = time.time()

        for dest_jctID in jct_obj.ids:
            if jct_obj.neigh_jcts[dest_jctID] != []:
                # TC checks all messages to its junctions that are at border between TCs
                num_ranges = int(jct_obj.t_max[dest_jctID] / LEN_RANGE)
                for origin_jctID in self.msgbox[dest_jctID].keys():
                    # For each message from origin_jctID to this dest_jctID
                    travel_time_msg = dict()
                    # It may receive more than one message during one timestep from the same or different junctions
                    num_msg = len(self.msgbox[dest_jctID][origin_jctID])
                    for ind_msg in range(0, num_msg):
                        if self.msgbox[dest_jctID][origin_jctID][0][0] == None:
                            # If empty inflow message, skip
                            pass
                        else:
                            # If received inflow profile message, store to generate arrivals
                            for in_laneID in self.msgbox[dest_jctID][origin_jctID][0][0].keys():
                                inflow_msg = self.msgbox[dest_jctID][origin_jctID][0][0][in_laneID][:]
                                # inflow_msg[0] is the reference time the arrivals are valid
                                # ranges_t_recvInflow represents the intervals between fixed length arrival ranges
                                ranges_t_recvInflow = [inflow_msg[0] + arr_range * LEN_RANGE for arr_range in
                                                       range(0, len(inflow_msg[1]))]
                                edg_obj.lane_inflow_profile[in_laneID][1] = [ACTUAL_T + time2updte,[]]
                                # Adjust the received inflow profile accordingly to the time it will be used
                                for arr_range in range(0, num_ranges):
                                    arr_range_t = ACTUAL_T + time2updte + arr_range * LEN_RANGE
                                    try:
                                        msg_arr_range = ranges_t_recvInflow.index(arr_range_t)
                                        edg_obj.lane_inflow_profile[in_laneID][1][1].append(
                                                                                      inflow_msg[1][msg_arr_range][:])
                                    except ValueError:
                                        # Not received for such arr. range time,
                                        # use equivalent information from last last t_max
                                        try:
                                            msg_arr_range = (num_ranges - 1) - arr_range
                                            edg_obj.lane_inflow_profile[in_laneID][1][1].append(
                                                                                    inflow_msg[1][msg_arr_range][:])
                                        except IndexError:
                                            # If not found any inflow for such time, used fixed inflow profile
                                            interv = edg_obj.lane_inflow_profile[in_laneID][0]["interv"]
                                            edg_obj.lane_inflow_profile[in_laneID][1][1].append(
                                                                        edg_obj.lane_inflow_profile
                                                                        [in_laneID][0][interv][:])

                        if self.msgbox[dest_jctID][origin_jctID][0][1] == None:
                            # If empty travel time message, skip.
                            pass
                        else:
                            for ing_egdID in self.msgbox[dest_jctID][origin_jctID][0][1].keys():
                                travel_time_msg.update({
                                    ing_egdID: self.msgbox[dest_jctID][origin_jctID][0][1][ing_egdID][:]})

                        if self.msgbox[dest_jctID][origin_jctID][0][2] == None:
                            # If empty queue message, skip.
                            pass
                        else:
                            # If received queue prediction message, store them.
                            for in_laneID in self.msgbox[dest_jctID][origin_jctID][0][2].keys():
                                queue_msg = self.msgbox[dest_jctID][origin_jctID][0][2][in_laneID][:]
                                ranges_t_recvQueue = [queue_msg[0] + arr_range * LEN_RANGE for arr_range in
                                                       range(0, len(queue_msg[1]))]
                                edg_obj.lane_queue_pred.update({in_laneID: [ACTUAL_T + time2updte,[]]})
                                # Adjust the information for each range accordingly to the time it will be used
                                for arr_range in range(0, num_ranges):
                                    arr_range_t = ACTUAL_T + time2updte + arr_range * LEN_RANGE
                                    try:
                                        msg_arr_range = ranges_t_recvQueue.index(arr_range_t)
                                        edg_obj.lane_queue_pred[in_laneID][1].append(queue_msg[1][msg_arr_range])
                                    except ValueError:
                                        # Not received for such arr range time,
                                        # use information from last t_max
                                        try:
                                            msg_arr_range = (num_ranges - 1) - arr_range
                                            edg_obj.lane_queue_pred[in_laneID][1].append(queue_msg[1][msg_arr_range])
                                        except IndexError:
                                            # If not found any queue length for such time assume no queue
                                            edg_obj.lane_queue_pred[in_laneID][1].append(0)

                                # Initialize available length of predicted queue
                                edg_obj.lane_aval_queue_len.update({in_laneID:
                                                                   [[max(edg_obj.lane_length[in_laneID]
                                                                         -
                                                                         edg_obj.lane_queue_pred[in_laneID][1][arr_range],
                                                                         0)
                                                                     for arr_range
                                                                     in range(0,
                                                                              len(edg_obj.lane_queue_pred[in_laneID][1]))]
                                                                    for cycle in range(0, MAX_T_MAX_CYCLE)]})

                        # Delete the message from origin_jctID to dest_jctID
                        del self.msgbox[dest_jctID][origin_jctID][0]

                    if time2updte == 0:
                        # Make sure the inflow profile ranges combine with the ACTUAL_T (time info is used)
                        for origin_jctID in jct_obj.commonFromNeigh_laneIDs[dest_jctID].keys():
                            for in_laneID in jct_obj.commonFromNeigh_laneIDs[dest_jctID][origin_jctID]:
                                try:
                                    if edg_obj.lane_inflow_profile[in_laneID][1][0] != ACTUAL_T:
                                        # Take the info from the range before ACTUAL_T and put as the last range
                                        # Assumes cyclic behaviour of the inflow profile
                                        edg_obj.lane_inflow_profile[in_laneID][1][1].append(
                                                                  edg_obj.lane_inflow_profile[in_laneID][1][1].pop(0))
                                        edg_obj.lane_inflow_profile[in_laneID][1][0] = ACTUAL_T
                                except IndexError:
                                    noMsgRcvedYet = 1

                    # Update table with travel times for the whole traffic network with the received travel time
                    edg_obj.updateTravelTimeDB(travel_time_msg, edg_obj.traffic_net_tt)

            else:
                pass

        end0 = time.time()
        timings[0][1] += end0 - start0


    def sendLdmDataMsgs(self,
                        jct_obj, # TC junction object
                        edg_obj, # TC edge object
                        tc_it,  # algorithm order interaction
                        jct_order, # Order of junctions to be explored. 1 or -1
                        lanes_order): # Lanes that follow the order of jct_order
        """Define messages to be sent to neighbouring junctions modelled by other TCs but in the same traffic network"""

        global last_tc_border_jct, tcs_set_to_rcv_tt, queue_msg, travel_time_msg
        start1 = time.time()

        # Define the border junctions (junction bordering other TCs)
        tc_border_jcts = [jct for jct in jct_obj.ids if jct_obj.neigh_jcts[jct] != []]
        if jct_order != 1:
            tc_border_jcts.reverse()

        if tc_it == TC_IT_MAX:
            last_tc_border_jct = tc_border_jcts[-1]
            # Init. the set with the TCs to receive this TC's edge travel times
            # To avoid sending again the same information to same TC more than once
            tcs_set_to_rcv_tt = set([])
            # Get this TC's edge travel times
            travel_time_msg = edg_obj.tc_tt.copy()
        else:
            # Queue and Travel Time Messages are shared only on the last interaction
            queue_msg = None
            travel_time_msg = None

        for origin_jctID in tc_border_jcts:
            # Prepare the dynamic inflow profile for the lanes following the order of junctions to be explored
            inflow_msg = edg_obj.updateDynLaneInflowProfile(jct_obj, lanes_order, origin_jctID)
            if tc_it == TC_IT_MAX:
                # Prepare the queue prediction of the incoming lane of origin_jctID which is out. lane of dest_jctID
                queue_msg = edg_obj.updateQueueLength(jct_obj, origin_jctID)

            # Prepare the message to junctions which have the inflow message part
            for dest_jctID in inflow_msg.keys():
                dest_jct_tc = self.tc_border_jcts["TC"][self.tc_border_jcts["IDs"].index(dest_jctID)]
                if queue_msg != None and queue_msg.has_key(dest_jctID) \
                and travel_time_msg != None and dest_jct_tc not in tcs_set_to_rcv_tt:
                    tcs_set_to_rcv_tt.add(dest_jct_tc)
                    self.msgbox[dest_jctID][origin_jctID].append([inflow_msg[dest_jctID].copy(),
                                                                  travel_time_msg.copy(),
                                                                  queue_msg[dest_jctID].copy()])

                elif queue_msg != None and queue_msg.has_key(dest_jctID):
                    self.msgbox[dest_jctID][origin_jctID].append([inflow_msg[dest_jctID].copy(),
                                                                  None,
                                                                  queue_msg[dest_jctID].copy()])

                elif travel_time_msg != None and dest_jct_tc not in tcs_set_to_rcv_tt:
                    tcs_set_to_rcv_tt.add(dest_jct_tc)
                    self.msgbox[dest_jctID][origin_jctID].append([inflow_msg[dest_jctID].copy(),
                                                                  travel_time_msg.copy(),
                                                                  None])
                else:
                    self.msgbox[dest_jctID][origin_jctID].append([inflow_msg[dest_jctID].copy(), None, None])

            # Prepare the message to junctions which don't have the inflow message part but the queue part instead
            if queue_msg != None:
                rem_dest_jctID = set(queue_msg.keys()).difference(set(inflow_msg.keys()))
                for dest_jctID in rem_dest_jctID:
                    dest_jct_tc = self.tc_border_jcts["TC"][self.tc_border_jcts["IDs"].index(dest_jctID)]
                    if travel_time_msg != None and dest_jct_tc not in tcs_set_to_rcv_tt:
                        tcs_set_to_rcv_tt.add(dest_jct_tc)
                        self.msgbox[dest_jctID][origin_jctID].append([None,
                                                                      travel_time_msg.copy(),
                                                                      queue_msg[dest_jctID].copy()])
                    else:
                        self.msgbox[dest_jctID][origin_jctID].append([None, None, queue_msg[dest_jctID].copy()])

            # Multicast this TC's edge travel times to other TCs that didn't received yet
            if travel_time_msg != None and origin_jctID == last_tc_border_jct:
                jcts_set_to_rcv_tt = set([])
                for jctID in self.msgbox[origin_jctID].keys():
                    jct_tc = self.tc_border_jcts["TC"][self.tc_border_jcts["IDs"].index(jctID)]
                    if jct_tc not in tcs_set_to_rcv_tt:
                        # Define the remaining border junctions and TCs to receive the travel times
                        jcts_set_to_rcv_tt.add(jctID)
                        tcs_set_to_rcv_tt.add(jct_tc)
                    if len(tcs_set_to_rcv_tt) == self.num_tcs - 1:
                        # If already set for all TCs in the traffic network, stop
                        break
                # Prepare the message to junctions which don't have the inflow and queue message parts
                for dest_jctID in jcts_set_to_rcv_tt:
                    self.msgbox[dest_jctID][origin_jctID].append([None, travel_time_msg.copy(), None])

        end1 = time.time()
        timings[1][1] += end1 - start1


class Junctions:
    """Junctions and their characteristics modelled by the TC"""

    def __init__(self, name):
        """Instances of the junctions object that represents the modelled junctions"""

        self.name = name # object name
        self.ids = [] # list of junction modelled by the TC
        self.lanesForwardOrder = dict() # lanes which their upstream junction and downstream junction is
                                         # the same to the order of self.ids. {jctID: [lane1,lane2,...]}
        self.lanesReverseOrder = dict() # same as above but against order. {jctID: [lane3,lane4,...]}
                                         # lanes from/to unknown junctions should be classified as both orders!
                                         # if lane connects to a border junction, it follows the order of TCs
        self.tlID = dict()   # traffic light id of each junction. {jctID: tlID}
        self.tlID_to_jctIDs = dict()  # junction ids of each traffic light. {tlID: [jct1, jct2,...]}
        self.neigh_jcts = dict() # border junctions of other TCs that are neighbouring a jctID. {jctID: [jct1,jct2,...]}
        self.commonToNeigh_laneIDs = dict() # outgoing lanes from jctID to a downstream border junction of other TC
                                            # {jctID: {neigh_jct1: [lane1,lane2,...], neigh_jct2: [lane3,lane4,...]}}
        self.commonFromNeigh_laneIDs = dict() # same as above but incoming lanes from border junction.
        self.jctType = dict() # junction type. {jctID: [jct_type1, jct_type2,...]}. jctType options:
        # "Coop." for cooperative junction (ie. it has RSU to communicate to CAVs)
        # "Mod. Non-C." for modelled junction not connected to the TC, ie. unsignilised, or not sharing queue length
        # "Mod. C." for connected modelled junction, sharing queue length and signal plans if signalized
        # "Conns." for junctions that don't have traffic control, ie. lane merge or lane split
        self.t_max = dict() # junction horizon time of planning. {jctID: int in seconds}
        self.t_max_cycle = dict() # the cycle of the prediction time when extrapolating results. {jctID: int}
        self.comm_range = dict() # communication range. {jctID: float}
        self.movgIDs = dict() # list of the junction's movement groups. {jctID: [movg1,movg2,...]}
        self.connIDs = dict() # list of the junction's connections. {jctID: [conn1,conn2,...]}
        self.in_laneIDs = dict() # list of the junction's incoming lanes. {jctID: set([lane1,lane2,...])}
        self.out_laneIDs = dict() # list of the junction's outgoing lanes. {jctID: set([lane1,lane2,...])}
        self.in_edgeIDs = dict() # list of the junction's incoming edges. {jctID: set([edge,edge2,...])}
        self.out_edgeIDs = dict() # list of the junction's outgoing edges. {jctID: set([edge1,edge2,...])}
        self.phases_green_plan = dict() # junction's green plan of movement groups for each phase
                                        # {jctID:[[[phase 1 movgIDs with permissive green],
                                        #          [phase 1 movgIDs with protected green]],
                                        #         [[phase 2 movgIDs with permissive green],
                                        #          [phase 2 movgIDs with protected green]], ...]}
        self.RSU_CAVs_data = dict() # data received from RSUs
        self.cavs_in_range = dict() # vehicles that have their position in one of the junction's incoming lanes
                                    # the RSU_id is the cooperative junction the vehicle was detected by its RSU
                                    # {jctID: {vehID: (RSU_id, in_laneID)}}
        self.arr_headways = dict() # junction's table to create arrivals based on inflow profile:
        # self.arr_headways[jctID][:, 0] = Time Interval Values of Headways given the STEP_LENGTH
        # self.arr_headways[jctID][:, 1] = Used according the headway probability distrib. Either PDF, CDF or 1 - CDF
        # self.arr_headways[jctID][:, 2] = Probability of headways within [k * STEP_LENGTH, (k + 1) * STEP_LENGTH]
        # where k = [1, ..., LEN_RANGE]


    def addJunction(self,
                    msgs_obj,  # message object of the traffic network
                    jctID,  # network Junction ID
                    jct_type,  # "Coop." whe it has RSU to communicate to CAVs
                               # "Mod. Non-C." when unsignilised or not sharing queue length
                               # "Mod. C." when signilised and sharing queue length as well as signal plans
                               # "Conns." for junctions that don't have traffic control, ie. lane merge or lane split
                    tlID=None,  # traffic light ID of the junction
                    neigh_jcts=[],  # border junction IDs direct neighbouring this junction. [neigh_jct1,neigh_jct2,...]
                    commonToNeigh_lanes=[], # common lanes with downstream neigh_jcts and this junction upstream
                                            # [[neigh_jct1_lane1,neigh_jct1_lane2],[neigh_jct2_lane4,neigh_jct2_lane3]]
                    commonFromNeigh_lanes=[], # as above, the indexes of neigh_jcts should match with the lanes list
                    lanes_forwardOrder=[], # list of lanes where their up and donwstream jct follows self.ids
                    lanes_reverseOrder=[], # same as above but opposite
                    t_max=STD_T_MAX,  # cycle time prediction in seconds
                    comm_range=STD_COMM_RANGE): # Communication range around the junction. In meters. (For Simulation)
        """Add a new junction to be modelled by the junction object of the TC"""

        self.ids.append(jctID)
        self.tlID.update({jctID: tlID})
        if tlID != None:
            try:
                self.tlID_to_jctIDs[tlID].append(jctID)
            except KeyError:
                self.tlID_to_jctIDs.update({tlID: [jctID]})

        self.neigh_jcts.update({jctID: neigh_jcts})
        self.commonToNeigh_laneIDs.update({jctID: dict()})
        self.commonFromNeigh_laneIDs.update({jctID: dict()})
        self.lanesForwardOrder.update({jctID: lanes_forwardOrder})
        self.lanesReverseOrder.update({jctID: lanes_reverseOrder})
        if neigh_jcts != []:
            # If junction has neighbouring junctions
            msgs_obj.tc_border_jcts["IDs"].append(jctID)
            msgs_obj.tc_border_jcts["TC"].append(self.name)
        for jct_ind,neigh_jct in enumerate(neigh_jcts):
            # Add lanes in common with neighbouring junctions
            try:
                self.commonToNeigh_laneIDs[jctID].update({neigh_jct: commonToNeigh_lanes[jct_ind]})
                self.commonFromNeigh_laneIDs[jctID].update({neigh_jct: commonFromNeigh_lanes[jct_ind]})
            except IndexError:
                self.commonToNeigh_laneIDs[jctID].update({neigh_jct: []})
                self.commonFromNeigh_laneIDs[jctID].update({neigh_jct: []})

        # Initialize junction object instances
        self.jctType.update({jctID: jct_type})
        self.t_max.update({jctID: t_max})
        self.t_max_cycle.update({jctID: None})
        self.comm_range.update({jctID: comm_range})
        self.movgIDs.update({jctID: []})
        self.connIDs.update({jctID: []})
        self.in_laneIDs.update({jctID: set()})
        self.out_laneIDs.update({jctID: set()})
        self.in_edgeIDs.update({jctID: set()})
        self.out_edgeIDs.update({jctID: set()})
        self.phases_green_plan.update({jctID: [[[],[]]]})
        self.RSU_CAVs_data.update({jctID: dict()})
        self.cavs_in_range.update({jctID: dict()})
        hdwy_0 = frange(STEP_LENGTH,MIN_HDWY + STEP_LENGTH,STEP_LENGTH)[-1]
        num_hdwy_interv = len(frange(hdwy_0, LEN_RANGE + STEP_LENGTH, STEP_LENGTH))
        self.arr_headways.update({jctID: np.ndarray(shape=(num_hdwy_interv,3), dtype=float)})
        self.arr_headways[jctID][:,0] = np.linspace(hdwy_0, LEN_RANGE, num_hdwy_interv)


    def getCAVinRange(self,
                      edg_obj,       # TC edge object
                      veh_obj):      # TC vehicle object
        """Define the detected junction, incoming lane and its junction for
         all CAVs running on the incoming lanes modelled by the TC"""

        start2 = time.time()

        if INTERFACE not in ("libsumo", "traci"):
            # Check getCAVdataViaLibsumo() and getCAVdataViaTraci() functions for pattern of self.RSU_CAVs_data
            LOGGER.debug("Get CAV data via other interface")

        for jctID in self.ids:
            # Init. cavs_in_range
            self.cavs_in_range.update({jctID: dict()})

        if INTERFACE == "libsumo":
            # If simulating and using libsumo, as context subscriptions are not working yet it is set that all junctions
            # have RSU and vehicles are already distributed per junctions according to the edges they are running,
            # this was done by function getCAVdataViaLibsumo
            RSU_ids = self.ids[:]
        else:
            RSU_ids = [jctID for jctID in self.ids if "Coop." in self.jctType[jctID]]

        global updated_jctID

        for RSU_id in RSU_ids:
            try:
                RSU_timestep_CAVs = self.RSU_CAVs_data[RSU_id].keys()
                for vehID in RSU_timestep_CAVs:
                    # Add CAVs to the set of all vehicles that are within the incoming (and internal) lanes of
                    # updated_jctID
                    try:
                        in_laneID = self.RSU_CAVs_data[RSU_id][vehID][LANE]
                        try:
                            edgeID = edg_obj.laneID_to_edgeID[in_laneID]
                            updated_jctID = edg_obj.inLaneID_to_jctID[in_laneID]
                            # I found updated_jctID, CAV on incoming lane modelled by this traffic controller (TC)
                            # Measure travel time of CAVs
                            # Calling measureCAVRealTravelTime() here makes that vehicles detected on a internal lanes
                            # will not be acounted as left the edge, only when arrived at the next one.
                            measureCAVRealTravelTime(edg_obj, veh_obj, vehID, edgeID, in_laneID)
                        except KeyError:
                            # CAV on junction internal lane or lane not modelled by this traffic controller (TC)
                            # Check if outgoing lane, then remove vehicle from veh_obj.cavs_metrics_samples
                            if edg_obj.outLaneID_to_jctID.has_key(in_laneID):
                                # It is an ongoing lane
                                measureCAVRealTravelTime(edg_obj, veh_obj, vehID, edgeID, in_laneID)
                                del veh_obj.cavs_metrics_samples["travel_times"][vehID]
                            # Check if internal lane of modelleded junction by this TC
                            # (in SUMO it contains the junction ID)
                            for updated_jctID in self.ids:
                                # Old way to check if connection is within updated_jctID
                                # ("_").join(in_laneID.split("_")[:-2])[1:] == updated_jctID
                                # New way to check if connection is within updated_jctID
                                if updated_jctID in in_laneID:
                                    # Junction is modelled by the TC.
                                    # Define vehicle in_lane as one of the out_lanes of the junction as vehicle is
                                    # in the middle of the junction.
                                    # First, get the next edge to check if vehicle will finish on the next edge
                                    if INTERFACE in ("libsumo", "traci"):
                                        veh_route = self.RSU_CAVs_data[RSU_id][vehID][ROUTE]
                                        route_ind = traci.vehicle.getRouteIndex(vehID)
                                        edgeIDs = self.out_edgeIDs[updated_jctID].intersection(
                                            set(veh_route[route_ind:route_ind+2]))
                                    else:
                                        LOGGER.debug("Get CAV route or next edge via other interface")

                                    if edgeIDs == set():
                                        # Not found outgoing edge on the route, vehicle's last edge (finishing)
                                        raise KeyError
                                    else:
                                        # Found outgoing edge on the route
                                        try:
                                            # Get out. lane of the current connection
                                            in_laneID = edg_obj.connID_to_outLaneID[in_laneID]
                                            # in_laneID = edg_obj.connID_to_outLaneID[traci.vehicle.getLaneID(vehID)]
                                        except KeyError:
                                            # Connection not found, assign any out. lane of the next edge
                                            in_laneID = edg_obj.edgeID_to_laneIDs[list(edgeIDs)[0]][0]
                                        # Check if the junction of the outgoing lane is modelled by the same TC
                                        # If don't work the line below it raises KeyError
                                        updated_jctID = edg_obj.inLaneID_to_jctID[in_laneID]
                                        break
                            else:
                                # Junction not modelled
                                raise KeyError

                        if self.cavs_in_range[updated_jctID].has_key(vehID) == False:
                            # Vehicle is added in the range of the junction which models vehicle's inc. lane
                            # and not on the detected junction that has RSU
                            self.cavs_in_range[updated_jctID].update({vehID: (RSU_id, in_laneID)})
                        else:
                            # Vehicle already added or the junction of the lane was not found
                            pass

                    except KeyError:
                        # Junction not modelled by this junction object.
                        jct_not_modelled = 1

            except (AttributeError, TypeError):
                no_vehicles = 1


        end2 = time.time()
        timings[2][1] += end2 - start2


    def getLanesOrder(self,
                      edg_obj, # TC edge object
                      jct_order): # Order of junctions to be explored following self.ids and order of TCs
        """Get the lanes order based on the sequence of junctions to be explored"""

        lanes_order = []
        jct_ids_order = self.ids[:]
        if jct_order == 1:
            # Make forward order
            for jctID in jct_ids_order:
                # As some lanes are classified both orders, it is used edg_obj.lane_arrs_done[lane] to avoid
                # putting the lane twice as in order
                lanes_order.extend([lane for lane in self.lanesForwardOrder[jctID]
                                    if edg_obj.lane_arrs_done[lane] == 0])
        else:
            # Make reverse order
            jct_ids_order.reverse()
            for jctID in jct_ids_order:
                lanes_order.extend([lane for lane in self.lanesReverseOrder[jctID]
                                    if edg_obj.lane_arrs_done[lane] == 0])

        return lanes_order



class MovementGroups:
    """Same idea of signal groups but as a generalization to work for unsignilised junctions.
    Movement groups of the junctions modelled by the TC"""

    # Used movement status:
    # a) permissive-Movement-Allowed (value 5) for each movement that need to yield, or
    # b) protected-Movement-Allowed (value 6) for each movement that has the preference

    # All movement status (but not used here, except 5 and 6):
    # MovementPhaseState = ["unavailable", # [0] This state is used for unknown or error.
    #                       "dark", # [1] The signal head is dark (unlit).
    #                       "stop-Then-Proceed", # [2] 'flashing red' in US, Stop vehicle at stop line,
    #                                                   do not proceed unless it is safe.
    #                       "stop-And-Remain", # [3] 'red light' in US
    #                       "pre-Movement", # [4] Red + Yellow, not used in the US, partly in REG-D,
    #                                                          not allowed to drive yet
    #                       "permissive-Movement-Allowed", # [5] 'permissive green' in US,
    #                                                            conflicting traffic may be present
    #                       "protected-Movement-Allowed", # [6] protected green' in US
    #                       "permissive-clearance", # [7] 'permissive yellow' in US,
    #                                                           conflicting traffic may be present
    #                       "protected-clearance", # [8] 'protected yellow' in US
    #                       "caution-Conflicting-Traffic"  # [9] 'flashing yellow/amber'
    #                                                           in US for extended periods of time
    #                       ]

    def __init__(self, name):
        """Instances of the movement groups object that represents when vehicles may cross"""

        self.name = name # object name
        self.ids = []   # list of movement groups of the junctions modelled by the TC. ID pattern: jctID_movgNum
        self.jctIDs = dict() # ID of the mov. group's junction. {movgID: jctID}
        self.movg_green_plan = dict() # Green plan of the mov. group. Example: {movgID:
                                      # [[begin times], > [1,30]
                                      #  [end times], > [26,40]
                                      #  [phase states] > phase_states = 5 (permissive) or 6 (protected green)
                                      #  [phases junction green_plan phase num], > [0,1]
                                      #  [red + amber times], > [1,0] if inexistent use zero
                                      #  [amber times]]} > [4,0] if inexistent use zero
                                      # each index represents times of a predicted phase
        self.probe_vehs_range = dict() # list of vehicles to get their travel times per inc. lane, mov. group and range
                                       # {in_laneID: {movg: [vehs. range1, vehs. range2, vehs. range3]}}
        self.movgID_to_connIDs = dict() # connection IDs of the Movement Group ID {movgID: connIDs}
        self.movgID_to_phaseNum = dict() # phase index of self.movg_green_plan[movgID] being used. {movgID: int}
        self.tlID_to_linkIndexes = dict() # It is tlight's linkIndexes and their respective movgID {tlID:
                                         # {linkIndex: movgID}}. LinkIndex the connection number within the junction
        self.movgID_to_inLaneIDs = dict() # incoming lanes IDs of the movement group. {movgID: [in_laneID1,in_laneID2]}
        self.movgID_to_outLaneIDs = dict() # outgoing lanes IDs of the movement group. {movgID: [out_laneID3,out_laneID4]}
        self.movgID_to_yield2movgIDs = dict() # other mov. groups the mov. group needs to yield. {movgID: list movgs2yield}


    def setMovGroupsDicts(self, jct_obj, edg_obj):
        """Initialize mov. groups object dictionaries"""

        for in_laneID in edg_obj.inLaneID_to_jctID.keys():
            num_ranges = int(jct_obj.t_max[edg_obj.inLaneID_to_jctID[in_laneID]] / LEN_RANGE)
            # Define possible mov. groups for each incoming lane
            self.probe_vehs_range.update({in_laneID: {movg: [[] for _ in range(0, num_ranges)]
                                                       for movg in edg_obj.inLaneID_to_movgIDs[in_laneID]}})


    def addMovGroup(self,
                    edg_obj,  # TC edge object
                    jctID,    # Respective Network Junction ID
                    movg_num, # LDM Definition 0...255 for signal group number, or arbitrary otherwise
                    connIDs,  # List of respective lane Connections. In SUMO it is the junction tag "via" attribute
                              # Ex: [conn0_of_MG0, conn1_of_MG0] e.g. certain lane through, while another right
                    jct_obj=None, # TC junction object
                    tlID = None,    # Respective traffic light ID
                    linkIndexs = [], # STRING link index of each connID in the traffic light.
                                     # In SUMO it is the junction tag "linkIndex" attribute. ["0","1",...]
                                     # Basically the number of the connection within the junction
                    add_neigh_tc_bool=0): # If 1, the info is regarding a neighbouring border junction, 0 otherwise
        """Add movement group to the movement group object"""

        movgID = str(jctID)+"_MG"+ str(movg_num)
        self.movgID_to_connIDs.update({movgID: connIDs})
        self.movgID_to_outLaneIDs.update({movgID: set()})
        if add_neigh_tc_bool == 0:
            self.ids.append(movgID)
            self.jctIDs.update({movgID: jctID})
            self.movg_green_plan.update({movgID: [[],[],[],[],[],[]]})
            jct_obj.movgIDs[jctID].append(movgID)
            self.movgID_to_phaseNum.update({movgID: 0})
            self.movgID_to_inLaneIDs.update({movgID: set()})
            self.movgID_to_yield2movgIDs.update({movgID: []})
            try:
                # Store the linkIndexs and their respective movgID for tlID
                for linkIndex in linkIndexs:
                    self.tlID_to_linkIndexes[tlID].update({linkIndex: movgID})
            except KeyError:
                # If there was not any information of link indexes for tlID, create a new key for tlID
                self.tlID_to_linkIndexes.update({tlID: dict()})
                # Store the linkIndexs and their respective movgID for tlID
                for linkIndex in linkIndexs:
                    self.tlID_to_linkIndexes[tlID].update({linkIndex: movgID})

        for conn in self.movgID_to_connIDs[movgID]:
            in_laneID = edg_obj.connID_to_inLaneID[conn]
            out_laneID = edg_obj.connID_to_outLaneID[conn]
            if add_neigh_tc_bool == 0:
                self.movgID_to_inLaneIDs[movgID].add(in_laneID)
                edg_obj.outLaneID_to_movgIDs[out_laneID].add(movgID)
                edg_obj.inLaneID_to_movgIDs[in_laneID].add(movgID)
            edg_obj.connID_to_movgID.update({conn: movgID})
            self.movgID_to_outLaneIDs[movgID].add(out_laneID)


    def setUnsignalizedJctMovGroupPlans(self,
                                        jct_obj, # TC junction object
                                        jctID,   # network junction ID
                                        movgID,  # Movement Group ID - jctID_movgNum
                                        phase_state): # 5 for permissive green or 6 for protected green
        """Static movement group plan, only for unsignised or connection junctions"""

        self.movg_green_plan.update({movgID: [[0], [float('inf')], [phase_state],[0], [0], [0]]})
        if phase_state == 5:
            # If permissive green
            jct_obj.phases_green_plan[jctID][0][0].append(movgID)
        elif phase_state == 6:
            # If protected green
            jct_obj.phases_green_plan[jctID][0][1].append(movgID)


    def getSignalizedJctMovGroupPlans(self,
                                      jct_obj): # TC junction object
        """Get TC's junctions planned signal plans during the time horizon of planning"""

        global links_protec, links_perm
        start5 = time.time()

        for tlID in jct_obj.tlID_to_jctIDs.keys():
            # If junction has traffic light, it has a signal plan.
            # Otherwise it uses setUnsignalizedJctMovGroupPlans function
            if INTERFACE in ("traci", "libsumo"):
                # It checks the green plan during prediction time t_max from ACTUAL_T, this allows fixed or dyn. plans
                # Any light change is a phase in SUMO (red, red + amber, green, amber = 4 phases), called as SUMO_PHASE,
                # while a REAL_PHASE is usually represented by its simultaneous green time to certain signal group
                SPaT = traci.trafficlight.getCompleteRedYellowGreenDefinition(tlID)
                phases_def = []
                phases_dur = []
                if SUMO_RELEASE > 1 or (SUMO_RELEASE == 1 and SUMO_PATCH >= 1):
                    # Get signal plan from SUMO for SUMO newer version (from 1.1.0)
                    for phase in SPaT[0].phases:
                        phases_def.append(phase.state)
                        phases_dur.append(phase.duration)
                else:
                    # Get signal plan from SUMO for SUMO older version (before 1.1.0)
                    SPaT = str(SPaT)
                    # Get current SUMO_PHASE
                    searched_str = 'currentPhaseIndex:'
                    pos_current_phase = SPaT.find(searched_str) + len(searched_str) + 1
                    # Initialize Variables
                    pos_def_init = 0
                    pos_dur_init = 0
                    pos_def_end = 0
                    def_searched_str = "phaseDef:"
                    len_def_str = len(def_searched_str)
                    dur_searched_str = "duration:"
                    len_dur_str = len(dur_searched_str)
                    # Read SUMO signal plan and get SUMO_PHASE begin and end times as well its duration
                    while pos_def_end != len(SPaT) - 2:
                        # Definition of the SUMO_PHASEs
                        pos_def = SPaT[pos_def_init:].find(def_searched_str)
                        pos_def_init += pos_def + 1
                        pos_def_beg = pos_def_init + len_def_str
                        pos_def_end = pos_def_init + SPaT[pos_def_init:].find("\n")
                        phases_def.append(SPaT[pos_def_beg:pos_def_end])
                        # Duration of the SUMO_PHASEs
                        pos_dur = SPaT[pos_dur_init:].find(dur_searched_str)
                        pos_dur_init += pos_dur + 1
                        pos_dur_beg = pos_dur_init + len_dur_str
                        pos_dur_end = pos_dur_init + SPaT[pos_dur_init:].find("\n")
                        converted_dur = float(SPaT[pos_dur_beg:pos_dur_end])
                        phases_dur.append(converted_dur)

                # Get linkIndexes movement status from the definition of the SUMO_PHASEs
                links_signal_phase = []
                for ph_num, ph_def in enumerate(phases_def):
                    # linkIndexes with permissive green movement status
                    links_perm = [link for link, _ in enumerate(ph_def)
                                  if _ == "g" and self.tlID_to_linkIndexes[tlID].has_key(link)]
                    # linkIndexes with protec green movement status
                    links_protec = [link for link, _ in enumerate(ph_def)
                                    if _ == "G" and self.tlID_to_linkIndexes[tlID].has_key(link)]
                    # linkIndexes with red + amber movement status
                    links_redAmber = [link for link, _ in enumerate(ph_def)
                                      if _ == "u" and self.tlID_to_linkIndexes[tlID].has_key(link)]
                    # linkIndexes with amber movement status
                    links_Amber = [link for link, _ in enumerate(ph_def)
                                   if _ == "y" and self.tlID_to_linkIndexes[tlID].has_key(link)]
                    links_signal_phase.append((links_perm, links_protec, links_redAmber, links_Amber))
            else:
                LOGGER.debug("Use here another interface to get signal timings."
                      "It requires the output with following format for the lists:"
                      "links_signal_phase[phase] = (links_perm, links_protec, links_redAmber, links_Amber)"
                      "phases_dur[phase] = float in seconds")

            # Initialize Instances
            for linkIndex in self.tlID_to_linkIndexes[tlID].keys():
                movgID = self.tlID_to_linkIndexes[tlID][linkIndex]
                self.movg_green_plan.update({movgID: [[], [], [], [], [], []]})

            for jctID in jct_obj.tlID_to_jctIDs[tlID]:
                jct_obj.phases_green_plan[jctID] = []
                next_is_new_phase = 1
                if INTERFACE in ("traci", "libsumo"):
                    # Get time to change current SUMO_PHASE
                    end_time = traci.trafficlight.getNextSwitch(tlID)
                    if SUMO_RELEASE > 1 or (SUMO_RELEASE == 1 and SUMO_PATCH >= 1):
                        # Get current phase for SUMO newer version (from 1.1.0)
                        current_phase = SPaT[0].currentPhaseIndex
                    else:
                        # Get time to change current SUMO_PHASE for SUMO older version (before 1.1.0)
                        current_phase = int(SPaT[pos_current_phase])
                else:
                    LOGGER.debug("Use there another interface to get current phase number")

                begin_time = end_time - phases_dur[current_phase]
                num_phases = len(phases_dur)
                # Store the signal plan and repeat it until the maximum time horizon of prediction
                while begin_time < (ACTUAL_T + jct_obj.t_max[jctID]):
                    try:
                        # Get current SUMO_PHASE green linkIndexes
                        links_perm = [perm_link for perm_link in links_signal_phase[current_phase][0][:]
                                          if self.tlID_to_linkIndexes[tlID][perm_link] in jct_obj.movgIDs[jctID]]
                        links_protec = [protec_link for protec_link in links_signal_phase[current_phase][1][:]
                                          if self.tlID_to_linkIndexes[tlID][protec_link] in jct_obj.movgIDs[jctID]]
                        # Check which linkIndexes have green status that was also green on the last SUMO_PHASE
                        # and also if the movgID is included in the junction
                        new_perm_links = [perm_link for perm_link in links_perm
                                          if self.tlID_to_linkIndexes[tlID][perm_link]
                                          not in jct_obj.phases_green_plan[jctID][-1][0]]
                        new_protec_links = [protec_link for protec_link in links_protec
                                            if self.tlID_to_linkIndexes[tlID][protec_link]
                                            not in jct_obj.phases_green_plan[jctID][-1][1]]
                        remaining_perm_links = [perm_link for perm_link in links_perm
                                                if self.tlID_to_linkIndexes[tlID][perm_link]
                                                in jct_obj.phases_green_plan[jctID][-1][0]]
                        remaining_protec_links = [protec_link for protec_link in links_protec
                                                  if self.tlID_to_linkIndexes[tlID][protec_link]
                                                  in jct_obj.phases_green_plan[jctID][-1][1]]
                    except IndexError:
                        # Initialize
                        new_perm_links = links_perm[:]
                        new_protec_links = links_protec[:]
                        remaining_perm_links = []
                        remaining_protec_links = []

                    if links_perm + links_protec != [] and (new_perm_links + new_protec_links != []
                                                            or next_is_new_phase == 1):
                        # Add a new REAL_PHASE that contains a green type of movement status for the mov. groups of jctID
                        next_is_new_phase = 0
                        jct_obj.phases_green_plan[jctID].append([[],[]])
                        links2consider = links_perm + links_protec
                        movgIDs_alreadySet = []
                        for linkIndex in links2consider:
                            movgID = self.tlID_to_linkIndexes[tlID][linkIndex]
                            if movgID not in movgIDs_alreadySet:
                                self.movg_green_plan[movgID][3].append(len(jct_obj.phases_green_plan[jctID]) - 1)
                                if linkIndex in links_perm:
                                    # Add movgID to the permissive green list for the REAL_PHASE
                                    jct_obj.phases_green_plan[jctID][-1][0].append(movgID)
                                else:
                                    # Add movgID to the protected green list for the REAL_PHASE
                                    jct_obj.phases_green_plan[jctID][-1][1].append(movgID)
                                # Define duration
                                self.movg_green_plan[movgID][0].append(begin_time)
                                self.movg_green_plan[movgID][1].append(end_time)
                                if linkIndex in links_signal_phase[current_phase][0]:
                                    # Assign permissive green to the movement group
                                    self.movg_green_plan[movgID][2].append(5)
                                if linkIndex in links_signal_phase[current_phase][1]:
                                    # Assign protected green to the movement group
                                    self.movg_green_plan[movgID][2].append(6)
                                if linkIndex in links_signal_phase[current_phase - 1][2]:
                                    # Get red + amber time of the movement group
                                    redAmber_t = phases_dur[current_phase - 1]
                                    self.movg_green_plan[movgID][4].append(redAmber_t)
                                else:
                                    self.movg_green_plan[movgID][4].append(0)
                                # Check amber time of the movement group (if exists)
                                try:
                                    amber_t = phases_dur[current_phase + 1]
                                    if linkIndex in links_signal_phase[current_phase + 1][3]:
                                        self.movg_green_plan[movgID][5].append(amber_t)
                                    else:
                                        # Not found amber time
                                        self.movg_green_plan[movgID][5].append(0)
                                except IndexError:
                                    # Last SUMO_PHASE, check amber time of first SUMO_PHASE
                                    amber_t = phases_dur[0]
                                    if linkIndex in links_signal_phase[0][3]:
                                        self.movg_green_plan[movgID][5].append(amber_t)
                                    else:
                                        # Not found amber time
                                        self.movg_green_plan[movgID][5].append(0)

                                movgIDs_alreadySet.append(movgID)
                            else:
                                pass

                    elif links_perm + links_protec == []:
                        # A SUMO_PHASE without green for the mov. groups of jctID, skip.
                        # Amber and red + amber SUMO_PHASEs are included when detected green SUMO_PHASE
                        next_is_new_phase = 1
                    else:
                        # A SUMO_PHASE that is a continuation of a REAL_PHASE for the mov. groups of jctID
                        next_is_new_phase = 0
                        links2consider = remaining_perm_links + remaining_protec_links
                        movgIDs_alreadySet = []
                        for linkIndex in links2consider:
                            movgID = self.tlID_to_linkIndexes[tlID][linkIndex]
                            if movgID not in movgIDs_alreadySet:
                                self.movg_green_plan[movgID][1][-1] = end_time
                            # Check amber time of the movement group (if exists)
                            try:
                                amber_t = phases_dur[current_phase + 1]
                                if linkIndex in links_signal_phase[current_phase + 1][3]:
                                    self.movg_green_plan[movgID][5][-1] = amber_t
                                else:
                                    # Not found amber time
                                    self.movg_green_plan[movgID][5][-1] = 0
                            except IndexError:
                                # Last SUMO_PHASE, check amber time of first SUMO_PHASE
                                amber_t = phases_dur[0]
                                if linkIndex in links_signal_phase[0][3]:
                                    self.movg_green_plan[movgID][5][-1] = amber_t
                                else:
                                    # Not found amber time
                                    self.movg_green_plan[movgID][5][-1] = 0

                    if current_phase < (num_phases - 1):
                        current_phase += 1
                    else:
                        # Last SUMO_PHASE, restart signal plan
                        current_phase = 0
                    begin_time = end_time
                    end_time += phases_dur[current_phase]

                    if end_time > ACTUAL_T + jct_obj.t_max[jctID]:
                        # Make end time of current phase as the end of prediction time
                        end_time = ACTUAL_T + jct_obj.t_max[jctID]

                for movgID in jct_obj.movgIDs[jctID]:
                    if self.movg_green_plan[movgID] == [[], [], [], [], [], []]:
                        # If cycle longer than the maximum prediction time (or no information for movgID)
                        # Make initial and final time the end of prediction time as protected green
                        # Vehicles will not move to cross because there is no green phase, but will move during red
                        self.movg_green_plan.update({movgID: [[ACTUAL_T + jct_obj.t_max[jctID]],
                                                              [ACTUAL_T + jct_obj.t_max[jctID]], [6], [None], [0], [0]]})
                    else:
                        # It is needed to add a non-crossing phase if mov. group has red signal but vehicle can move due
                        # vehicle ahead having green (when same lane has different mov. group)
                        self.movg_green_plan[movgID][0].append(ACTUAL_T + jct_obj.t_max[jctID])
                        self.movg_green_plan[movgID][1].append(ACTUAL_T + jct_obj.t_max[jctID])
                        self.movg_green_plan[movgID][2].append(6)
                        self.movg_green_plan[movgID][3].append(None)
                        self.movg_green_plan[movgID][4].append(0)
                        self.movg_green_plan[movgID][5].append(0)

        end5 = time.time()
        timings[5][1] += end5 - start5



class Edges:
    """Edges and their characteristics modelled by the TC"""

    # Definition of standard values
    std_lane_max_speed = 13.89 # Maximum standard speed on lane. In m/s. This is 50 km/h
    # Base ave. critical gap based on direction, if minor/major road and major num lanes. In sec
    # Used when calculating connection capacity.
    base_ave_cr_gap = {"2-lanes/left/major": 4.1,
                       "4-lanes/left/major": 4.1,
                       "2-lanes/right/minor": 6.2,
                       "4-lanes/right/minor": 6.9,
                       "2-lanes/through/minor": 6.5,
                       "4-lanes/through/minor": 6.5,
                       "2-lanes/left/minor": 7.1,
                       "4-lanes/left/minor": 7.5}
    std_cr_gap = 4.1 # standard critical gap. Used when comparing each estimated gap when modelling each vehicle
    # Base follow up time queued veh. after a leader based on direction and if minor/major road. In sec
    # Used when calculating connection capacity.
    base_follow_up = {"left/major": 2.2,
                      "right/minor": 3.3,
                      "through/minor": 4.0,
                      "left/minor": 3.5}
    std_follow_up = 2.2 # standard follow up time. Used when comparing each estimated gap when modelling each vehicle
    std_cap = 1       # Std. saturation flow. In veh/s/connection. Default is 3600 veh/h/connection.


    def __init__(self, name):
        """Instances of the edge object that represents the modelled edges, lanes and connections"""

        # lane_ind on the descriptions below represent the lane index on the veh_obj.lane_route of the veh_num
        # veh_num is the index of the vehID in the list of all vehicles veh_obj.ids_all
        self.name = name # object name
        self.edge_ids = []   # list of edges modelled by the TC
        self.lane_ids = []  # list of lanes modelled by the TC
        self.conn_ids = []  # list of connections modelled by the TC
        self.tc_edge_movgs_tt = dict()  # edges and mov. group last travel times. Ex.: {edgeID: {movgID: float}}
        self.tc_tt = dict()  # incoming edge travel times. {in_edgID: (ACTUAL_T, arr_ranges_TT)}
        self.traffic_net_tt = dict()  # edge travel times of the traffic network. {in_edgID: (ACTUAL_T, arr_ranges_TT)}
        self.lane_must_stop = dict()  # define if vehicles must stop at the stop line before yielding {laneID: bool}
        self.lane_arrs_done = dict()  # define if arrivals on that lane were already done {laneID: bool}
        self.stop_est_probes = dict()  # bool to stop estimating probes for the lane {in_laneID: bool}
        self.inLaneStopped_dueOutQueue = dict()  # bool to define that a vehicle on a incoming lane will not move due to
        # its outgoing lane full queue length {in_laneID: bool}
        self.outLane_vehs_atr_arr_range = dict() # arriving attrs of vehs. {out_laneID: [(veh1_arr_t, veh1_arr_speed),
                                                                                       # (veh2_arr_t, veh2_arr_speed)]}
        self.CAVsAlreadyOnLaneBalance = dict() # balance between previously added CAVs that were not replacing
                                               # a non-CAV already on one lane of the edge. Used to replace non-CAVs
                                               # already on lane by CAVs and over generate vehicles {edgeID: int >= 0}
        self.veh_already_disharged = [] # list of vehicles that were already discharged (even if didn't cross stop line)
                                        # at current REAL_PHASE. [(veh_num1,veh1_lane_ind),(veh_num2,veh2_lane_ind)]
        self.lane_discharging_vehs_phase = dict() # define the vehicles expected to discharge at current REAL_PHASE
                                                # per lane {laneID: [(veh_num1,veh1_lane_ind),(veh_num2,veh2_lane_ind)]}
        self.lane_last_discharged_veh_ind = dict() # last discharged veh. from lane_discharging_vehs_phase. Ex:
                                                   # {laneID: veh_ind}
        self.yield2lane_last_oppVeh = dict() # last opposite veh. on a preferencial lane that is needed to yield to
                                             # from lane_discharging_vehs_phase. {in_laneID: {pref_lane: oppVeh_ind}}
        self.edgeID_to_laneIDs = dict() # conversion from edge to lane. {edgeID: [laneID1,laneID2]}
        self.laneID_to_edgeID = dict() # conversion from edge to lane. {laneID: edgeID}
        self.inLaneID_to_connIDs = dict() # connection IDs of the incoming lane {in_laneID: [connID1,connID2]}
        self.outLaneID_to_connIDs = dict() # connection IDs of the outgoing lane {out_laneID: [connID1,connID2]}
        self.lane_length = dict() # lane length. {laneID: float > 0}
        self.edge_accel_adj = dict() # adjustement to be multiplied by the acceleration {edgeID: float > 0}
        self.edge_decel_adj = dict() # adjustement to be multiplied by the acceleration {edgeID: float > 0}
        self.lane_max_speed = dict() # lane maximum speed. {laneID: float > 0}. In m/s
        self.conn_cap = dict() # connection capacity per its movgID phase. {connID: [fixed_cap, [dyn_cap_ph1, dyn_cap_ph2]]}
        self.conn_ave_cr_gap = dict() # connection ave. critical gap (needed pref. traffic gap) in sec. Ex:
                             # {connID: {"adj": [float per veh. type], adjusted according to context (e.g. minor left)
                             # "std": [float per veh. type]}}, standard value specifically for each real estimated gap
        self.conn_follow_up = dict() # connection needed pref. traffic gap if queued after a leader. Ex:
                            # {connID: {"adj": [float per veh. type], adjusted according to context (e.g. minor left)
                            # "std": [float per veh. type]}}, standard value specifically for each real estimated gap
        self.lane_inflow_profile = dict() # define the inflow profile of arriving vehicles
        # params. for the arrival range: num_vehs, mean_speed, mean_hdwy, and if medium flow also std_dev_hdwy
        # Dyn. params. only for border jcts and for each arr_range. {in_laneID: [{interv:[fixed_params]}
        #                                                                       [ACTUAL_T,
        #                                                                       [dyn_params_per_range]]]}
        self.lane_blockage_dist = dict() # distance where lane is blocked and remaining time of closure
        # if not given rem. time (but the distance), it will not calculate the lane travel time. Default [None, None]
        # {in_laneID: [dist_stopline_meters, rem_t_wait_sec]}
        self.closure_affected_edgeLanes = {"edges": set([]), "lanes": set([])} # edges and lanes affected by lane closures
        self.unique_inLane2outEdge = dict() # a lane that is the only predecessor of certain outgoing edge.
                                            # it is used for transfering in_lane queue and out_lane delays
                                            # to the out_edge
                                            # {in_laneID: {out_edgeID1, out_edgeID2,...}}
        self.lane_queue_pred = dict() # prediction of queue length within t_max {in_laneID: [ACTUAL_T,
                                                                                     #  [queue_len_per_range]]}
        self.lane_aval_queue_len = dict()  # calculation of available length of queue per cycle per range and cycle
                                      # {in_laneID: [[aval_len_per_range] per_tmax_cycle]}
        self.lane_queue_tLimitRanges = dict() # end time of lanes queue prediction intervals
        self.lane_cavs_metrics = dict() # lane metrics of CAVs,
                                        # 1) queue length at time measured by CAV read bumber distance to stop line
                                        # 2) last veh. travel time per mov. group of the lane with the time CAV arrived
                                        # and the CAV measured travel time. Ex:
                                        # {in_laneID: {"queue_len": {"by_CAV": dist, "by_LLR_pred": dist}
                                        #              "travel_times": {movgID: (time_CAV_arrived,
                                        #                                        prevTravelTime
                                        #                                        CAVtravelTime)}}}
        self.vehs_on_lane = dict() # vehicles that are and will be on the lane {in_laneID: [(veh_num1,veh1_lane_ind),
                                                                                           # (veh_num2,veh2_lane_ind)]}}
        self.already_on_lane = dict() # only vehicles that are on the lane at the update of the algorithm.
                                                                            # {in_laneID: [(veh_num1,veh1_lane_ind),
                                                                                          # (veh_num2,veh2_lane_ind)]}}
        self.inLaneID_to_movgIDs = dict() # movgIDs of the inc. lane {in_laneID: [movgID1, movgID2]}
        self.outLaneID_to_movgIDs = dict() # movgIDs of the out. lane {out_laneID: [movgID3, movgID4]}
        self.connID_to_jctID = dict() # jctID of the connection {connID: jctID}
        self.connID_to_inLaneID = dict() # in_laneID of the connection {connID: in_laneID}
        self.connID_to_outLaneID = dict() # out_laneID of the connection {connID: out_laneID}
        self.connID_to_movgID = dict() # movgID of the connection {connID: movgID}
        self.lane_num_beg_vehs = dict() # number of vehicles beginning their trip at the lane in one range per distance.
        # when no distance defined, use "random" and 1 for the probability. {laneID: {interv:(float >= 0, ["random"], [1])}}
        # or the distances (meters) and their probability (use same index). {laneID: {interv:(float >= 0, [10,20],[0.5,0.5])}}
        self.last_edg_prob = dict() # probability that a vehicle will finish its trip at the edge {edgeID: {interval: [0,1]}}
        self.next_edg_prob = dict() # each next edge probability from the edge (turning rate)
        # {edgeID: {interval: {next_edgeID: [0,1]}}}. It must sum 1 all probabilities
        # Note that if possible to turn to next edge it must be inserted a key for a next edge even the probability is zero.
        self.next_edg_radius = dict() # each next edge turning radius from the edge (in meters),
                                      # {edgeID: {next_edgeID: float in meters}}.
                                      # if straight, then use float("inf")
        self.conn_length = dict() # connection length (length of SUMO internal lane). {connID: float > 0}. In meters
        self.conn_direction = dict() # connection direction ("r", "s", "l"). {connID: string}
        self.inLaneID_to_jctID = dict() # jctID of the incoming lane {in_laneID: jctID}
        self.outLaneID_to_jctID = dict() # jctID of the outgoing lane {out_laneID: jctID}
        self.conn_phase_deps = dict() # connection departures per movgID and movgID's green phase num
        # {movgID: {connID: [(ph1_veh_num1,veh1_lane_ind),(ph1_veh_num2,veh2_lane_ind)
        #                    (ph2_veh_num3,veh3_lane_ind),(ph2_veh_num4,veh4_lane_ind)]}}
        self.connID_to_yield2connIDs = dict() # other connections the connection needs to yield to. Ex:
                                              # {connID: list conns2yield}


    def setEdgesAttrbInterval(self,
                              ACTUAL_T): # Simulation Time
        """Define the interval for the time-dependent edge/lane attributes"""

        start37 = time.time()

        # Because the settings of turning rates the intervals for next_edg_prob is the same for each edge,
        # it is not needed to define for each edge (but it can be set that is different edge)
        for edgeID in self.next_edg_prob.keys():
            if self.next_edg_prob[edgeID].has_key(ACTUAL_T):
                self.next_edg_prob[edgeID]["interv"] = ACTUAL_T

        for edgeID in self.last_edg_prob.keys():
            if self.last_edg_prob[edgeID].has_key(ACTUAL_T):
                self.last_edg_prob[edgeID]["interv"] = ACTUAL_T

        for in_laneID in self.lane_inflow_profile.keys():
            if self.lane_inflow_profile[in_laneID][0] != () \
            and self.lane_inflow_profile[in_laneID][0].has_key(ACTUAL_T):
                self.lane_inflow_profile[in_laneID][0]["interv"] = ACTUAL_T

        for in_laneID in self.lane_num_beg_vehs.keys():
            if self.lane_num_beg_vehs[in_laneID].has_key(ACTUAL_T):
                self.lane_num_beg_vehs[in_laneID]["interv"] = ACTUAL_T

        end37 = time.time()
        timings[37][1] += end37 - start37


    def setEdgesDicts(self,
                      jct_obj,    # TC junction object
                      movg_obj):  # Mov. group junction object
        """Initialize edge object dictionaries"""

        for movgID in movg_obj.movgID_to_connIDs.keys():
            # Define possible mov. groups for each connection
            self.conn_phase_deps.update({movgID: {conn: [] for conn in movg_obj.movgID_to_connIDs[movgID]}})

        for in_laneID in self.inLaneID_to_movgIDs.keys():
            for movgID in self.inLaneID_to_movgIDs[in_laneID]:
                self.lane_cavs_metrics[in_laneID]["travel_times"].update({movgID: (BEGIN_T + STEP_LENGTH,
                                                                                   np.nan, np.nan)})
                in_edgeID = self.laneID_to_edgeID[in_laneID]
                self.tc_edge_movgs_tt[in_edgeID].update({movgID: np.nan})

        num_ranges = dict()
        all_lanes_neigh = dict()
        for jctID in jct_obj.ids:
            num_ranges.update({jctID: int(jct_obj.t_max[jctID] / LEN_RANGE)})
            all_lanes_neigh.update({jctID: set([])})
            # Define all lanes that are commom with a border junction of other TCs
            for neigh_jct in jct_obj.commonToNeigh_laneIDs[jctID].keys():
                all_lanes_neigh[jctID].update(jct_obj.commonToNeigh_laneIDs[jctID][neigh_jct])
            for neigh_jct in jct_obj.commonFromNeigh_laneIDs[jctID].keys():
                all_lanes_neigh[jctID].update(jct_obj.commonFromNeigh_laneIDs[jctID][neigh_jct])

        for laneID in self.lane_ids:
            try:
                jctID = self.inLaneID_to_jctID[laneID]
                # Initialize queue prediction and its time intervals, and lane_blockage_dist for all incoming lanes
                if laneID in all_lanes_neigh[jctID]:
                    # If common lane with border junction, the intervals of queue prediction is LEN_RANGE
                    self.lane_queue_pred.update({laneID: [BEGIN_T + STEP_LENGTH,
                                                     [0 for rge in range(0, num_ranges[jctID])]]})
                    self.lane_aval_queue_len.update({laneID: [[self.lane_length[laneID]
                                                          for arr_range in range(0, len(self.lane_queue_pred[laneID][1]))]
                                                         for cycle in range(0, MAX_T_MAX_CYCLE)]})
                    self.lane_blockage_dist.update({laneID: [None, None]})
                    self.lane_queue_tLimitRanges.update({laneID: [BEGIN_T + rge * LEN_RANGE
                                                               for rge in range(1, num_ranges[jctID] + 1)]})
                else:
                    # If lane modelled only by the TC, use STEP_LENGTH precision
                    self.lane_queue_pred.update({laneID: [BEGIN_T + STEP_LENGTH, [0 for rge in frange(0,
                                                                                      jct_obj.t_max[jctID],
                                                                                      QUEUE_PRED_RES)]]})
                    self.lane_aval_queue_len.update({laneID: [[self.lane_length[laneID]
                                                          for arr_range in range(0, len(self.lane_queue_pred[laneID][1]))]
                                                         for cycle in range(0, MAX_T_MAX_CYCLE)]})

                    self.lane_blockage_dist.update({laneID: [None, None]})
                    self.lane_queue_tLimitRanges.update({laneID: [BEGIN_T + rge * STEP_LENGTH
                                                               for rge in frange(1, jct_obj.t_max[jctID] + 1,
                                                                                QUEUE_PRED_RES)]})
            except KeyError:
                jctID = self.outLaneID_to_jctID[laneID]
                # Initialize for outgoing lane only if common lane to border junction
                if laneID in all_lanes_neigh[jctID]:
                    self.lane_queue_pred.update({laneID: [BEGIN_T + STEP_LENGTH,
                                                     [0 for rge in range(0, num_ranges[jctID])]]})
                    self.lane_aval_queue_len.update({laneID: [[self.lane_length[laneID]
                                                          for arr_range in range(0, len(self.lane_queue_pred[laneID][1]))]
                                                         for cycle in range(0, MAX_T_MAX_CYCLE)]})
                    self.lane_blockage_dist.update({laneID: [None, None]})
                    self.lane_queue_tLimitRanges.update({laneID: [BEGIN_T + rge * LEN_RANGE
                                                               for rge in range(1, num_ranges[jctID] + 1)]})

        # Create a list of outgoing edges for each lane of certain incoming edge:
        # a) incoming edge has more than one lane
        # b) at least one lane leads to a different out. edge from other lanes (different pattern of next edges per lane)
        # c) outgoing edge receive only from this sending edge (only one incoming edge)
        for edgeID in self.edge_ids:
            if len(self.edgeID_to_laneIDs[edgeID]) > 1:
                global diff_pattern, outEdge_diff_edge
                diff_pattern = 0
                outEdge_diff_edge = 0
                for in_laneID in self.edgeID_to_laneIDs[edgeID]:
                    self.unique_inLane2outEdge.update({in_laneID: set([])})
                    if outEdge_diff_edge == 0:
                        for connID in self.inLaneID_to_connIDs[in_laneID]:
                            out_edgeID = self.laneID_to_edgeID[self.connID_to_outLaneID[connID]]
                            self.unique_inLane2outEdge[in_laneID].add(out_edgeID)
                            out_lanes = self.edgeID_to_laneIDs[out_edgeID]
                            for out_lane in out_lanes:
                                for out_edge_conn in self.outLaneID_to_connIDs[out_lane]:
                                    if self.connID_to_inLaneID[out_edge_conn] not in self.edgeID_to_laneIDs[edgeID]:
                                        # If out_edge has a connection that the incoming lane comes from an edge
                                        # other than edgeID, so in_laneID is not unique to out_edge
                                        outEdge_diff_edge = 1
                                        break
                                if outEdge_diff_edge == 1:
                                    # in_laneID is not unique to out_edge, stop all verifications as it failed to pass
                                    # one of the conditions
                                    break

                            if outEdge_diff_edge == 1:
                                break
                        try:
                            if last_pattern != self.unique_inLane2outEdge[in_laneID]:
                                # When the in_laneID has multiple connections and at least one out. edge is different
                                # then it has a different pattern, continue storing out edges and use diff_pattern = 1
                                # to not delete its out edges
                                diff_pattern = 1
                        except NameError:
                            # If there was not a value for last_pattern, initialize it
                            last_pattern = self.unique_inLane2outEdge[in_laneID]
                    else:
                        break

                if diff_pattern == 0 or outEdge_diff_edge == 1:
                    # If at least one of the conditions of unique in_laneID is not fulfiled,
                    # then delete out edges of in_laneID
                    for in_laneID in self.edgeID_to_laneIDs[edgeID]:
                        del self.unique_inLane2outEdge[in_laneID]

                del last_pattern


    def addEdgeLanes(self,
                     jct_obj,
                     edgeID,                                   # network edge ID
                     laneIDs,                                  # network lane IDs
                     edge_type=None,                           # incoming or outgoing
                     lanes_length=None,                        # in Meters
                     lane_num_beg_vehs=tuple(),                # {interv:(num_veh_arr_range, list_dists, list_probs)}
                     last_edg_prob=None,                       # probability vehicle will finish its trip on edgeID
                                                               # {interval: [0,1]}
                     next_edg_prob=None,                       # next edges probabilities (turning rates from edgeID)
                                                               # {interval: {next_edgeID: [0,1]}}
                     next_edg_radius=None,                     # next edges turning radius from edgeID
                                                               # {interval: {next_edgeID: float in meters}}
                     edge_accel_adj=STD_ACCEL_ADJUST,          # edge acceleration adjustment
                     edge_decel_adj=STD_DECEL_ADJUST,          # edge deceleration adjustment
                     lanes_max_speed = std_lane_max_speed,     # max speed in m/s
                     lane_must_stop = 0,                       # bool if lane has obligatory stop sign
                     lanes_inflow_profile = [tuple(),tuple()], # [{interv:[fixed_params]}, tuple()]
                                                               # params.: num_vehs, mean_speed, mean_hdwy,
                                                               # and if medium flow also std_dev_hdwy
                     add_neigh_tc_bool=0): # If 1, the info is regarding a neighbouring border junction, 0 otherwise
        """Add an edge and its lanes to the edge object"""

        if edgeID not in self.edge_ids:
            if add_neigh_tc_bool == 0:
                self.edge_ids.append(edgeID)
                self.tc_edge_movgs_tt.update({edgeID: dict()})
                self.edge_accel_adj.update({edgeID: edge_accel_adj})
                self.edge_decel_adj.update({edgeID: edge_decel_adj})
                self.CAVsAlreadyOnLaneBalance.update({edgeID: 0})
                self.next_edg_prob.update({edgeID: next_edg_prob})
                self.last_edg_prob.update({edgeID: last_edg_prob})
                self.next_edg_radius.update({edgeID: next_edg_radius})
            self.edgeID_to_laneIDs.update({edgeID: []})
            for lane_ind,laneID in enumerate(laneIDs):
                if add_neigh_tc_bool == 0:
                    self.lane_ids.append(laneID)
                    self.vehs_on_lane.update({laneID: []})
                self.lane_arrs_done.update({laneID: 0})
                self.edgeID_to_laneIDs[edgeID].append(laneID)
                self.laneID_to_edgeID.update({laneID: edgeID})
                self.inLaneID_to_connIDs.update({laneID: []})
                self.outLaneID_to_connIDs.update({laneID: []})
                if add_neigh_tc_bool == 0:
                    try:
                        self.lane_length.update({laneID: lanes_length[lane_ind]})
                    except TypeError:
                        self.lane_length.update({laneID: lanes_length})
                    try:
                        self.lane_max_speed.update({laneID: lanes_max_speed[lane_ind]})
                    except TypeError:
                        self.lane_max_speed.update({laneID: lanes_max_speed})
        else:
            laneIDs = self.edgeID_to_laneIDs[edgeID]

        if add_neigh_tc_bool == 0 and "incoming" in edge_type:
            for laneID in laneIDs:
                self.inLaneID_to_connIDs.update({laneID: []})
                self.stop_est_probes.update({laneID: 0})
                self.inLaneStopped_dueOutQueue.update({laneID: 0})
                self.inLaneID_to_jctID.update({laneID: None})
                self.lane_num_beg_vehs.update({laneID: lane_num_beg_vehs})
                self.lane_inflow_profile.update({laneID: lanes_inflow_profile})
                self.already_on_lane.update({laneID: []})
                self.inLaneID_to_movgIDs.update({laneID: set()})
                self.lane_last_discharged_veh_ind.update({laneID: None})
                self.yield2lane_last_oppVeh.update({laneID: []})
                self.lane_discharging_vehs_phase.update({laneID: []})
                self.lane_must_stop.update({laneID: lane_must_stop})
                if "Mod. Non-C." in jct_obj.jctType[jctID]:
                    self.lane_cavs_metrics.update({laneID: {"queue_len": {"by_CAV": None, "by_LLR_pred": None},
                                                            "travel_times": dict()}})
                else:
                    self.lane_cavs_metrics.update({laneID: {"travel_times": dict()}})
        if add_neigh_tc_bool == 0 and "outgoing" in edge_type:
            for laneID in laneIDs:
                self.outLaneID_to_connIDs.update({laneID: []})
                self.outLaneID_to_movgIDs.update({laneID: set()})
                self.outLaneID_to_jctID.update({laneID: None})


    def addConnection(self,
                      veh_obj,           # TC vehicles object
                      connID,            # network Connection ID, in SUMO it is the "via" attribute
                      in_laneID,         # respective network lane ID starting from in_laneID
                      out_laneID,        # respective network lane ID finishing at out_laneID
                      jct_obj=None,      # TC junction object
                      dir=None,          # TC connection direction ("r", "s", "l")
                      jctID=None,        # respective network junction ID between in_laneID and out_laneID
                      length = 0,        # length of the connection in meters
                      cap = std_cap,     # fixed sat. flow, also for left-turns (without conflict flow), in veh/s
                      adj_ave_cr_gaps=std_cr_gap,     # list/tuple with adjusted (for capacity calc.) ave critical gap
                                                      # in sec for each index of vehicle type
                      adj_follow_up_t=std_follow_up,  # list/tuple with adjusted (for capacity calc.) follow up time
                                                      # in sec for each index of vehicle type
                      std_ave_cr_gaps=std_cr_gap,     # list/tuple with standard (to compare with each estimated gap)
                                                      # ave critical gap in sec for each index of vehicle type
                      std_follow_up_t=std_follow_up,  # list/tuple with standard (to compare with each estimated gap)
                                                      # follow up time in sec for each index of vehicle type
                      add_neigh_tc_bool = 0): # If 1, the info is regarding a neighbouring border junction, 0 otherwise
        """Add a connection to the edge object"""

        self.inLaneID_to_connIDs[in_laneID].append(connID)
        self.outLaneID_to_connIDs[out_laneID].append(connID)
        self.connID_to_inLaneID.update({connID: in_laneID})
        self.connID_to_outLaneID.update({connID: out_laneID})
        self.connID_to_movgID.update({connID: None})
        if add_neigh_tc_bool == 0:
            self.conn_ids.append(connID)
            self.connID_to_jctID.update({connID: jctID})
            self.conn_length.update({connID: length})
            self.conn_direction.update({connID: dir})
            self.conn_cap.update({connID: [cap, []]}) # index 0 is fixed capacity while index 1 is dynamic capacity
            self.conn_ave_cr_gap.update({connID: {"adj": None, "std": None}})
            self.conn_follow_up.update({connID: {"adj": None, "std": None}})
            if type(adj_ave_cr_gaps) == list or type(adj_ave_cr_gaps) == tuple:
                # If inputed list of different values for each vehicle type
                self.conn_ave_cr_gap[connID]["adj"] = adj_ave_cr_gaps
            else:
                # If adj_ave_cr_gaps is a float/int value, assume this value for all vehicle types
                self.conn_ave_cr_gap[connID]["adj"] = [adj_ave_cr_gaps] * len(veh_obj.vType_vals["ids"])
            if type(adj_follow_up_t) == list or type(adj_follow_up_t) == tuple:
                # If inputed list of different values for each vehicle type
                self.conn_follow_up[connID]["adj"] = adj_follow_up_t
            else:
                # If adj_follow_up_t is a float/int value, assume this value for all vehicle types
                self.conn_follow_up[connID]["adj"] = [adj_follow_up_t] * len(veh_obj.vType_vals["ids"])

            if type(std_ave_cr_gaps) == list or type(std_ave_cr_gaps) == tuple:
                # If inputed list of different values for each vehicle type
                self.conn_ave_cr_gap[connID]["std"] = std_ave_cr_gaps
            else:
                # If std_ave_cr_gaps is a float/int value, assume this value for all vehicle types
                self.conn_ave_cr_gap[connID]["std"] = [std_ave_cr_gaps] * len(veh_obj.vType_vals["ids"])
            if type(std_follow_up_t) == list or type(std_follow_up_t) == tuple:
                # If inputed list of different values for each vehicle type
                self.conn_follow_up[connID]["std"] = std_follow_up_t
            else:
                # If std_follow_up_t is a float/int value, assume this value for all vehicle types
                self.conn_follow_up[connID]["std"] = [std_follow_up_t] * len(veh_obj.vType_vals["ids"])
            self.inLaneID_to_jctID.update({in_laneID: jctID})
            self.outLaneID_to_jctID.update({out_laneID: jctID})
            jct_obj.connIDs[jctID].append(connID)
            jct_obj.in_laneIDs[jctID].add(in_laneID)
            jct_obj.out_laneIDs[jctID].add(out_laneID)
            in_edgeID = self.laneID_to_edgeID[in_laneID]
            jct_obj.in_edgeIDs[jctID].add(in_edgeID)
            out_edgeID = self.laneID_to_edgeID[out_laneID]
            jct_obj.out_edgeIDs[jctID].add(out_edgeID)


    def addConnYield2Conns(self,
                           movg_obj, # TC mov. group object
                           connID, # network connection ID
                           yield2conns=[]): # other connIDs this connection needs to yield (in case of permissive green)
        """Define the connIDs and movgIDs the connID and its movgIDd need to yield when permissive green"""

        movgID = self.connID_to_movgID[connID]
        yield2movgs = set([])
        if yield2conns != []:
            for conn in yield2conns:
                yield2movgs.add(self.connID_to_movgID[conn])
            # If movement group has conflicts, then define connections that it needs to yield.
            movg_obj.movgID_to_yield2movgIDs.update({movgID: yield2movgs})  # Mov. Group to Yield
            self.connID_to_yield2connIDs.update({connID: yield2conns})
        else:
            movg_obj.movgID_to_yield2movgIDs.update({movgID: set([])})
            self.connID_to_yield2connIDs.update({connID: []})


    def setClosedLanes(self,
                       closed_lanes,   # List with the IDs of the lanes that are being closed
                       dist_block):    # Distance of the closure
        """Define the lanes that were closed and also the lane and edges (if all lanes) that no vehicle
        will take because they only lead to a closed lane"""

        global continue_find_affected_edges, continue_find_affected_lanes
        continue_find_affected_edges = 0
        continue_find_affected_lanes = 0

        for lane2close in closed_lanes:
            try:
                if dist_block == -1:
                    closing_dist = self.lane_length[lane2close]
                else:
                    closing_dist = dist_block
                self.lane_blockage_dist[lane2close][0] = closing_dist
            except KeyError:
                skip = 1

        blocked_edges = set([])
        for lane2close in closed_lanes:
            try:
                blocked_edges.add(self.laneID_to_edgeID[lane2close])
                # Add closed lane as affected
                self.closure_affected_edgeLanes["lanes"].add(lane2close)
                continue_find_affected_lanes = 1
            except KeyError:
                skip = 1

        for blocked_edge in blocked_edges:
            lanes_blocked = []
            for parallel_lane in self.edgeID_to_laneIDs[blocked_edge]:
                if parallel_lane in closed_lanes:
                    lanes_blocked.append(1)
            if len(self.edgeID_to_laneIDs[blocked_edge]) == len(lanes_blocked):
                # If all lanes of the blocked_edge are closed, then add edge as affected by the closure
                self.closure_affected_edgeLanes["edges"].add(blocked_edge)
                continue_find_affected_edges = 1

        # Recursively try to find edges that send vehicles only to edges already set as affected by closure
        # The closure_affected_edgeLanes is useful to not include in vehicle's route edges and lanes in this list
        while continue_find_affected_edges == 1:
            continue_find_affected_edges = 0
            edges_to_consider = [edge for edge in self.next_edg_prob.keys()
                                 if edge not in self.closure_affected_edgeLanes["edges"]]

            for edgeID in edges_to_consider:
                interv_next = edg_objs[tc_ind].next_edg_prob[edgeID]["interv"]
                try:
                    possible_next_edges = self.next_edg_prob[edgeID][interv_next].keys()
                    edges_lead_affect_closed = [edge for edge in possible_next_edges
                                                if edge in self.closure_affected_edgeLanes["edges"]]
                    if len(possible_next_edges) == len(edges_lead_affect_closed):
                        # If all possible next edges from edgeID are set as affected by closure
                        # Then the edgeID will be also affected by the closure of other lanes
                        self.closure_affected_edgeLanes["edges"].add(edgeID)
                        self.closure_affected_edgeLanes["lanes"].update(self.edgeID_to_laneIDs[edgeID])
                        continue_find_affected_edges = 1
                except KeyError:
                    skip = 1

        while continue_find_affected_lanes == 1:
            # Add lanes that only lead to the closed lane as affected also (vehicles will change lane or rarely use them)
            continue_find_affected_lanes = 0
            new_affected_lanes = []
            for affected_lane in self.closure_affected_edgeLanes["lanes"]:
                try:
                    lanesLeading2affected = [self.connID_to_inLaneID[connLeading2affected]
                                             for connLeading2affected in self.outLaneID_to_connIDs[affected_lane]
                                             if self.connID_to_inLaneID[connLeading2affected]
                                             not in self.closure_affected_edgeLanes["lanes"]]
                    for laneLeading2affected in lanesLeading2affected:
                        for connID in self.inLaneID_to_connIDs[laneLeading2affected]:
                            if self.connID_to_outLaneID[connID] not in self.closure_affected_edgeLanes["lanes"] \
                            and self.connID_to_outLaneID[connID] not in new_affected_lanes:
                                break
                        else:
                            # If all possible next lanes of laneLeading2affected (lane leading to a already set
                            # affected lane by closure) are also affected, so the laneLeading2affected lane will be
                            # also affected
                            new_affected_lanes.append(laneLeading2affected)
                            continue_find_affected_lanes = 1
                except KeyError:
                    skip = 1

            self.closure_affected_edgeLanes["lanes"].update(new_affected_lanes)


    def clearClosedLanes(self,
                         opened_lanes):  # List with the IDs of the lanes that are being reopen
        """Define the lanes and edges that vehicles will take again because lead to a recently opened lane"""

        global continue_find_affected_edges, continue_find_affected_lanes
        continue_find_affected_edges = 0
        continue_find_affected_lanes = 0

        for lane2close in LANES2CLOSE:
            try:
                self.lane_blockage_dist[lane2close][0] = None
            except KeyError:
                skip = 1

        new_unaffected_lanes = set([])
        unblocked_edges = set([])
        for lane2open in opened_lanes:
            try:
                self.closure_affected_edgeLanes["edges"].remove(self.laneID_to_edgeID[lane2open])
                self.closure_affected_edgeLanes["lanes"].remove(lane2open)
                new_unaffected_lanes.add(lane2open)
                unblocked_edges.add(self.laneID_to_edgeID[lane2open])
                continue_find_affected_edges = 1
                continue_find_affected_lanes = 1
            except KeyError:
                skip = 1

        # Recursively try to find edges that vehicles may use again because edges will not be affected by any closure
        while continue_find_affected_edges == 1:
            continue_find_affected_edges = 0
            edges_to_consider = self.closure_affected_edgeLanes["edges"].copy()
            for edgeID in edges_to_consider:
                try:
                    interv_next = edg_objs[tc_ind].next_edg_prob[edgeID]["interv"]
                    possible_next_edges = set(self.next_edg_prob[edgeID][interv_next].keys())
                    if possible_next_edges.intersection(unblocked_edges) != set():
                        self.closure_affected_edgeLanes["edges"].remove(edgeID)
                        unblocked_edges.add(edgeID)
                        continue_find_affected_edges = 1
                except KeyError:
                    skip = 1

        while continue_find_affected_lanes == 1:
            # Removes lanes that only lead to the closed lanes as affected also (vehicles will change lane or rarely use them)
            continue_find_affected_lanes = 0
            lanes_to_consider = self.closure_affected_edgeLanes["lanes"].copy()
            for affected_lane in lanes_to_consider:
                try:
                    for connID in self.inLaneID_to_connIDs[affected_lane]:
                        for possible_next_lane in self.connID_to_outLaneID[connID]:
                            if possible_next_lane in new_unaffected_lanes:
                                new_unaffected_lanes.add(affected_lane)
                                self.closure_affected_edgeLanes["lanes"].remove(affected_lane)
                                continue_find_affected_lanes = 1
                                break
                except KeyError:
                    skip = 1


    def estDelayConnCap(self,
                        jct_obj,        # TC junctions object
                        movg_obj,       # TC mov. groups object
                        veh_obj,        # TC vehicles object
                        jctID,          # network junction of the connection or mov. group
                        cycle,          # number of cycles ahead to be extrapolated from actual t_max limitation
                        phase_num,      # phase index (number) of the movg of conn at the departure
                        endGreen_t,     # end time of green phase of the movg of conn
                        stopLine_arr_t, # stop line arriving time
                        ref_hdwy_t,     # headway with the vehicle ahead or begin green time (if first veh), in sec
                        conn=None,      # connection ID to be estimated the delay. If not given, input movg and laneID
                        movg=None,      # network movgID to select which connIDs to estimate the average delay
                        laneID=None):   # network in_laneID to filter connections that starts on this lane
        """"Estimate the delay that a vehicle will experience when taking a connection,
        or the average delay considering each connection with inc. lane as laneID and movg"""

        start54 = time.time()

        if movg != None:
            # If not specified the connection (i.e. estimating delay for the queueing model)
            edgeID = self.laneID_to_edgeID[laneID]
            interv_next = self.next_edg_prob[edgeID]["interv"]
            connIDs = [conn for conn in movg_obj.movgID_to_connIDs[movg]
                       if laneID == self.connID_to_inLaneID[conn]
                       and self.connID_to_outLaneID[conn] not in self.closure_affected_edgeLanes["lanes"]
                       and self.next_edg_prob[edgeID][interv_next][self.laneID_to_edgeID[self.connID_to_outLaneID[conn]]]
                       != 0]
            conn_prob = [self.next_edg_prob[edgeID][interv_next][self.laneID_to_edgeID[self.connID_to_outLaneID[conn]]]
                         for conn in connIDs]
            discount_maxCap = 1
        else:
            # If specified the connection (i.e. estimating delay for the traffic model extrapolating the plannig horizon)
            connIDs = [conn]
            conn_prob = [1]
            discount_maxCap = 0

        conns_pref_wait = []
        for connID in connIDs:
            if self.conn_cap[connID][1][phase_num] == []:
                # If dynamic capacity not defined for the phase_num, update it
                self.updateDynConnCapacity(jct_obj, movg_obj, veh_obj, connID)

            if self.conn_cap[connID][1][phase_num] >= 0:
                try:
                    # If opposite flow is below high flow, the waiting time is 1/cap
                    conns_pref_wait.append((1. / self.conn_cap[connID][1][phase_num]))
                except (ZeroDivisionError, FloatingPointError):
                    # No capacity, it will delay until end of green
                    conns_pref_wait.append(max(endGreen_t - stopLine_arr_t, 0))
                # The delay needs to consider the headway with the vehicle ahead or begin of green time (if first vehicle)
                interv_deps = max(stopLine_arr_t - (ref_hdwy_t - discount_maxCap * (1. / self.conn_cap[connID][0])), 0)
                conns_pref_wait[-1] = max(conns_pref_wait[-1] - interv_deps, 0)
            else:
                # If preferential flow is above high flow, the waiting time is the time the last vehicle will discharge
                # The minus capacity means the flow is high and it has returned the time the last opposite vehicle
                # on the preferencial lane will have its length crossed the stop line
                interv_deps = max(((-self.conn_cap[connID][1][phase_num]) + jct_obj.t_max[jctID] * cycle)
                                  - stopLine_arr_t, 0)
                conns_pref_wait.append(interv_deps)

        end54 = time.time()
        timings[54][1] += end54 - start54

        if conns_pref_wait:
        
            try:
                return np.average(conns_pref_wait, weights=conn_prob)
            except (FloatingPointError, ZeroDivisionError) as e:
                LOGGER.error('cannot compute weighted average', exc_info=True)
                LOGGER.debug('conns_pref_wait=' + str(conns_pref_wait))
                try:
                    return np.mean(conns_pref_wait)
                except FloatingPointError as e:
                    LOGGER.error('neither mean can be computed', exc_info=True)
                    return 100000.0
        else:
            # No preferred connection wait means no delay
            return 0.0


    def updateDynConnCapacity(self,
                              jct_obj, # TC junction object
                              movg_obj, # TC movement group object
                              veh_obj, # TC vehicles object
                              connID): # network connection ID
        """Calculate the dynamic capacity of the connection when yielding to preferential lanes.
        When more than 1 lane, returns the lowest capacity found amongst the preferential lanes"""

        start6 = time.time()

        global capacity, preferential_flow
        movgID = self.connID_to_movgID[connID]
        jctID = movg_obj.jctIDs[movgID]
        in_laneID = self.connID_to_inLaneID[connID]
        out_laneID = self.connID_to_outLaneID[connID]
        in_edgID = self.laneID_to_edgeID[in_laneID]
        out_edgID = self.laneID_to_edgeID[out_laneID]
        num_green_phases = len(movg_obj.movg_green_plan[movgID][2])
        if num_green_phases > 1:
            # Don't take non-crossing phase (last one)
            num_green_phases -= 1

        for phase in range(0,num_green_phases):
            capacity = self.conn_cap[connID][0]  # fixed capacity (without yielding)
            if movg_obj.movgID_to_yield2movgIDs[movgID] == set([]) or movg_obj.movg_green_plan[movgID][2][phase] == 6:
                # If movement group of the connection don't need to yield
                # or it has protected green on the analysing phase the capacity is the fixed one
                pass
            else:
                for conflt_movg in movg_obj.movgID_to_yield2movgIDs[movgID]:
                    # For each conflicting movement group
                    for conflt_phase in range(0,len(movg_obj.movg_green_plan[conflt_movg][2])):
                        if (movg_obj.movg_green_plan[conflt_movg][1][conflt_phase] >
                            movg_obj.movg_green_plan[movgID][0][phase]) and \
                           (movg_obj.movg_green_plan[conflt_movg][0][conflt_phase] <
                            movg_obj.movg_green_plan[movgID][1][phase]):
                            # If end of green time of conflt_movg is after begin of permissive green of movgID
                            yield2conns = (y2conn for y2conn in self.connID_to_yield2connIDs[connID]
                                           if self.connID_to_movgID[y2conn] == conflt_movg)
                            deps_per_dur = {"num_vehs": [], "phase_dur": []}
                            yield_conns_mean_speed = []
                            preferential_flow = 0
                            for conn2yield in yield2conns:
                                try:
                                    # Flow is the number of vehicles that crossed the stop line by the green duration
                                    num_vehs = []
                                    for veh_num,lane_ind in self.conn_phase_deps[conflt_movg][conn2yield][conflt_phase]:
                                        num_vehs.append(veh_obj.len[veh_num] / veh_obj.vType_vals["length"][0])
                                        yield_conns_mean_speed.append(veh_obj.phase_end_speed[veh_num][lane_ind])

                                    if movg_obj.movg_green_plan[conflt_movg][1][conflt_phase] == float("inf"):
                                        # If unsignalized junction, phase duration is t_max
                                        phase_dur = float(jct_obj.t_max[jctID])
                                    else:
                                        # If ignalized junction, get movgID phase duration
                                        phase_dur = float(movg_obj.movg_green_plan[conflt_movg][1][conflt_phase] -
                                                          movg_obj.movg_green_plan[conflt_movg][0][conflt_phase])

                                    deps_per_dur["num_vehs"].append(sum(num_vehs))
                                    deps_per_dur["phase_dur"].append(phase_dur)
                                except (ZeroDivisionError, FloatingPointError):
                                    skip = 1

                            try:
                                preferential_flow += sum(deps_per_dur["num_vehs"]) / sum(deps_per_dur["phase_dur"])
                            except:
                                preferential_flow = 0
                            if preferential_flow > MIN_FLOW:
                                # If flow is too low, assume no difference of capacity, so if above minimum flow
                                speed_adj = 0
                                # Average critical gap of all vehicle types
                                mean_vehs_ave_cr_gap = round(np.average(self.conn_ave_cr_gap[connID]["adj"],
                                                                        weights=veh_obj.vType_vals["probability"]), 1)
                                if mean_vehs_ave_cr_gap > 0:
                                    # Adjust critical gap based on preferential traffic mean speed
                                    # (base values are for 50 km/h so any lower/higher speeds require different gaps)
                                    if self.lane_must_stop[in_laneID] == 1:
                                        # For a difference of 50 km/h = 13.89 m/s increase 0.5 seconds
                                        conn_mean_speed = 0
                                        speed_adj += 0.5
                                    else:
                                        # Adjust critical gap based on subject traffic mean speed
                                        # (base values are for 50 km/h so any lower/higher speeds require different gaps)
                                        conn_mean_speed = np.nanmean([veh_obj.phase_end_speed[veh_num][lane_ind]
                                                                      for veh_num,lane_ind
                                                                      in self.conn_phase_deps[movgID][connID][phase]])
                                        if np.isnan(conn_mean_speed):
                                            conn_turning_speed = math.sqrt(FRICTION_COEF * GRAVITY_ACC
                                                                           * self.next_edg_radius[in_edgID][out_edgID])
                                            conn_mean_speed = min(conn_turning_speed, self.lane_max_speed[in_laneID])

                                        # For a difference of 50 km/h = 13.89 m/s increase 0.5 seconds
                                        speed_adj += ((13.89 - conn_mean_speed) * 0.5) / 13.89

                                    yield_conns_mean_speed = np.nanmean(yield_conns_mean_speed)
                                    if np.isnan(yield_conns_mean_speed):
                                        # No vehicle to yield, so don't adjust
                                        pass
                                    else:
                                        if self.conn_direction[connID] == "r":
                                            # If left or right turn movements
                                            if conn_mean_speed == 0:
                                                # If vehicle at subject connection must stop, change 1 second per 40 km/h
                                                speed_adj += ((yield_conns_mean_speed - 13.89) * 1) / 11.11
                                            else:
                                                speed_adj += ((yield_conns_mean_speed - 13.89) * 0.5) / 11.11
                                        else:
                                            # If left turn or through movements
                                            if conn_mean_speed == 0:
                                                # If vehicle at subject connection must stop, change 1.5 second per 40 km/h
                                                speed_adj += ((yield_conns_mean_speed - 13.89) * 1.5) / 11.11
                                            else:
                                                speed_adj += ((yield_conns_mean_speed - 13.89) * 1) / 11.11

                                    mean_vehs_ave_cr_gap += speed_adj

                                    # Average critical gap and follow up time is the average of all vehicle types
                                    mean_vehs_follow_up = round(np.average(self.conn_follow_up[connID]["adj"],
                                                                           weights=veh_obj.vType_vals["probability"]), 1)
                                    if preferential_flow < LOW_FLOW:
                                        # Assuming Negative Exponential Distribution
                                        capacity = min(capacity,
                                                       (preferential_flow
                                                        * math.exp((- preferential_flow) * mean_vehs_ave_cr_gap)) /
                                                       float(1 - math.exp((- preferential_flow) * mean_vehs_follow_up)))
                                    elif preferential_flow < MID_FLOW:
                                        # Assuming a Dichotomized Distribution (some vehicles in platoon and some not)
                                        capacity = min(capacity,
                                                   (1 - preferential_flow * MIN_HDWY) * ((preferential_flow * math.exp(
                                                    (- preferential_flow) * (mean_vehs_ave_cr_gap - MIN_HDWY))) /
                                                    (1 - math.exp((- preferential_flow) * mean_vehs_follow_up))))
                                    else:
                                        # If high flow, the capacity will return the time
                                        # the last vehicle will cross the junction (negative time)
                                        for conn2yield in yield2conns:
                                            last_veh_dep,lane_ind = self.conn_phase_deps[conflt_movg] \
                                                                                        [conn2yield][conflt_phase][-1]
                                            capacity = min(-veh_obj.len_crossed_t[last_veh_dep][lane_ind], capacity)

                        else:
                            # No overlap of signal timings between movgID and conflt_movg
                            pass

            # The final capacity is constrained by the lowest capacity of the movement's group connection
            self.conn_cap[connID][1][phase] = capacity

        end6 = time.time()
        timings[6][1] += end6 - start6


    def adjQueuePred(self,
                     jct_obj): # TC junctions object
        """Adjust the queue prediction and time ranges with the ACTUAL_T"""

        start55 = time.time()

        for laneID in self.lane_queue_pred.keys():
            if self.lane_queue_pred[laneID][0] != ACTUAL_T:
                # If reference time of lane queue prediction is different from actual time,
                # move values to match with the time intervals the ranges are valid, like a stack
                self.lane_queue_pred[laneID][0] = ACTUAL_T
                self.lane_queue_pred[laneID][1].append(self.lane_queue_pred[laneID][1].pop(0))
                for cycle in range(0, MAX_T_MAX_CYCLE):
                    self.lane_aval_queue_len[laneID][cycle].append(self.lane_aval_queue_len[laneID][cycle].pop(0))
            try:
                # Get junction based on the lane
                jctID = self.inLaneID_to_jctID[laneID]
            except KeyError:
                # If lane is only outgoing lane of the modelled lanes by the traffic controller
                jctID = self.outLaneID_to_jctID[laneID]
            num_queue_ranges = len(self.lane_queue_tLimitRanges[laneID])
            if num_queue_ranges == (jct_obj.t_max[jctID] / QUEUE_PRED_RES):
                # If lane modelled by the traffic controller, use QUEUE_PRED_RES as resolution
                self.lane_queue_tLimitRanges.update({laneID: [ACTUAL_T + rge * QUEUE_PRED_RES
                                                           for rge in range(1, num_queue_ranges + 1)]})
            else:
                # If lane not modelled by the traffic controller, use LEN_RANGE as resolution
                self.lane_queue_tLimitRanges.update({laneID: [ACTUAL_T + rge * LEN_RANGE
                                                           for rge in range(1, num_queue_ranges + 1)]})

        end55 = time.time()
        timings[55][1] += end55 - start55


    def getVehAlreadyOnLane(self,
                            jct_obj, # TC junctions object
                            veh_obj): # TC vehicles object
        """Get the expected vehicles running on the incoming lane at ACTUAL_T given estimation of
        position, speed and time from last arrival range (update)"""

        start7 = time.time()

        for jctID in jct_obj.ids:
            for in_laneID in jct_obj.in_laneIDs[jctID]:
                sum_Already = 0
                copied_already_on_lane = self.already_on_lane[in_laneID][:]
                for veh_num, lane_ind in copied_already_on_lane:
                    if "CAV" not in veh_obj.vClass[veh_num]:
                        # CAVs are updated on function edg_obj.genCAVOnLane
                        in_edgeID = self.laneID_to_edgeID[in_laneID]
                        if self.CAVsAlreadyOnLaneBalance[in_edgeID] > 0:
                            # If more CAVs were added than already on lane, remove already on lane vehicles
                            self.CAVsAlreadyOnLaneBalance[in_edgeID] -= 1
                            veh_obj.vehs2del.add(veh_num)
                            for l_ind, laneID in enumerate(veh_obj.lane_route[veh_num]):
                                try:
                                    self.vehs_on_lane[laneID].remove((veh_num, l_ind))
                                except ValueError:
                                    veh_not_on_lane = 1
                                try:
                                    self.already_on_lane[laneID].remove((veh_num, l_ind))
                                except (ValueError, KeyError):
                                    veh_not_on_lane = 1

                            # For Counting Number of Vehicles
                            if NOGUI == False:
                                tStep_area_AlreadyNotGenCAVBalance.append(1)

                        else:
                            # For Counting Number of Vehicles
                            if NOGUI == False:
                                sum_Already += 1

                            # The position, speed, and time the vehicle will start for the first phase is
                            # the estimated time by the last arrival range
                            veh_obj.edge_route[veh_num] = veh_obj.edge_route[veh_num][lane_ind:]
                            veh_obj.phase_start_dist[veh_num][lane_ind] = veh_obj.update_dist[veh_num][lane_ind]
                            veh_obj.phase_start_speed[veh_num][lane_ind] = veh_obj.update_speed[veh_num][lane_ind]
                            veh_obj.phase_start_t[veh_num][lane_ind] = veh_obj.update_t[veh_num][lane_ind]
                            veh_obj.delAttrPrevRouteEdgs(jct_obj, self, veh_num, veh_obj.vClass[veh_num], lane_ind)

                # For Counting Number of Vehicles
                if NOGUI == False:
                    tStep_area_alreadyOnLane.append((in_laneID, sum_Already))


        end7 = time.time()
        timings[7][1] += end7 - start7


    def updateDynLaneInflowProfile(self,
                                   jct_obj,     # TC junctions object
                                   lanes_order, # list of lanes following the order of junctions to be explored
                                   jctID):      # network junction ID
        """If reason is store, store vehicle arrivals and speeds on outgoing lanes.
        If reason is retrieve, defines the outflow profile on each outgoing lane with common lane of the jctID"""

        start9 = time.time()

        global last_v_ind

        # Initialize
        inflow_msg = dict()
        num_ranges = int(jct_obj.t_max[jctID] / LEN_RANGE)
        interval_arr_range = [[] for _ in range(0, num_ranges)]
        ranges_t_limit = [ACTUAL_T + arr_range * LEN_RANGE for arr_range in
                          range(1, num_ranges + 1)]
        all_dest_jcts = [jct for jct in jct_obj.neigh_jcts[jctID]
                         if jct_obj.commonToNeigh_laneIDs[jctID].has_key(jct)]
        for dest_jctID in all_dest_jcts:
            inflow_msg.update({dest_jctID: {outLane: [ACTUAL_T, interval_arr_range[:]]
                                            for outLane in jct_obj.commonToNeigh_laneIDs[jctID][dest_jctID]
                                            if outLane in lanes_order}})
            for out_laneID in inflow_msg[dest_jctID].keys():
                # Sort vehicles by their outgoing lane arrival time on the retrieve part it takes the headways
                sorted_outLane_vehs_atr_arr_range = sorted(self.outLane_vehs_atr_arr_range[out_laneID],
                                                           key=lambda sort_it: (sort_it[1]))
                # Initializing
                last_v_ind = -1
                self.outLane_vehs_atr_arr_range[out_laneID] = [[] for _ in range(0, num_ranges)]
                num_veh_arr_range = [[] for _ in range(0, num_ranges)]
                flow_arr_range = [[] for _ in range(0, num_ranges)]
                mean_speed_arr_range = [[] for _ in range(0, num_ranges)]
                mean_hdwy_arr_range = [[] for _ in range(0, num_ranges)]
                std_dev_hdwy_arr_range = [[] for _ in range(0, num_ranges)]
                for arr_range in range(0, num_ranges):
                    # If arrival on outgoing lane is before the end of the arrival range,
                    # append vehicle to arrivals during the range
                    for last_v_ind, (veh_num, veh_arr_t, veh_arr_speed) \
                    in enumerate(sorted_outLane_vehs_atr_arr_range[(last_v_ind + 1):]):
                        if veh_arr_t < ranges_t_limit[arr_range]:
                            self.outLane_vehs_atr_arr_range[out_laneID][arr_range].append((veh_arr_t, veh_arr_speed))
                        else:
                            break
                    # Number of vehicles to be sent during the arr. range
                    num_veh_arr_range[arr_range] = len(self.outLane_vehs_atr_arr_range[out_laneID][arr_range])
                    if num_veh_arr_range[arr_range] > 0:
                        all_hdwys = []
                        # Headway of first vehicle in the arr. range must be compared to the begin of arr. range
                        all_hdwys.append(self.outLane_vehs_atr_arr_range[out_laneID][arr_range][0][0] -
                                         (ACTUAL_T + (arr_range * LEN_RANGE)))
                        # For the others, compare their arrival on next edge
                        for veh_ind,_ in enumerate(self.outLane_vehs_atr_arr_range[out_laneID][arr_range][1:]):
                            all_hdwys.append(self.outLane_vehs_atr_arr_range[out_laneID][arr_range][veh_ind + 1][0]
                                             - self.outLane_vehs_atr_arr_range[out_laneID][arr_range][veh_ind][0])
                        # mean headway of vehicles during arr. range.
                        mean_hdwy_arr_range[arr_range] =  np.mean(all_hdwys)
                        # standard deviation of headways of vehicles during arr. range
                        std_dev_hdwy_arr_range[arr_range] = np.std(all_hdwys)
                        # average speed of vehicles during arr. range.
                        mean_speed_arr_range[arr_range] = np.mean([veh_attrs[1]
                                                                   for veh_attrs
                                                                   in self.outLane_vehs_atr_arr_range
                                                                   [out_laneID][arr_range]])
                        # Flow of vehicles during arr. range, in veh/s
                        flow_arr_range[arr_range] = num_veh_arr_range[arr_range] / LEN_RANGE
                    else:
                        flow_arr_range[arr_range] = 0
                        mean_speed_arr_range[arr_range] = 0
                        mean_hdwy_arr_range[arr_range] =  0
                        std_dev_hdwy_arr_range[arr_range] = 0

                    if flow_arr_range[arr_range] >= LOW_FLOW and flow_arr_range[arr_range] < MID_FLOW:
                        # Medium Flow of Vehicles, create the flow profile using Pearson Type III Distribution
                        inflow_msg[dest_jctID][out_laneID][1][arr_range] = (round(num_veh_arr_range[arr_range],2),
                                                                            round(mean_speed_arr_range[arr_range],2),
                                                                            round(mean_hdwy_arr_range[arr_range],2),
                                                                            round(std_dev_hdwy_arr_range[arr_range],2))
                    else:
                        # Low Flow of Vehicles, create the flow profile using Negative Exponential Distribution
                        # or
                        # High Flow of Vehicles, create the flow profile using Normal Distribution
                        inflow_msg[dest_jctID][out_laneID][1][arr_range] = (round(num_veh_arr_range[arr_range],2),
                                                                            round(mean_speed_arr_range[arr_range],2),
                                                                            round(mean_hdwy_arr_range[arr_range],2))
        # Delete inflow message for the destination junctions without values
        all_dest_jcts = inflow_msg.keys()
        for dest_jctID in all_dest_jcts:
            if inflow_msg[dest_jctID] == dict():
                del inflow_msg[dest_jctID]

        end9 = time.time()
        timings[9][1] += end9 - start9

        return inflow_msg


    def estArrivals(self,
                    jct_obj, # TC junctions object
                    movg_obj, # TC junctions object
                    veh_obj, # TC junctions object
                    lanes_order): # list of lanes following the order of junctions to be explored
        """Estimate future vehicles arriving based on information of inflow profile"""

        global num_veh_arr_range

        start10 = time.time()

        for jctID in jct_obj.ids:
            num_ranges = int(jct_obj.t_max[jctID] / LEN_RANGE)
            for in_laneID in jct_obj.in_laneIDs[jctID]:
                in_edgeID = self.laneID_to_edgeID[in_laneID]

                if in_laneID in lanes_order:
                    # Estimate vehicle arrivals and generate arriving vehicles on lane
                    # Parameters of vehicle arrivals based on upstream intersections of the order
                    self.lane_arrs_done[in_laneID] = 1

                    for veh_num, lane_ind in self.vehs_on_lane[in_laneID]:
                        # For vehicles already known to be on the lane
                        if veh_obj.in_lane_arr_dist[veh_num][lane_ind] == self.lane_length[in_laneID]:
                            # If future arrival, reset phase start values and crossing bool if arrival after ACTUAL_T
                            if veh_obj.in_lane_arr_t[veh_num][lane_ind] >= ACTUAL_T:
                                veh_obj.crossing_bool[veh_num][lane_ind] = None

                                veh_obj.detailed_tts[veh_num][lane_ind] = []

                                veh_obj.phase_start_t[veh_num][lane_ind] = veh_obj.in_lane_arr_t[veh_num][lane_ind]
                                veh_obj.phase_start_dist[veh_num][lane_ind] = veh_obj.in_lane_arr_dist[veh_num][lane_ind]
                                veh_obj.phase_start_speed[veh_num][lane_ind] = veh_obj.in_lane_arr_speed[veh_num][lane_ind]
                        else:
                            # If previous arrival, always reset phase start values and crossing bool
                            veh_obj.crossing_bool[veh_num][lane_ind] = None

                            veh_obj.detailed_tts[veh_num][lane_ind] = []

                            veh_obj.phase_start_t[veh_num][lane_ind] = veh_obj.update_t[veh_num][lane_ind]
                            veh_obj.phase_start_dist[veh_num][lane_ind] = veh_obj.update_dist[veh_num][lane_ind]
                            veh_obj.phase_start_speed[veh_num][lane_ind] = veh_obj.update_speed[veh_num][lane_ind]


                    if self.lane_inflow_profile[in_laneID][0] != () and ACTUAL_T >= BEGIN_ARRIVALS:
                        # Initializing
                        interv = edg_objs[tc_ind].lane_inflow_profile[in_laneID][0]["interv"]
                        num_veh_arr_range = [[] for _ in range(0, num_ranges)]
                        flow_arr_range = [[] for _ in range(0, num_ranges)]
                        mean_speed_arr_range = [[] for _ in range(0, num_ranges)]
                        mean_hdwy_arr_range = [[] for _ in range(0, num_ranges)]
                        std_dev_hdwy_arr_range = [[] for _ in range(0, num_ranges)]

                        # For counting number o vehicles
                        isCommonLane = False
                        for neight_jct in jct_obj.commonFromNeigh_laneIDs[jctID]:
                            if in_laneID in jct_obj.commonFromNeigh_laneIDs[jctID][neight_jct]:
                                isCommonLane = True
                                break

                        for arr_range in range(0, num_ranges):
                            cumul_hdwys = ACTUAL_T + (arr_range * LEN_RANGE)
                            try:
                                # a) update dynamic inflow given info by upstream junction
                                num_veh_arr_range[arr_range] = self.lane_inflow_profile[in_laneID][1][1][arr_range][0]
                                mean_speed_arr_range[arr_range] = self.lane_inflow_profile[in_laneID][1][1][arr_range][1]
                                mean_hdwy_arr_range[arr_range] = self.lane_inflow_profile[in_laneID][1][1][arr_range][2]
                            except IndexError:
                                # b) using inputed fixed inflow profile
                                # For fixed inflow profile, stop arrivals if queue prediction is full
                                num_veh_arr_range[arr_range] = self.lane_inflow_profile[in_laneID][0][interv][0]
                                mean_speed_arr_range[arr_range] = self.lane_inflow_profile[in_laneID][0][interv][1]
                                mean_hdwy_arr_range[arr_range] = self.lane_inflow_profile[in_laneID][0][interv][2]

                            if 0 < num_veh_arr_range[arr_range] < 1:
                                # Round the number of vehicles using binomial distribution with probability for decimals
                                num_veh_arr_range[arr_range] = np.random.binomial(1, num_veh_arr_range[arr_range])

                            if num_veh_arr_range[arr_range] >= 1:

                                start40 = time.time()

                                flow_arr_range[arr_range] = num_veh_arr_range[arr_range] / LEN_RANGE
                                # Round the number of vehicles using binomial distribution with the decimal as probability
                                int_num_veh = math.floor(num_veh_arr_range[arr_range])
                                dec_num_veh = num_veh_arr_range[arr_range] - int_num_veh
                                if dec_num_veh > 0:
                                    dec_num_veh = np.random.binomial(1, dec_num_veh)
                                num_veh_arr_range[arr_range] = int(int_num_veh + dec_num_veh)
                                jct_obj.arr_headways[jctID][:, 1] = 0
                                jct_obj.arr_headways[jctID][:, 2] = 0

                                if flow_arr_range[arr_range] < LOW_FLOW:
                                    # Low Flow of Vehicles, flow profile use Negative Exponential Distribution
                                    # Store the probability of headways equal or greater than the initial headway hdwy_0
                                    # i.e 1 - F(hdwy), where F is the CDF (Cumulative Distributive Function)
                                    # which represents de prob. of headways shorten than hdwy
                                    # As it is the shortest headway possible, the probability is 100%
                                    jct_obj.arr_headways[jctID][0, 1] = 1
                                    for h_ind, hdwy in enumerate(jct_obj.arr_headways[jctID][1:, 0]):
                                        # Store the probability of headways greater than the current headway hdwy
                                        # minus the minimum headway MIN_HDWY = 1 - F(hdwy)
                                        jct_obj.arr_headways[jctID][h_ind + 1, 1] = np.exp(-flow_arr_range[arr_range]
                                                                                           * (hdwy - MIN_HDWY))
                                        # Store the probability of headways between
                                        # current headway hdwy and previous headway
                                        jct_obj.arr_headways[jctID][h_ind, 2] = jct_obj.arr_headways[jctID][h_ind, 1] \
                                                                                - jct_obj.arr_headways[jctID][h_ind + 1,
                                                                                                              1]

                                elif flow_arr_range[arr_range] < MID_FLOW:
                                    # Medium Flow of Vehicles, flow profile uses Pearson Type III Distribution
                                    mean_hdwy_arr_range[arr_range] = max(mean_hdwy_arr_range[arr_range], MIN_HDWY)
                                    try:
                                        std_dev_hdwy_arr_range[arr_range] = self.lane_inflow_profile[in_laneID][1][1] \
                                                                                                    [arr_range][3]
                                    except IndexError:
                                        std_dev_hdwy_arr_range[arr_range] = self.lane_inflow_profile[in_laneID][0] \
                                                                                                    [interv][3]

                                    global shape_factor
                                    try:
                                        shape_factor = (mean_hdwy_arr_range[arr_range] - MIN_HDWY) / \
                                                       round(std_dev_hdwy_arr_range[arr_range], 1)
                                    except (ZeroDivisionError, FloatingPointError, ValueError):
                                        shape_factor = 0

                                    if shape_factor != 0:
                                        flow_rate = shape_factor / (mean_hdwy_arr_range[arr_range] - MIN_HDWY)
                                        gamma_func_shape = math.gamma(shape_factor)

                                        # Store the probability density function (PDF) of headways equal than
                                        # the initial headway hdwy_0 minus the minimum headway MIN_HDWY
                                        hdwy_0 = jct_obj.arr_headways[jctID][0, 0]
                                        try:
                                            calc_result = max((flow_rate / gamma_func_shape)
                                                         * (flow_rate * (hdwy_0 - MIN_HDWY) ** (shape_factor - 1))
                                                         * math.exp(-flow_rate * (hdwy_0 - MIN_HDWY)), 0)
                                            jct_obj.arr_headways[jctID][0, 1] = calc_result
                                        except (FloatingPointError, ValueError):
                                            # Since a hdwy_0 lower than or equal to the minimum headway the expression
                                            # wouldn't be able to be evaluated, it just approximates it to zero
                                            jct_obj.arr_headways[jctID][0, 1] = 0

                                        use_short_intervs = 1
                                        for h_ind, hdwy in enumerate(jct_obj.arr_headways[jctID][1:, 0]):
                                            # Store the probability density function (PDF) of headways equal than
                                            # the current headway hdwy minus the minimum headway MIN_HDWY
                                            # Using shorter intervals as the calculation is based on a linear line
                                            # between headways, even 1 second is not enough short. Then, is is summed
                                            # the probabilities of such very short intervals for the used longer interval
                                            # PS.: short intervals are used only when diff. between cumul of short
                                            # and longer intervals is less than 5%, to avoid too many calculations
                                            global result,cumul_result
                                            if use_short_intervs == 1:
                                                prev_result = jct_obj.arr_headways[jctID][h_ind, 1]
                                                cumul_result = 0
                                                for short_hdwy in frange(hdwy_0 + 0.1,hdwy + 0.1,0.1):
                                                    result = max((flow_rate / gamma_func_shape)
                                                                  * ((flow_rate * (short_hdwy - MIN_HDWY))
                                                                     ** (shape_factor - 1))
                                                                  * math.exp(-flow_rate * (short_hdwy - MIN_HDWY)), 0)

                                                    cumul_result += ((prev_result + result) / 2) * 0.1
                                                    prev_result = result
                                                else:
                                                     jct_obj.arr_headways[jctID][h_ind + 1, 1] = result
                                            else:
                                                result = max((flow_rate / gamma_func_shape)
                                                             * ((flow_rate * (hdwy - MIN_HDWY))
                                                                ** (shape_factor - 1))
                                                             * math.exp(-flow_rate * (hdwy - MIN_HDWY)), 0)
                                                jct_obj.arr_headways[jctID][h_ind + 1, 1] = result

                                            direct_result = ((jct_obj.arr_headways[jctID][h_ind, 1] +
                                                   jct_obj.arr_headways[jctID][h_ind + 1, 1]) / 2) * (hdwy - hdwy_0)

                                            # Store the probability of headways between
                                            # current headway hdwy and previous headway
                                            if use_short_intervs == 1 and abs(cumul_result - direct_result) >= 0.01:
                                                jct_obj.arr_headways[jctID][h_ind, 2] = cumul_result
                                            else:
                                                use_short_intervs = 0
                                                jct_obj.arr_headways[jctID][h_ind, 2] = direct_result

                                            cumul_prob = round(sum(jct_obj.arr_headways[jctID][:h_ind + 1, 2]),2)
                                            if cumul_prob >= 1:
                                                remaining_prob = 1 - sum(jct_obj.arr_headways[jctID][:h_ind, 2])
                                                jct_obj.arr_headways[jctID][h_ind, 2] = max(remaining_prob, 0)
                                                jct_obj.arr_headways[jctID][h_ind, 2] = min(jct_obj.arr_headways[jctID]
                                                                                            [h_ind, 2], 1)
                                                break
                                            hdwy_0 += (hdwy - hdwy_0)

                                if flow_arr_range[arr_range] >= MID_FLOW or (flow_arr_range[arr_range] > LOW_FLOW
                                                                             and shape_factor == 0):
                                    if flow_arr_range[arr_range] >= MID_FLOW:
                                        # High Flow of Vehicles, create the flow profile using Normal Distribution
                                        mean_hdwy_arr_range[arr_range] = max(mean_hdwy_arr_range[arr_range], MIN_HDWY)
                                        std_dev_hdwy = (mean_hdwy_arr_range[arr_range] - MIN_HDWY) / 3
                                    else:
                                        # When flow_arr_range[arr_range] < MID_FLOW
                                        # and std_dev_hdwy_arr_range[arr_range] == 0
                                        # or mean_hdwy_arr_range[arr_range] == MIN_HDWY
                                        std_dev_hdwy = 0

                                    # Store the probability of headways equal or shorter than
                                    # the initial headway hdwy_0  minus the minimum headway MIN_HDWY
                                    # ie. CDF
                                    hdwy_0 = jct_obj.arr_headways[jctID][0, 0]
                                    if std_dev_hdwy == 0 and hdwy_0 == round(mean_hdwy_arr_range[arr_range]):
                                        jct_obj.arr_headways[jctID][0, 1] = 1
                                        jct_obj.arr_headways[jctID][0, 2] = 1
                                    else:
                                        if std_dev_hdwy != 0:
                                            jct_obj.arr_headways[jctID][0, 1] = \
                                                st.norm.cdf(((hdwy_0 - mean_hdwy_arr_range[arr_range]) / std_dev_hdwy))
                                        else:
                                            jct_obj.arr_headways[jctID][0, 1] = 0

                                        for h_ind, hdwy in enumerate(jct_obj.arr_headways[jctID][1:, 0]):
                                            # Store the probability of headways equal or shorter than
                                            # the current headway hdwy minus the minimum headway MIN_HDWY
                                            # ie. CDF
                                            if std_dev_hdwy != 0:
                                                jct_obj.arr_headways[jctID][h_ind + 1, 1] = \
                                                    st.norm.cdf(((hdwy - mean_hdwy_arr_range[arr_range]) / std_dev_hdwy))
                                            elif std_dev_hdwy == 0 and hdwy == round(mean_hdwy_arr_range[arr_range]):
                                                jct_obj.arr_headways[jctID][h_ind + 1, 1] = 1
                                                jct_obj.arr_headways[jctID][h_ind, 2] = 1
                                                break
                                            else:
                                                jct_obj.arr_headways[jctID][h_ind + 1, 1] = 0

                                            # Store the probability of headways between
                                            # current headway hdwy and previous headway
                                            jct_obj.arr_headways[jctID][h_ind, 2] = \
                                                max(jct_obj.arr_headways[jctID][h_ind + 1, 1] - \
                                                jct_obj.arr_headways[jctID][h_ind, 1], 0)

                                            cumul_prob = round(sum(jct_obj.arr_headways[jctID][:h_ind + 1, 2]),2)
                                            if cumul_prob == 1:
                                                remaining_prob = 1 - sum(jct_obj.arr_headways[jctID][:h_ind, 2])
                                                jct_obj.arr_headways[jctID][h_ind, 2] = max(remaining_prob, 0)
                                                jct_obj.arr_headways[jctID][h_ind, 2] = min(jct_obj.arr_headways[jctID]
                                                                                            [h_ind, 2], 1)
                                                break

                                # Define the headway of each vehicle given the number of vehicles and
                                # the probability of each headway following the chosen distribution
                                jct_obj.arr_headways[jctID][-1, 2] = 0
                                jct_obj.arr_headways[jctID][-1, 2] += 1 - sum(jct_obj.arr_headways[jctID][:, 2])
                                hdwys = np.random.choice(jct_obj.arr_headways[jctID][:, 0], num_veh_arr_range[arr_range],
                                                         replace=True,
                                                         p=jct_obj.arr_headways[jctID][:, 2])

                                # Adjust headways to don't be longer than the LEN_RANGE
                                adjusted_len_range = max(sum(hdwys), LEN_RANGE)
                                adjusted_hdwy = [(hdy_val * LEN_RANGE) / adjusted_len_range
                                                 for hdy_ind,hdy_val in enumerate(hdwys)]

                                end40 = time.time()
                                timings[40][1] += end40 - start40

                                start41 = time.time()

                                for arr_veh_num, hdwy in enumerate(adjusted_hdwy):
                                    # time step vehicle will enter the incoming lane
                                    cumul_hdwys += hdwy
                                    # vehID is the arrival range, edge that it enters and time step entered
                                    vehID = "arr" + str(arr_range) + "." + str(arr_veh_num) + \
                                            "/" + in_laneID + "/" + str(round(cumul_hdwys,1))
                                    # generate vehicle
                                    veh_num = len(veh_obj.ids_all)
                                    veh_obj.ids_all.append(vehID)
                                    veh_obj.setVehParams(veh_num)
                                    veh_obj.defVehNextEdgeJct(movg_obj, self, veh_num, 0, jctID, first_edgeID=in_edgeID)
                                    veh_obj.defVehConnMovgNextLane(jct_obj, self, veh_num, jctID,
                                                                   0, self.laneID_to_edgeID[in_laneID], cumul_hdwys)
                                    veh_obj.defVehAttrOnLane(self, veh_num, 0,
                                                             cumul_hdwys, mean_speed_arr_range[arr_range],
                                                             self.lane_length[in_laneID])

                                    # For counting number o vehicles
                                    if isCommonLane == False and cumul_hdwys - ACTUAL_T < LEN_RANGE:
                                        LLR_nexttStep_area_updGen.append(1)

                                end41 = time.time()
                                timings[41][1] += end41 - start41

                            else:
                                # No vehicles
                                num_veh_arr_range[arr_range] = 0

                        # For Counting Number of All Vehicles
                        # Account vehicles arriving during actual time step
                        if NOGUI == False:
                            tStep_area_newArrivals.append(sum(num_veh_arr_range))
                else:

                    start43 = time.time()

                    # For Counting Number of All Vehicles
                    for veh_num, lane_ind in self.vehs_on_lane[in_laneID]:
                        if veh_obj.in_lane_arr_dist[veh_num][lane_ind] == self.lane_length[in_laneID]:
                            # Account vehicles arriving but from previous time step
                            if NOGUI == False and veh_obj.ids_all[veh_num] not in tStep_area_alreadyAccounted:
                                tStep_area_previousArrivals.append(1)
                                tStep_area_alreadyAccounted.append(veh_obj.ids_all[veh_num])

                    end43 = time.time()
                    timings[43][1] += end43 - start43

        end10 = time.time()
        timings[10][1] += end10 - start10


    def defProbes(self,
                  jct_obj,     # TC junction object
                  movg_obj,    # TC mov. group object
                  veh_obj,     # TC vehicles object
                  jctID,       # network junction ID to be defined the probe vehicles
                  in_laneID):  # network lane ID to be defined the probe vehicles
        """Define the probe vehicles that will be used to collect travel times"""

        start42 = time.time()

        num_ranges = int(jct_obj.t_max[jctID] / LEN_RANGE)
        ranges_t_limit = [ACTUAL_T + arr_range * LEN_RANGE for arr_range in range(0, num_ranges + 1)]

        start52 = time.time()

        if in_laneID not in self.closure_affected_edgeLanes["lanes"]:
            for veh_num,lane_ind in self.vehs_on_lane[in_laneID]:
                for movgID in veh_obj.movgID[veh_num][lane_ind]:
                    # Don't set vehicle as probe if its movgID doesn't have green time during t_max,
                    # if it is not a future arrival, and if it arrives on the edge before
                    # or after the prediction horizon or before ACTUAL_T
                    if movg_obj.movg_green_plan[movgID][0] != movg_obj.movg_green_plan[movgID][1] \
                    and veh_obj.in_lane_arr_dist[veh_num][lane_ind] == self.lane_length[in_laneID]  \
                    and veh_obj.in_lane_arr_t[veh_num][lane_ind] < ACTUAL_T + jct_obj.t_max[jctID] \
                    and veh_obj.in_lane_arr_t[veh_num][lane_ind] >= ACTUAL_T:
                        # Define the suitable arrival range
                        for arr_range in range(1, num_ranges + 1):
                            if ranges_t_limit[arr_range - 1] \
                                    <= veh_obj.in_lane_arr_t[veh_num][lane_ind] \
                                    < ranges_t_limit[arr_range]:
                                movg_obj.probe_vehs_range[in_laneID][movgID][arr_range - 1].append((veh_num, lane_ind))
                                veh_obj.probe_arr_range[veh_num][lane_ind] = arr_range - 1
                                break

            end52 = time.time()
            timings[52][1] += end52 - start52

            start53 = time.time()

            # Add artificial vehicles to estimate lane characteristcs (travel time)
            if in_laneID not in self.closure_affected_edgeLanes["lanes"]:
                for lane_movg in self.inLaneID_to_movgIDs[in_laneID]:
                    if movg_obj.movg_green_plan[lane_movg][0] != movg_obj.movg_green_plan[lane_movg][1]:
                        for arr_range in range(0, num_ranges):
                            if movg_obj.probe_vehs_range[in_laneID][lane_movg][arr_range] == []:
                                # Add artificial vehicles as probe vehicles to estimate travel time during each arr. range
                                in_edgeID = self.laneID_to_edgeID[in_laneID]
                                arr_time = ACTUAL_T + (arr_range * LEN_RANGE)
                                probe_vehID = "artProbe/" + in_laneID + "/" + str(arr_range)
                                probe_veh_num = len(veh_obj.ids_all)
                                veh_obj.ids_all.append(probe_vehID)
                                veh_obj.setVehParams(probe_veh_num, veh_class="Art. Probe")
                                veh_obj.defVehNextEdgeJct(movg_obj, self, probe_veh_num, 0, jctID, in_laneID,
                                                          lane_movg, in_edgeID)
                                veh_obj.defVehConnMovgNextLane(jct_obj, self, probe_veh_num, jctID, 0,
                                                               in_edgeID, arr_time,
                                                               next_lane=in_laneID, next_movgID=lane_movg)
                                veh_obj.defVehAttrOnLane(self, probe_veh_num,
                                                         0, arr_time, self.lane_max_speed[in_laneID],
                                                         self.lane_length[in_laneID])

                                movg_obj.probe_vehs_range[in_laneID][lane_movg][arr_range].append((probe_veh_num, 0))
                                veh_obj.probe_arr_range[probe_veh_num][0] = arr_range
                            else:
                                # Already vehicle travel time or probe vehicle to estimate travel time
                                pass

            end53 = time.time()
            timings[53][1] += end53 - start53
        else:
            # Not possible to estimate travel time
            self.stop_est_probes[in_laneID] = 1

        end42 = time.time()
        timings[42][1] += end42 - start42


    def estDepartures(self,
                             jct_obj,      # TC junction object
                             movg_obj,     # TC mov. group object
                             veh_obj,      # TC vehicles object
                             tc_it,        # algorithm order interaction
                             jct_order,    # order of junctions, 1 or -1
                             lanes_order): # lanes with same order of junctions to be explored
        """Estimates the trip vehicles will take and the departures of each vehicle for each junction"""

        global phase_green_movgs,remaining_probe_to_discharge

        start11 = time.time()

        # reset vehicles already on lane
        for jctID in jct_obj.ids:
            for in_laneID in jct_obj.in_laneIDs[jctID]:
                if in_laneID in lanes_order:
                    self.already_on_lane[in_laneID] = []

        # define the junction sequence based on junction order
        jct_seq = jct_obj.ids[:]
        if jct_order == -1:
            jct_seq.reverse()

        for jctID in jct_seq:
            num_ranges = int(jct_obj.t_max[jctID] / LEN_RANGE)

            # The outgoing laneIDs to have attributes stores are the ones that send to a neighbouring junction
            all_toNeigh_lanes = set([])
            for dest_jctID in jct_obj.commonToNeigh_laneIDs[jctID]:
                all_toNeigh_lanes.update(jct_obj.commonToNeigh_laneIDs[jctID][dest_jctID])
            all_toNeigh_lanes.intersection_update(set(lanes_order))
            for out_laneID in all_toNeigh_lanes:
                self.outLane_vehs_atr_arr_range.update({out_laneID: []})

            for in_laneID in jct_obj.in_laneIDs[jctID]:
                # reset instances
                if tc_it == 1:
                    self.stop_est_probes[in_laneID] = 0
                for movgID in self.inLaneID_to_movgIDs[in_laneID]:
                    num_greens = len(movg_obj.movg_green_plan[movgID][2])
                    conns = [conn for conn in movg_obj.movgID_to_connIDs[movgID]
                             if conn in self.inLaneID_to_connIDs[in_laneID]]
                    for connID in conns:
                        self.conn_phase_deps[movgID][connID] = [[] for phase in range(0, num_greens)]
                    if tc_it == 1:
                        movg_obj.probe_vehs_range[in_laneID][movgID] = [[] for _ in range(0, num_ranges)]

                if in_laneID in lanes_order:
                    # vehicles on a lane which is in the order will be discharged again, also those from new arrivals
                    # then reset following instances to get values from new dicharging estimations
                    self.lane_queue_pred.update({in_laneID: [ACTUAL_T, [0 for arr_range
                                                                    in range(0,
                                                                             len(self.lane_queue_pred[in_laneID][1]))]]})
                    self.inLaneStopped_dueOutQueue.update({in_laneID: 0})
                    for connID in self.inLaneID_to_connIDs[in_laneID]:
                        movgID = self.connID_to_movgID[connID]
                        num_green_phases = len(movg_obj.movg_green_plan[movgID][2])
                        if num_green_phases > 1:
                            # Don't take non-crossing phase (last one)
                            num_green_phases -= 1
                        self.conn_cap[connID][1] = [[] for phase in range(0, num_green_phases)]

                if USE_PROBES and tc_it == TC_IT_MAX:
                    # Probes are defined only in the last interaction of the algorithm
                    self.defProbes(jct_obj, movg_obj, veh_obj, jctID, in_laneID)

            num_phases = len(jct_obj.phases_green_plan[jctID])
            remaining_probe_to_discharge = 1
            jct_obj.t_max_cycle[jctID] = 0
            while remaining_probe_to_discharge == 1:
                # Extrapolate the departures estimation after the t_max prediction horizon
                # into more cycle until all probe vehicles are discharged
                for phase in range(0,num_phases):
                    # For each phase of the junction within t_max
                    phase_green_movgs = jct_obj.phases_green_plan[jctID][phase][0][:] \
                                        + jct_obj.phases_green_plan[jctID][phase][1][:]
                    for movgID in phase_green_movgs:
                        # The phase number of the movgIDs that have green on the current junction phase is
                        # the phase index of the current junction phase
                        movg_obj.movgID_to_phaseNum.update({movgID: movg_obj.movg_green_plan[movgID][3].index(phase)})

                    actual_phase_begin = movg_obj.movg_green_plan[phase_green_movgs[0]][0] \
                                        [movg_obj.movgID_to_phaseNum[phase_green_movgs[0]]]

                    for movgID in (movg for movg in jct_obj.movgIDs[jctID] if movg not in phase_green_movgs):
                        num_movg_phases = len(movg_obj.movg_green_plan[movgID][0])
                        for phase_num in range(0, num_movg_phases):
                            if movg_obj.movg_green_plan[movgID][0][phase_num] > actual_phase_begin:
                                # The phase number of the movgIDs that don't have green on the phase is
                                # the movgID phase number that starts after the start of the current junction phase
                                movg_obj.movgID_to_phaseNum.update({movgID: phase_num})
                                break

                    # Define all lanes that have a mov. group which gets green light on curren phase
                    lanes_phase_green_movgs = set([])
                    for green_movg in phase_green_movgs:
                        for laneID in movg_obj.movgID_to_inLaneIDs[green_movg]:
                            lanes_phase_green_movgs.add(laneID)

                    # select vehicles that will get green on the phase and order them
                    discharging_vehs_phase = veh_obj.orderVehLane(jct_obj, movg_obj, self, jctID,
                                                                  lanes_phase_green_movgs)
                    # classify them based on their lane
                    self.lane_discharging_vehs_phase.update({inlane: [(veh, v_lane_ind, v_movgID)
                                                                       for (veh, v_lane_ind, v_movgID)
                                                                       in discharging_vehs_phase
                                                                 if (veh, v_lane_ind)
                                                                 in self.vehs_on_lane[inlane]
                                                                 and veh_obj.vClass[veh] != "Art. Probe"]
                                                                 for inlane in lanes_phase_green_movgs})
                    # define vehicles already discharged
                    self.veh_already_disharged = [(veh, v_lane_ind)
                                                  for (veh, v_lane_ind, _) in discharging_vehs_phase
                                                  if veh_obj.crossing_bool[veh][v_lane_ind] == 1]

                    # define vehicles that still have to be discharged
                    veh_to_discharge = [(veh, v_lane_ind, v_movgID)
                                        for (veh, v_lane_ind, v_movgID) in discharging_vehs_phase
                                        if veh_obj.crossing_bool[veh][v_lane_ind] in (0, None)]

                    for veh_num, lane_ind, movgID in veh_to_discharge:
                        veh_obj.defVehMovement(jct_obj, movg_obj, self, jctID, veh_num, lane_ind, movgID,
                                               phase_green_movgs, lanes_order, all_toNeigh_lanes, tc_it)

                    nonCrossing_discharged_vehs = [(veh,l_ind) for veh,l_ind,_ in veh_to_discharge
                                                   if veh_obj.crossing_bool[veh][l_ind] in (0, None)]
                    for veh_num, lane_ind in nonCrossing_discharged_vehs:
                        # If vehicle will not cross during the cycle, set the final position, speed, and time
                        # for the begin of the next green phase of this mov. group (cycle)
                        veh_obj.phase_start_dist[veh_num][lane_ind] = veh_obj.phase_end_dist[veh_num][lane_ind]
                        veh_obj.phase_start_speed[veh_num][lane_ind] = veh_obj.phase_end_speed[veh_num][lane_ind]
                        veh_obj.phase_start_t[veh_num][lane_ind] = veh_obj.phase_end_t[veh_num][lane_ind]

                else:
                    remaining_probe_to_discharge = 0
                    if USE_PROBES == False or movg_obj.movg_green_plan[phase_green_movgs[0]][1][0] == float("inf"):
                        # All vehicles modelled, not necessary another cycle
                        pass
                    else:
                        for in_laneID in jct_obj.in_laneIDs[jctID]:
                            if jct_obj.t_max_cycle[jctID] == (MAX_T_MAX_CYCLE - 1) \
                            or self.stop_est_probes[in_laneID] == 1:
                                # If reached MAX_T_MAX_CYCLE or already set to stop estimating departures of probes,
                                # skip calculations for all remaining probes to cross to avoid infinity loops
                                empty_ranges = [[] for rg in range(0, num_ranges)]
                                self.stop_est_probes[in_laneID] = 1
                                for movg in movg_obj.probe_vehs_range[in_laneID].keys():
                                    if movg_obj.probe_vehs_range[in_laneID][movg] != empty_ranges:
                                        for rge in range(0, num_ranges):
                                            if movg_obj.probe_vehs_range[in_laneID][movg][rge] != []:
                                                num_del = 0
                                                copied_probes = movg_obj.probe_vehs_range[in_laneID][movg][rge][:]
                                                for v_ind, (veh, l_ind) in enumerate(copied_probes):
                                                    if veh_obj.crossing_bool[veh][l_ind] in (0, None):
                                                        del movg_obj.probe_vehs_range[in_laneID][movg][rge][v_ind
                                                                                                            - num_del]
                                                        num_del += 1
                                                        try:
                                                            veh_obj.probe_arr_range[veh][l_ind] = None
                                                        except TypeError:
                                                            skip = 1
                            else:
                                # If t_max cycle is below MAX_T_MAX_CYCLE, check if all probes crossed the stop line,
                                # if not model another t_max cycle
                                probes = []
                                for movg_probes in movg_obj.probe_vehs_range[in_laneID].values():
                                    for range_probes in movg_probes:
                                        if range_probes != []:
                                            probes.extend(range_probes)

                                for veh_num, lane_ind in probes:
                                    if veh_obj.crossing_bool[veh_num][lane_ind] in (0, None):
                                        remaining_probe_to_discharge = 1
                                        break

                                if remaining_probe_to_discharge == 1:
                                    break

                    if remaining_probe_to_discharge == 1:
                        jct_obj.t_max_cycle[jctID] += 1


            for in_laneID in jct_obj.in_laneIDs[jctID]:
                # The vehicles on lane will not account probe vehicles
                self.vehs_on_lane[in_laneID] = [(veh, l_ind) for veh, l_ind in self.vehs_on_lane[in_laneID]
                                                 if veh_obj.vClass[veh] != "Art. Probe"]

                # Initialize available length of predicted queue
                if in_laneID in lanes_order:
                    self.lane_aval_queue_len.update({in_laneID: [[max(self.lane_length[in_laneID]
                                                                 - self.lane_queue_pred[in_laneID][1][arr_range], 0)
                                                                  for arr_range in
                                                                  range(0, len(self.lane_queue_pred[in_laneID][1]))]
                                                                 for cycle in range(0, MAX_T_MAX_CYCLE)]})

        end11 = time.time()
        timings[11][1] += end11 - start11


    def estEdgeTravelTime(self,
                          jct_obj,  # TC junction object
                          veh_obj,  # TC vehicle object
                          movg_obj, # TC mov. group object
                          tc_it):   # algorithm order interaction
        """Estimate the travel time of each edge by taking average travel time of each vehicle on the edge,
        if there is no vehicle, it uses a queue model based on arrivals and departures"""

        start12 = time.time()

        if tc_it == TC_IT_MAX:
            # Reset TC area edge travel times
            self.tc_tt = dict()
            for jctID in jct_obj.ids:
                num_ranges = int(jct_obj.t_max[jctID] / LEN_RANGE)
                for in_edgID in jct_obj.in_edgeIDs[jctID]:
                    self.tc_tt.update({in_edgID: [ACTUAL_T, [0 for _ in range(0, num_ranges)]]})

            for jctID in jct_obj.ids:
                # Initialize
                num_ranges = int(jct_obj.t_max[jctID] / LEN_RANGE)
                ranges_t_limit = [ACTUAL_T + arr_range * LEN_RANGE for arr_range in range(1, num_ranges + 1)]
                for in_edgID in jct_obj.in_edgeIDs[jctID]:
                    interv_next = self.next_edg_prob[in_edgID]["interv"]
                    arr_ranges_TT = [[[] for _ in self.edgeID_to_laneIDs[in_edgID]] for _ in range(0, num_ranges)]
                    arr_ranges_qDelay = [[[] for _ in self.edgeID_to_laneIDs[in_edgID]] for _ in range(0, num_ranges)]
                    # Initialize accumulated number of vehicles arrived per range
                    num_veh_range = [[1 for _ in self.edgeID_to_laneIDs[in_edgID]] for _ in range(0, num_ranges)]
                    # Get total number of vehicles during whole t_max, but at least 1 vehicles to estimate the time
                    # if an additional vehicle would arrive within t_max (planning horizon)
                    num_veh_lane = [(0 if (in_laneID in self.closure_affected_edgeLanes["lanes"]
                                           and self.lane_blockage_dist[in_laneID][1] == None)
                                     else max(1, len(self.vehs_on_lane[in_laneID])))
                                    for lane,in_laneID in enumerate(self.edgeID_to_laneIDs[in_edgID])]

                    for lane,in_laneID in enumerate(self.edgeID_to_laneIDs[in_edgID]):
                        # Edge travel time is the weighted average of each lane's travel time by number of vehicles,
                        # which is the average of the travel times for each of its mov. group

                        global calc_lane_discharge_t

                        if num_veh_lane[lane] == 0:
                            # If no vehicle (due closure affected lane) don't calculate travel time, use high number
                            for arr_range in range(0, num_ranges):
                                arr_ranges_TT[arr_range][lane].append(np.nan)
                                arr_ranges_qDelay[arr_range][lane].append(0)
                        else:
                            calc_lane_discharge_t = 0
                            set_blockage_tt_lane = 0
                            if in_laneID not in self.closure_affected_edgeLanes["lanes"]:
                                # Lane is not affected by the lanes closure
                                if USE_PROBES == False or self.unique_inLane2outEdge.has_key(in_laneID):
                                    # If in_laneID is the only lane that send vehicles to a certain edge, it will need
                                    # to use deterministic queue model to transfer queue delays from in_laneID to the
                                    # outgoing edge.
                                    calc_lane_discharge_t = 1
                                else:
                                    for lane_movg in self.inLaneID_to_movgIDs[in_laneID]:
                                        if [] in movg_obj.probe_vehs_range[in_laneID][lane_movg]:
                                            # If no probe vehicle for a mov. group of the lane it will be needed to use
                                            # the deterministic queue model
                                            calc_lane_discharge_t = 1
                                            break
                            else:
                                # Lane is affected by lane closures
                                set_blockage_tt_lane = 1

                            if calc_lane_discharge_t == 1:
                                # Initialize
                                zero_ranges = [0 for _ in range(0, num_ranges)]
                                vehs_arrivals_range = zero_ranges[:]
                                vehs_deps_range = zero_ranges[:]
                                arrivals_flow_range = zero_ranges[:]
                                deps_flow_range = zero_ranges[:]
                                queue_range = zero_ranges[:]
                                total_delay_queue = zero_ranges[:]
                                outLane_utilization = [dict() for _ in range(0, num_ranges)]
                                outLane_delay = [dict() for _ in range(0, num_ranges)]
                                all_connIDs = [conn for conn in self.inLaneID_to_connIDs[in_laneID]
                                               if self.connID_to_outLaneID[conn]]
                                all_out_lanes = set([])
                                for conn in all_connIDs:
                                    all_out_lanes.update(self.edgeID_to_laneIDs[
                                                             self.laneID_to_edgeID[self.connID_to_outLaneID[conn]]])
                                for arr_range in range(0, num_ranges):
                                    outLane_delay[arr_range].update({out_lane: 0 for out_lane in all_out_lanes})
                                    # At least 1 vehicle for the outLane utilization,
                                    # because in case there is no vehicle it will still account the delay
                                    outLane_utilization[arr_range].update({out_lane: 1 for out_lane in all_out_lanes})
                                # Account initial queue, vehicles arrivals, and departures during t_max
                                for veh_num, lane_ind in self.vehs_on_lane[in_laneID]:
                                    if veh_obj.vClass[veh_num] != "Art. Probe" \
                                    and veh_obj.crossing_bool[veh_num][lane_ind] != None:
                                        # Define vehicle arrival time as the time it would arrive at maximum speed and
                                        # without any delay
                                        veh_arr_stopline = veh_obj.in_lane_arr_t[veh_num][lane_ind] + \
                                                           veh_obj.in_lane_arr_dist[veh_num][lane_ind] \
                                                           / (self.lane_max_speed[in_laneID]
                                                              * veh_obj.speed_factor[veh_num])

                                        # If arrival on outgoing lane is before the end of the arrival range,
                                        # append vehicle to arrivals during the range
                                        for arr_range in range(0, num_ranges):
                                            if veh_arr_stopline < ranges_t_limit[arr_range]:
                                                vehs_arrivals_range[arr_range] += 1
                                                break

                                        if veh_obj.phase_end_t[veh_num][lane_ind] != None \
                                        and veh_obj.crossing_bool[veh_num][lane_ind] == 1:
                                            # Account vehicle departure only if crosses the stop line and on
                                            # the proper range
                                            veh_dep_stopline = max(veh_obj.phase_end_t[veh_num][lane_ind],
                                                                   veh_arr_stopline)
                                            for arr_range in range(0, num_ranges):
                                                if veh_dep_stopline < ranges_t_limit[arr_range]:
                                                    vehs_deps_range[arr_range] += 1
                                                    try:
                                                        out_lane = veh_obj.lane_route[veh_num][lane_ind + 1]
                                                    except IndexError:
                                                        out_lane = 0
                                                        if len(veh_obj.lane_route) < veh_num:
                                                            LOGGER.error('ERROR: Trying to access veh_obj.lane_route[{:d}] but len(veh_obj.lane_route)={:d}'.format(veh_num, len(veh_obj.lane_route)))
                                                        else:
                                                            lr = veh_obj.lane_route[veh_num]
                                                            if len(lr) < lane_ind + 1:
                                                                LOGGER.error('ERROR: Trying to access veh_obj.lane_route[{0:d}][{2:d}] but len(veh_obj.lane_route[{0:d}])={1:d}'.format(veh_num, len(lr), lane_ind + 1))
                                                            else:
                                                                LOGGER.error('ERROR: Unknown index error accessing veh_obj.lane_route[{:d}][{:d}]'.format(veh_num, lane_ind + 1))
                                                                raise
                                                    outLane_utilization[arr_range][out_lane] += 1
                                                    break

                                # estimate the queue, and number of arrivals and departures per range for the in_laneID
                                for arr_range in range(0, num_ranges):
                                    arrivals_flow_range[arr_range] = sum(vehs_arrivals_range[:arr_range + 1]) / LEN_RANGE
                                    deps_flow_range[arr_range] = sum(vehs_deps_range[:arr_range + 1]) / LEN_RANGE
                                    queue_range[arr_range] = sum(vehs_arrivals_range[:arr_range + 1])\
                                                             - sum(vehs_deps_range[:arr_range + 1])
                                    queue_flow_range = round(arrivals_flow_range[arr_range] -
                                                             deps_flow_range[arr_range], 2)
                                    total_delay_queue[arr_range] = round(queue_flow_range * (LEN_RANGE ** 2), 2)

                                    num_veh_range[arr_range][lane] += vehs_arrivals_range[arr_range]
                                    sum_vehs_outLane = float(sum(outLane_utilization[arr_range].values()))
                                    for out_lane in all_out_lanes:
                                        outLane_utilization[arr_range][out_lane] = \
                                            outLane_utilization[arr_range][out_lane] / sum_vehs_outLane

                            # Estimate travel time for the lane mov. group per arr. range
                            for lane_movg in self.inLaneID_to_movgIDs[in_laneID]:
                                if USE_PROBES == False or self.unique_inLane2outEdge.has_key(in_laneID) \
                                or [] in movg_obj.probe_vehs_range[in_laneID][lane_movg]:
                                    # Define possible out lanes (to check delay due queue spill back and signal capacity)
                                    connIDs = [conn for conn in self.inLaneID_to_connIDs[in_laneID]
                                               if self.connID_to_movgID[conn] == lane_movg
                                               and self.connID_to_outLaneID[conn]
                                               not in self.closure_affected_edgeLanes["lanes"]]
                                    possible_out_lanes = set([])
                                    for conn in connIDs:
                                        possible_out_lanes.update(
                                            self.edgeID_to_laneIDs[
                                                self.laneID_to_edgeID[self.connID_to_outLaneID[conn]]])
                                    queue_info_out_lanes = [out_lane for out_lane in possible_out_lanes
                                                            if self.lane_queue_pred.has_key(out_lane)]

                                for arr_range in range(0, num_ranges):
                                    if USE_PROBES == False or self.unique_inLane2outEdge.has_key(in_laneID) \
                                    or movg_obj.probe_vehs_range[in_laneID][lane_movg][arr_range] == []:
                                        # If no probe vehicle on the arrival range for the mov. group of the inc. lane,
                                        # use the queue model below using the queue, arrivals, and departures above

                                        start13 = time.time()

                                        if set_blockage_tt_lane == 0 and possible_out_lanes != set([]):
                                            # If lane not blocked and there is at least one possible outgoing lane,
                                            # estimate travel time based on queue model

                                            global suitable_cycle, suitable_rge, suitable_ph, suitable_end_t, \
                                                   suitable_beg_t, lane_length_tt, green_phase_wait, delay_queue, \
                                                   delay_spillback, delay_conn, jct_crossing_t

                                            # Calculate free flow travel time
                                            lane_length_tt = (self.lane_length[in_laneID] /
                                                              self.lane_max_speed[in_laneID])

                                            # Estimating arriving time at stop line if veh. arrives at the begin of range
                                            arr_t_vert_queue = ACTUAL_T + (arr_range * LEN_RANGE) + lane_length_tt

                                            if movg_obj.movg_green_plan[lane_movg][1][-1] != float("inf"):
                                                # If it is a signalized junction
                                                # Find suitable cycle, range and phase when arriving at the queue
                                                ph_cycle = 0
                                                found_ph = 0
                                                num_green_phases = len(movg_obj.movg_green_plan[lane_movg][0])
                                                if num_green_phases > 1:
                                                    end_green_phases = movg_obj.movg_green_plan[lane_movg][1][:-1]
                                                else:
                                                    end_green_phases = movg_obj.movg_green_plan[lane_movg][1][:]
                                                while found_ph == 0:
                                                    if arr_t_vert_queue <= end_green_phases[-1] + (jct_obj.t_max[jctID]
                                                                                                   * ph_cycle):
                                                        suitable_cycle = [ph_cycle]
                                                        for ph, ph_end_t in enumerate(end_green_phases):
                                                            if arr_t_vert_queue < ph_end_t \
                                                                                  + (jct_obj.t_max[jctID] * ph_cycle):
                                                                found_ph = 1
                                                                suitable_ph = [ph]
                                                                suitable_end_t = [ph_end_t
                                                                                  + (jct_obj.t_max[jctID] * ph_cycle)]
                                                                suitable_beg_t = [movg_obj.movg_green_plan[lane_movg]
                                                                                  [0][ph]
                                                                                  + (jct_obj.t_max[jctID] * ph_cycle)]

                                                                for rge in range(0, num_ranges):
                                                                    if suitable_beg_t[0] < ACTUAL_T \
                                                                                          + ((rge + 1) * LEN_RANGE) \
                                                                                          + (jct_obj.t_max[jctID]
                                                                                             * suitable_cycle[0]):
                                                                        suitable_rge = [rge]
                                                                        break
                                                                break

                                                    elif ph_cycle == MAX_T_MAX_CYCLE - 1:
                                                        # If reached MAX_T_MAX_CYCLE, use latest values
                                                        found_ph = 1
                                                        suitable_cycle = [ph_cycle]
                                                        suitable_ph = [len(end_green_phases) - 1]
                                                        suitable_end_t = [arr_t_vert_queue]
                                                        suitable_beg_t = [arr_t_vert_queue]
                                                        suitable_rge = [num_ranges - 1]
                                                    else:
                                                        ph_cycle += 1

                                                # Estimate the delay due waiting for green phase for an arriving
                                                # additional veh
                                                green_phase_wait = max(suitable_beg_t[0] - arr_t_vert_queue, 0)
                                            else:
                                                # If unsignalized junction
                                                # Find suitable cycle and range when arriving at the queue
                                                found_rge = 0
                                                cycle = 0
                                                end_cycle_t = ACTUAL_T + jct_obj.t_max[jctID]
                                                while found_rge == 0:
                                                    if arr_t_vert_queue <= end_cycle_t * (1 + cycle):
                                                        suitable_cycle = [cycle]
                                                        for rge in range(0, num_ranges):
                                                            if arr_t_vert_queue < ACTUAL_T + ((rge + 1) * LEN_RANGE) \
                                                            + (jct_obj.t_max[jctID] * suitable_cycle[0]):
                                                                suitable_rge = [rge]
                                                                found_rge = 1
                                                                break
                                                    elif cycle == MAX_T_MAX_CYCLE - 1:
                                                        # If reached MAX_T_MAX_CYCLE, use latest values
                                                        suitable_rge = [num_ranges - 1]
                                                        suitable_cycle = [cycle]
                                                        found_rge = 1
                                                    else:
                                                        cycle += 1
                                                suitable_end_t = [ACTUAL_T + jct_obj.t_max[jctID] * (suitable_cycle[0]
                                                                                                     + 1)]
                                                green_phase_wait = 0

                                            # Estimate Queue Delay (delay to discharge all vehicles ahead 1 additional one)
                                            if total_delay_queue[suitable_rge[0]] > 0:
                                                # Use range when queue dissipates or capacity of discharge
                                                try:
                                                    # Find range when queue dissipates from suitable range within same
                                                    # cycle
                                                    end_queue_rge = suitable_rge[0] \
                                                                    + queue_range[suitable_rge[0]:].index(0)
                                                    delay_queue = (sum(total_delay_queue[suitable_rge[0]:
                                                                                         end_queue_rge + 1])
                                                                   / (arrivals_flow_range[end_queue_rge] * LEN_RANGE))
                                                except ValueError:
                                                    try:
                                                        # If not found range when the queue dissipates within same cycle,
                                                        # assume periodicity for next cycle and find it from the
                                                        # beginning of the t_max cycle, considering the range has
                                                        # acumul. arrival flow
                                                        end_queue_rge = [rge for rge,q_rge in enumerate(queue_range)
                                                                         if q_rge == 0
                                                                         and arrivals_flow_range[rge] > 0][0]
                                                        delay_queue = ((sum(total_delay_queue[suitable_rge[0]:])
                                                                       / (arrivals_flow_range[-1]
                                                                          * LEN_RANGE))
                                                                       + (sum(total_delay_queue[:end_queue_rge + 1])
                                                                       / (arrivals_flow_range[end_queue_rge]
                                                                          * LEN_RANGE)))
                                                    except IndexError:
                                                        # If not found when queue dissipates within entire cycle,
                                                        # estimate based on capacity of discharge within a t_max cycle
                                                        # and queue length at the suitable range
                                                        global rem_veh_next_cycles, rge_dissipates_queue

                                                        if movg_obj.movg_green_plan[lane_movg][1][-1] != float("inf"):
                                                            # If signalized, check time sum of green time per cycle and
                                                            # average capacity of discharge within one cycle to estimate
                                                            # the number of needed cycles
                                                            num_green_ph = len(movg_obj.movg_green_plan[lane_movg][0]) \
                                                                           - 1
                                                            ave_cap = []
                                                            conn_prob = []
                                                            if not connIDs:
                                                                LOGGER.warning('connIDs is empty!')
                                                            if num_green_ph <= 0:
                                                                LOGGER.warning('num_green_ph <= 0, connIDs=' + str(connIDs))
                                                            for connID in connIDs:
                                                                out_edgID = self.laneID_to_edgeID[
                                                                            self.connID_to_outLaneID[connID]]
                                                                for phase_num in range(0, num_green_ph):
                                                                    if self.conn_cap[connID][1][phase_num] == []:
                                                                    # If dynamic capacity not defined for the phase_num,
                                                                    # update it
                                                                        self.updateDynConnCapacity(jct_obj, movg_obj,
                                                                                                   veh_obj, connID)
                                                                    ave_cap.append(self.conn_cap[connID][1][phase_num])
                                                                    conn_prob.append(
                                                                        self.next_edg_prob[in_edgID][interv_next]
                                                                                          [out_edgID])

                                                            cycle_green_dur = [max(movg_obj.movg_green_plan[lane_movg]
                                                                                   [1][ph]
                                                                                   - movg_obj.movg_green_plan[lane_movg]
                                                                                   [0][ph]
                                                                                   - STD_ACC_T, 0)
                                                                               for ph in range(0, num_green_ph)]

                                                            try:
                                                                ave_cycle_cap = np.average(ave_cap, weights=conn_prob)
                                                            except ZeroDivisionError:
                                                                LOGGER.warning('num_green_ph={:d}, connIDs={:s}, ave_cap={:s}, conn_prob={:s}'.format(num_green_ph, str(connIDs), str(ave_cap), str(conn_prob)))
                                                                if ave_cap:
                                                                    ave_cycle_cap = np.mean(ave_cap)
                                                                else:
                                                                    ave_cycle_cap = 0
                                                                    
                                                            cycle_deps = float(ave_cycle_cap * sum(cycle_green_dur))
                                                            phases_deps = [ave_cycle_cap * cycle_green_dur[ph]
                                                                           for ph in range(0, num_green_ph)]

                                                            ph_aval_dur = max(suitable_end_t[0] - max(arr_t_vert_queue,
                                                                                                      suitable_beg_t[0])
                                                                              - STD_ACC_T, 0)
                                                            
                                                            if num_green_ph > 0:
                                                                try:
                                                                    phases_deps[suitable_ph[0]] = int((phases_deps[
                                                                                                           suitable_ph[0]]
                                                                                               * ph_aval_dur)
                                                                                               / cycle_green_dur[
                                                                                                          suitable_ph[0]])
                                                                except (ZeroDivisionError,FloatingPointError):
                                                                    phases_deps[suitable_ph[0]] = 0

                                                                rem_veh_next_cycles = max(queue_range[suitable_rge[0]]
                                                                                          - phases_deps[suitable_ph[0]], 0)
                                                                for ph in range(suitable_ph[0] + 1, num_green_ph):
                                                                    rem_veh_next_cycles = max(rem_veh_next_cycles
                                                                                              - phases_deps[ph], 0)
                                                                    if rem_veh_next_cycles == 0:
                                                                        for rge in range(suitable_rge[0], num_ranges):
                                                                            if movg_obj.movg_green_plan[lane_movg][0][ph] \
                                                                            < ACTUAL_T + ((rge + 1) * LEN_RANGE) \
                                                                            + (jct_obj.t_max[jctID] * suitable_cycle[0]):
                                                                                rge_dissipates_queue = rge
                                                                                break

                                                                        break
                                                                else:
                                                                    rge_dissipates_queue = num_ranges - 1
                                                        else:
                                                            # If unsignalized check the queue at arrival and the
                                                            # departures per cycles to estimate number of needed cycles
                                                            num_green_ph = len(movg_obj.movg_green_plan[lane_movg][0])
                                                            cycle_deps = round(sum(vehs_deps_range),0)
                                                            rem_veh_next_cycles = max(queue_range[suitable_rge[0]]
                                                                                      - vehs_deps_range[suitable_rge[0]],
                                                                                      0)
                                                            for rge in range(suitable_rge[0] + 1, num_ranges):
                                                                rem_veh_next_cycles = max(rem_veh_next_cycles
                                                                                          - vehs_deps_range[rge], 0)
                                                                if rem_veh_next_cycles == 0:
                                                                    rge_dissipates_queue = rge
                                                                    break
                                                            else:
                                                                rge_dissipates_queue = rge

                                                        try:
                                                            # Estimate how many cycles will be needed
                                                            needed_cycles = rem_veh_next_cycles / cycle_deps
                                                        except (ZeroDivisionError, FloatingPointError):
                                                            # No vehicle departures, assume the number of phases as the
                                                            # queue during the suitable range (1 veh per t_max cycle)
                                                            if num_green_ph > 0:
                                                                needed_cycles = rem_veh_next_cycles / num_green_ph
                                                            else:
                                                                needed_cycles = 1

                                                        delay_queue = ((sum(total_delay_queue[suitable_rge[0]:
                                                                                              rge_dissipates_queue + 1])
                                                                        + sum(total_delay_queue) * needed_cycles) /
                                                                       (arrivals_flow_range[-1] * LEN_RANGE))
                                            else:
                                                delay_queue = 0

                                            # Estimate delay due full outgoing lane for the additional vehicle
                                            global moved2out_lane, rge_ind, delayed_starting_t, \
                                                   out_queue_range, out_queue_cycle

                                            for out_lane in possible_out_lanes:
                                                if out_lane in queue_info_out_lanes:
                                                    delayed_starting_t = arr_t_vert_queue + green_phase_wait \
                                                                         + delay_queue
                                                    time_resolution = self.lane_queue_tLimitRanges[out_lane][1]\
                                                                      - self.lane_queue_tLimitRanges[out_lane][0]
                                                    out_queue_t_max = self.lane_queue_tLimitRanges[out_lane][-1] \
                                                                      + time_resolution \
                                                                      - self.lane_queue_tLimitRanges[out_lane][0]
                                                    out_queue_cycle = [0]
                                                    found_cycle = 0
                                                    # Find suitable cycle for delayed_starting_t
                                                    while found_cycle == 0:
                                                        if delayed_starting_t < ACTUAL_T + out_queue_t_max + (
                                                                out_queue_t_max * out_queue_cycle[0]) \
                                                                or out_queue_cycle[0] == (MAX_T_MAX_CYCLE - 1):
                                                            found_cycle = 1
                                                        else:
                                                            cycle = out_queue_cycle[0]
                                                            del out_queue_cycle[0]
                                                            out_queue_cycle.append(cycle + 1)

                                                    for range_i, range_t in enumerate(
                                                            self.lane_queue_tLimitRanges[out_lane]):
                                                        # define the interval range of the queue prediction
                                                        # to consider to check the queue length
                                                        if delayed_starting_t < range_t + out_queue_t_max \
                                                                                * out_queue_cycle[0]:
                                                            out_queue_range = [range_i]
                                                            break
                                                    else:
                                                        found_cycle = 0
                                                    moved2out_lane = 0
                                                    rge_ind = 0
                                                    while moved2out_lane == 0:
                                                        if found_cycle == 1:
                                                            for range_i, range_t in enumerate(
                                                                    self.lane_queue_tLimitRanges[out_lane]
                                                                    [out_queue_range[0]:]):
                                                                rge_ind = out_queue_range[0] + range_i
                                                                if self.lane_queue_pred[out_lane][1][rge_ind] != 0 \
                                                                and veh_obj.vType_vals["min_gap_d"][0] \
                                                                + veh_obj.vType_vals["length"][0] \
                                                                > self.lane_aval_queue_len[out_lane] \
                                                                    [out_queue_cycle[0]][rge_ind]:
                                                                    # delay departure if out. lane queue cannot take
                                                                    # vehicle's length
                                                                    delayed_starting_t = range_t \
                                                                                         + (out_queue_t_max
                                                                                            * out_queue_cycle[0])
                                                                else:
                                                                    # vehicle has space in the outgoing lane queue
                                                                    moved2out_lane = 1
                                                                    outLane_delay[arr_range].update(
                                                                        {out_lane: delayed_starting_t
                                                                                   - (arr_t_vert_queue
                                                                                      + green_phase_wait
                                                                                      + delay_queue)})
                                                                    break

                                                        if moved2out_lane == 0:
                                                            # if checked queue along whole cycle and vehicle couldn't
                                                            # go to outgoing lane, try on the next cycle
                                                            cycle = out_queue_cycle[0]
                                                            del out_queue_cycle[0]
                                                            out_queue_cycle.append(cycle + 1)
                                                            del out_queue_range[0]
                                                            out_queue_range.append(0)
                                                        else:
                                                            # Reset range
                                                            del out_queue_range[0]
                                                            out_queue_range.append(rge_ind)
                                                        if out_queue_cycle[0] >= MAX_T_MAX_CYCLE \
                                                        and moved2out_lane == 0:
                                                            # If reached MAX_T_MAX_CYCLE and vehicle couldn't go to
                                                            # outgoing lane, use half of the average time between
                                                            # time instants where new new vehicles can enter
                                                            # the out lane queue considering the queue prediction
                                                            last_q_aval = 1
                                                            first_full_range = 0
                                                            delay_inRanges = []
                                                            for q_range, q_pred in enumerate(
                                                                    self.lane_queue_pred[out_lane][1]):
                                                                q_aval_range = self.lane_length[out_lane] - q_pred \
                                                                               - (veh_obj.vType_vals["min_gap_d"][0]
                                                                                  + veh_obj.vType_vals["length"][0])

                                                                if last_q_aval < 0 and q_aval_range >= 0:
                                                                    delay_inRanges.append(q_range - first_full_range)
                                                                elif q_aval_range < 0 and last_q_aval >= 0:
                                                                    first_full_range = q_range
                                                                else:
                                                                    pass

                                                                last_q_aval = q_aval_range
                                                            else:
                                                                if q_aval_range < 0:
                                                                    delay_inRanges.append(q_range
                                                                                          + 1 - first_full_range)

                                                            try:
                                                                if delay_inRanges[0] == \
                                                                len(self.lane_queue_pred[out_lane][1]):
                                                                    # full queue during whole t_max cycle,
                                                                    # assume needed 1 t_max cycle
                                                                    outLane_delay[arr_range].update(
                                                                        {out_lane: out_queue_t_max})
                                                                else:
                                                                    outLane_delay[arr_range].update(
                                                                        {out_lane: (float(np.mean(delay_inRanges)
                                                                                          * time_resolution) / 2)})

                                                            except (FloatingPointError,IndexError):
                                                                # No full queue during whole t_max cycle
                                                                outLane_delay[arr_range].update({out_lane: 0})
                                                            break
                                                else:
                                                    # Outgoing lane has no information of queue prediction
                                                    outLane_delay[arr_range].update({out_lane: 0})

                                            out_delay = np.array(outLane_delay[arr_range].values())
                                            weights = np.array(outLane_utilization[arr_range].values())
                                            indices = [(True if out_lane in all_out_lanes
                                                                and outLane_delay[arr_range].values()
                                                                [out_lane_ind] > 0 else False)
                                                       for out_lane_ind,out_lane
                                                       in enumerate(outLane_delay[arr_range].keys())]
                                            # The outgoing lane delay is the weighted average based on lane utilization
                                            try:
                                                delay_spillback = np.average(out_delay[indices],
                                                                             weights=weights[indices])
                                            except ZeroDivisionError:
                                                delay_spillback = 0

                                            # Estimate Connection Delay For an additional vehicle
                                            if movg_obj.movg_green_plan[lane_movg][1][-1] != float("inf"):
                                                # If signalized junction, the reference headway time is the begin of
                                                # green time of the suitable phase
                                                phase_num = suitable_ph[0]
                                                ref_hdwy_t = suitable_beg_t[0]
                                            else:
                                                # If unsignalized junction, the reference headway time is veh. arriving
                                                # time already
                                                phase_num = 0
                                                ref_hdwy_t = arr_t_vert_queue + green_phase_wait \
                                                            + delay_queue + delay_spillback

                                            # Estimate the connection delay
                                            delay_conn = self.estDelayConnCap(jct_obj, movg_obj, veh_obj, jctID,
                                                                              suitable_cycle[0], phase_num,
                                                                              suitable_end_t[0], arr_t_vert_queue +
                                                                              green_phase_wait + delay_queue
                                                                              + delay_spillback, ref_hdwy_t,
                                                                              movg=lane_movg, laneID=in_laneID)

                                            # Estimate junction crossing time of the additional vehicle
                                            crossing_times = []
                                            for connID in connIDs:
                                                out_edgID = self.laneID_to_edgeID[self.connID_to_outLaneID[connID]]
                                                turning_speed = min(math.sqrt(FRICTION_COEF * GRAVITY_ACC
                                                                          * self.next_edg_radius[in_edgID][out_edgID]),
                                                                    self.lane_max_speed[in_laneID])
                                                crossing_times.append(self.conn_length[connID] / float(turning_speed))
                                                jct_crossing_t = np.mean(crossing_times)

                                            # The total travel time is the sum o lane length travel time and all delays
                                            total_tt = lane_length_tt + green_phase_wait + delay_queue \
                                                       + delay_spillback + delay_conn + jct_crossing_t

                                            # If vehicle would experience any delay, add a stanrdard acceleration time
                                            if green_phase_wait > 1 or delay_conn > 1 or delay_spillback > 1:
                                                total_tt += STD_ACC_T

                                            # Store the total travel time
                                            arr_ranges_TT[arr_range][lane].append(total_tt)
                                            if arr_range == 0:
                                                self.tc_edge_movgs_tt[in_edgID][lane_movg] = np.ceil(total_tt)

                                            # Store the queue and outgoing lane delay for transferring to artificial
                                            # edges. This avoids edge with lanes that have very different travel times
                                            # because the direction
                                            arr_ranges_qDelay[arr_range][lane].append(delay_queue + delay_spillback)
                                            try:
                                                for out_edgID in self.unique_inLane2outEdge[in_laneID]:
                                                    for out_laneID in possible_out_lanes:
                                                        self.tc_tt[out_edgID][1][arr_range] += np.ceil((delay_queue
                                                                                                 * outLane_utilization
                                                                                                 [arr_range][out_laneID])
                                                                                                + outLane_delay
                                                                                                  [arr_range][out_laneID])
                                            except KeyError:
                                                arr_ranges_qDelay[arr_range][lane][-1] = 0

                                        elif self.lane_blockage_dist[in_laneID][1] != None:
                                            # If lane is blocked and the expected time to reopen is given
                                            lane_length_tt = (self.lane_length[in_laneID] /
                                                             self.lane_max_speed[in_laneID])

                                            rem_t_wait = self.lane_blockage_dist[in_laneID][1] \
                                                         - (ACTUAL_T + arr_range + LEN_RANGE)
                                            arr_ranges_TT[arr_range][lane].append(rem_t_wait + lane_length_tt)
                                            if arr_range == 0:
                                                self.tc_edge_movgs_tt[in_edgID][lane_movg] = np.ceil(rem_t_wait
                                                                                                     + lane_length_tt)
                                            arr_ranges_qDelay[arr_range][lane].append(0)
                                        else:
                                            # If lane is blocked or no possible outgoing lane, and the remaining
                                            # travel time is NOT given, don't consider the travel time on the lane
                                            arr_ranges_TT[arr_range][lane].append(np.nan)
                                            if arr_range == 0:
                                                self.tc_edge_movgs_tt[in_edgID][lane_movg] = np.nan
                                            arr_ranges_qDelay[arr_range][lane].append(0)

                                        end13 = time.time()
                                        timings[13][1] += end13 - start13

                                    else:
                                        # Use travel times estimated by probe vehicles
                                        arr_ranges_qDelay[arr_range][lane].append(0)
                                        if arr_range == 0:
                                            self.tc_edge_movgs_tt[in_edgID][lane_movg] = 0
                                            num_probes = \
                                                float(len(movg_obj.probe_vehs_range[in_laneID][lane_movg][arr_range]))
                                        for veh_num, lane_ind in movg_obj.probe_vehs_range[in_laneID][lane_movg][arr_range]:
                                            arr_ranges_TT[arr_range][lane].append(veh_obj.travel_time[veh_num][lane_ind])
                                            if arr_range == 0:
                                                self.tc_edge_movgs_tt[in_edgID][lane_movg] += np.ceil(
                                                    veh_obj.travel_time[veh_num][lane_ind] / num_probes)

                        # Travel times for the lane are the average of each estimation for each movement group
                        for arr_range in range(0, num_ranges):
                            estTravelTime = np.nanmean(arr_ranges_TT[arr_range][lane])
                            lane_CAVs_tts = []
                            CAV_tts_arrived_t = 0
                            for movgID in self.inLaneID_to_movgIDs[in_laneID]:
                                # But consider the travel time of last CAV on the lane for using each of its mov. groups
                                CAVtravelTime = self.lane_cavs_metrics[in_laneID]["travel_times"][movgID][2]
                                prevTravelTime = self.lane_cavs_metrics[in_laneID]["travel_times"][movgID][1]
                                adjusted_CAVtravelTime = (CAVtravelTime * estTravelTime) / prevTravelTime
                                lane_CAVs_tts.append(adjusted_CAVtravelTime)
                                if self.lane_cavs_metrics[in_laneID]["travel_times"][movgID][0] > CAV_tts_arrived_t:
                                    CAV_tts_arrived_t = self.lane_cavs_metrics[in_laneID]["travel_times"][movgID][0]
                            # The estimated travel times are a weighted average of CAV last travel time and estimated
                            # by deterministic queue model
                            rge_tts = np.array([np.nanmean(lane_CAVs_tts), estTravelTime])
                            # The weight is based on how many t_max cyles before the CAV had arrived,
                            # the newest the higher weight. Decreasing weight by intervals of 5% per past cycle
                            beginTime = ACTUAL_T + arr_range * LEN_RANGE
                            if np.isnan(rge_tts[0]):
                                # If no lane travel time by CAVs, make weight zero
                                CAV_tts_weight = 0
                            else:
                                CAV_tts_weight = max(1 - int((beginTime - CAV_tts_arrived_t) / jct_obj.t_max[jctID])
                                                     * 0.05, 0)
                            weights = np.array([CAV_tts_weight, 1 - CAV_tts_weight])
                            indices = ~np.isnan(rge_tts)
                            if True in indices:
                                arr_ranges_TT[arr_range][lane] = np.ceil(np.average(rge_tts[indices],
                                                                              weights=weights[indices]))
                            else:
                                arr_ranges_TT[arr_range][lane] = np.nan
                            # The estimated queueing delay is only based on the deterministic estimation (or using)
                            # CAV queue speed data when available
                            arr_ranges_qDelay[arr_range][lane] = np.mean(arr_ranges_qDelay[arr_range][lane])

                    # Travel times on the edge are the weighted average of lane TTs based on number of vehicles
                    for arr_range in range(0, num_ranges):
                        if 1 in [1 for val in arr_ranges_TT[arr_range] if np.isnan(val) == False]:
                            # If at least one travel time information for the arrival range
                            rge_tts = np.array(arr_ranges_TT[arr_range])
                            rge_qdelays = np.array(arr_ranges_qDelay[arr_range])
                            weights = np.array(num_veh_range[arr_range])
                            indices = ~np.isnan(rge_tts)
                            # Travel times are weighted average using the accumulated number of vehicles per range
                            arr_ranges_TT[arr_range] = np.ceil(np.average(rge_tts[indices],
                                                                          weights=weights[indices]))
                            arr_ranges_qDelay[arr_range] = np.ceil(np.average(rge_qdelays[indices],
                                                                              weights=weights[indices]))
                        else:
                            # All lanes of the edge are affected by lane closures use a high number
                            arr_ranges_TT[arr_range] = 10000
                            arr_ranges_qDelay[arr_range] = 0

                    # Ensure FIFO rule and Store the Value
                    for arr_range in range(0, num_ranges):
                        if arr_range > 0:
                            # The departure time of a vehicle arriving at arr_range cannot be earlier than the departure
                            # of a vehicle that arrived at an earlier arrival range.
                            arr_ranges_TT[arr_range] = (max((arr_range - 1) * LEN_RANGE
                                                           + arr_ranges_TT[arr_range - 1],
                                                           arr_range * LEN_RANGE + arr_ranges_TT[arr_range])
                                                        - arr_range * LEN_RANGE)

                        # If transfered travel time to another out going edge, decrease the delays from total travel time
                        self.tc_tt[in_edgID][1][arr_range] += (arr_ranges_TT[arr_range]
                                                               - arr_ranges_qDelay[arr_range])

                        # For measuring accuracy of travel time prediction
                        beginTime = ACTUAL_T + arr_range * LEN_RANGE
                        if in_edgID not in edg_obj.closure_affected_edgeLanes["edges"] \
                        and beginTime >= BEGIN_T + WARM_UP_T + STEP_LENGTH and beginTime <= END_T + STEP_LENGTH:
                            # Acoount Estimated Travel Times, notice the total travel times (including all delays)
                            # is stored, even if transfered travel times
                            tts_accur_vals[in_edgID]["estimation"][beginTime].append(arr_ranges_TT[arr_range])

            # Update travel time in the edg_obj.traffic_net_tt database
            self.updateTravelTimeDB(self.tc_tt, self.traffic_net_tt)

            end12 = time.time()
            timings[12][1] += end12 - start12


    def updateTravelTimeDB(self,
                           edg_tt_data, # travel times to be inserted in the format
                                        # {edgeID: (begin_arr_range, arr_ranges_TT)}
                           tt_db):      # database with edge travel times
        """Update database with edge travel times with received or calculated travel times"""

        start3 = time.time()

        for edgeID in edg_tt_data.keys():
            begin_arr_range = edg_tt_data[edgeID][0]
            arr_ranges_TT = edg_tt_data[edgeID][1]
            tt_db.update({edgeID: (begin_arr_range, arr_ranges_TT)})

        end3 = time.time()
        timings[3][1] += end3 - start3


    def updateQueueLength(self,
                          jct_obj, # TC junction object
                          jctID):  # network junction ID
        """Prepare the message with queue lengths"""

        start14 = time.time()

        # Initialize
        queue_msg = dict()
        num_ranges = int(jct_obj.t_max[jctID] / LEN_RANGE)
        interval_arr_range = [[] for _ in range(0, num_ranges)]
        for dest_jctID in jct_obj.neigh_jcts[jctID]:
            # Queue message: {dest_jctID: {inLane: [ACTUAL_T, queue_per_range]}}
            queue_msg.update({dest_jctID: {inLane: [ACTUAL_T, interval_arr_range[:]] for inLane in
                                           jct_obj.commonFromNeigh_laneIDs[jctID][dest_jctID]}})
            for in_laneID in jct_obj.commonFromNeigh_laneIDs[jctID][dest_jctID]:
                # Get queue prediction for lanes that come from a border junctions
                for arr_range in range(0, num_ranges):
                    queue_msg[dest_jctID][in_laneID][1][arr_range] = self.lane_queue_pred[in_laneID][1][arr_range]

        end14 = time.time()
        timings[14][1] += end14 - start14

        return queue_msg



class Vehicles:
    """Vehicles and their characteristics modelled by the TC"""

    stop_line_mingap = 1 # Minimum stop line min gap distance (when no vehcle in front). In meters.
    thresh_stopped = 0.1 # Min speed in m/s to consider stopped

    def __init__(self, name):
        self.name = name              # object name

        # Definition of parameters for each vehicle type
        # First values correspond to the standard parameters for vehicle types without info
        self.vType_vals = {"ids": ["PC_STD"],
                           "probability": [0],
                           "acc": [2.9],
                           "decel": [7.5],
                           "min_gap_t": [1],
                           "length": [4.3],
                           "min_gap_d": [2.5],
                           "reaction_t": [STD_REACTION_T]}

        # The instances below use list indexes to correspond the vehicle number on self.ids_all. self.instance[veh_num]
        self.ids_all = []             # list of all vehicles modelled by the TC
        self.vClass = []              # veh. class: "CAV/solitary", "CAV/leader", "CAV/follower", "Mod.", or "Art. Probe"
        self.vType = []               # vehicle type with its index: e.g.: [(1, PC_HATCH_EU4), (2, LDV_STD), ...]
        self.accel_cap = []           # vehicle acceleration capability. In m/s^2
        self.decel_ave = []           # vehicle deceleration average. In m/s^2
        self.min_gap_d = []           # veh. min. dist. gap with the rear bumber of vehicle in front when stopped. In m
        self.min_gap_t = []           # veh. min. time gap with the rear bumber of vehicle in front when moving. In s
        self.len = []                 # vehicle length. In meters
        self.speed_factor = []        # speed factor to define vehicle desired speed. Float
        self.sharing_route_bool = []  # boolean if vehicle shares its route to RSUs.
        self.probe_arr_range = []     # define which arrival range the vehicle is a probe, if not a probe use None

        # Additionally, each vehicle may have inner lists for each entry of its ids_all index, corresponding to a
        # value specific for position on its route (which is represented on edge_route,lane_route and jct_route)
        # self.instance[veh_num][lane_ind] or self.instance[veh_num][edge_ind] or self.instance[veh_num][jct_ind]
        self.edge_route = []         # lane vehicle is running on
        self.lane_route = []         # lane vehicle is running on
        self.jct_route = []          # vehicle route passing junctions
        self.movgID = []             # list of veh. mov. group(s) to use at the junction (jct_ind)/inc. lane(lane_ind)
        self.in_lane_arr_speed = []  # vehicle arrival speed on incoming lane of a junction. In m/s
        self.in_lane_arr_t = []      # vehicle arrival time on incoming lane of a junction. In s
        self.in_lane_arr_dist = []   # vehicle arrival distance on incoming lane of a junction. In s
                                     # if future arrival, it corresponds to the lane length,
                                     # otherwise where the vehicle is generated
        self.phase_start_t = []      # veh. start moving time at analyzing phase of its lane_ind mov. group. In s
        self.phase_end_t = []        # as above but the end time, either when it crosses the stop line or end of green
        self.phase_start_dist = []   # veh. start moving dist. at analyzing phase of its lane_ind mov. group. In m
        self.phase_end_dist = []     # as above but the final dist. of the movement within the phase
        self.phase_start_speed = []  # veh. start speed at analyzing phase of its lane_ind mov. group. In m/s
        self.phase_end_speed = []    # as above but the final speed of the movement within the phase
        self.len_crossed_t = []      # vehicle time taken to cross its whole length. In s
        self.update_dist = []        # vehicle dist. at the time of next algorithm update
        self.update_speed = []       # vehicle speed at the time of next algorithm update
        self.update_t = []           # vehicle time at the time of next algorithm update
        self.travel_time = []        # vehicle travel time from arriving its in_lane to arriving to its out_lane. In s
        self.connection = []         # vehicle connection to take at the junction
        self.crossing_bool = []      # boolean value to indicate if vehicle has crossed the stop line at the junction

        self.detailed_tts = []

        # Other Instances
        self.vehs2del = set([])   # set of vehicles set to be deleted at last algorithm order interaction
        self.cavs_metrics_samples = {"travel_times": dict()} # samples of lane metrics of CAVs
                                                             # and previous travel time estimated by LLR for movgID. Ex:
                                                             # {"travel_times": {vehID: [edgeID_arrived,
                                                             #                           time_CAV_arrived,
                                                             #                           in_laneID_arrived,
                                                             #                           {movgID:
                                                             #                            prevTravelTime]}}}

    def ResetVehsOnLane(self,
                        jct_obj,    # TC junction object
                        edg_obj,    # TC edge object
                        tc_it,      # algorithm order interaction
                        jct_order): # order of junctions to be explored. 1 or -1
        """Delete non-CAVs that will be running on a lane which will be in the list of lanes order on the next
        algorithm lanes order interaction. Also delete vehicles on self.vehs2del if last interaction of the algorithm
        update, and all artificial probe vehicles"""

        start15 = time.time()
        start16 = time.time()

        all_veh_num = set(range(0, len(self.ids_all)))
        vehs_to_remove_from_lanes = set([])
        vehs_to_remain = set([])
        alreadyOnLaneVehs = set([])
        lanes2reset = set()

        if tc_it == TC_IT_MAX:
            # ff last interaction of an algorithm update
            for jctID in jct_obj.ids:
                for laneID in jct_obj.in_laneIDs[jctID]:
                    # Reset lane arrivals
                    edg_obj.lane_arrs_done[laneID] = 0

            for laneID in (lane for lane in edg_obj.lane_ids if edg_obj.inLaneID_to_jctID.has_key(lane) == False):
                # reset the list of vehicles on all lanes modelled by the TC which are only outgoing lanes
                # such lanes either send outflow to border junctions or are the end of modelled area
                edg_obj.vehs_on_lane[laneID] = []

        # lane to reset its vehicles and delete them will be the one which will be in the order on the next
        # algorithm order interaction
        if (TC_IT_MAX == 2 and ((jct_order == 1 and tc_it == TC_IT_MAX) or (jct_order == -1 and tc_it != TC_IT_MAX))) \
            or (TC_IT_MAX == 1 and jct_order == -1):
            for jctID in jct_obj.ids:
                lanes2reset.update([lane for lane in jct_obj.in_laneIDs[jctID]
                                    if lane in jct_obj.lanesForwardOrder[jctID] if edg_obj.lane_arrs_done[lane] == 0])
        else:
            for jctID in jct_obj.ids:
                lanes2reset.update([lane for lane in jct_obj.in_laneIDs[jctID]
                                    if lane in jct_obj.lanesReverseOrder[jctID] if edg_obj.lane_arrs_done[lane] == 0])

        for jctID in jct_obj.ids:
            for laneID in lanes2reset:
                # Delete vehicles on lanes to reset
                vehs_to_remove_from_lanes.update([veh for veh,_ in edg_obj.vehs_on_lane[laneID]])

            for laneID in set(edg_obj.lane_ids).difference(lanes2reset):
                # Vehicle on lanes that will not be in order on the next interaction should remain, except Art. Probes
                vehs_to_remain.update([veh for veh,_ in edg_obj.vehs_on_lane[laneID]
                                           if self.vClass[veh] != "Art. Probe"])

            for laneID in jct_obj.in_laneIDs[jctID]:
                # Vehicles that will be already on a lane on the next algorithm update should remain
                alreadyOnLaneVehs.update([veh for veh,_ in edg_obj.already_on_lane[laneID]])

        if tc_it == TC_IT_MAX:
            vehs_to_remove_from_lanes.update([veh for veh in all_veh_num if self.vClass[veh] == "Art. Probe"])
            all_vehs_all_inLanes = set([])
            for jctID in jct_obj.ids:
                for laneID in jct_obj.in_laneIDs[jctID]:
                    # Get all vehicles on incoming lanes
                    all_vehs_all_inLanes.update([veh for veh,_ in edg_obj.vehs_on_lane[laneID]])

            # delete vehicles that are only on outgoing lanes or none
            vehs_to_remove_from_lanes.update(all_veh_num.difference(all_vehs_all_inLanes))

        vehs_to_remain.update(alreadyOnLaneVehs)
        # Don't delete and remove vehicles set to remain
        vehs_to_remove_from_lanes.difference_update(vehs_to_remain)
        # Delete vehicles on lanes to reset, except CAVs. This is to avoid reinserting CAVs that will be
        # detected again on the next algorithm update. If they are not detected they go to self.vehs2del
        vehs_to_del = {veh for veh in vehs_to_remove_from_lanes if "CAV" not in self.vClass[veh]}
        if tc_it == TC_IT_MAX:
            # If last iteration also delete CAVs not detected at ACTUAL_T or vehicles set to be deleted because of
            # estimated queue length less than modelled already on lane
            vehs_to_remove_from_lanes.update(self.vehs2del)
            vehs_to_del.update(self.vehs2del)
            self.vehs2del = set([])
        # Sort vehicles to del in order for the mapping map_old2new_num work properly
        vehs_to_del = sorted(list(vehs_to_del))

        # For Counting Number of All Vehicles
        if NOGUI == False:
            if tc_it == 1:
                tStep_area_beforeIT1.append(len(self.ids_all))
            else:
                tStep_area_beforeIT2.append(len(self.ids_all))

        end16 = time.time()
        timings[16][1] += end16 - start16

        start17 = time.time()

        # Mapping of new vehicle number given which vehicles will be deleted
        map_old2new_num = [veh_num - len([veh for veh in vehs_to_del if veh < veh_num]) for veh_num in all_veh_num]

        end17 = time.time()
        timings[17][1] += end17 - start17

        for old_veh_num in vehs_to_del:
            veh_num = map_old2new_num[old_veh_num]
            # Delete vehicle
            self.delVehAttrs(veh_num)

        start19 = time.time()

        # For Counting Number of All Vehicles
        if NOGUI == False:
            if tc_it == 1:
                tStep_area_afterIT1.append(len(self.ids_all))
            else:
                tStep_area_afterIT2.append(len(self.ids_all))

        if tc_it != TC_IT_MAX:
            # as some vehicles were deleted on this interaction, update veh. numbers on self.vehs2del
            copied_vehs2del = self.vehs2del.copy()
            self.vehs2del = set([])
            for old_veh_num in copied_vehs2del:
                self.vehs2del.add(map_old2new_num[old_veh_num])

        for laneID in edg_obj.lane_ids:
            # update vehicles on the instance vehs_on_lane
            copied_vehs_on_lane = edg_obj.vehs_on_lane[laneID][:]
            num_deleted = 0
            for veh_ind,(veh_num, lane_ind) in enumerate(copied_vehs_on_lane):
                if vehs_to_remove_from_lanes.__contains__(veh_num) or \
                (tc_it == TC_IT_MAX and veh_num not in alreadyOnLaneVehs and laneID not in lanes2reset):
                    # delete vehicle if is set to be removed from lanes
                    # or vehicle is based on information of the last interaction
                    del edg_obj.vehs_on_lane[laneID][veh_ind - num_deleted]
                    num_deleted += 1
                else:
                    # update veh. number on the lane
                    edg_obj.vehs_on_lane[laneID][veh_ind - num_deleted] = (map_old2new_num[veh_num], lane_ind)

        for jctID in jct_obj.ids:
            for laneID in jct_obj.in_laneIDs[jctID]:
                # update vehs. numbers of the instance with vehicles already on lane
                copied_already_on_lane = edg_obj.already_on_lane[laneID][:]
                for veh_ind, (veh_num,lane_ind) in enumerate(copied_already_on_lane):
                    edg_obj.already_on_lane[laneID][veh_ind] = (map_old2new_num[veh_num], lane_ind)


        end19 = time.time()
        timings[19][1] += end19 - start19

        end15 = time.time()
        timings[15][1] += end15 - start15


    def delVehAttrs(self,
                    veh_num,                         # vehicle number, index of vehID on self.ids_all
                    edge_ind=None):                  # index of vehicle's route
        """Delete vehicles attributes for a specific index of its route, or all info of the vehicle"""

        start18 = time.time()


        if edge_ind != None:
            try:
                del self.lane_route[veh_num][edge_ind]
            except IndexError:
                noValueForInd = 1

            try:
                del self.jct_route[veh_num][edge_ind]
            except IndexError:
                noValueForInd = 1

            try:
                del self.in_lane_arr_t[veh_num][edge_ind]
                del self.in_lane_arr_speed[veh_num][edge_ind]
                del self.in_lane_arr_dist[veh_num][edge_ind]
                del self.phase_start_t[veh_num][edge_ind]
                del self.phase_end_t[veh_num][edge_ind]
                del self.phase_start_dist[veh_num][edge_ind]
                del self.phase_end_dist[veh_num][edge_ind]
                del self.phase_start_speed[veh_num][edge_ind]
                del self.phase_end_speed[veh_num][edge_ind]
                del self.travel_time[veh_num][edge_ind]
                del self.len_crossed_t[veh_num][edge_ind]
                del self.crossing_bool[veh_num][edge_ind]
                del self.update_dist[veh_num][edge_ind]
                del self.update_t[veh_num][edge_ind]
                del self.update_speed[veh_num][edge_ind]

                del self.detailed_tts[veh_num][edge_ind]

            except IndexError:
                noValueForInd = 1

            try:
                del self.connection[veh_num][edge_ind]
            except IndexError:
                noValueForInd = 1

            try:
                del self.movgID[veh_num][edge_ind]
                del self.probe_arr_range[veh_num][edge_ind]
            except IndexError:
                noValueForInd = 1
        else:
            del self.ids_all[veh_num]
            del self.vClass[veh_num]
            del self.vType[veh_num]
            del self.accel_cap[veh_num]
            del self.decel_ave[veh_num]
            del self.min_gap_d[veh_num]
            del self.min_gap_t[veh_num]
            del self.len[veh_num]
            del self.speed_factor[veh_num]
            del self.sharing_route_bool[veh_num]
            del self.edge_route[veh_num]
            del self.lane_route[veh_num]
            del self.jct_route[veh_num]
            del self.in_lane_arr_t[veh_num]
            del self.in_lane_arr_speed[veh_num]
            del self.in_lane_arr_dist[veh_num]
            del self.phase_start_t[veh_num]
            del self.phase_end_t[veh_num]
            del self.phase_start_dist[veh_num]
            del self.phase_end_dist[veh_num]
            del self.phase_start_speed[veh_num]
            del self.phase_end_speed[veh_num]
            del self.travel_time[veh_num]
            del self.len_crossed_t[veh_num]
            del self.crossing_bool[veh_num]
            del self.update_dist[veh_num]
            del self.update_t[veh_num]
            del self.update_speed[veh_num]
            del self.connection[veh_num]
            del self.movgID[veh_num]
            del self.probe_arr_range[veh_num]

            del self.detailed_tts[veh_num]

        end18 = time.time()
        timings[18][1] += end18 - start18


    def genCAVOnLane(self,
                     jct_obj,  # TC junction object
                     movg_obj, # TC mov. group object
                     edg_obj): # TC edge object
        """Generate CAVs running on one of the incoming lanes modelled by the TC"""

        global e_ind, nonCAV_num, nonCAV_edge_ind
        start20 = time.time()

        # Get vehicle numbers of the detected CAVs from the previous algorithm update
        all_prev_upd_CAVs = set([veh for veh in range(0, len(self.ids_all)) if "CAV" in self.vClass[veh]])
        # Initialize
        detectedCAVs = set([])

        for jctID in jct_obj.ids:
            for vehID in jct_obj.cavs_in_range[jctID].keys():
                # Generate vehicles based on the incoming lane they are running on
                RSU_id = jct_obj.cavs_in_range[jctID][vehID][0]
                in_laneID = jct_obj.cavs_in_range[jctID][vehID][1]
                actual_edgeID_route = edg_obj.laneID_to_edgeID[in_laneID]
                if in_laneID == jct_obj.RSU_CAVs_data[RSU_id][vehID][LANE]:
                    CAVnotOnInternalLane = 1
                else:
                    # Vehicle on junction internal lane
                    CAVnotOnInternalLane = 0
                # edg_obj.lane_length[in_laneID] - (SUMO_dist * CAVnotOnInternalLane) because if vehicle on an internal
                # lane, then in_laneID is the suitable outgoing lane of the junction and CAVnotOnInternalLane == 0,
                # so the distance is the lane length, while if otherwise as SUMO uses distance from the beginning of
                # the lane (SUMO_dist), then use the remaining of this distance as distance to stop line
                CAV_dist = edg_obj.lane_length[in_laneID] - (jct_obj.RSU_CAVs_data[RSU_id][vehID][POSITION]
                                                             * CAVnotOnInternalLane)
                # Add GNSS Position Error
                CAV_dist = max(min(CAV_dist + np.random.normal(0, GNSS_ERROR),
                                   edg_obj.lane_length[in_laneID] - jct_obj.RSU_CAVs_data[RSU_id][vehID][LENGTH]), 0)
                try:
                    # Try to find if vehicle is already detected
                    veh_num = self.ids_all.index(vehID)
                    new_veh = 0
                except ValueError:
                    # Vehicle not detected before, try to substitute for a vehicle already on lane of same edge
                    last_dist_diff = float("inf")
                    entered_loop = 0
                    veh_num = -1
                    nonCAVs_already_on_lane = []
                    lane_num = []
                    laneIDs = [in_laneID] + [lane for lane in edg_obj.edgeID_to_laneIDs[actual_edgeID_route]
                                              if lane != in_laneID]
                    while nonCAVs_already_on_lane == [] and len(lane_num) < len(laneIDs):
                        # find non-CAVs to be substituted by the CAV on one of the lanes on the same edge
                        nonCAVs_already_on_lane = [(nonCAV,l_ind)
                                                   for nonCAV,l_ind in edg_obj.already_on_lane[laneIDs[len(lane_num)]]
                                                   if "CAV" not in self.vClass[nonCAV]]
                        lane_num.append(1)

                    for nonCAV_ind, (nonCAV_num,nonCAV_edge_ind) in enumerate(nonCAVs_already_on_lane):
                        entered_loop = 1
                        # Substitute the vehicle wich has the lowest difference of distance between the non-CAV and CAV
                        dist_diff = abs(CAV_dist - self.update_dist[nonCAV_num][nonCAV_edge_ind])
                        if dist_diff > last_dist_diff:
                            # longer distance difference, chose the previous non-CAV to be substituted
                            prev_nonCAV_num,_ = nonCAVs_already_on_lane[nonCAV_ind - 1]
                            veh_num = prev_nonCAV_num
                            break
                        else:
                            # the difference of distance is lower, go to next non-CAV on the lane
                            last_dist_diff = dist_diff
                            veh_num = -1
                    else:
                        if entered_loop == 1 and veh_num == -1:
                            # compare distances to all non-CAVs, take the last one to substitute
                            veh_num = nonCAV_num

                    if veh_num == -1:
                        # If not found any already on lane to substitute, insert new CAV
                        # and add the instance that balances the number of already on lane and CAVs
                        veh_num = len(self.ids_all)
                        self.ids_all.append(vehID)
                        new_veh = 1
                        if ACTUAL_T > BEGIN_T + STEP_LENGTH:
                            edg_obj.CAVsAlreadyOnLaneBalance[actual_edgeID_route] += 1
                    else:
                        # For Counting Number of Vehicles
                        if NOGUI == False:
                            tStep_area_CAVreplacedAlready.append(1)

                        # Substituted a non-CAV already on lane vehicle
                        self.ids_all[veh_num] = vehID
                        new_veh = 0

                    # Set CAV fixed parameters
                    # Assuming that vehicle type name contains either leader or follower when in platoon mode
                    if "leader" in jct_obj.RSU_CAVs_data[RSU_id][vehID][TYPE]:
                        veh_class = "CAV/leader"
                    elif "follower" in jct_obj.RSU_CAVs_data[RSU_id][vehID][TYPE]:
                        veh_class = "CAV/follower"
                    else:
                        veh_class = "CAV/solitary"
                    self.setVehParams(veh_num, veh_class,
                                         jct_obj.RSU_CAVs_data[RSU_id][vehID][TYPE],
                                         jct_obj.RSU_CAVs_data[RSU_id][vehID][LENGTH],
                                         jct_obj.RSU_CAVs_data[RSU_id][vehID][MINGAPD],
                                         jct_obj.RSU_CAVs_data[RSU_id][vehID][MINGAPT],
                                         jct_obj.RSU_CAVs_data[RSU_id][vehID][ACC],
                                        (lambda decel: decel if decel <= 0 else -decel)
                                        (jct_obj.RSU_CAVs_data[RSU_id][vehID][DEC]),
                                         jct_obj.RSU_CAVs_data[RSU_id][vehID][SPEEDFACTOR])

                prev_route_edge_ind = None
                rem_route = []
                if self.sharing_route_bool[veh_num] == 1:
                    # If CAV shares its route to RSU
                    try:
                        # Find route index of the actual edge the CAV is running on
                        prev_route_inds = [e_ind for e_ind, edgeID in enumerate(self.edge_route[veh_num])
                                           if edgeID == actual_edgeID_route]
                        if prev_route_inds == []:
                            # If not found actual edge on CAV previous route
                            raise IndexError
                        elif len(prev_route_inds) == 1:
                            # If actual edge appears only once on the route
                            prev_route_edge_ind = prev_route_inds[0]
                        else:
                            # if actual edge appears more than once on the route, define the route index based on the
                            # ACTUAL_T and the expected start time of the movement to cross the stop line
                            best_match_edge = (float("inf"),None)
                            for e_ind in prev_route_inds:
                                diff_time = abs(self.phase_start_t[veh_num][e_ind] - ACTUAL_T)
                                if diff_time < best_match_edge[0]:
                                    best_match_edge = (diff_time, e_ind)
                            prev_route_edge_ind = e_ind
                    except IndexError:
                        # No edge_route yet or not found actual edge of the route
                        skip = 1

                    # get CAV remaining route
                    if INTERFACE in ("traci", "libsumo"):
                        rem_route = jct_obj.RSU_CAVs_data[RSU_id][vehID][ROUTE][traci.vehicle.getRouteIndex(vehID):]
                    else:
                        LOGGER.debug("Use there another interface to get vehicle's remaining edges of its route")
                    copied_rem_route = rem_route[:]
                    former_edgeID = copied_rem_route[0]
                    for edge_ind, next_edgeID in enumerate(copied_rem_route[1:]):
                        # Get only the remaining edges of the route which are modelled by the traffic controller
                        try:
                            interv = edg_obj.next_edg_prob[former_edgeID]["interv"]
                            if edg_obj.next_edg_prob[former_edgeID][interv].has_key(next_edgeID) == False:
                                raise KeyError
                            former_edgeID = next_edgeID
                        except (KeyError, AttributeError):
                            rem_route = copied_rem_route[:edge_ind + 1]
                            break
                    if actual_edgeID_route != rem_route[0]:
                        # If the actual edge is in fact the outgoing edge of the vehicle which is on an internal lane
                        rem_route = rem_route[1:]
                    try:
                        # If edge route instance of the CAV is already set, update it
                        self.edge_route[veh_num] = list(rem_route)
                    except IndexError:
                        self.edge_route.append(list(rem_route))
                else:
                    # if CAV don't share its route
                    if len(self.edge_route) > veh_num:
                        if actual_edgeID_route in self.edge_route[veh_num]:
                            # Actual edge in the route, make route starting from actual edge
                            prev_route_edge_ind = self.edge_route[veh_num].index(actual_edgeID_route)
                            self.edge_route[veh_num] = self.edge_route[veh_num][prev_route_edge_ind:]
                        else:
                            # Actual edge not in the route, reset route
                            self.edge_route[veh_num] = [actual_edgeID_route]

                if len(self.crossing_bool) > veh_num:
                    # if CAV's attributes already set, delete attributes of the route index before actual edge
                    self.delAttrPrevRouteEdgs(jct_obj, edg_obj, veh_num, "CAV", prev_route_edge_ind)

                # Define vehicle attributes for its route edge/lane/junction
                self.defVehNextEdgeJct(movg_obj, edg_obj, veh_num, 0, jctID, first_edgeID=actual_edgeID_route,
                                       new_veh=new_veh)
                self.defVehConnMovgNextLane(jct_obj, edg_obj, veh_num, jctID, 0,
                                            edg_obj.laneID_to_edgeID[in_laneID], ACTUAL_T, new_veh=1)
                self.defVehAttrOnLane(edg_obj, veh_num, 0, ACTUAL_T,
                                          jct_obj.RSU_CAVs_data[RSU_id][vehID][SPEED], CAV_dist, new_veh=new_veh)

                # Store vehicle mov. group for measuring CAV travel times
                try:
                    chosen_movgID = np.random.choice(self.movgID[veh_num][0])
                    self.cavs_metrics_samples["travel_times"][vehID][3] = \
                        {chosen_movgID: self.cavs_metrics_samples["travel_times"][vehID][3][chosen_movgID]}
                except (KeyError, ValueError):
                    # CAV removed because it is on an internal lane or finishing its trip on the edge
                    skip = 1

                detectedCAVs.add(veh_num)
                if len(rem_route) == 1:
                    # CAV is on the last edge of its trip within TC area, remove from the system
                    # When changed to keys instead of list with veh. number, then just don't include the vehicle,
                    # or remove it when substituting an already on lane vehicle
                    self.vehs2del.add(veh_num)
                    try:
                        for lane_ind, laneID in enumerate(self.lane_route[veh_num]):
                            try:
                                edg_obj.vehs_on_lane[laneID].remove((veh_num, lane_ind))
                            except ValueError:
                                veh_not_on_lane = 1
                            try:
                                edg_obj.already_on_lane[laneID].remove((veh_num, lane_ind))
                            except (ValueError, KeyError):
                                veh_not_on_lane = 1
                    except IndexError:
                        noLaneRouteYet = 1
                else:
                    # CAV not sharing route or CAV in a edge where it can go to another edge modelled by the same TC
                    edg_obj.already_on_lane[self.lane_route[veh_num][0]].append((veh_num, 0))

        not_detected_CAVs = all_prev_upd_CAVs.difference(detectedCAVs)
        # CAVs from the last algorithm update that were not detect at this algorithm update are set to be deleted
        # and they are removed from the lane instances
        self.vehs2del.update(not_detected_CAVs)
        for veh_num in not_detected_CAVs:
            for lane_ind,laneID in enumerate(self.lane_route[veh_num]):
                try:
                    edg_obj.vehs_on_lane[laneID].remove((veh_num,lane_ind))
                except ValueError:
                    veh_not_on_lane = 1
                try:
                    edg_obj.already_on_lane[laneID].remove((veh_num,lane_ind))
                except (ValueError, KeyError):
                    veh_not_on_lane = 1

        end20 = time.time()
        timings[20][1] += start20 - end20


    def genStartingVehOnLane(self,
                             jct_obj,  # TC junction object
                             movg_obj, # TC mov. group object
                             edg_obj): # TC edge object
        """Generate queueing vehicles or those beginning their trips on the lane"""

        global last_veh_ind, front_dist, back_dist, start_speed, influenced
        start22 = time.time()

        for jctID in jct_obj.ids:
            # Initialize for queue length (in number of vehicles)
            queue_num_diff = {laneID: 0 for laneID in jct_obj.in_laneIDs[jctID]}
            LLR_vehs_queueing = dict()
            for in_laneID in jct_obj.in_laneIDs[jctID]:
                # Define the number of LLR queueing vehicles
                LLR_vehs_queueing.update({in_laneID: [(veh, l_ind) for veh, l_ind in edg_obj.already_on_lane[in_laneID]
                                          if self.phase_start_speed[veh][l_ind] <= self.thresh_stopped]})
                # Get the queue of incoming lane
                if "Mod. C." in jct_obj.jctType[jctID]:
                    if INTERFACE in ("traci", "libsumo"):
                        # Get queue length (number of vehicle) from SUMO adding certain error of measurement
                        queue_num_diff[in_laneID] += getInLaneQueueFromSUMO(jct_obj, edg_obj, self, jctID, in_laneID)
                    else:
                        LOGGER.debug("Use here the output of other queue length estimation sub-system "
                              "(must be in number of vehicles)")
                else:
                    # base_queue is the predicted queue at the moment the CAV defined the queue length
                    base_queue = edg_obj.lane_cavs_metrics[in_laneID]["queue_len"]["by_LLR_pred"]
                    # comparing_queue is the predicted queue at ACTUAL_T
                    comparing_queue = edg_obj.lane_queue_pred[in_laneID][1][0]
                    try:
                        diff_LLR_queue = comparing_queue - base_queue
                        queue_num_diff[in_laneID] += int((edg_obj.lane_cavs_metrics[in_laneID]["queue_len"]
                                                          ["by_CAV"] + diff_LLR_queue)
                                                         / self.vType_vals["length"][0]) - 1
                        # - 1 to avoid adding a vehicle that was used to measure the queue and if base and comparing
                        # are zero, no vehicle has left
                    except TypeError:
                        skip = 1

                # Reduce vehicles expected to be on the lane that are halted during actual time
                queue_num_diff[in_laneID] -= len(LLR_vehs_queueing[in_laneID])

            for in_laneID in jct_obj.in_laneIDs[jctID]:
                # For CAVs, if the real lane is other than in_laneID,
                # then subtract the queue num of this lane and add on the other
                lane_CAVs = [veh for veh,_ in LLR_vehs_queueing[in_laneID] if "CAV" in self.vClass[veh]]
                for veh_num in lane_CAVs:
                    CAVs_real_lane = jct_obj.cavs_in_range[jctID][self.ids_all[veh_num]][1]
                    if CAVs_real_lane != in_laneID:
                        try:
                            queue_num_diff[CAVs_real_lane] -= 1
                            queue_num_diff[in_laneID] += 1
                        except KeyError:
                            not_found_lane = 1

            for in_laneID in jct_obj.in_laneIDs[jctID]:
                if queue_num_diff[in_laneID] < 0:
                    # Excess of modelled queueing vehicles at this algorithm update,
                    # make queueing non-CAVs as running vehicles from the first vehicle which shouldn't be queueing
                    sorted_notQueueing = sorted([(veh, l_ind, self.phase_start_dist[veh][l_ind])
                                                 for veh, l_ind in edg_obj.already_on_lane[in_laneID]
                                                 if self.phase_start_speed[veh][l_ind] > self.thresh_stopped],
                                                key=lambda sort_it: (sort_it[2]))
                    sorted_queueing = sorted([(veh, l_ind, self.phase_start_dist[veh][l_ind])
                                              for veh, l_ind in LLR_vehs_queueing[in_laneID]],
                                             key=lambda sort_it: (sort_it[2]))
                    sorted_vehs_lane = sorted([(veh, l_ind, self.phase_start_dist[veh][l_ind])
                                               for veh, l_ind in edg_obj.already_on_lane[in_laneID]],
                                              key=lambda sort_it: (sort_it[2]))
                    try:
                        if len(sorted_notQueueing) > 0:
                            # Running vehicles on the lane
                            first_running_dist = sorted_notQueueing[0][2]
                            sorted_queueing_after_running = [(veh, l_ind, self.phase_start_dist[veh][l_ind])
                                                             for veh, l_ind in sorted_queueing
                                                             if self.phase_start_dist[veh][l_ind]
                                                             > first_running_dist]
                            if len(sorted_queueing_after_running) > 0:
                                # Vehicles are moving before or within queue, make first vehs. in the queue to be running
                                # Get the speed of the first vehicle running before the queue,
                                # vehicles to be set running will have their speeds based on this vehicle's speed
                                first_running_num = sorted_notQueueing[0][0]
                                first_running_lane_ind = sorted_notQueueing[0][0]
                                running_speed = self.phase_start_speed[first_running_num][first_running_lane_ind]
                                vehs_to_set_running = [(veh, l_ind) for veh, l_ind in sorted_queueing
                                                       if "CAV" not in self.vClass[veh]][:-queue_num_diff[in_laneID]]
                                change_dist = 0
                                maximum_dist = 0
                                first_running_ind = None
                                for v_ind, (veh_num, lane_ind, _) in enumerate(sorted_vehs_lane):
                                    if veh_num == first_running_num:
                                        first_running_ind = v_ind
                                    try:
                                        if maximum_dist == 1:
                                            raise
                                        else:
                                            if (veh_num, lane_ind) in vehs_to_set_running:
                                                # Increase speed o vehicle to be set as running if in front of first
                                                # running, or decrease if after first running vehicle
                                                queue_num_diff[in_laneID] += 1
                                                if first_running_ind == None:
                                                    set_speed = running_speed \
                                                                + (self.accel_cap[veh_num]
                                                                   * self.vType_vals["reaction_t"][0] *
                                                                   ((first_running_dist
                                                                    - self.phase_start_dist[veh_num][lane_ind])
                                                                    / self.vType_vals["length"][0]))

                                                else:
                                                    set_speed = set_speed \
                                                                + (self.decel_ave[veh_num]
                                                                   * self.vType_vals["reaction_t"])
                                                if change_dist == 0:
                                                    first_to_change = 1
                                                    change_dist = 1
                                                else:
                                                    first_to_change = 0
                                            else:
                                                set_speed = self.phase_start_speed[veh_num][lane_ind]
                                            if change_dist == 1:
                                                # Change distances to fit all vehicles
                                                if v_ind == 0:
                                                    for movgID in edg_obj.inLaneID_to_movgIDs[in_laneID]:
                                                        if movg_obj.movg_green_plan[movgID][0][0] <= ACTUAL_T:
                                                            front_dist = self.phase_start_dist[veh_num][lane_ind]
                                                            break
                                                else:
                                                    if first_to_change == 1:
                                                        front_dist = back_dist + self.min_gap_d[veh_num] \
                                                                     + (-set_speed ** 2 /
                                                                        (2 * self.decel_ave[veh_num]))
                                                    else:
                                                        front_dist = back_dist + self.min_gap_d[veh_num] \
                                                                     + self.min_gap_t[veh_num] * set_speed

                                                back_dist = front_dist + self.len[veh_num]
                                                self.phase_start_dist[veh_num][lane_ind] = front_dist
                                                self.phase_start_speed[veh_num][lane_ind] = set_speed
                                                if back_dist > edg_obj.lane_length[in_laneID]:
                                                    maximum_dist = 1
                                                    raise
                                            else:
                                                back_dist = self.phase_start_dist[veh_num][lane_ind] + self.len[veh_num]
                                    except:
                                        # If not enough space, set to delete non-CAV that could not fit
                                        # Remove vehicle to be deleted from the lanes and add it to the vehs2del set
                                        self.vehs2del.add(veh_num)
                                        for l_ind, laneID in enumerate(self.lane_route[veh_num]):
                                            try:
                                                edg_obj.vehs_on_lane[laneID].remove((veh_num, l_ind))
                                            except ValueError:
                                                veh_not_on_lane = 1
                                            try:
                                                edg_obj.already_on_lane[laneID].remove((veh_num, l_ind))
                                            except (ValueError, KeyError):
                                                veh_not_on_lane = 1

                                        # For counting number of all vehicles
                                        if NOGUI == False:
                                            tStep_area_QueueBalance.append(-1)
                            else:
                                # Vehicles are moving after the queue
                                raise
                        else:
                            # No running vehicle
                            raise
                    except:
                        # Make last vehs. in the queue to be running
                        if len(sorted_notQueueing) > 0:
                            running_speed = np.mean([self.phase_start_speed[veh][l_ind]
                                                     for veh,l_ind,_ in sorted_notQueueing])
                        else:
                            num_LLR_queueing = len(LLR_vehs_queueing[in_laneID])
                            rho_max = 1 / (self.vType_vals["length"][0] + self.vType_vals["min_gap_d"][0])
                            rho = (num_LLR_queueing + queue_num_diff[in_laneID]) / edg_obj.lane_length[in_laneID]
                            running_speed = edg_obj.lane_max_speed[in_laneID] * (1 - rho / rho_max)
                        vehs_to_set_running = [(veh, l_ind) for veh, l_ind, _ in sorted_queueing
                                               if "CAV" not in self.vClass[veh]][queue_num_diff[in_laneID]:]
                        maximum_dist = 0
                        for v_ind,(veh_num, lane_ind) in enumerate(vehs_to_set_running):
                            try:
                                if maximum_dist == 1:
                                    raise
                                else:
                                    # The distance of the starting vehicle cannot overlap with a expected vehicle,
                                    # try to find an empty space to put the vehicle between last and next vehicles
                                    queue_num_diff[in_laneID] += 1
                                    if v_ind == 0:
                                        try:
                                            if sorted_vehs_lane[0][0] == veh_num:
                                                for movgID in edg_obj.inLaneID_to_movgIDs[in_laneID]:
                                                    if movg_obj.movg_green_plan[movgID][0][0] <= ACTUAL_T:
                                                        front_dist = self.phase_start_dist[veh_num][lane_ind]
                                                        break
                                                else:
                                                    raise
                                            else:
                                                raise
                                        except:
                                            front_dist = self.phase_start_dist[veh_num][lane_ind] \
                                                         + self.min_gap_d[veh_num] \
                                                         + (-running_speed ** 2 / (2 * self.decel_ave[veh_num]))
                                    else:
                                        front_dist = back_dist + self.min_gap_d[veh_num] \
                                                     + self.min_gap_t[veh_num] * running_speed

                                    back_dist = front_dist + self.len[veh_num]
                                    self.phase_start_dist[veh_num][lane_ind] = front_dist
                                    self.phase_start_speed[veh_num][lane_ind] = running_speed
                                    if back_dist > edg_obj.lane_length[in_laneID]:
                                        maximum_dist = 1
                                        raise
                            except:
                                # If not enough space, set to delete non-CAV that could not fit
                                # Remove vehicle to be deleted from the lanes and add it to the vehs2del set
                                self.vehs2del.add(veh_num)
                                for l_ind, laneID in enumerate(self.lane_route[veh_num]):
                                    try:
                                        edg_obj.vehs_on_lane[laneID].remove((veh_num, l_ind))
                                    except ValueError:
                                        veh_not_on_lane = 1
                                    try:
                                        edg_obj.already_on_lane[laneID].remove((veh_num, l_ind))
                                    except (ValueError, KeyError):
                                        veh_not_on_lane = 1

                                # For counting number of all vehicles
                                if NOGUI == False:
                                    tStep_area_QueueBalance.append(-1)

                elif queue_num_diff[in_laneID] > 0:
                    # Missing queueing vehicles at this algorithm update, make running non-CAVs as queueing vehicles
                    # from the first vehicle which should be queueing
                    sorted_notQueueing = sorted([(veh, l_ind, self.phase_start_dist[veh][l_ind])
                                                 for veh, l_ind in edg_obj.already_on_lane[in_laneID]
                                                 if self.phase_start_speed[veh][l_ind] > self.thresh_stopped],
                                                key=lambda sort_it: (sort_it[2]))
                    vehs_queueing_back_dist = sorted([self.phase_start_dist[veh][l_ind] + self.len[veh]
                                                      for veh, l_ind in LLR_vehs_queueing[in_laneID]])

                    for veh_ind,(veh_num,lane_ind,_) in enumerate(sorted_notQueueing):
                        if "CAV" not in self.vClass[veh_num]:
                            queue_num_diff[in_laneID] -= 1
                            try:
                                self.phase_start_dist[veh_num][lane_ind] = vehs_queueing_back_dist[-1] \
                                                                           + self.min_gap_d[veh_num]
                            except IndexError:
                                # No queueing vehicle, use minimum stop line minimum gap
                                self.phase_start_dist[veh_num][lane_ind] = self.stop_line_mingap
                            self.phase_start_speed[veh_num][lane_ind] = 0
                            vehs_queueing_back_dist.append(self.phase_start_dist[veh_num][lane_ind] + self.len[veh_num])

                        if queue_num_diff[in_laneID] == 0:
                            break

            for in_laneID in jct_obj.in_laneIDs[jctID]:
                # Initialize for vehicles starting their trip on lane
                input_starting = [(), ()]
                maximum_queue = 0
                in_edgeID = edg_obj.laneID_to_edgeID[in_laneID]
                # (num. of vehs., list with initial distances, probability of each distance, reason of starting)
                input_starting[0] = (max(queue_num_diff[in_laneID], 0), [self.stop_line_mingap], [1], "queue")

                # Get vehicles that begin their trip on this incoming lane
                interv = edg_obj.lane_num_beg_vehs[in_laneID]["interv"]
                begTrip_veh_num = edg_obj.lane_num_beg_vehs[in_laneID][interv][0]
                if begTrip_veh_num > 0:
                    # Get distances vehicles start their trip
                    init_dists = edg_obj.lane_num_beg_vehs[in_laneID][interv][1][:]
                    init_dist_probs = edg_obj.lane_num_beg_vehs[in_laneID][interv][2][:]
                    # Get integer part of begTrip_veh_num and for decimal part round using binomial distribution
                    # with sucess rate the decimal part
                    int_starting = math.floor(begTrip_veh_num)
                    dec_starting = begTrip_veh_num - int_starting
                    if dec_starting > 0:
                        dec_starting = np.random.binomial(1, dec_starting)
                    begTrip_veh_num = int_starting + dec_starting
                    # Initialize number of vehicles beginning their trip with CAVs
                    begTrip_veh_num_wCAV = int(begTrip_veh_num)
                    if begTrip_veh_num_wCAV > 0 and edg_obj.CAVsAlreadyOnLaneBalance[in_edgeID] > 0:
                        # Reduce the number of vehicles beginning its trip by the number of previously added CAVs
                        # which were added not substituting an already on lane vehicle
                        begTrip_veh_num_noCAV = begTrip_veh_num_wCAV - edg_obj.CAVsAlreadyOnLaneBalance[in_edgeID]
                        edg_obj.CAVsAlreadyOnLaneBalance[in_edgeID] -= begTrip_veh_num_wCAV
                        edg_obj.CAVsAlreadyOnLaneBalance[in_edgeID] = max(edg_obj.CAVsAlreadyOnLaneBalance[in_edgeID], 0)
                    else:
                        # No vehicles beginning their trip or no previous added CAVs to replace
                        begTrip_veh_num_noCAV = begTrip_veh_num_wCAV

                    if begTrip_veh_num_noCAV >= 1:
                        # (num. of vehs., list with initial distances, probability of each distance, reason of starting)
                        input_starting[1] = (begTrip_veh_num_noCAV, init_dists, init_dist_probs, "beg")
                    else:
                        pass

                for input_ind,input_set in enumerate(input_starting):
                    # First add queueing vehicles and then vehicles beginning their trip on lane
                    if input_set != tuple():
                        num_starting_vehs = input_set[0]
                        init_dists = input_set[1]
                        init_dist_probs = input_set[2]
                        reason = input_set[3]
                        if num_starting_vehs > 0 and maximum_queue == 0:
                            # Get the vehicles expected to be on lane (to be modelled)
                            expected_veh_info = sorted([[veh,
                                                         self.phase_start_dist[veh][l_ind],
                                                         self.phase_start_dist[veh][l_ind] + self.len[veh],
                                                         None,  # available gap distance to previous vehicle
                                                         self.phase_start_speed[veh][l_ind]]
                                                        for veh, l_ind in edg_obj.already_on_lane[in_laneID]],
                                                       key=lambda sort_it: (sort_it[1]))
                            last_veh_ind = -1
                            try:
                                # Calculate gap distance to previous vehicle
                                veh = expected_veh_info[0][0]
                                v_speed = expected_veh_info[0][4]
                                expected_veh_info[0][3] = expected_veh_info[0][1] - (self.min_gap_d[veh]
                                                                                          + (self.min_gap_t[veh]
                                                                                             * v_speed))
                                if expected_veh_info[0][2] >= edg_obj.lane_length[in_laneID]:
                                    # If first vehicle back distance of vehicle reaches the lane length stop adding vehs.
                                    maximum_queue = 1
                                else:
                                    # If not, calculate gap distance for each vehicle to be added
                                    copied_expected_veh_info = expected_veh_info[1:]
                                    for veh_ahead_ind,(veh,v_front,v_back,_,v_speed) in enumerate(copied_expected_veh_info):
                                        v_ahead_back = expected_veh_info[veh_ahead_ind][2]
                                        expected_veh_info[veh_ahead_ind + 1][3] = (v_front - (self.min_gap_d[veh]
                                                                                              + (self.min_gap_t[veh]
                                                                                                 * v_speed))) \
                                                                                  - v_ahead_back
                                        if v_back >= edg_obj.lane_length[in_laneID]:
                                            maximum_queue = 1
                                            break
                            except IndexError:
                                skip = 1

                            if maximum_queue == 0:
                                for start_veh_num in range(1, num_starting_vehs + 1):
                                    # For counting number of all vehicles
                                    if NOGUI == False:
                                        tStep_area_QueueBalance.append(1)

                                    # Set vehicle fixed parameters
                                    veh_num = len(self.ids_all)
                                    self.setVehParams(veh_num)
                                    # Define vehicle attributes of its upcoming lane/edge/junction route
                                    self.defVehNextEdgeJct(movg_obj, edg_obj, veh_num, 0, jctID,
                                                           first_in_laneID=in_laneID, first_edgeID=in_edgeID)
                                    self.defVehConnMovgNextLane(jct_obj, edg_obj, veh_num, jctID, 0,
                                                                edg_obj.laneID_to_edgeID[in_laneID], ACTUAL_T,
                                                                next_lane=in_laneID)
                                    edg_obj.already_on_lane[in_laneID].append((veh_num, 0))
                                    # Define vehicle starting distance (front distance)
                                    if init_dists != ["random"]:
                                        if reason == "beg" or start_veh_num == 1:
                                            # Front distance of vehicle beginning its trip on lane or first queueing
                                            # vehicle, define its starting distance using initial distances
                                            front_dist = np.random.choice(init_dists, p=init_dist_probs)
                                        else:
                                            # When queuing vehicle,
                                            # front dist is based on back distance of last vehicle in the queue
                                            front_dist = back_dist + self.min_gap_d[veh_num]
                                    else:
                                        # If initial distances are random
                                        max_init_dist = max(int(edg_obj.lane_length[in_laneID]
                                                                - (self.min_gap_d[veh_num] + self.len[veh_num])), 0)
                                        if max_init_dist <= 1:
                                            front_dist = 0
                                        else:
                                            front_dist = np.random.choice(range(0, max_init_dist))

                                    # Define vehicle starting rear distance
                                    back_dist = front_dist + self.len[veh_num]

                                    # Define vehicle ID
                                    if reason == "beg":
                                        # ID is based on the incoming lane and distance it begins its trip
                                        vehID = reason + str(front_dist) + "/" + in_laneID + "/" + str(ACTUAL_T)
                                    else:
                                        # ID is based on the incoming lane and position in the order of queuing vehicles
                                        vehID = reason + str(start_veh_num) + "/" + in_laneID + "/" + str(ACTUAL_T)
                                    self.ids_all.append(vehID)

                                    # Initialize
                                    if reason == "beg":
                                        try:
                                            # If vehicle beginning trip on lane find the closest vehicle ahead
                                            last_veh_ind = [veh_ind for veh_ind,(_,_,v_back,_,v_speed)
                                                            in enumerate(expected_veh_info)
                                                            if v_back + self.min_gap_d[veh_num]
                                                               + self.min_gap_t[veh_num] * v_speed <= front_dist][-1]
                                        except IndexError:
                                            last_veh_ind = -1

                                        if last_veh_ind > -1:
                                            # Initialize speed to estimate needed space if a vehicle ahead
                                            speed_last_veh = expected_veh_info[last_veh_ind][4]
                                        else:
                                            # Initialize speed to estimate needed space if no vehicle ahead
                                            speed_last_veh = 4
                                    else:
                                        # if queuing vehicle, initial speed is always zero
                                        speed_last_veh = 0

                                    influenced = 0
                                    # The distance of the starting vehicle cannot overlap with a expected vehicle,
                                    # try to find an empty space to put the vehicle between last and next vehicles
                                    needed_space = self.len[veh_num] + self.min_gap_d[veh_num] \
                                                   + self.min_gap_t[veh_num] * speed_last_veh
                                    try:
                                        space_next_veh = expected_veh_info[last_veh_ind + 1][3]
                                    except IndexError:
                                        space_next_veh = float("inf")

                                    while needed_space > space_next_veh:
                                        last_veh_ind += 1
                                        if reason == "beg":
                                            speed_last_veh = expected_veh_info[last_veh_ind][4]
                                            needed_space = self.len[veh_num] + self.min_gap_d[veh_num] \
                                                           + self.min_gap_t[veh_num] * speed_last_veh
                                        try:
                                            space_next_veh = expected_veh_info[last_veh_ind + 1][3]
                                        except IndexError:
                                            space_next_veh = float("inf")
                                        # if vehicle start distance is changed due a vehicle ahead, it means vehicle
                                        # is influenced by a vehicle ahead
                                        influenced = 1
                                    else:
                                        try:
                                            next_veh_num = expected_veh_info[last_veh_ind + 1][0]
                                            front_next_veh = expected_veh_info[last_veh_ind + 1][1]
                                            speed_next_veh = expected_veh_info[last_veh_ind + 1][4]
                                            if reason == "beg" \
                                            and back_dist > front_next_veh - (self.min_gap_d[next_veh_num]
                                                           + self.min_gap_t[next_veh_num] * speed_next_veh):
                                                # if vehicle is begginng its trip and is influenced by a vehicle after it,
                                                # its speed should be the same of vehicle after it
                                                start_speed = speed_next_veh
                                                # and vehicle after it is following it
                                                front_dist = (front_next_veh - (self.min_gap_d[next_veh_num]
                                                              + self.min_gap_t[next_veh_num] * speed_next_veh)) \
                                                              - self.len[veh_num]
                                                # Update the minimum front distance for the next queueing vehicle
                                                back_dist = front_dist + self.len[veh_num]
                                            else:
                                                raise IndexError
                                        except IndexError:
                                            if influenced == 1:
                                                # if vehicle is influenced by a vehicle ahead,
                                                if reason == "beg":
                                                    # then its speed should be the same of vehicle ahead
                                                    start_speed = expected_veh_info[last_veh_ind][4]
                                                else:
                                                    # if queueing, speed is always zero
                                                    start_speed = 0
                                                # its distance is constrained by the vehicle ahead
                                                back_last_veh = expected_veh_info[last_veh_ind][2]
                                                front_dist = back_last_veh + self.min_gap_d[veh_num] \
                                                             + self.min_gap_t[veh_num] * start_speed
                                                # Update the minimum front distance for the next queueing vehicle
                                                back_dist = front_dist + self.len[veh_num]
                                            else:
                                                if reason == "beg":
                                                    # if vehicle is not influenced, set an arbritary start speed (in m/s)
                                                    start_speed = 4
                                                else:
                                                    # if queueing, speed is always zero
                                                    start_speed = 0

                                    # Define the gap distance between adding vehicle with the vehicle in front (ahead)
                                    if last_veh_ind > -1:
                                        avalDist2vehAhead = (front_dist - (self.min_gap_d[veh_num]
                                                                           + (self.min_gap_t[veh_num] * start_speed))) \
                                                            - expected_veh_info[last_veh_ind][2]
                                    else:
                                        avalDist2vehAhead = front_dist - (self.min_gap_d[veh_num]
                                                                          + (self.min_gap_t[veh_num] * start_speed))

                                    try:
                                        next_veh_num = expected_veh_info[last_veh_ind + 1][0]
                                        front_next_veh = expected_veh_info[last_veh_ind + 1][1]
                                        speed_next_veh = expected_veh_info[last_veh_ind + 1][4]
                                        # Update the gap distance of vehicle behind the newly added one starting
                                        expected_veh_info[last_veh_ind + 1][3] = (front_next_veh
                                                                                  - (self.min_gap_d[next_veh_num]
                                                                                     + (self.min_gap_t[next_veh_num]
                                                                                        * speed_next_veh))) \
                                                                                  - back_dist
                                    except IndexError:
                                        skip = 1
                                    # Add vehicle information in the list of expected vehicles on lane
                                    expected_veh_info.insert(last_veh_ind + 1, [veh_num,
                                                                                front_dist,
                                                                                back_dist,
                                                                                avalDist2vehAhead,
                                                                                start_speed])

                                    # Define vehicle attributes on the current in_laneID
                                    self.defVehAttrOnLane(edg_obj, veh_num, 0, ACTUAL_T, start_speed, front_dist)

                                    if back_dist >= edg_obj.lane_length[in_laneID]:
                                        # if adding a vehicle and the queue is full, stop adding more vehicles
                                        maximum_queue = 1
                                        break


        end22 = time.time()
        timings[22][1] += end22 - start22


    def delAttrPrevRouteEdgs(self,
                             jct_obj,               # TC junction object
                             edg_obj,               # TC edge object
                             veh_num,               # vehicle number (vehID index of self.ids_all)
                             veh_class,             # "CAV/solitary", "CAV/leader", "CAV/follower", "Mod.", or "Art. Probe"
                             prev_route_edge_ind):  # route index of actual edgeID
        """Delete route attributes before certain lane/edge route index and remove vehicle on the lane instances of
        the index and check if remaining route is still the same, if not also delete attributes of later route indexes."""

        start23 = time.time()

        copied_lane_route = self.lane_route[veh_num][:]
        copied_jct_route = self.jct_route[veh_num][:]
        # The previous route will be set as the longest list of lanes or junctions to avoid any missing route
        if len(self.lane_route[veh_num]) >= len(self.jct_route[veh_num]):
            prev_route = self.lane_route[veh_num][:]
        else:
            prev_route = self.jct_route[veh_num][:]
        new_lane_ind = 0
        if prev_route_edge_ind == None:
            # if not given prev_route_edge_ind, delete the attributes of all previous route indexes
            delete_next_inds = 1
        else:
            delete_next_inds = 0
        num_deleted_inds = 0
        try:
            connID = self.connection[veh_num][new_lane_ind]
        except IndexError:
            connID = None
        for prev_route_ind, _ in enumerate(prev_route):
            # get laneID and jctID of the previous route
            try:
                laneID = copied_lane_route[prev_route_ind]
            except IndexError:
                laneID = None
            try:
                jctID = copied_jct_route[prev_route_ind]
            except IndexError:
                jctID = ""
            try:
                if delete_next_inds == 1:
                    # if 1, delete the route index attributes without any other condition
                    raise IndexError
                elif prev_route_ind < prev_route_edge_ind:
                    # delete the route index attributes because before actual edgeID route index
                    raise IndexError
                elif prev_route_ind == prev_route_edge_ind and "CAV" in veh_class:
                    # previous route index the same as actual edgeID index, if CAV remove CAV from the lane instances
                    # in order to avoid CAV present on parallel lanes
                    new_lane_ind += 1
                    raise IndexError
                elif self.sharing_route_bool[veh_num] == 1 and \
                    ((laneID != None and laneID not in edg_obj.edgeID_to_laneIDs[self.edge_route[veh_num][new_lane_ind]])
                     or (connID != None and edg_obj.connID_to_outLaneID[connID]
                         not in edg_obj.edgeID_to_laneIDs[self.edge_route[veh_num][new_lane_ind + 1]])
                     or (jctID != "" and self.edge_route[veh_num][new_lane_ind] not in jct_obj.in_edgeIDs[jctID])):
                    # previous route index after actual edgeID index, delete attributes only if:
                    # a) vehicle is sharing route, because if not it doesn't change future route
                    # b) vehicle's lane on the index doesn't match with the edge
                    # c) outgoing lane of the connection for the index doesn't go to a lane of next edge
                    # d) the edge not match with the incoming edges of the junction
                    delete_next_inds = 1
                    raise IndexError
                else:
                    # update veh_num and lane_ind on the lane instances
                    try:
                        veh_ind = edg_obj.vehs_on_lane[laneID].index((veh_num,prev_route_ind))
                        edg_obj.vehs_on_lane[laneID][veh_ind] = (veh_num, new_lane_ind)
                    except (ValueError, KeyError):
                        veh_not_on_lane = 1
                    try:
                        veh_ind = edg_obj.already_on_lane[laneID].index((veh_num,prev_route_ind))
                        edg_obj.already_on_lane[laneID][veh_ind] = (veh_num, new_lane_ind)
                    except (ValueError, KeyError):
                        veh_not_on_lane = 1

                    # update new route index based on each index deleted
                    new_lane_ind += 1
                    try:
                        connID = self.connection[veh_num][new_lane_ind]
                    except IndexError:
                        connID = None
            except (IndexError, KeyError):
                if prev_route_ind != prev_route_edge_ind:
                    # delete attributes of the route index, but when CAV don't delete
                    self.delVehAttrs(veh_num, prev_route_ind - num_deleted_inds)
                    num_deleted_inds += 1
                # remove vehicle from lane instances
                try:
                    edg_obj.vehs_on_lane[laneID].remove((veh_num,prev_route_ind))
                except (ValueError, KeyError):
                    veh_not_on_lane = 1
                try:
                    edg_obj.already_on_lane[laneID].remove((veh_num,prev_route_ind))
                except (ValueError, KeyError):
                    veh_not_on_lane = 1

        end23 = time.time()
        timings[23][1] += end23 - start23


    def setVehParams(self,
                     veh_num,             # vehicle number (vehID index of self.ids_all)
                     veh_class=None,      # "CAV/solitary", "CAV/leader", "CAV/follower", "Mod.", or "Art. Probe"
                     veh_type=None,       # vehicle type
                     veh_len=None,        # vehicle length. In meters
                     min_gap_d=None,      # vehicle expected minimum gap when stopped. In meters
                     min_gap_t=None,      # vehicle minimum gap time. In seconds
                     accel_cap=None,      # vehicle acceleration capability. In m/s^2
                     decel_ave=None,      # vehicle deceleration average. In m/s^2
                     speed_factor=None):  # speed factor to define vehicle desired speed. Float
        """Set parameters of generated vehicles,
        for CAVs it uses values got from the vehicle,
        for non-CAVs it uses standard values based on vehicle type"""

        start24 = time.time()

        if veh_class != None and "CAV" in veh_class:
            # If CAV, get the vehicle type index based on the vehicle type
            try:
                type_ind = self.vType_vals["ids"].index(veh_type)
            except ValueError:
                type_ind = 0

            # get if vehicle shares or not its route with RSU
            if INTERFACE in ("traci", "libsumo"):
                # FOR SIMULATION PURPOSES define if vehicle shares its route based on probability to sharing route
                sharing_bool = np.random.binomial(1, ROUTE_SHARE_PROB)
            else:
                LOGGER.debug("Use here the interface that defines if vehicle shares or not its route with RSU")
        else:
            # If nonCAV, define vehicle type, randomly chosen based on the probability of the vehicle type
            veh_type = np.random.choice(self.vType_vals["ids"], p=self.vType_vals["probability"])
            type_ind = self.vType_vals["ids"].index(veh_type)
            if veh_class == None:
                veh_class = "Mod."
            # std. parameters set based on vehicle type
            veh_len = self.vType_vals["length"][type_ind]
            min_gap_d = self.vType_vals["min_gap_d"][type_ind]
            min_gap_t = self.vType_vals["min_gap_t"][type_ind]
            accel_cap = self.vType_vals["acc"][type_ind]
            decel_ave = -self.vType_vals["decel"][type_ind]
            speed_factor = 1
            sharing_bool = 0

        if veh_num == len(self.vType):
            # append template new veh
            self.vClass.append(veh_class)
            self.vType.append((type_ind, veh_type))
            self.len.append(veh_len)
            self.min_gap_d.append(min_gap_d)
            self.min_gap_t.append(min_gap_t)
            self.accel_cap.append(accel_cap)
            self.decel_ave.append(decel_ave)
            self.speed_factor.append(speed_factor)
            self.sharing_route_bool.append(sharing_bool)
        else:
            # update template
            self.vClass[veh_num] = veh_class
            self.vType[veh_num] = (type_ind, veh_type)
            self.len[veh_num] = veh_len
            self.min_gap_d[veh_num] = min_gap_d
            self.min_gap_t[veh_num] = min_gap_t
            self.accel_cap[veh_num] = accel_cap
            self.decel_ave[veh_num] = decel_ave
            self.speed_factor[veh_num] = speed_factor
            self.sharing_route_bool[veh_num] = sharing_bool



        end24 = time.time()
        timings[24][1] += end24 - start24


    def defVehAttrOnLane(self,
                         edg_obj,            # TC edge object
                         veh_num,            # vehicle number (vehID index of self.ids_all)
                         lane_ind,           # vehicle route index
                         time_pos,           # time vehicle is at distance position
                         speed_pos,          # speed vehicle is at distance position
                         dist_pos,           # distance to stop line vehicle is at the position (within lane or begin)
                         update_time=None,   # time vehicle is at distance position of actual or next algorithm update
                         update_speed=None,  # speed vehicle is at distance position of actual or next algorithm update
                         update_dist=None,   # distance vehicle is at distance position of actual or next algorithm update
                         new_veh=1,          # boolean if it is a new vehicle
                         tc_it=None):        # algorithm order interaction
        """Define vehicle attributes for the route index (lane_ind) inserted"""

        start21 = time.time()

        time_pos = round(time_pos, 2)
        speed_pos = round(speed_pos, 2)
        dist_pos = round(dist_pos, 2)
        try:
            update_time = round(update_time, 2)
            update_speed = round(update_speed, 2)
            update_dist = round(update_dist, 2)
        except:
            NoneVals = 1
        laneID = self.lane_route[veh_num][lane_ind]
        if (veh_num, lane_ind) not in edg_obj.vehs_on_lane[laneID]:
            edg_obj.vehs_on_lane[laneID].append((veh_num, lane_ind))

        if new_veh == 0 and laneID not in lanes_order and update_time == None:
            # If laneID not in lanes_order, it means it was estimate before its arrival or it will be in the next
            # interaction. Therefore, if the update values are None, remove for already_on_lane which could have
            # added on previous interaction and now it was not anymore
            try:
                edg_obj.already_on_lane[laneID].remove((veh_num, lane_ind))
            except ValueError:
                skip = 1

        if new_veh == 1:
            # append template new veh
            # Initialize
            self.phase_end_t.append([None])
            self.phase_end_speed.append([None])
            self.phase_end_dist.append([None])
            self.len_crossed_t.append([None])
            self.travel_time.append([None])
            self.crossing_bool.append([None])

            self.detailed_tts.append([[]])

            # define the arrival time on the lane
            self.in_lane_arr_t.append([time_pos])
            # define the arrival distance on the lane
            self.in_lane_arr_dist.append([dist_pos])
            # define the arrival speed on the lane
            self.in_lane_arr_speed.append([speed_pos])
            # define values the vehicle will of the position vehicle will start move for a future green phase
            self.phase_start_t.append([time_pos])
            self.phase_start_speed.append([speed_pos])
            self.phase_start_dist.append([dist_pos])
            # define values the vehicle will of the position vehicle will start move on the next algorithm update
            if dist_pos != edg_obj.lane_length[laneID]:
                self.update_dist.append([dist_pos])
                self.update_t.append([time_pos])
                self.update_speed.append([speed_pos])
            else:
                self.update_dist.append([update_dist])
                self.update_t.append([update_time])
                self.update_speed.append([update_speed])
        else:
            # update template
            if len(self.crossing_bool[veh_num]) - 1 >= lane_ind:
                if new_veh == 1 or tc_it != TC_IT_MAX or laneID in lanes_order or update_time != None:
                    # define the arrival time on the lane
                    self.in_lane_arr_t[veh_num][lane_ind] = time_pos
                    # define the arrival distance on the lane
                    self.in_lane_arr_dist[veh_num][lane_ind] = dist_pos
                    # define the arrival speed on the lane
                    self.in_lane_arr_speed[veh_num][lane_ind] = speed_pos
                    # define values the vehicle will of the position vehicle will start move for a future green phase
                    self.phase_start_t[veh_num][lane_ind] = time_pos
                    self.phase_start_speed[veh_num][lane_ind] = speed_pos
                    self.phase_start_dist[veh_num][lane_ind] = dist_pos
                    self.crossing_bool[veh_num][lane_ind] = None

                    self.detailed_tts[veh_num][lane_ind] = []

                    # define values the vehicle will of the position vehicle will start move on the next algorithm update
                    if dist_pos != edg_obj.lane_length[laneID]:
                        self.update_dist[veh_num][lane_ind] = dist_pos
                        self.update_t[veh_num][lane_ind] = time_pos
                        self.update_speed[veh_num][lane_ind] = speed_pos
                    else:
                        self.update_dist[veh_num][lane_ind] = update_dist
                        self.update_t[veh_num][lane_ind] = update_time
                        self.update_speed[veh_num][lane_ind] = update_speed
            else:
                # append template for a known veh
                # Initialize
                self.phase_end_t[veh_num].append(None)
                self.phase_end_speed[veh_num].append(None)
                self.phase_end_dist[veh_num].append(None)
                self.len_crossed_t[veh_num].append(None)
                self.travel_time[veh_num].append(None)
                self.crossing_bool[veh_num].append(None)

                self.detailed_tts[veh_num].append([])

                # define the arrival time on the lane
                self.in_lane_arr_t[veh_num].append(time_pos)
                # define the arrival distance on the lane
                self.in_lane_arr_dist[veh_num].append(dist_pos)
                # define the arrival speed on the lane
                self.in_lane_arr_speed[veh_num].append(speed_pos)
                # define values the vehicle will of the position vehicle will start move for a future green phase
                self.phase_start_t[veh_num].append(time_pos)
                self.phase_start_speed[veh_num].append(speed_pos)
                self.phase_start_dist[veh_num].append(dist_pos)
                # define values the vehicle will of the position vehicle will start move on the next algorithm update
                if dist_pos != edg_obj.lane_length[laneID]:
                    self.update_dist[veh_num].append(dist_pos)
                    self.update_t[veh_num].append(time_pos)
                    self.update_speed[veh_num].append(speed_pos)
                else:
                    self.update_dist[veh_num].append(update_dist)
                    self.update_t[veh_num].append(update_time)
                    self.update_speed[veh_num].append(update_speed)


        end21 = time.time()
        timings[21][1] += end21 - start21


    def defVehNextEdgeJct(self,
                          movg_obj,               # TC mov. group object
                          edg_obj,                # TC edge object
                          veh_num,                # vehicle number (vehID index of self.ids_all)
                          actual_edge_ind,        # vehicle route index
                          jctID,                  # network junction ID
                          first_in_laneID=None,  # if new vehicle and it has fixed inc. lane
                          first_movgID=None,      # if new vehicle and it has fixed inc. lane and movgID
                          first_edgeID=None,      # if new vehicle, also needs its edgeID
                          new_veh=1):             # boolean if it is a new vehicle
        """Choose vehicle's next edges and junctions using turning rates"""

        start35 = time.time()

        if new_veh == 1 and self.sharing_route_bool[veh_num] == 0:
            # first edge where vehicle begins its route
            self.edge_route.append([first_edgeID])
        if new_veh == 1:
            # first junction where vehicle begins its route
            self.jct_route.append([jctID])
        elif self.jct_route[veh_num] == []:
            # previously known vehicle, but reseted its parameters
            self.jct_route[veh_num].append(jctID)
        elif self.jct_route[veh_num][0] == None:
            # previously known vehicle, but reseted its parameters
            self.jct_route[veh_num][0] = jctID

        if self.vClass[veh_num] != "Art. Probe":
            # if not an artificial probe vehicle, always define 2 edges ahead the actual one
            num_edges_ahead = 2
        else:
            num_edges_ahead = 1

        for edge_num in range(0, num_edges_ahead):
            # index of next edges ahead
            edge_ind = actual_edge_ind + edge_num
            try:
                current_edg = self.edge_route[veh_num][edge_ind]
                try:
                    next_edge = self.edge_route[veh_num][edge_ind + 1]
                    if next_edge == None:
                        none_next_edge = 1
                    else:
                        none_next_edge = 0
                except IndexError:
                    none_next_edge = 0
            except IndexError:
                current_edg = None
                none_next_edge = 1

            if current_edg != None and none_next_edge == 0:
                if (self.sharing_route_bool[veh_num] == 0 and len(self.edge_route[veh_num]) == edge_ind + 1):
                    # next edge by turning rates
                    interv_last = edg_obj.last_edg_prob[current_edg]["interv"]
                    interv_next = edg_obj.next_edg_prob[current_edg]["interv"]
                    if edge_ind == 0 or np.random.binomial(1,edg_obj.last_edg_prob[current_edg][interv_last]) == 0:
                        # Vehicle will continue to a next edge
                        if edge_num == 0 and first_in_laneID != None:
                            # if it is a starting vehicle, constrain possible next edges based on actual lane
                            possible_nextEdges = set([])
                            if first_movgID == None:
                                # Consider all possible connections
                                for possible_conn in edg_obj.inLaneID_to_connIDs[first_in_laneID]:
                                    possible_nextEdges.add(edg_obj.laneID_to_edgeID[edg_obj.connID_to_outLaneID[possible_conn]])
                            else:
                                # Consider possible connections given probe vehicle already defined movement group
                                for possible_conn in edg_obj.inLaneID_to_connIDs[first_in_laneID]:
                                    if possible_conn in movg_obj.movgID_to_connIDs[first_movgID]:
                                        possible_nextEdges.add(edg_obj.laneID_to_edgeID[edg_obj.connID_to_outLaneID[possible_conn]])
                        else:
                            # not a starting vehicle, no constrain of next edges
                            try:
                                possible_nextEdges = edg_obj.next_edg_prob[current_edg][interv_next].keys()
                            except AttributeError:
                                possible_nextEdges = []

                        possible_nextEdges = list(possible_nextEdges)
                        if possible_nextEdges != []:
                            nextEdges_probs = []
                            for nextEdge in possible_nextEdges:
                                # The next edge is defined based on the probability of each next edge (turning rates)
                                if nextEdge in edg_obj.closure_affected_edgeLanes["edges"] \
                                or nextEdge in self.edge_route[veh_num]:
                                    # If the next edge only leades to a blocked lane, vehicle will not use it,
                                    # and if next junction of the possible connection is already on the route,
                                    # set prob. to zero.
                                    nextEdges_probs.append(0)
                                else:
                                    nextEdges_probs.append(edg_obj.next_edg_prob[current_edg][interv_next][nextEdge])
                            # Setting probability of a connection to zero makes necessary to increase all others
                            sum_prob = sum(nextEdges_probs)
                            if sum_prob != 1.0:
                                remaining_prob = 1.0 - sum_prob
                                non_zero_probs = [nextEdge_prob_ind
                                                  for nextEdge_prob_ind, nextEdge_prob in enumerate(nextEdges_probs)
                                                  if nextEdge_prob != 0]
                                if non_zero_probs == [] and edge_ind == 0:
                                    # First edge and no probability to go to another edge,
                                    # distribute equally to possible connection
                                    distr_prob = remaining_prob / len(nextEdges_probs)
                                    for nextEdge_prob_ind,_ in enumerate(nextEdges_probs):
                                        nextEdges_probs[nextEdge_prob_ind] += distr_prob
                                else:
                                    # distribute the remaining probability among the non-zero probability edges
                                    try:
                                        distr_prob = remaining_prob / len(non_zero_probs)
                                        for nextEdge_prob_ind in non_zero_probs:
                                            nextEdges_probs[nextEdge_prob_ind] += distr_prob
                                    except (ZeroDivisionError, FloatingPointError):
                                        # When no other edge alternative (veh. don't go further)
                                        conns_probs_remains_same = 1
                            try:
                                next_edge = np.random.choice(possible_nextEdges, p=nextEdges_probs)
                            except ValueError:
                                # When no other edge alternative (veh. don't go further)
                                next_edge = None
                                if len(self.edge_route[veh_num]) == 1:
                                    # But if new vehicle, choose any of the alternative, even if they are blocked
                                    # In order to be able to insert the new vehicle and create the congestion
                                    next_edge = np.random.choice(possible_nextEdges)
                        else:
                            # no possible next edge
                            next_edge = None
                        try:
                            # define the next junction based on the next edge
                            out_laneID = edg_obj.edgeID_to_laneIDs[next_edge][0]
                            next_jct = edg_obj.inLaneID_to_jctID[out_laneID]
                        except (KeyError,TypeError):
                            # if next edge is not an incoming edge, none junction
                            next_jct = None

                        self.edge_route[veh_num].append(next_edge)
                        try:
                            self.jct_route[veh_num][edge_ind + 1] = next_jct
                        except IndexError:
                            self.jct_route[veh_num].append(next_jct)
                    else:
                        # Vehicle will finish its route at actual edge
                        next_edge = None
                        next_jct = None
                else:
                    # Next edge already defined (specially when sharing route), define next junction
                    try:
                        next_edge = self.edge_route[veh_num][edge_ind + 1]
                        out_laneID = edg_obj.edgeID_to_laneIDs[next_edge][0]
                        try:
                            next_jct = edg_obj.inLaneID_to_jctID[out_laneID]
                        except KeyError:
                            next_jct = None
                        try:
                            self.jct_route[veh_num][edge_ind + 1] = next_jct
                        except IndexError:
                            self.jct_route[veh_num].append(next_jct)
                    except IndexError:
                        noNextEdge = 1
            else:
                # The edge or the next one is already None
                pass

        end35 = time.time()
        timings[35][1] += end35 - start35


    def defVehConnMovgNextLane(self,
                               jct_obj,            # TC junction object
                               edg_obj,            # TC edge object
                               veh_num,            # vehicle number (vehID index of self.ids_all)
                               jctID,              # network junction ID
                               in_edge_ind,        # vehicle route index
                               next_edge,          # next edgeID
                               lane_decision_t,    # time vehicle will decide which next lane to take on next edge
                               current_lane=None,  # current laneID (if not a new vehicle)
                               current_movgID=None,# current movgID (if not a new vehicle)
                               next_lane=None,     # if new vehicle and it is already set the first lane
                               next_movgID=None,   # if new vehicle and it is already set the first movgID
                               new_veh=1):         # boolean if it is a new vehicle
        """Choose vehicle connection and next mov. group based on next edge, queue length prediction and lane blockage"""

        start36 = time.time()

        global out_queue_range, out_queue_cycle, rem_ranges

        if new_veh == 1:
            try:
                after_next_edge = self.edge_route[veh_num][1]
            except IndexError:
                after_next_edge = None
        else:
            # if not inserting the vehicle, but estimating its departure, use the current analyzing t_max cycle
            try:
                after_next_edge = self.edge_route[veh_num][in_edge_ind + 2]
            except IndexError:
                after_next_edge = None

        try:
            # Define possible outgoing lanes, only those not affected by lane closure
            not_affected_outLanes = [edg_obj.connID_to_outLaneID[conn] for conn in edg_obj.inLaneID_to_connIDs[next_lane]
                                     if edg_obj.connID_to_outLaneID[conn] not in edg_obj.closure_affected_edgeLanes["lanes"]]
        except KeyError:
            not_affected_outLanes = []

        if next_lane != None and not_affected_outLanes != []:
            # if given first lane (next lane) and at least one outgoing lane from next lane leads to a lane not affected
            # by closed lanes
            if next_movgID == None:
                # but not the first movgID, check possible movgIDs that will
                # lead to a outgoing lane which belongs to the after next edge (the next edge is the first edge)
                possible_next_movgIDs = set([])
                for conn in edg_obj.inLaneID_to_connIDs[next_lane]:
                    if edg_obj.connID_to_outLaneID[conn] in edg_obj.edgeID_to_laneIDs[after_next_edge]:
                        possible_next_movgIDs.add(edg_obj.connID_to_movgID[conn])
            else:
                # next_movgID is already defined
                possible_next_movgIDs = [next_movgID]
            best_next_lane_attr = (None, list(possible_next_movgIDs), next_lane, None)
        else:
            # if not set the next lane (or first lane when a new vehicle)
            possible_next_lanes = []
            if after_next_edge != None:
                # Define the cycle after prediction horizon time to consider when checking the outgoing lane queue
                lane_for_check_queue_range = edg_obj.edgeID_to_laneIDs[next_edge][0]
                time_resolution = edg_obj.lane_queue_tLimitRanges[lane_for_check_queue_range][1] \
                                  - edg_obj.lane_queue_tLimitRanges[lane_for_check_queue_range][0]
                out_queue_t_max = edg_obj.lane_queue_tLimitRanges[lane_for_check_queue_range][-1] + time_resolution \
                                  - edg_obj.lane_queue_tLimitRanges[lane_for_check_queue_range][0]
                if new_veh == 1:
                    out_queue_cycle = [0]
                else:
                    # if not inserting the vehicle, but estimating its departure, use the current analyzing t_max cycle
                    out_queue_cycle = [int((jct_obj.t_max_cycle[jctID] * jct_obj.t_max[jctID]) / out_queue_t_max)]

                found_cycle = 0
                # Define suitable cycle when deciding next lane
                while found_cycle == 0:
                    if lane_decision_t < ACTUAL_T + out_queue_t_max + (out_queue_t_max * out_queue_cycle[0]) \
                    or out_queue_cycle[0] == (MAX_T_MAX_CYCLE - 1):
                        found_cycle = 1
                    else:
                        cycle = out_queue_cycle[0]
                        del out_queue_cycle[0]
                        out_queue_cycle.append(cycle + 1)

                rem_ranges = 0
                for range_ind, range_t in enumerate(edg_obj.lane_queue_tLimitRanges[lane_for_check_queue_range]):
                    # define the suitable interval range of the queue prediction to consider to check the queue length
                    if lane_decision_t < range_t + out_queue_t_max * out_queue_cycle[0]:
                        out_queue_range = [range_ind]
                        rem_ranges = len(edg_obj.lane_queue_tLimitRanges[lane_for_check_queue_range]) - (range_ind)
                        break
                else:
                    out_queue_range = [range_ind]
                    rem_ranges = 1

                for parallel_next_lane in edg_obj.edgeID_to_laneIDs[next_edge]:
                    # Parallel outgoing lanes because SUMO sometimes doesn't have connections
                    # from one incoming lane to each parallel out. lane
                    for possible_next_conn in edg_obj.inLaneID_to_connIDs[parallel_next_lane]:
                        if edg_obj.connID_to_outLaneID[possible_next_conn] in edg_obj.edgeID_to_laneIDs[after_next_edge]:
                            # define as a possible next lane if there is a connection from the next lane to a lane
                            # of the after next edge and next lane is not affected by lane closure
                            # it sums the aval_len_queue to capture future available space as well
                            bool_closure_affected = (lambda affected_lanes:
                                                     1 if parallel_next_lane in affected_lanes else 0) \
                                                    (edg_obj.closure_affected_edgeLanes["lanes"])
                            possible_next_lanes.append([parallel_next_lane,
                                                        bool_closure_affected,
                                                        rem_ranges * edg_obj.lane_length[parallel_next_lane] -
                                                        sum(edg_obj.lane_aval_queue_len[parallel_next_lane]
                                                        [out_queue_cycle[0]][out_queue_range[0]:]),
                                                        None])
                            break

                num_possible_lanes = len(possible_next_lanes)
                # use a random coeficient to choose a the next lanes when they all have same queue prediction
                random_coef = np.random.choice(range(0, num_possible_lanes), num_possible_lanes, replace=False)
                for p_lane,_ in enumerate(possible_next_lanes):
                    possible_next_lanes[p_lane][3] = random_coef[p_lane]
                # order possible next lanes, take first the one with no lane affected by lane closures,
                # then shortest queue length and
                # then lowest random coeficient assigned to each possible lane
                ordered_possible_next_lanes = sorted(possible_next_lanes, key=lambda sort_it: (sort_it[1],
                                                                                               sort_it[2],
                                                                                               sort_it[3]))
                next_lane = ordered_possible_next_lanes[0][0]
                try:
                    # Try to get the connection from the current lane to the next_lane
                    current_connID = [conn for conn in edg_obj.inLaneID_to_connIDs[current_lane]
                                      if edg_obj.connID_to_outLaneID[conn] == next_lane
                                      and edg_obj.connID_to_movgID[conn] == current_movgID][0]
                except IndexError:
                    # In case there is no connection to the possible out. lane, assumes lane change
                    current_connID = np.random.choice([conn for conn in edg_obj.inLaneID_to_connIDs[current_lane]
                                                       if edg_obj.connID_to_outLaneID[conn]
                                                       in edg_obj.edgeID_to_laneIDs[next_edge]
                                                       and edg_obj.connID_to_movgID[conn] == current_movgID])
                except KeyError:
                    # No current lane, no current connection (when new vehicle)
                    current_connID = None

                possible_next_movgIDs = set([])
                for conn in edg_obj.inLaneID_to_connIDs[next_lane]:
                    # define the possible next movgIDs as the ones that have incoming lane as the next lane
                    if edg_obj.connID_to_outLaneID[conn] in edg_obj.edgeID_to_laneIDs[after_next_edge]:
                        # if the outgoing lane of the possible connection goes to one lane of after next edge,
                        # define the movgID
                        possible_next_movgIDs.add(edg_obj.connID_to_movgID[conn])

                best_next_lane_attr = (current_connID, list(possible_next_movgIDs), next_lane)
            else:
                # if after next edge is unknown, put vehicle on any lane of the next edge
                out_queue_range = [None]
                possible_next_lanes.extend(edg_obj.edgeID_to_laneIDs[next_edge][:])
                next_lane = np.random.choice(possible_next_lanes)
                try:
                    # Try to get the connection from the current lane to the next_lane using current_movgID
                    current_connID = [conn for conn in edg_obj.inLaneID_to_connIDs[current_lane]
                                      if edg_obj.connID_to_outLaneID[conn] == next_lane
                                      and edg_obj.connID_to_movgID[conn] == current_movgID][0]
                except IndexError:
                    # No direct connection, chose any going to next_edge using current_movgID
                    current_connID = np.random.choice([conn for conn in edg_obj.inLaneID_to_connIDs[current_lane]
                                                       if edg_obj.connID_to_outLaneID[conn]
                                                       in edg_obj.edgeID_to_laneIDs[next_edge]
                                                       and edg_obj.connID_to_movgID[conn] == current_movgID])
                except KeyError:
                    # If thre is no current lane (when new vehicle that has no after next edge)
                    current_connID = None

                best_next_lane_attr = (current_connID, None, next_lane)

        # Store defined values
        try:
            # Replace connection already defined
            self.connection[veh_num][in_edge_ind] = best_next_lane_attr[0]
        except IndexError:
            try:
                # Add for a new edge for a known vehicle
                self.connection[veh_num].append(best_next_lane_attr[0])
            except IndexError:
                # Add for a new vehicle
                self.connection.append([best_next_lane_attr[0]])

        try:
            if new_veh == 1:
                # If new vehicle define first movgID
                self.movgID[veh_num][in_edge_ind] = best_next_lane_attr[1]
            else:
                # If not new vehicle, define movgID of next edge
                self.movgID[veh_num][in_edge_ind + 1] = best_next_lane_attr[1]
        except IndexError:
            try:
                # For a new edge for a known vehicle
                self.movgID[veh_num].append(best_next_lane_attr[1])
                self.probe_arr_range[veh_num].append(None)
            except IndexError:
                # For a new vehicle
                self.movgID.append([best_next_lane_attr[1]])
                self.probe_arr_range.append([None])

        try:
            if new_veh == 1:
                # If new vehicle the next lane is the first lane
                lane_ind = in_edge_ind
            else:
                lane_ind = in_edge_ind + 1
            prev_out_lane = self.lane_route[veh_num][lane_ind]
            # Replace next lane already defined
            self.lane_route[veh_num][lane_ind] = best_next_lane_attr[2]
            if best_next_lane_attr[2] != prev_out_lane:
                # If vehicle lane route was on a different lane, remove vehicle from the different lane instances
                try:
                    edg_obj.vehs_on_lane[prev_out_lane].remove((veh_num,lane_ind))
                except ValueError:
                    veh_not_on_lane = 1
                try:
                    edg_obj.already_on_lane[prev_out_lane].remove((veh_num,lane_ind))
                except (ValueError, KeyError):
                    veh_not_on_lane = 1

        except IndexError:
            try:
                # For a new edge of a known vehicle
                self.lane_route[veh_num].append(best_next_lane_attr[2])
            except IndexError:
                # For a new vehicle
                self.lane_route.append([best_next_lane_attr[2]])

        end36 = time.time()
        timings[36][1] += end36 - start36


    def orderVehLane(self,
                     jct_obj,                   # TC junction object
                     movg_obj,                  # TC mov. group object
                     edg_obj,                   # TC edge object
                     jctID,                     # network junction ID
                     lanes_phase_green_movgs):  # lanes that have a mov. group with green light on actual jctID phase
        """Order detected vehicles based on their distances to stop line (if already on lane)
        or their arrival time otherwise"""

        start25 = time.time()

        arrivals_onLane_to_sort = []
        already_onLane_to_sort = []
        for in_laneID in lanes_phase_green_movgs:
            # Initialize
            edg_obj.lane_last_discharged_veh_ind.update({in_laneID: -1})
            edg_obj.yield2lane_last_oppVeh.update({in_laneID: {pref_lane: 0 for pref_lane in lanes_phase_green_movgs}})
            for veh_num, lane_ind in edg_obj.vehs_on_lane[in_laneID]:
                add_cycle_time = (jct_obj.t_max[jctID] * jct_obj.t_max_cycle[jctID])
                # according to how mov. groups are assigned, vehicle may have more than 1 possible mov. group
                # this happens specially for the case when an outgoing lane of an edge with parallel lanes is used only
                # for one mov. group while the other lanes for other movg. groups
                # then, due the way the algorithm assumes lane change, the vehicle may have more possibilities
                # (suitable_movgID, phase_num, endGreen_t)
                movg_earlier_t = (None, None, float("inf"))
                for movgID in self.movgID[veh_num][lane_ind]:
                    phase_num = movg_obj.movgID_to_phaseNum[movgID]
                    endGreen_t = movg_obj.movg_green_plan[movgID][1][phase_num] + add_cycle_time
                    if endGreen_t <= movg_earlier_t[2]:
                        suitable_movgID = movgID
                        movg_earlier_t = (suitable_movgID, phase_num, endGreen_t)
                # the suitable mov. group is the one that starts/ends the earliest (next green phase)
                suitable_movgID = movg_earlier_t[0]
                phase_num = movg_earlier_t[1]
                endGreen_t = movg_earlier_t[2]
                amber_t = movg_obj.movg_green_plan[suitable_movgID][5][phase_num]
                if jct_obj.t_max_cycle[jctID] == 0:
                    # if first t_max cycle, there is no previous phase than 0
                    prev_phase_num = (lambda ph_num: None if ph_num == 0 else ph_num - 1)(phase_num)
                else:
                    # if cycles after first one, define the add cycle time based on the movgID phase number
                    if phase_num == 0 and len(movg_obj.movg_green_plan[suitable_movgID][1]) > 1:
                        # if movgID phase num is 0, use cycle time of previous cycle and the last phase num
                        add_cycle_time = jct_obj.t_max[jctID] * (jct_obj.t_max_cycle[jctID] - 1)
                        prev_phase_num = phase_num - 2
                    else:
                        prev_phase_num = phase_num - 1
                # Get end time of previous phase green and amber time
                try:
                    prev_phase_endGreen_t = movg_obj.movg_green_plan[suitable_movgID][1][prev_phase_num] + add_cycle_time
                    prev_amber_t = movg_obj.movg_green_plan[suitable_movgID][5][prev_phase_num]
                except TypeError:
                    prev_phase_endGreen_t = ACTUAL_T
                    prev_amber_t = 0

                if (self.crossing_bool[veh_num][lane_ind] in (0, None)
                    and self.phase_start_t[veh_num][lane_ind] < endGreen_t + amber_t) \
                or (self.crossing_bool[veh_num][lane_ind] == 1
                    and self.phase_end_t[veh_num][lane_ind] > prev_phase_endGreen_t + prev_amber_t
                    and self.phase_end_t[veh_num][lane_ind] <= endGreen_t + amber_t):

                    if self.in_lane_arr_dist[veh_num][lane_ind] != edg_obj.lane_length[in_laneID] \
                    or self.in_lane_arr_t[veh_num][lane_ind] < ACTUAL_T:
                        already_onLane_to_sort.append((veh_num, lane_ind, suitable_movgID,
                                                       self.in_lane_arr_t[veh_num][lane_ind],
                                                       self.phase_start_dist[veh_num][lane_ind]))
                    else:
                        arrivals_onLane_to_sort.append((veh_num, lane_ind, suitable_movgID,
                                                        self.in_lane_arr_t[veh_num][lane_ind], None))

        # order by arrival time and distance
        already_onLane_order_veh = sorted(already_onLane_to_sort, key=lambda sort_it: (sort_it[3], sort_it[4]))
        # order by arrival time
        arrivals_onLane_order_veh = sorted(arrivals_onLane_to_sort, key=lambda sort_it: sort_it[3])
        # the order of discharge if first by the already on lane vehicles by their time and distance and then
        # by the arriving ones later based on their arrival time
        order_veh = already_onLane_order_veh + arrivals_onLane_order_veh
        sorted_data = [(v_num, v_lane_ind, v_movgID) for (v_num, v_lane_ind, v_movgID,_ ,_) in order_veh]

        end25 = time.time()
        timings[25][1] += end25 - start25
        return sorted_data


    def getTimeSpeedDistOfThreshDist(self,
                                     start_time,       # initial time in s
                                     max_final_time,   # maximum final time in seconds
                                     start_dist,       # initial distance in meters
                                     final_dist,       # final distance in meters
                                     max_speed,        # maximum speed in m/s
                                     final_speed,      # final speed in m/s
                                     start_speed,      # initial speed in m/s
                                     acc,              # maximum acceleration in m/s^2
                                     decel,            # perceived deceleration in m/s^2
                                     reaction_t):      # reaction time in s
        """Return the threshold distance and its speed and time considering vehicle will accelerate till max_speed
        and then brake, ie. the minimum distance the vehicle needs to achieve v2 given a distance it will maintain
        from the obstacle in front of it"""

        start26 = time.time()

        available_time = max_final_time - start_time
        if start_speed > max_speed:
            start_speed = max_speed
        if final_dist >= start_dist:
            thresh_dist = final_dist
            thresh_dist_speed = final_speed
            thresh_dist_t = start_time
        elif start_time >= max_final_time or reaction_t >= available_time:
            thresh_dist = start_dist
            thresh_dist_speed = start_speed
            thresh_dist_t = start_time
        else:
            # Check if reaction time + decel. movement reaches final speed
            thresh_dist_reaction = (start_speed * reaction_t) \
                                   + ((final_speed ** 2 - start_speed ** 2) / (2 * decel)) + final_dist
            if final_speed < start_speed and thresh_dist_reaction >= start_dist:
                # vehicle is braking already or vehicle will try to accelerate but right after reaction will
                # start braking already
                thresh_dist = start_dist
                thresh_dist_speed = start_speed
                thresh_dist_t = start_time
            else:
                # Check if accelerating till max_speed reaches final speed
                time_acc_max = (max_speed - start_speed) / float(acc)
                dist_acc_max = (max_speed ** 2 - start_speed ** 2) / (2 * float(acc))
                if final_speed < max_speed:
                    thresh_dist = max_speed * reaction_t \
                                  + ((final_speed ** 2 - max_speed ** 2) / (2 * decel)) + final_dist
                else:
                    thresh_dist = max_speed * reaction_t + final_dist
                if dist_acc_max + thresh_dist <= start_dist - final_dist \
                and time_acc_max + reaction_t < available_time:
                    # Vehicle will be able to accelerate until its maximum speed and then decelerate to reach
                    # final_speed and final_dist (this also implies some possible constant movement)
                    rem_dist = start_dist - (dist_acc_max + thresh_dist)
                    thresh_dist_speed = max_speed
                    thresh_dist_t = start_time + (thresh_dist_speed - start_speed) / acc \
                                    + (rem_dist / thresh_dist_speed)
                else:
                    # Vehicle will accelerate until certain speed below desired speed and then start braking,
                    # determine the distance of accelerated movement (to reach desired speed)
                    # but that cannot be higher than the maximum displacement
                    if time_acc_max + reaction_t >= available_time \
                    and dist_acc_max + thresh_dist <= start_dist - final_dist:
                        max_time_acc = min((max_speed - start_speed) / float(acc), available_time)
                        dist_acc = min((max_speed ** 2 - start_speed ** 2) / (2 * float(acc)),
                                       start_dist - final_dist,
                                       start_speed * max_time_acc + (1 / 2.) * acc * max_time_acc ** 2)
                        thresh_dist_speed = math.sqrt(start_speed ** 2 + 2 * acc * dist_acc)
                        time_acc = (thresh_dist_speed - start_speed) / float(acc)
                        max_time_const = available_time - time_acc
                        dist_const = min(start_dist - final_dist - dist_acc, max_time_const * thresh_dist_speed)
                        try:
                            time_const = dist_const / float(thresh_dist_speed)
                        except (ZeroDivisionError, FloatingPointError):
                            time_const = 0
                        thresh_dist_t = start_time + time_acc + time_const
                        thresh_dist = start_dist - dist_acc - dist_const
                    elif final_speed < start_speed:
                        a = decel - acc
                        b = 2 * reaction_t * acc * decel
                        c = final_speed ** 2 * acc - start_speed ** 2 * decel \
                            - ((start_dist - final_dist) * 2 * acc * decel)

                        thresh_dist_speed = max((- b - math.sqrt(b ** 2 - 4 * a * c)) / (2 * a),
                                                (- b + math.sqrt(b ** 2 - 4 * a * c)) / (2 * a))
                        thresh_dist = start_dist - ((thresh_dist_speed ** 2 - start_speed ** 2) / (2 * acc))
                        thresh_dist_t = start_time + (thresh_dist_speed - start_speed) / acc

                        if thresh_dist_t > max_final_time:
                            # If not enough time,
                            # check the maximum speed and dist. vehicle will achieve at max_final_time
                            thresh_dist_t = max_final_time
                            dist_acc = start_speed * available_time + (1 / 2.) * acc * available_time ** 2
                            thresh_dist = start_dist - dist_acc
                            thresh_dist_speed = math.sqrt(start_speed ** 2 + 2 * acc * dist_acc)
                    else:
                        # Check the maximum speed and dist. vehicle will achieve constrained by max_final_time
                        # and available distance
                        max_time_acc = min((final_speed - start_speed) / float(acc), available_time)
                        dist_acc = min((final_speed ** 2 - start_speed ** 2) / (2 * float(acc)),
                                       start_dist - final_dist,
                                       start_speed * max_time_acc + (1 / 2.) * acc * max_time_acc ** 2)
                        thresh_dist_speed = math.sqrt(start_speed ** 2 + 2 * acc * dist_acc)
                        time_acc = (thresh_dist_speed - start_speed) / float(acc)
                        max_time_const = available_time - time_acc
                        dist_const = min(start_dist - final_dist - dist_acc, max_time_const * thresh_dist_speed)
                        try:
                            time_const = dist_const / float(thresh_dist_speed)
                        except (ZeroDivisionError, FloatingPointError):
                            time_const = 0
                        thresh_dist_t = start_time + time_acc + time_const
                        thresh_dist = start_dist - dist_acc - dist_const

        end26 = time.time()
        timings[26][1] += end26 - start26

        return round(thresh_dist_speed, 2), round(thresh_dist_t, 2), round(thresh_dist, 2)


    def getTimeSpeedDistOfAccConstMov(self,
                                      start_time,           # initial time in seconds
                                      max_final_time,       # maximum final time in seconds
                                      start_dist,           # maximum displacement in meters
                                      max_speed,            # maximum (desired) speed in m/s
                                      start_speed,          # initial speed in m/s
                                      acc):                 # maximum acceleration in m/s^2
        """"Return the final time, speed and distance of an accelerated + constant movement constrained by time,
         maximum displacement and vehicle's maximum speed"""

        start27 = time.time()

        if start_speed > max_speed:
            start_speed = max_speed
        available_time = max_final_time - start_time
        if available_time <= 0 or start_dist <= 0:
            final_speed = start_speed
            final_dist = max(start_dist,0)
            final_time = start_time
        else:
            # The time of accelerated movement cannot be higher than the available time
            try:
                max_time_acc = min((max_speed - start_speed) / float(acc), available_time)
                # Determine the distance of accelerated movement (to reach desired speed)
                # that cannot be higher than the maximum displacement
                dist_acc = min((max_speed ** 2 - start_speed ** 2) / (2 * float(acc)), start_dist,
                               start_speed * max_time_acc + (1 / 2.) * acc * max_time_acc ** 2)
                # Speed at reaching the end of distance with accelerated movement
                final_speed = math.sqrt(start_speed ** 2 + 2 * acc * dist_acc)
                # Time of accelerated movement (if no change of speed, it results to 0)
                time_acc = (final_speed - start_speed) / float(acc)
            except (ZeroDivisionError, FloatingPointError):
                max_time_acc = 0
                dist_acc = 0
                time_acc = 0
                final_speed = start_speed

            # Determine the available time for constant movement
            max_time_const = available_time - time_acc
            # Determine the distance of constant movement
            dist_const = min(start_dist - dist_acc, max_time_const * max_speed)
            # Time of constant movement
            try:
                time_const = dist_const / float(final_speed)
            except (ZeroDivisionError, FloatingPointError):
                time_const = 0
            # Time to stop line is the sum of accelerated and constant movement
            final_time = start_time + time_acc + time_const
            # Final distance of the accelerated and constant movement
            final_dist = start_dist - dist_acc - dist_const

        end27 = time.time()
        timings[27][1] += end27 - start27

        return round(final_speed, 2), round(final_time, 2), round(final_dist, 2)


    def getTimeSpeedDistOfDecelMov(self,
                                   start_time,      # initial time in seconds
                                   max_final_time,  # maximum final time in seconds
                                   start_speed,     # initial speed in m/s
                                   decel,           # maximum acceleration in m/s^2
                                   reaction_t,      # reaction time in s
                                   max_dist,        # initial distance in meters
                                   min_dist):       # minimum distance in meters
        """"Return the distance, final speed and deceleration time of a decelerating movement when reach
        certain decelerating time which is either constrained by initial/final time or the time needed for to stop"""

        start28 = time.time()

        available_time = max_final_time - start_time
        if min_dist > max_dist:
            if min_dist > self.stop_line_mingap:
                max_dist = min_dist
            # There is no restriction by distance because the max_dist is the distance to totally brake already
            # The deceleration time will be the minimum between time to get speed zero and available time for stopping
            # Therefore if short distance but enough time, it will be able to stop
            final_speed = 0
            final_time = start_time
            final_dist = max_dist
        elif 0 < available_time < reaction_t:
            # If available time is shorter than the reaction time, just accounts the reaction time movement
            final_speed = start_speed
            final_time = min(start_time + reaction_t, max_final_time)
            mov_time = final_time - start_time
            final_dist = max(max_dist - (start_speed * mov_time), min_dist)
        elif available_time < 0:
            final_speed = start_speed
            final_time = start_time
            final_dist = max_dist
        else:
            # if available time allows deceleration movement
            decel_time = min(available_time - reaction_t, -(start_speed / decel))
            final_speed = start_speed + decel * decel_time
            final_time = start_time + reaction_t + decel_time
            final_dist = max(max_dist - (start_speed * reaction_t + ((final_speed ** 2)
                                                                     - (start_speed ** 2)) / (2 * decel)), min_dist)

        end28 = time.time()
        timings[28][1] += end28 - start28

        return round(final_time, 2), round(final_speed, 2), round(final_dist, 2)


    def defVehMovement(self,
                       jct_obj,             # TC junction object
                       movg_obj,            # TC mov. group object
                       edg_obj,             # TC edge object
                       jctID,               # network junction ID
                       veh_num,             # vehicle number as the index of self.ids_all
                       lane_ind,            # vehicle route index
                       movgID,              # movement Group ID - jctID_movgNum
                       phase_green_movgs,   # mov. groups which have green light at current analyzed junction phase
                       lanes_order,         # lanes with same order of junctions to be explored
                       all_toNeigh_lanes,   # lanes that lead to all neighbouring border junctions
                       tc_it):              # algorithm order interaction
        """Define vehicle movement through the route"""

        start29 = time.time()

        in_laneID = self.lane_route[veh_num][lane_ind]
        if (veh_num, lane_ind) not in edg_obj.veh_already_disharged:
            edg_obj.veh_already_disharged.append((veh_num, lane_ind))
            if in_laneID in lanes_order:
                # as lane order vehicles are always set to discharge, store the index of the last one to discharge
                # on the incoming lane
                veh_ind = edg_obj.lane_last_discharged_veh_ind[in_laneID] + 1
                if self.vClass[veh_num] != "Art. Probe":
                    # artificial probe vehicles are not stored, only modelled vehicles and CAVs
                    edg_obj.lane_last_discharged_veh_ind[in_laneID] = veh_ind
            else:
                # if vehicle on a lane which is not in the order, it needs to find the vehicle which is in front of it
                if self.vClass[veh_num] != "Art. Probe":
                    veh_ind = edg_obj.lane_discharging_vehs_phase[in_laneID].index((veh_num, lane_ind, movgID))
                else:
                    # artificial probe vehicles are not on lane_discharging_vehs_phase,
                    # find its suitable leader vehicle
                    veh_ind = 0
                    for v_ind,(veh, l_ind, _) in enumerate(edg_obj.lane_discharging_vehs_phase[in_laneID]):
                        if self.in_lane_arr_t[veh][l_ind] <= self.in_lane_arr_t[veh_num][lane_ind]:
                            veh_ind = v_ind
                        else:
                            break
            if veh_ind > 0:
                # Get attributes of vehicle in front (ahead)
                veh_ahead_num = edg_obj.lane_discharging_vehs_phase[in_laneID][veh_ind - 1][0]
                veh_ahead_lane_ind = edg_obj.lane_discharging_vehs_phase[in_laneID][veh_ind - 1][1]
                veh_ahead_movgID = edg_obj.lane_discharging_vehs_phase[in_laneID][veh_ind - 1][2]
                veh_ahead_will_cross = self.crossing_bool[veh_ahead_num][veh_ahead_lane_ind]
            else:
                # First vehicle to be discharged on lane
                veh_ahead_num = -1
                veh_ahead_lane_ind = None
                veh_ahead_movgID = None
                veh_ahead_will_cross = 1

            veh_probe_arr_rge = self.probe_arr_range[veh_num][lane_ind]
            if (edg_obj.stop_est_probes[in_laneID] == 1 and veh_probe_arr_rge != None
                and movg_obj.probe_vehs_range[in_laneID][movgID][veh_probe_arr_rge] != []):
                # Make all vehicles on this lane from this vehicle's arr. range till the end as not probe vehicles
                num_ranges = int(jct_obj.t_max[jctID] / LEN_RANGE)
                for movg in movg_obj.probe_vehs_range[in_laneID].keys():
                    for rge in range(veh_probe_arr_rge, num_ranges):
                        if movg_obj.probe_vehs_range[in_laneID][movg][rge] != []:
                            for veh,l_ind in movg_obj.probe_vehs_range[in_laneID][movg][rge]:
                                try:
                                    self.probe_arr_range[veh][l_ind] = None
                                except TypeError:
                                    skip = 1
                            movg_obj.probe_vehs_range[in_laneID][movg][rge] = []

            # Estimate the start and end position, speed, and time of the vehicle during the cycle of
            # the analyzing green phase
            if edg_obj.stop_est_probes[in_laneID] == 0 \
            or (edg_obj.stop_est_probes[in_laneID] == 1 and self.vClass[veh_num] != "Art. Probe"):
                # Discharge vehicle
                (self.crossing_bool[veh_num][lane_ind],
                 out_lane_arr_t,
                 out_lane_arr_speed) = self.estVehConnDepartureTime(jct_obj, movg_obj, edg_obj, jctID,
                                                                     veh_num, veh_ahead_num, veh_ahead_will_cross,
                                                                     lane_ind, veh_ahead_lane_ind, movgID,
                                                                     veh_ahead_movgID, in_laneID, phase_green_movgs,
                                                                     lanes_order, all_toNeigh_lanes, tc_it)
            else:
                # If edg_obj.stop_est_probes[in_laneID] == 1 and self.vClass[veh_num] == "Art. Probe"
                # Don't discharge vehicle
                self.crossing_bool[veh_num][lane_ind] = 0

            if self.crossing_bool[veh_num][lane_ind] == 1:
                # if discharged vehicle will cross the stop line during this junction phase
                connID = self.connection[veh_num][lane_ind]
                if self.vClass[veh_num] != "Art. Probe" and jct_obj.t_max_cycle[jctID] == 0:
                    # Append vehicle to the respective departure during the analyzing phase using the movement group
                    edg_obj.conn_phase_deps[movgID][connID][movg_obj.movgID_to_phaseNum[movgID]].append((veh_num,
                                                                                                         lane_ind))
                    out_laneID = self.lane_route[veh_num][lane_ind + 1]
                    try:
                        next_jct = self.jct_route[veh_num][lane_ind + 1]
                    except IndexError:
                        next_jct = None

                    if in_laneID in lanes_order:
                        # For measuring accuracy of departures modelling
                        in_edgeID = edg_obj.laneID_to_edgeID[in_laneID]
                        rge = int((self.phase_end_t[veh_num][lane_ind] - ACTUAL_T) / LEN_RANGE)
                        rge_begTime = ACTUAL_T + rge * LEN_RANGE
                        if rge_begTime >= BEGIN_T + WARM_UP_T + STEP_LENGTH and rge_begTime <= END_T + STEP_LENGTH:
                            try:
                                deps_accur_vals[in_edgeID]["estimation"][rge_begTime][rge] += 1
                            except IndexError:
                                AfterPlanningHorizon = 1

                        # For counting number of all vehicles
                        if NOGUI == False and self.phase_end_t[veh_num][lane_ind] - ACTUAL_T < LEN_RANGE:
                            LLR_nexttStep_area_updDeps.append(1)

                        # only define the attributes of the next route index if this lane route index is in lanes order
                        # if it would be against order, it would redefine attributes that were already defined on the
                        # same algorithm update
                        if next_jct != None:
                            try:
                                after_next_edge = self.edge_route[veh_num][lane_ind + 2]
                            except IndexError:
                                after_next_edge = None
                            if after_next_edge != None:
                                # If vehicle would be on an internal lane (within the junction) at the next algorithm
                                # update, then set vehicle as would be already on lane of its outgoing lane
                                next_update_veh_crossing = self.phase_end_t[veh_num][lane_ind] \
                                                           < ACTUAL_T  + LEN_RANGE \
                                                            and out_lane_arr_t \
                                                                >= ACTUAL_T + LEN_RANGE \
                                                            and (veh_num,lane_ind) \
                                                                 not in edg_obj.already_on_lane[in_laneID] \
                                                            and (veh_num,lane_ind + 1) \
                                                                 not in edg_obj.already_on_lane[out_laneID]
                                # if this is already last algorithm order interaction and vehicle outgoing lane is
                                # against actual order as well as it would arrive before next algorithm update,
                                # then put on the already on lane to avoid deleting vehicle
                                this_update_outLane_not_order = out_laneID not in lanes_order \
                                                                 and tc_it == TC_IT_MAX \
                                                                 and out_lane_arr_t \
                                                                     < ACTUAL_T + LEN_RANGE \
                                                                 and (veh_num,lane_ind + 1) \
                                                                      not in edg_obj.already_on_lane[out_laneID]

                                # define vehicle next edge and junction
                                self.defVehNextEdgeJct(movg_obj, edg_obj, veh_num, lane_ind + 1, next_jct, new_veh=0)
                                if (next_update_veh_crossing == 1 or this_update_outLane_not_order == 1):
                                    # set route attributes and arrival on outgoing lane with update values
                                    # as well as update time as algorithm next update time
                                    edg_obj.already_on_lane[out_laneID].append((veh_num,lane_ind + 1))
                                    self.defVehAttrOnLane(edg_obj, veh_num, lane_ind + 1,
                                                          out_lane_arr_t,
                                                          out_lane_arr_speed,
                                                          edg_obj.lane_length[out_laneID],
                                                          out_lane_arr_t,
                                                          out_lane_arr_speed,
                                                          edg_obj.lane_length[out_laneID], new_veh=0, tc_it=tc_it)
                                else:
                                    # set arrival on outgoing lane without update values
                                    # (arrival after next algorithm update time)
                                    self.defVehAttrOnLane(edg_obj, veh_num, lane_ind + 1,
                                                          out_lane_arr_t,
                                                          out_lane_arr_speed,
                                                          edg_obj.lane_length[out_laneID], new_veh=0, tc_it=tc_it)
                            else:
                                # Vehicle will finish its route inside the area modelled by the TC
                                # For Counting Number of All Vehicles
                                if NOGUI == False:
                                    tStep_area_finishing.append(1)
                        else:
                            # Vehicle will finish its route because leaving the area modelled by the TC
                            if out_laneID in lanes_order:
                                # Add to out. lane only if lanes order because it is when it sends message to border jct
                                edg_obj.vehs_on_lane[out_laneID].append((veh_num, lane_ind + 1))

                                # Account vehicles leaving from outgoing lane to neighbouring border junctions
                                if out_laneID in all_toNeigh_lanes:
                                    edg_obj.outLane_vehs_atr_arr_range[out_laneID].append((veh_num, out_lane_arr_t,
                                                                                           out_lane_arr_speed))


        end29 = time.time()
        timings[29][1] += end29 - start29


    def estRefNextArrRange(self,
                           edg_obj,         # TC edge object
                           laneID,          # network lane ID
                           veh_num,         # vehicle number (vehID index of self.ids_all)
                           lane_ind,        # vehicle route index
                           event_t,         # initial event (e.g. accelerating) time, in s
                           event_dist,      # initial event (e.g. accelerating) distance, in m
                           event_speed,     # initial event (e.g. accelerating) speed, in m/s
                           acc,             # acceleration/deceleration, in m/s^2
                           reaction_t,      # reaction time in s
                           added_already,   # if vehicle is added already in the list of already on lane
                           max_speed=None,  # maximum speed (desired speed), in m/s
                           min_dist=0):     # minimum distance, in m
        """Estimate the distance and speed vehicle will have at the time of next algorithm update"""

        start30 = time.time()

        if acc >= 0:
            # Accelerating movement
            (mov_speed, mov_t, mov_dist) = self.getTimeSpeedDistOfAccConstMov(
                                                                              event_t,
                                                                              ACTUAL_T + LEN_RANGE,
                                                                              event_dist,
                                                                              max_speed,
                                                                              event_speed,
                                                                              acc)
        else:
            # Decelerating movement
            (mov_t, mov_speed, mov_dist) = self.getTimeSpeedDistOfDecelMov(
                                                                           event_t,
                                                                           ACTUAL_T + LEN_RANGE,
                                                                           event_speed,
                                                                           acc,
                                                                           reaction_t,
                                                                           event_dist,
                                                                           min_dist)
        if mov_t == ACTUAL_T + LEN_RANGE:
            self.update_speed[veh_num][lane_ind] = mov_speed
            self.update_dist[veh_num][lane_ind] = mov_dist
            self.update_t[veh_num][lane_ind] = mov_t
            if added_already == 0:
                edg_obj.already_on_lane[laneID].append((veh_num,lane_ind))
                added_already = 1

        end30 = time.time()
        timings[30][1] += end30 - start30

        return added_already


    def estVehConnDepartureTime(self,
                                jct_obj,              # TC junction object
                                movg_obj,             # TC mov. group object
                                edg_obj,              # TC edge object
                                jctID,                # network junction ID
                                veh_num,              # vehicle number (vehID index of self.ids_all)
                                veh_ahead_num,        # discharged vehicle ahead number (vehID index of self.ids_all)
                                vehAhead_will_cross,  # boolean if vehicle ahead will cross the stop line
                                lane_ind,             # vehicle route index
                                veh_ahead_lane_ind,   # discharged vehicle ahead route index
                                movgID,               # movement Group ID - jctID_movgNum
                                veh_ahead_movgID,     # movement Group ID - jctID_movgNum of vehicle ahead
                                in_laneID,            # network lane ID
                                phase_green_movgs,    # mov. groups which have green light at analyzing junction phase
                                lanes_order,          # lanes with same order of junctions to be explored
                                all_toNeigh_lanes,    # lanes that lead to all neighbouring border junctions
                                tc_it):               # algorithm order interaction
        """Estimate vehicle's time, speed and distance to stop line of the movement during the analyzing green phase
        of its mog. group, as well as the its length crosses the stop line, travel time and store
        its queue back distance"""

        start31 = time.time()

        start32 = time.time()

        start46 = time.time()

        start49 = time.time()

        global turning_speed, veh_was_in_already, veh_in_already, set_already_thresDist, \
               delayed_starting_t

        # Get values
        in_edgeID = edg_obj.laneID_to_edgeID[in_laneID]
        out_edgeID = self.edge_route[veh_num][lane_ind + 1]
        phase_num = movg_obj.movgID_to_phaseNum[movgID]
        redAmber_t = movg_obj.movg_green_plan[movgID][4][phase_num]
        amber_t = movg_obj.movg_green_plan[movgID][5][phase_num]
        movg_status = movg_obj.movg_green_plan[movgID][2][phase_num]
        beginGreen_t = movg_obj.movg_green_plan[movgID][0][phase_num] + (jct_obj.t_max[jctID]
                                                                         * jct_obj.t_max_cycle[jctID])
        if movgID in phase_green_movgs:
            endGreen_t = movg_obj.movg_green_plan[movgID][1][phase_num] + (jct_obj.t_max[jctID]
                                                                           * jct_obj.t_max_cycle[jctID])
        else:
            endGreen_t = beginGreen_t - redAmber_t

        firstVeh_bool = (lambda aheadID: 1 if aheadID == -1 else 0)(veh_ahead_num)
        # Estimate vehicle turning speed based on in_edgeID and out_edgeID
        turning_speed = math.sqrt(FRICTION_COEF * GRAVITY_ACC * edg_obj.next_edg_radius[in_edgeID][out_edgeID]) \
                        * min(1, self.speed_factor[veh_num])

        # Define if vehicle was already set as already on lane
        if self.phase_start_t[veh_num][lane_ind] < ACTUAL_T + LEN_RANGE and self.vClass[veh_num] != "Art. Probe" \
        and in_laneID in lanes_order:
            set_already_thresDist = 0
            if (veh_num, lane_ind) not in edg_obj.already_on_lane[in_laneID]:
                veh_in_already = 0
                veh_was_in_already = 0
            else:
                veh_in_already = 1
                veh_was_in_already = 1
        else:
            veh_in_already = None
            veh_was_in_already = None
            set_already_thresDist = None

        end49 = time.time()
        timings[49][1] += end49 - start49

        start50 = time.time()

        # Estimate the movement with no deceleration by calculating the remaining distance until start braking
        if firstVeh_bool == 1:
            # If there is a not a discharging vehicle ahead during the phase,
            # the vehicle distance to completely stop is stop_line_mingap, in meters.
            begin_speed = 0
            if edg_obj.lane_blockage_dist[in_laneID][0] != None:
                # if lane blocked at certain distance, vehicle will stop there
                begin_dist = max(edg_obj.lane_blockage_dist[in_laneID][0], self.stop_line_mingap)
            elif self.phase_start_dist[veh_num][lane_ind] < self.stop_line_mingap:
                begin_dist = self.phase_start_dist[veh_num][lane_ind]
            else:
                begin_dist = self.stop_line_mingap
        else:
            # If there is a discharging vehicle ahead during the analyzing phase
            # Uses the distance the vehicle ahead will stop at the begining of its movement towards the stop line
            # (which is the start of its next green phase)
            dist_to_vehAhead = self.phase_start_dist[veh_ahead_num][veh_ahead_lane_ind] + self.len[veh_ahead_num] \
                               + self.min_gap_d[veh_num] \
                               + (self.min_gap_t[veh_num] * self.phase_start_speed[veh_ahead_num][veh_ahead_lane_ind])
            begin_speed = self.phase_start_speed[veh_ahead_num][veh_ahead_lane_ind]
            if edg_obj.lane_blockage_dist[in_laneID][0] != None:
                begin_dist = max(edg_obj.lane_blockage_dist[in_laneID][0], dist_to_vehAhead)
            else:
                begin_dist = dist_to_vehAhead

        # get the distance vehicle will have to start braking to stop completely
        (start_threshDist_speed,
         start_threshDist_t,
         start_threshDist_dist) = self.getTimeSpeedDistOfThreshDist(self.phase_start_t[veh_num][lane_ind],
                                                                    endGreen_t,
                                                                    self.phase_start_dist[veh_num][lane_ind],
                                                                    begin_dist,
                                                                    self.speed_factor[veh_num]
                                                                    * edg_obj.lane_max_speed[in_laneID],
                                                                    begin_speed,
                                                                    self.phase_start_speed[veh_num][lane_ind],
                                                                    self.accel_cap[veh_num]
                                                                    * edg_obj.edge_accel_adj[in_edgeID],
                                                                    self.decel_ave[veh_num]
                                                                    * edg_obj.edge_decel_adj[in_edgeID],
                                                                    self.vType_vals["reaction_t"]
                                                                            [self.vType[veh_num][0]])

        # If vehicle is closer than the distance threshold:
        # start_threshDist_dist = self.phase_start_dist[veh_num][lane_ind]

        # And it is braking already, stopped or accelerating to cross
        # start_threshDist_t = self.phase_start_t[veh_num][lane_ind]
        # start_threshDist_speed = self.phase_start_speed[veh_num][lane_ind]

        self.detailed_tts[veh_num][lane_ind].append((start_threshDist_t,
                                                     start_threshDist_t - self.phase_start_t[veh_num][lane_ind],
                                                     "movement from phase start until begin thres. dist."))

        # If the time vehicle will reach the threshold dist. is longer than the begin of next algorithm update,
        # then use this "checkpoint" as reference for initial values for the vehicle at next algorithm update
        if self.phase_start_t[veh_num][lane_ind] < ACTUAL_T + LEN_RANGE \
        and start_threshDist_t >= ACTUAL_T + LEN_RANGE and veh_in_already == 0:
            set_already_thresDist = 1
            veh_in_already = self.estRefNextArrRange(edg_obj,
                                                     in_laneID,
                                                     veh_num,
                                                     lane_ind,
                                                     self.phase_start_t[veh_num][lane_ind],
                                                     self.phase_start_dist[veh_num][lane_ind],
                                                     self.phase_start_speed[veh_num][lane_ind],
                                                     self.accel_cap[veh_num] * edg_obj.edge_accel_adj[in_edgeID],
                                                     self.vType_vals["reaction_t"][self.vType[veh_num][0]],
                                                     veh_in_already,
                                                     max_speed=self.speed_factor[veh_num]
                                                               * edg_obj.lane_max_speed[in_laneID])

        end50 = time.time()
        timings[50][1] += end50 - start50

        end46 = time.time()
        timings[46][1] += end46 - start46

        start47 = time.time()

        if start_threshDist_speed > begin_speed and start_threshDist_speed > self.thresh_stopped \
        and ((firstVeh_bool == 1 and start_threshDist_t < beginGreen_t - redAmber_t) or
        (firstVeh_bool == 0 and start_threshDist_t < self.phase_start_t[veh_ahead_num][veh_ahead_lane_ind])
        or edg_obj.lane_blockage_dist[in_laneID][0] != None):
            # If vehicle's begin speed is lower than start_threshDist_speed and it would arrive at threshold distance
            # before the begin of green or the start of the vehicle ahead, or lane is blocked and vehicle is not stopped
            # Then it will start braking at distance threshold
            # and start accelerating again when begin green or vehicle ahead start moving
            if edg_obj.lane_blockage_dist[in_laneID][0] != None:
                # If lane closed, end of stopping time is the end of planning horizon
                end_stopping_t = ACTUAL_T + jct_obj.t_max[jctID]
            elif firstVeh_bool == 1:
                # If first vehicle, it will stop when vehicle get's right-of-way (green light)
                end_stopping_t = beginGreen_t - redAmber_t
            else:
                # If not first vehicle, it will stop decelerating when vehicle ahead starts moving
                # Notice that later a reaction time is included after the end of stopping time
                end_stopping_t = self.phase_start_t[veh_ahead_num][veh_ahead_lane_ind]

            # Estimate the final position and speed when change to accelerate again
            # (the constant movement during reaction time is included already on the distance threshold calculation)
            (final_decel_t, final_decel_speed, final_decel_dist) = self.getTimeSpeedDistOfDecelMov(
                                                        start_threshDist_t,
                                                        end_stopping_t,
                                                        start_threshDist_speed,
                                                        self.decel_ave[veh_num] * edg_obj.edge_decel_adj[in_edgeID],
                                                        self.vType_vals["reaction_t"][self.vType[veh_num][0]],
                                                        start_threshDist_dist,
                                                        begin_dist)
            # Make "checkpoint" for next algorithm update if suitable
            if start_threshDist_t < ACTUAL_T + LEN_RANGE \
            and final_decel_t >= ACTUAL_T + LEN_RANGE and veh_in_already == 0:
                veh_in_already = self.estRefNextArrRange(edg_obj,
                                                         in_laneID,
                                                         veh_num,
                                                         lane_ind,
                                                         start_threshDist_t,
                                                         start_threshDist_dist,
                                                         start_threshDist_speed,
                                                         self.decel_ave[veh_num] * edg_obj.edge_decel_adj[in_edgeID],
                                                         self.vType_vals["reaction_t"][self.vType[veh_num][0]],
                                                         veh_in_already,
                                                         min_dist=begin_dist)

            self.detailed_tts[veh_num][lane_ind].append((final_decel_t,
                                                         final_decel_t
                                                         - start_threshDist_t,
                                                         "decel. for ahead start movement or green"))

            # Calculate the distance travelled at constant speed during the reaction time of re-acceleration
            # but constrained by the minimum distance
            dist_const = max(min(final_decel_speed * self.vType_vals["reaction_t"][self.vType[veh_num][0]],
                             final_decel_dist - begin_dist), 0)
            if final_decel_dist - dist_const >= 0:

                self.detailed_tts[veh_num][lane_ind].append((end_stopping_t
                                                             + self.vType_vals["reaction_t"][self.vType[veh_num][0]],
                                                             end_stopping_t
                                                             + self.vType_vals["reaction_t"][self.vType[veh_num][0]]
                                                             - final_decel_t,
                                                             "waiting for ahead start movement or green"))

                self.phase_start_t[veh_num][lane_ind] = end_stopping_t \
                                                        + self.vType_vals["reaction_t"][self.vType[veh_num][0]]
                if dist_const != 0:
                    # Estimate the time vehicle will depart at position when it gets green or vehicle ahead starts moving
                    self.phase_start_dist[veh_num][lane_ind] = round(final_decel_dist - dist_const, 2)
                else:
                    # vehicle could stop before end_stopping_t
                    self.phase_start_dist[veh_num][lane_ind] = final_decel_dist
            else:
                # Constant distance is even longer than the remaining to cross
                self.phase_start_dist[veh_num][lane_ind] = 0
                # Estimate the time vehicle will depart at position when it reaches the stop line
                fraction_reaction_t = (final_decel_dist
                                       * self.vType_vals["reaction_t"][self.vType[veh_num][0]]) / dist_const

                self.detailed_tts[veh_num][lane_ind].append((end_stopping_t + fraction_reaction_t,
                                                             end_stopping_t + fraction_reaction_t
                                                             - final_decel_t,
                                                             "waiting for ahead start movement or green"))

                self.phase_start_t[veh_num][lane_ind] = round(end_stopping_t + fraction_reaction_t, 2)

            self.phase_start_speed[veh_num][lane_ind] = final_decel_speed

            start44 = time.time()

            # Account Distance the vehicle stopped to the queue prediction
            if final_decel_speed <= self.thresh_stopped and self.vClass[veh_num] != "Art. Probe" \
                                    and jct_obj.t_max_cycle[jctID] <= 1:
                back_queue = self.phase_start_dist[veh_num][lane_ind] + self.len[veh_num]
                if back_queue <= edg_obj.lane_length[in_laneID]:
                    # Only account distance if within lane length, this avoids queues that never discharge and cause
                    # lanes sending to this lane to not allow any vehicle to go on the next algorithm update
                    suitable_ranges = [range_ind
                                       for range_ind, range_t in enumerate(edg_obj.lane_queue_tLimitRanges[in_laneID])
                                       if final_decel_t < range_t <= self.phase_start_t[veh_num][lane_ind]]
                    for suitable_range in suitable_ranges:
                        edg_obj.lane_queue_pred[in_laneID][1][suitable_range] = np.ceil(back_queue)

            end44 = time.time()
            timings[44][1] += end44 - start44

            # Make "checkpoint" for next algorithm update if suitable
            if final_decel_t < ACTUAL_T + LEN_RANGE \
            and self.phase_start_t[veh_num][lane_ind] >= ACTUAL_T + LEN_RANGE and veh_in_already == 0:
                veh_in_already = 1
                self.update_dist[veh_num][lane_ind] = self.phase_start_dist[veh_num][lane_ind]
                self.update_t[veh_num][lane_ind] = ACTUAL_T + LEN_RANGE
                self.update_speed[veh_num][lane_ind] = self.phase_start_speed[veh_num][lane_ind]
                edg_obj.already_on_lane[in_laneID].append((veh_num, lane_ind))
        else:
            # If vehicle would arrive at thresh. dist. after the time to switch to green or vehicle ahead starts moving,
            # or its begin speed is higher than start_threshDist_speed
            # vehicle will not decelerate then
            if start_threshDist_speed > self.thresh_stopped:
                # If moving when achieving threshold distance
                self.phase_start_t[veh_num][lane_ind] = start_threshDist_t
                self.phase_start_dist[veh_num][lane_ind] = start_threshDist_dist
                self.phase_start_speed[veh_num][lane_ind] = start_threshDist_speed
            else:
                # If vehicle is already stopped at threshold distance, check the time it will start accelerating again
                if firstVeh_bool == 1 and endGreen_t == float("inf") \
                and edg_obj.lane_blockage_dist[in_laneID][0] == None:
                    # If first vehicle and the junction is not traffic controlled, the starting time is the time
                    # to reach threshhold distane
                    starting_t = start_threshDist_t
                else:
                    # If not first vehicle or it is traffic light controlled junctions, uses vehicle in front or
                    # begin green time to decide when vehicle starts
                    if edg_obj.lane_blockage_dist[in_laneID][0] != None:
                        if edg_obj.lane_blockage_dist[in_laneID][1] != None:
                            # If lane is closed and known when it will be open, starting time is time to reopen
                            starting_t = edg_obj.lane_blockage_dist[in_laneID][1]
                        else:
                            # If lane is closed and not known when it will be open, starting time is at the end of t_max
                            starting_t = ACTUAL_T + jct_obj.t_max[jctID]
                    elif firstVeh_bool == 1 and start_threshDist_t < beginGreen_t - redAmber_t:
                        # Starting time is when vehicle gets green time
                        starting_t = beginGreen_t - redAmber_t + self.vType_vals["reaction_t"][self.vType[veh_num][0]]
                    elif firstVeh_bool == 1:
                        # Starting time is when vehicle reaches its threshold distance
                        starting_t = start_threshDist_t
                    else:
                        # Starting time is when vehicle ahead moves or vehicle reaches its threshold distance
                        starting_t = max(start_threshDist_t,
                                         self.phase_start_t[veh_ahead_num][veh_ahead_lane_ind] + self.min_gap_t[veh_num])

                    start44 = time.time()

                    # Account distance the vehicle stopped to the queue prediction
                    if self.vClass[veh_num] != "Art. Probe" and jct_obj.t_max_cycle[jctID] <= 1:
                        back_queue = self.phase_start_dist[veh_num][lane_ind] + self.len[veh_num]
                        if back_queue <= edg_obj.lane_length[in_laneID]:
                            # Only account distance if within lane length, this avoids queues that never discharge and
                            # cause lanes sending to this lane to not allow any vehicle to go on the next algorithm upd.
                            suitable_ranges = [range_ind
                                               for range_ind, range_t
                                               in enumerate(edg_obj.lane_queue_tLimitRanges[in_laneID])
                                               if self.phase_start_t[veh_num][lane_ind] < range_t <= starting_t]
                            for suitable_range in suitable_ranges:
                                edg_obj.lane_queue_pred[in_laneID][1][suitable_range] = np.ceil(back_queue)

                    end44 = time.time()
                    timings[44][1] += end44 - start44

                # Make "checkpoint" for next algorithm update if suitable
                if self.phase_start_t[veh_num][lane_ind] < ACTUAL_T + LEN_RANGE \
                and starting_t >= ACTUAL_T + LEN_RANGE and veh_in_already == 0:
                    veh_in_already = 1
                    self.update_dist[veh_num][lane_ind] = self.phase_start_dist[veh_num][lane_ind]
                    self.update_t[veh_num][lane_ind] = ACTUAL_T + LEN_RANGE
                    self.update_speed[veh_num][lane_ind] = self.phase_start_speed[veh_num][lane_ind]
                    edg_obj.already_on_lane[in_laneID].append((veh_num, lane_ind))
                else:
                    pass

                self.detailed_tts[veh_num][lane_ind].append((starting_t, starting_t
                                                             - start_threshDist_t,
                                                             "already stopped, waiting from green or ahead to move"))

                self.phase_start_t[veh_num][lane_ind] = starting_t

        end47 = time.time()
        timings[47][1] += end47 - start47

        start48 = time.time()

        # Correct the time and speed at phase start constrained by the values of the vehicle ahead
        if firstVeh_bool == 0 and edg_obj.lane_blockage_dist[in_laneID][0] == None:
            min_phase_start_t = self.phase_start_t[veh_ahead_num][veh_ahead_lane_ind] + self.min_gap_t[veh_num]
            if self.phase_start_t[veh_num][lane_ind] < min_phase_start_t:
                # Vehicle acceleration and maximum speed influenced by the vehicle ahead
                stored_phase_start_t = self.phase_start_t[veh_num][lane_ind]
                stored_phase_start_dist = self.phase_start_dist[veh_num][lane_ind]
                stored_phase_start_speed = self.phase_start_speed[veh_num][lane_ind]

                self.detailed_tts[veh_num][lane_ind].append((min_phase_start_t, min_phase_start_t
                                                             - self.phase_start_t[veh_num][lane_ind],
                                                             "correction due veh. ahead " + str(
                                                             (veh_ahead_num, veh_ahead_lane_ind)) +
                                                             " start movement modelling"))

                self.phase_start_t[veh_num][lane_ind] = min_phase_start_t
                self.phase_start_dist[veh_num][lane_ind] = round(self.phase_start_dist[veh_ahead_num][veh_ahead_lane_ind]
                                                           + self.len[veh_ahead_num] + self.min_gap_d[veh_num]
                                                           + (self.min_gap_t[veh_num]
                                                              * self.phase_end_speed[veh_ahead_num][veh_ahead_lane_ind]),
                                                                 2)
                self.phase_start_speed[veh_num][lane_ind] = min(self.phase_start_speed[veh_num][lane_ind],
                                                                self.phase_start_speed[veh_ahead_num][veh_ahead_lane_ind])
                veh_acc = min(self.accel_cap[veh_ahead_num],
                              self.accel_cap[veh_num]) * edg_obj.edge_accel_adj[in_edgeID]
                veh_max_speed = min(self.speed_factor[veh_ahead_num], self.speed_factor[veh_num]) \
                                * edg_obj.lane_max_speed[in_laneID]

                # Account distance the vehicle stopped to the queue prediction
                if self.vClass[veh_num] != "Art. Probe" and jct_obj.t_max_cycle[jctID] <= 1\
                and self.phase_start_speed[veh_num][lane_ind] <= self.thresh_stopped:
                    back_queue = self.phase_start_dist[veh_num][lane_ind] + self.len[veh_num]
                    if back_queue <= edg_obj.lane_length[in_laneID]:
                        # Only account distance if within lane length, this avoids queues that never discharge and
                        # cause lanes sending to this lane to not allow any vehicle to go on the next algorithm upd.
                        suitable_ranges = [range_ind
                                           for range_ind, range_t
                                           in enumerate(edg_obj.lane_queue_tLimitRanges[in_laneID])
                                           if stored_phase_start_t < range_t <= self.phase_start_t[veh_num][lane_ind]]
                        for suitable_range in suitable_ranges:
                            edg_obj.lane_queue_pred[in_laneID][1][suitable_range] = np.ceil(back_queue)

                # Make "checkpoint" for next algorithm update if suitable
                if (veh_in_already == 1 and veh_was_in_already == 0
                    and (set_already_thresDist == 0 or start_threshDist_t == self.phase_start_t[veh_num][lane_ind])) \
                        or (veh_in_already == 0 and stored_phase_start_t < ACTUAL_T + LEN_RANGE
                            and self.phase_start_t[veh_num][lane_ind] >= ACTUAL_T + LEN_RANGE):

                    if (veh_ahead_num, veh_ahead_lane_ind) in edg_obj.already_on_lane[in_laneID]:
                        # If known update values of vehicle ahead, the update values are based on the vehicle ahead
                        self.update_t[veh_num][lane_ind] = ACTUAL_T + LEN_RANGE
                        influenced_update_dist = round(self.update_dist[veh_ahead_num][veh_ahead_lane_ind]
                                                  + self.len[veh_ahead_num]
                                                  + self.min_gap_d[veh_num]
                                                  + (self.min_gap_t[veh_num]
                                                     * self.update_speed[veh_ahead_num][veh_ahead_lane_ind]), 2)
                        if veh_in_already == 0 or influenced_update_dist >= self.update_dist[veh_num][lane_ind]:
                            # If vehicle was not added to the already on lane list or it is influenced,
                            # it cannot be higher than the vehicle in front
                            self.update_speed[veh_num][lane_ind] = self.update_speed[veh_ahead_num][veh_ahead_lane_ind]
                            self.update_dist[veh_num][lane_ind] = influenced_update_dist
                    else:
                        # Not known update values of vehicle ahead, but known the corrected start values, use them
                        # for next algorithm update
                        if self.phase_start_speed[veh_num][lane_ind] <= self.thresh_stopped:
                            self.update_t[veh_num][lane_ind] = ACTUAL_T + LEN_RANGE
                            self.update_dist[veh_num][lane_ind] = self.phase_start_dist[veh_num][lane_ind]
                            self.update_speed[veh_num][lane_ind] = self.phase_start_speed[veh_num][lane_ind]
                        else:
                            # If not stopped at start of movement,
                            # first define the update values with constrained values
                            veh_in_already = self.estRefNextArrRange(edg_obj,
                                                                     in_laneID,
                                                                     veh_num,
                                                                     lane_ind,
                                                                     stored_phase_start_t,
                                                                     stored_phase_start_dist,
                                                                     stored_phase_start_speed,
                                                                     veh_acc,
                                                                     self.vType_vals["reaction_t"]
                                                                     [self.vType[veh_num][0]],
                                                                     veh_in_already,
                                                                     max_speed=veh_max_speed)
                            if veh_in_already == 0:
                                # If vehicle still would arrive before next algorithm update,
                                # then assume the vehicle would have its phase end values at next update
                                self.update_t[veh_num][lane_ind] = ACTUAL_T + LEN_RANGE
                                self.update_dist[veh_num][lane_ind] = self.phase_start_dist[veh_num][lane_ind]
                                self.update_speed[veh_num][lane_ind] = self.phase_start_speed[veh_num][lane_ind]

                    if veh_in_already == 0:
                        edg_obj.already_on_lane[in_laneID].append((veh_num, lane_ind))
                    veh_was_in_already = 1  # needed to avoid entering in the correction of phase_end values
                    veh_in_already = 1
            else:
                # Too far from vehicle ahead, no influence of vehicle ahead
                veh_acc = self.accel_cap[veh_num] * edg_obj.edge_accel_adj[in_edgeID]
                veh_max_speed = self.speed_factor[veh_num] * edg_obj.lane_max_speed[in_laneID]
        else:
            # First vehicle, no influence of vehicle ahead, or if lane blocked, don't correct.
            veh_acc = self.accel_cap[veh_num] * edg_obj.edge_accel_adj[in_edgeID]
            veh_max_speed = self.speed_factor[veh_num] * edg_obj.lane_max_speed[in_laneID]

        end32 = time.time()
        timings[32][1] += end32 - start32

        end48 = time.time()
        timings[48][1] += end48 - start48

        start33 = time.time()

        if movgID in phase_green_movgs and edg_obj.lane_blockage_dist[in_laneID][0] == None \
        and (veh_ahead_movgID == None or veh_ahead_movgID in phase_green_movgs):
            # Estimate time and speed to stop line considering acceleration/deceleration from the time vehicle
            # starts its movement to cross constrained by the end of green time and phase_end_dist of vehicle ahead
            if firstVeh_bool == 1 or vehAhead_will_cross == 1:
                end_dist = 0
            else:
                end_dist = round(min(self.phase_start_dist[veh_num][lane_ind],
                                     self.phase_end_dist[veh_ahead_num][veh_ahead_lane_ind]
                                     + self.len[veh_ahead_num]
                                     + self.min_gap_d[veh_num]
                                     + (self.min_gap_t[veh_num]
                                     * self.phase_end_speed[veh_ahead_num][veh_ahead_lane_ind])), 2)

            # Estimate moviment to stop line containing acceleration or deceleration movement
            if edg_obj.lane_must_stop[in_laneID] == 1:
                veh_min_speed = 0
            else:
                if firstVeh_bool == 0:
                    veh_min_speed = min(turning_speed, veh_max_speed,
                                        self.phase_end_speed[veh_ahead_num][veh_ahead_lane_ind])
                else:
                    veh_min_speed = min(turning_speed, veh_max_speed)

            if self.phase_start_speed[veh_num][lane_ind] > veh_min_speed:
                # If obligatory to stop at stop line or if needed slower speed due turning or vehicle ahead,
                # vehicle will accelerate and/or have constant movement until it start decelerating until end_dist,
                (end_threshDist_speed,
                 end_threshDist_t,
                 end_threshDist_dist) = self.getTimeSpeedDistOfThreshDist(self.phase_start_t[veh_num][lane_ind],
                                                                          endGreen_t,
                                                                          self.phase_start_dist[veh_num][lane_ind],
                                                                          end_dist,
                                                                          veh_max_speed,
                                                                          veh_min_speed,
                                                                          self.phase_start_speed[veh_num][lane_ind],
                                                                          veh_acc,
                                                                          self.decel_ave[veh_num]
                                                                          * edg_obj.edge_decel_adj[in_edgeID],
                                                                          self.vType_vals["reaction_t"]
                                                                                      [self.vType[veh_num][0]])

                # Make "checkpoint" for next algorithm update if suitable
                if self.phase_start_t[veh_num][lane_ind] < ACTUAL_T + LEN_RANGE \
                        and end_threshDist_t >= ACTUAL_T + LEN_RANGE and veh_in_already == 0:
                    veh_in_already = self.estRefNextArrRange(edg_obj,
                                                             in_laneID,
                                                             veh_num,
                                                             lane_ind,
                                                             self.phase_start_t[veh_num][lane_ind],
                                                             self.phase_start_dist[veh_num][lane_ind],
                                                             self.phase_start_speed[veh_num][lane_ind],
                                                             veh_acc,
                                                             self.vType_vals["reaction_t"][self.vType[veh_num][0]],
                                                             veh_in_already,
                                                             max_speed=veh_max_speed)

                self.detailed_tts[veh_num][lane_ind].append((end_threshDist_t, end_threshDist_t
                                                             - self.phase_start_t[veh_num][lane_ind],
                                                             "acc/const. mov. before mov. to cross stop line "
                                                             "with decel."))

                if firstVeh_bool == 0 and ((veh_min_speed == self.phase_end_speed[veh_ahead_num][veh_ahead_lane_ind]
                and end_threshDist_t > self.phase_end_t[veh_ahead_num][veh_ahead_lane_ind])
                                           or turning_speed >= veh_max_speed):
                    # If vehicle ahead was constraining the final speed but the end_threshDist_t is after end mov. of
                    # vehicle ahead, then no deceleration movement happens.
                    (self.phase_end_speed[veh_num][lane_ind],
                     self.phase_end_t[veh_num][lane_ind],
                     end_mov_dist) = self.getTimeSpeedDistOfAccConstMov(self.phase_start_t[veh_num][lane_ind],
                                                                        endGreen_t,
                                                                        self.phase_start_dist[veh_num][lane_ind]
                                                                        - end_dist,
                                                                        veh_max_speed,
                                                                        self.phase_start_speed[veh_num][lane_ind],
                                                                        veh_acc)

                    self.detailed_tts[veh_num][lane_ind].append((self.phase_end_t[veh_num][lane_ind],
                                                                 self.phase_end_t[veh_num][lane_ind]
                                                                 - self.phase_start_t[veh_num][lane_ind],
                                                                 "movement to cross stop line without decel."))

                    # End dist. is the minimum possible end_dist plus the distance of the acc./constant mov. end_mov_dist
                    self.phase_end_dist[veh_num][lane_ind] = round(end_dist + end_mov_dist, 2)

                    # Make "checkpoint" for next algorithm update if suitable
                    if self.phase_start_t[veh_num][lane_ind] < ACTUAL_T + LEN_RANGE \
                    and self.phase_end_t[veh_num][lane_ind] >= ACTUAL_T + LEN_RANGE and veh_in_already == 0:
                        veh_in_already = self.estRefNextArrRange(edg_obj,
                                                                 in_laneID,
                                                                 veh_num,
                                                                 lane_ind,
                                                                 self.phase_start_t[veh_num][lane_ind],
                                                                 self.phase_start_dist[veh_num][lane_ind],
                                                                 self.phase_start_speed[veh_num][lane_ind],
                                                                 veh_acc,
                                                                 self.vType_vals["reaction_t"][self.vType[veh_num][0]],
                                                                 veh_in_already,
                                                                 max_speed=veh_max_speed)
                else:
                    # If vehicle's turning speed is constraining the final speed
                    if end_threshDist_t < endGreen_t:
                        # If reached end_threshDist_dist, vehicle will start decelerating
                        (self.phase_end_t[veh_num][lane_ind],
                         self.phase_end_speed[veh_num][lane_ind],
                         self.phase_end_dist[veh_num][lane_ind]) = self.getTimeSpeedDistOfDecelMov(end_threshDist_t,
                                                                         endGreen_t,
                                                                         end_threshDist_speed,
                                                                         self.decel_ave[veh_num]
                                                                         * edg_obj.edge_decel_adj[in_edgeID],
                                                                         self.vType_vals["reaction_t"][
                                                                             self.vType[veh_num][0]],
                                                                         end_threshDist_dist,
                                                                         end_dist)

                        self.detailed_tts[veh_num][lane_ind].append((self.phase_end_t[veh_num][lane_ind],
                                                                     self.phase_end_t[veh_num][lane_ind]
                                                                     - end_threshDist_t,
                                                                     "mov. to cross stop line with decel."))

                        # Make "checkpoint" for next algorithm update if suitable
                        if end_threshDist_t < ACTUAL_T + LEN_RANGE \
                                and self.phase_end_t[veh_num][lane_ind] >= ACTUAL_T + LEN_RANGE and veh_in_already == 0:
                            veh_in_already = self.estRefNextArrRange(edg_obj,
                                                                     in_laneID,
                                                                     veh_num,
                                                                     lane_ind,
                                                                     end_threshDist_t,
                                                                     end_threshDist_dist,
                                                                     end_threshDist_speed,
                                                                     self.decel_ave[veh_num]
                                                                     * edg_obj.edge_decel_adj[in_edgeID],
                                                                     self.vType_vals["reaction_t"][
                                                                         self.vType[veh_num][0]],
                                                                     veh_in_already,
                                                                     min_dist=end_dist)
                    else:
                        # Vehicle didn't reach end_threshDist_dist, the modelling finishes
                        self.phase_end_speed[veh_num][lane_ind] = end_threshDist_speed
                        self.phase_end_dist[veh_num][lane_ind] = end_threshDist_dist
                        self.phase_end_t[veh_num][lane_ind] = end_threshDist_t
            else:
                # If no need to slow down, vehicle will accelerate and/or have constant speed until stop line
                (self.phase_end_speed[veh_num][lane_ind],
                 self.phase_end_t[veh_num][lane_ind],
                 end_mov_dist) = self.getTimeSpeedDistOfAccConstMov(self.phase_start_t[veh_num][lane_ind],
                                                                    endGreen_t,
                                                                    self.phase_start_dist[veh_num][lane_ind] - end_dist,
                                                                    veh_max_speed,
                                                                    self.phase_start_speed[veh_num][lane_ind],
                                                                    veh_acc)

                self.detailed_tts[veh_num][lane_ind].append((self.phase_end_t[veh_num][lane_ind],
                                                             self.phase_end_t[veh_num][lane_ind]
                                                             - self.phase_start_t[veh_num][lane_ind],
                                                             "movement to cross stop line without decel."))

                # End dist. is the minimum possible end_dist plus the distance of the acc./constant mov. end_mov_dist
                self.phase_end_dist[veh_num][lane_ind] = round(end_dist + end_mov_dist, 2)

                # Make "checkpoint" for next algorithm update if suitable
                if self.phase_start_t[veh_num][lane_ind] < ACTUAL_T + LEN_RANGE \
                and self.phase_end_t[veh_num][lane_ind] >= ACTUAL_T + LEN_RANGE and veh_in_already == 0:
                    veh_in_already = self.estRefNextArrRange(edg_obj,
                                                             in_laneID,
                                                             veh_num,
                                                             lane_ind,
                                                             self.phase_start_t[veh_num][lane_ind],
                                                             self.phase_start_dist[veh_num][lane_ind],
                                                             self.phase_start_speed[veh_num][lane_ind],
                                                             veh_acc,
                                                             self.vType_vals["reaction_t"][self.vType[veh_num][0]],
                                                             veh_in_already,
                                                             max_speed=veh_max_speed)

            if firstVeh_bool == 0:
                # Correct the time and speed at phase end constrained by the values of the vehicle ahead
                if end_dist == 0:
                    # If end_dist == 0, then the min_phase_end_t is the time vehicle ahead crosses stop line
                    # plus the time for it to cross its whole length and the gap distance in time and also gap time
                    try:
                        if self.len[veh_ahead_num] / self.phase_end_speed[veh_ahead_num][veh_ahead_lane_ind] < STD_ACC_T:
                            # If vehicle not at higher accelerating speeds,
                            min_phase_end_t = round(self.phase_end_t[veh_ahead_num][veh_ahead_lane_ind] \
                                                + self.len[veh_ahead_num] \
                                                / self.phase_end_speed[veh_ahead_num][veh_ahead_lane_ind] \
                                                + self.min_gap_d[veh_ahead_num] \
                                                / self.phase_end_speed[veh_num][lane_ind] \
                                                + self.min_gap_t[veh_num], 2)
                        else:
                            # Zero speed, use standard accelerating time
                            raise ZeroDivisionError
                    except ZeroDivisionError:
                        # If vehicle at low accelerating/decelerating speed, use time of accelerating time based on
                        # vehicle's length normalized by the standard vehicle lentgh
                        min_phase_end_t = self.phase_end_t[veh_ahead_num][veh_ahead_lane_ind] \
                                            + STD_ACC_T * (self.len[veh_ahead_num] / self.vType_vals["length"][0]) \
                                            + self.min_gap_t[veh_num]
                else:
                    # If end_dist > 0, then it considers the minimum distance to vehicle ahead, and the distance should
                    # be at the time of phase_end_t of vehicle ahead
                    min_phase_end_t = self.phase_end_t[veh_ahead_num][veh_ahead_lane_ind]

                if self.phase_end_t[veh_num][lane_ind] < min_phase_end_t:
                    # Vehicle acceleration and maximum speed influenced by the vehicle ahead
                    stored_phase_end_t = self.phase_end_t[veh_num][lane_ind]
                    self.phase_end_t[veh_num][lane_ind] = min_phase_end_t
                    self.phase_end_speed[veh_num][lane_ind] = min(
                        self.phase_end_speed[veh_ahead_num][veh_ahead_lane_ind],
                        self.phase_end_speed[veh_num][lane_ind])

                    self.detailed_tts[veh_num][lane_ind].append((self.phase_end_t[veh_num][lane_ind],
                                                                 self.phase_end_t[veh_num][lane_ind]
                                                                 - stored_phase_end_t,
                                                                 "correction due veh. ahead " + str(
                                                             (veh_ahead_num, veh_ahead_lane_ind)) +
                                                             " end movement modelling"))


                    veh_acc = min(veh_acc, self.accel_cap[veh_num] * edg_obj.edge_accel_adj[in_edgeID])
                    veh_max_speed = min(veh_max_speed, self.speed_factor[veh_ahead_num]
                                        * edg_obj.lane_max_speed[in_laneID])

                    # Account distance the vehicle stopped to the queue prediction
                    if self.vClass[veh_num] != "Art. Probe" and jct_obj.t_max_cycle[jctID] <= 1 \
                    and self.phase_end_speed[veh_num][lane_ind] <= self.thresh_stopped:
                        back_queue = self.phase_end_dist[veh_num][lane_ind] + self.len[veh_num]
                        if back_queue <= edg_obj.lane_length[in_laneID]:
                            # Only account distance if within lane length, this avoids queues that never discharge and
                            # cause lanes sending to this lane to not allow any vehicle to go on the next algorithm upd.
                            suitable_ranges = [range_ind
                                               for range_ind, range_t in
                                               enumerate(edg_obj.lane_queue_tLimitRanges[in_laneID])
                                               if stored_phase_end_t < range_t <= self.phase_end_t[veh_num][lane_ind]]
                            for suitable_range in suitable_ranges:
                                edg_obj.lane_queue_pred[in_laneID][1][suitable_range] = np.ceil(back_queue)

                    # Make "checkpoint" for next algorithm update if suitable
                    if (veh_in_already == 1 and veh_was_in_already == 0
                        and self.phase_start_t[veh_num][lane_ind] < ACTUAL_T + LEN_RANGE) \
                    or (veh_in_already == 0 and stored_phase_end_t < ACTUAL_T + LEN_RANGE
                         and self.phase_end_t[veh_num][lane_ind] >= ACTUAL_T + LEN_RANGE):

                        if (veh_ahead_num, veh_ahead_lane_ind) in edg_obj.already_on_lane[in_laneID]:
                            # If known update values of vehicle ahead, the update values are based on the vehicle ahead
                            self.update_t[veh_num][lane_ind] = ACTUAL_T + LEN_RANGE
                            influenced_update_dist = round(self.update_dist[veh_ahead_num][veh_ahead_lane_ind]
                                                      + self.len[veh_ahead_num]
                                                      + self.min_gap_d[veh_num]
                                                      + (self.min_gap_t[veh_num]
                                                         * self.update_speed[veh_ahead_num][veh_ahead_lane_ind]), 2)
                            if veh_in_already == 0 or influenced_update_dist >= self.update_dist[veh_num][lane_ind]:
                                # If vehicle was not added to the already on lane list or it is influenced,
                                # it cannot be higher than the vehicle in front
                                self.update_speed[veh_num][lane_ind] = self.update_speed[veh_ahead_num][veh_ahead_lane_ind]
                                self.update_dist[veh_num][lane_ind] = influenced_update_dist
                        else:
                            # Not known update values of vehicle ahead, but known the corrected end values, use them
                            # for next algorithm update
                            if self.phase_end_speed[veh_num][lane_ind] <= self.thresh_stopped:
                                self.update_t[veh_num][lane_ind] = ACTUAL_T + LEN_RANGE
                                self.update_dist[veh_num][lane_ind] = self.phase_end_dist[veh_num][lane_ind]
                                self.update_speed[veh_num][lane_ind] = self.phase_end_speed[veh_num][lane_ind]
                            else:
                                # If not stopped at end of movement,
                                # first define the update values with constrained values
                                veh_in_already = self.estRefNextArrRange(edg_obj,
                                                                         in_laneID,
                                                                         veh_num,
                                                                         lane_ind,
                                                                         self.phase_start_t[veh_num][lane_ind],
                                                                         self.phase_start_dist[veh_num][lane_ind]
                                                                         - end_dist,
                                                                         min(self.phase_start_speed[veh_num][lane_ind],
                                                                             self.phase_start_speed[veh_ahead_num]
                                                                             [veh_ahead_lane_ind]),
                                                                         veh_acc,
                                                                         self.vType_vals["reaction_t"]
                                                                         [self.vType[veh_num][0]],
                                                                         veh_in_already,
                                                                         max_speed=veh_max_speed)
                                if veh_in_already == 0:
                                    # If vehicle still would arrive before next algorithm update,
                                    # then assume the vehicle would have its phase end values at next update
                                    self.update_t[veh_num][lane_ind] = ACTUAL_T + LEN_RANGE
                                    self.update_dist[veh_num][lane_ind] = self.phase_end_dist[veh_num][lane_ind]
                                    self.update_speed[veh_num][lane_ind] = self.phase_end_speed[veh_num][lane_ind]

                        if veh_in_already == 0:
                            edg_obj.already_on_lane[in_laneID].append((veh_num, lane_ind))
                        veh_was_in_already = 1  # needed to avoid entering in the correction of phase_end values
                        veh_in_already = 1

            if self.phase_end_t[veh_num][lane_ind] > endGreen_t + amber_t:
                # If able to cross after end green time and amber time, vehicle will not cross
                veh_will_cross = 0
            elif self.phase_end_t[veh_num][lane_ind] >= endGreen_t and self.phase_end_dist[veh_num][lane_ind] >= 0:
                # If the time at crossing the stop line is after than the end of green time,
                # analyse if vehicle might cross during amber time (assuming maximum one vehicle crossing during amber)
                already_using_amber_t = self.phase_end_t[veh_num][lane_ind] - endGreen_t
                phase_end_dist_amber = max(self.phase_end_dist[veh_num][lane_ind]
                                           - (amber_t - already_using_amber_t)
                                           * self.phase_end_speed[veh_num][lane_ind], 0)

                if self.phase_end_speed[veh_num][lane_ind] > self.thresh_stopped \
                and phase_end_dist_amber == 0 and (firstVeh_bool == 1
                                                   or (vehAhead_will_cross == 1
                                                       and self.phase_end_t[veh_ahead_num][veh_ahead_lane_ind]
                                                       < endGreen_t)):
                    # First vehicle during amber will pass if vehicle ahead passes before end of green green
                    veh_will_cross = 1
                    used_amber = self.phase_end_dist[veh_num][lane_ind] / self.phase_end_speed[veh_num][lane_ind]

                    # Make "checkpoint" for next algorithm update if suitable
                    if self.phase_end_t[veh_num][lane_ind] < ACTUAL_T + LEN_RANGE \
                    and self.phase_end_t[veh_num][lane_ind] + used_amber >= ACTUAL_T + LEN_RANGE and veh_in_already == 0:
                        veh_in_already = self.estRefNextArrRange(edg_obj,
                                                                 in_laneID,
                                                                 veh_num,
                                                                 lane_ind,
                                                                 self.phase_end_t[veh_num][lane_ind],
                                                                 self.phase_end_dist[veh_num][lane_ind],
                                                                 self.phase_end_speed[veh_num][lane_ind],
                                                                 veh_acc,
                                                                 self.vType_vals["reaction_t"][self.vType[veh_num][0]],
                                                                 veh_in_already,
                                                                 max_speed=veh_max_speed)

                    # Add the used amber time at phase_end_t
                    self.phase_end_t[veh_num][lane_ind] += used_amber

                    self.detailed_tts[veh_num][lane_ind].append((self.phase_end_t[veh_num][lane_ind],
                                                                 used_amber,
                                                                 "added time of used amber"))

                    self.phase_end_dist[veh_num][lane_ind] = 0
                else:
                    # Vehicle will not cross during amber time
                    veh_will_cross = 0
            elif self.phase_end_dist[veh_num][lane_ind] == 0 and (firstVeh_bool == 1 or vehAhead_will_cross == 1):
                # Normal condition of crossing before end of green time when vehicle ahead also had crossed
                veh_will_cross = 1
            else:
                # Any other ways, vehicle doesn't cross the stop line
                veh_will_cross = 0

            if veh_will_cross == 1:
                # Define next lane and its movgID as well as the connID from this lane to next lane
                self.defVehConnMovgNextLane(jct_obj, edg_obj, veh_num, jctID, lane_ind,
                                            self.edge_route[veh_num][lane_ind + 1],
                                            self.phase_end_t[veh_num][lane_ind], current_lane=in_laneID,
                                            current_movgID=movgID, new_veh=0)
                start45 = time.time()
                global moved2out_lane, rge_ind

                # Delay vehicle departure if queue prediction on the lane reaches its maximum length
                delayed_starting_t = self.phase_end_t[veh_num][lane_ind]
                out_lane = self.lane_route[veh_num][lane_ind + 1]
                moved2out_lane = 0
                rge_ind = 0
                if out_queue_range[0] != None:
                    if edg_obj.inLaneStopped_dueOutQueue[in_laneID] == 0:
                        # Define outgoing lane time resolution
                        time_resolution = edg_obj.lane_queue_tLimitRanges[out_lane][1] \
                                          - edg_obj.lane_queue_tLimitRanges[out_lane][0]
                        # Define outgoing lane planning horizon
                        out_queue_t_max = edg_obj.lane_queue_tLimitRanges[out_lane][-1] \
                                          + time_resolution - edg_obj.lane_queue_tLimitRanges[out_lane][0]
                        while moved2out_lane == 0 and delayed_starting_t <= endGreen_t + amber_t:
                            for range_i, range_t \
                            in enumerate(edg_obj.lane_queue_tLimitRanges[out_lane][out_queue_range[0]:]):
                                # out_queue_range is the first range to look for the queue onwards
                                rge_ind = out_queue_range[0] + range_i
                                if edg_obj.lane_queue_pred[out_lane][1][rge_ind] != 0 \
                                and self.min_gap_d[veh_num] + self.len[veh_num] \
                                > edg_obj.lane_aval_queue_len[out_lane][out_queue_cycle[0]][rge_ind]:
                                    # delay vehicle's departure if out. lane queue cannot take vehicle's length
                                    delayed_starting_t = range_t + (out_queue_t_max * out_queue_cycle[0])
                                    if delayed_starting_t > endGreen_t + amber_t:
                                        # if delayed starting after the end of green time + amber time, stop delay
                                        # calculation, this will set veh_will_cross = 0
                                        break
                                else:
                                    # Add new vehicle at the queue available length, just to avoid sending vehicles
                                    # to a full queue to be predicted when estimating arrivals of out_lane
                                    if delayed_starting_t < endGreen_t \
                                    or (firstVeh_bool == 1
                                         or (vehAhead_will_cross == 1
                                             and self.phase_end_t[veh_ahead_num][veh_ahead_lane_ind] < endGreen_t)):
                                        edg_obj.lane_aval_queue_len[out_lane][out_queue_cycle[0]][rge_ind] -= \
                                        (self.min_gap_d[veh_num] + self.len[veh_num])
                                        moved2out_lane = 1
                                    break

                            if moved2out_lane == 0:
                                # if checked queue along whole cycle and vehicle couldn't
                                # go to outgoing lane, try on the next cycle
                                cycle = out_queue_cycle[0]
                                del out_queue_cycle[0]
                                out_queue_cycle.append(cycle + 1)
                                del out_queue_range[0]
                                out_queue_range.append(0)
                            else:
                                # Reset range
                                del out_queue_range[0]
                                out_queue_range.append(rge_ind)

                            if out_queue_cycle[0] == MAX_T_MAX_CYCLE and moved2out_lane == 0 \
                            and delayed_starting_t <= endGreen_t + amber_t:
                                # If reached MAX_T_MAX_CYCLE and vehicle couldn't go to
                                # outgoing lane, set that incoming lane is stopped due outgoing queue full
                                edg_obj.inLaneStopped_dueOutQueue[in_laneID] = 1
                            if edg_obj.inLaneStopped_dueOutQueue[in_laneID] == 1:
                                break

                    if edg_obj.inLaneStopped_dueOutQueue[in_laneID] == 1:
                        # If full queue prediction for the whole prediction time,
                        # make all vehicles on this lane from this vehicle's range till the end as not probe
                        # vehicles, it is not possible to calculate travel times for the vehicles after this one
                        edg_obj.stop_est_probes[in_laneID] = 1
                        if self.probe_arr_range[veh_num][lane_ind] != None:
                            num_ranges = int(jct_obj.t_max[jctID] / LEN_RANGE)
                            for rge in range(self.probe_arr_range[veh_num][lane_ind], num_ranges):
                                for movg in movg_obj.probe_vehs_range[in_laneID].keys():
                                    if movg_obj.probe_vehs_range[in_laneID][movg][rge] != []:
                                        for veh,l_ind in movg_obj.probe_vehs_range[in_laneID][movg][rge]:
                                            try:
                                                self.probe_arr_range[veh][l_ind] = None
                                            except TypeError:
                                                skip = 1
                                        movg_obj.probe_vehs_range[in_laneID][movg][rge] = []
                else:
                    # if after next edge is unknown, there is no information of queue, don't consider queue length
                    moved2out_lane = 1

                end45 = time.time()
                timings[45][1] += end45 - start45

                if moved2out_lane == 0:
                    # Vehicle will not cross due to full queue length on the next edge,
                    veh_will_cross = 0
                else:
                    # if vehicle will cross to outgoing lane, decrease the departure speed due the delay
                    braking_t = delayed_starting_t - self.phase_end_t[veh_num][lane_ind]

                    if braking_t > 0:
                        self.detailed_tts[veh_num][lane_ind].append((delayed_starting_t, delayed_starting_t
                                                                     - self.phase_end_t[veh_num][lane_ind],
                                                                     "delay due full outgoing lane queue"))

                    if braking_t > 0 and self.phase_end_speed[veh_num][lane_ind] > self.thresh_stopped:
                        copied_phase_end_speed = self.phase_end_speed[veh_num][lane_ind]
                        (_,
                        self.phase_end_speed[veh_num][lane_ind],
                        _) = self.getTimeSpeedDistOfDecelMov(0, braking_t,
                                                             self.phase_end_speed[veh_num][lane_ind],
                                                             self.decel_ave[veh_num]
                                                             * edg_obj.edge_decel_adj[in_edgeID],
                                                             self.vType_vals["reaction_t"][self.vType[veh_num][0]],
                                                             0, 0)

                        # Make "checkpoint" for next algorithm update if suitable
                        if self.phase_end_t[veh_num][lane_ind] < ACTUAL_T + LEN_RANGE \
                        and delayed_starting_t >= ACTUAL_T + LEN_RANGE and veh_in_already == 0:
                            veh_in_already = 1
                            self.update_dist[veh_num][lane_ind] = self.phase_end_dist[veh_num][lane_ind]
                            self.update_t[veh_num][lane_ind] = ACTUAL_T + LEN_RANGE
                            self.update_speed[veh_num][lane_ind] = round(max(copied_phase_end_speed \
                                                                       + self.decel_ave[veh_num] \
                                                                       * ((ACTUAL_T + LEN_RANGE)
                                                                           - self.phase_end_t[veh_num][lane_ind]), 0), 2)
                            edg_obj.already_on_lane[in_laneID].append((veh_num, lane_ind))

                            veh_will_cross =  1

                # Account distance the vehicle stopped to the queue prediction
                if self.vClass[veh_num] != "Art. Probe" and jct_obj.t_max_cycle[jctID] <= 1 \
                and delayed_starting_t != self.phase_end_t[veh_num][lane_ind]:
                    back_queue = self.phase_start_dist[veh_num][lane_ind] + self.len[veh_num]
                    if back_queue <= edg_obj.lane_length[in_laneID]:
                        # Only account distance if within lane length, this avoids queues that never discharge and
                        # cause lanes sending to this lane to not allow any vehicle to go on the next algorithm upd.
                        suitable_ranges = [range_ind
                                           for range_ind, range_t in
                                           enumerate(edg_obj.lane_queue_tLimitRanges[in_laneID])
                                           if self.phase_start_t[veh_num][lane_ind] < range_t <= delayed_starting_t]
                        for suitable_range in suitable_ranges:
                            edg_obj.lane_queue_pred[in_laneID][1][suitable_range] = np.ceil(back_queue)

                self.phase_end_t[veh_num][lane_ind] = delayed_starting_t
            else:
                # Vehicle will not reach stop line within end of green time or vehicle ahead will not cross
                # veh_will_cross = 0 already
                pass
        else:
            # movgID or veh_ahead_movgID not in phase_green_movgs
            # the end values of the phase are the updated start values representing the time vehicle will start moving
            # on the next suitable green phase
            veh_will_cross = 0
            self.phase_end_speed[veh_num][lane_ind] = self.phase_start_speed[veh_num][lane_ind]
            self.phase_end_t[veh_num][lane_ind] = self.phase_start_t[veh_num][lane_ind]
            self.phase_end_dist[veh_num][lane_ind] = self.phase_start_dist[veh_num][lane_ind]

        end33 = time.time()
        timings[33][1] += end33 - start33
        start34 = time.time()

        # Delay vehicle's departure if it needs to yield to preferential traffic
        global wait_t_pref
        movs2yield = [conflt_movg for conflt_movg in movg_obj.movgID_to_yield2movgIDs[movgID]
                      if conflt_movg in phase_green_movgs]
        if veh_will_cross == 0 or movg_status == 6 or movs2yield == [] \
        or self.phase_end_t[veh_num][lane_ind] >= endGreen_t:
            # 1) The vehicle will not cross anyway,
            # 2) The movement phase is protected green,
            # Then don't consider the opposite (preferential) flow (also not obligatory to stop)
            # 3) This also happens when vehicle cannot brake in time and cross in amber
            # (implies it will not wait for preferential traffic)
            wait_t_pref = 0
        else:
            # The phase of the movement group is not protected and the vehicle would cross but having to yield
            if jct_obj.t_max_cycle[jctID] == 0:
                # If known vehicles departing during the phase,
                # Then use the time of vehicle on preferential traffic crosses the stop line
                conns2yield = [yield_conn
                               for yield_conn in edg_obj.connID_to_yield2connIDs[self.connection[veh_num][lane_ind]]
                               if edg_obj.connID_to_movgID[yield_conn] in phase_green_movgs]
                lanes2yield = set([edg_obj.connID_to_inLaneID[yield_conn] for yield_conn in conns2yield])

                # Adjust critical gap based on subject traffic mean speed
                # (base values are for 50 km/h so any lower/higher speeds require different gaps)
                # For a difference of 50 km/h = 13.89 m/s increase 0.5 seconds
                ego_veh_speed_adj = ((13.89 - self.phase_end_speed[veh_num][lane_ind]) * 0.5) / 13.89

                # Initialize
                ref_beg_cross = self.phase_end_t[veh_num][lane_ind]
                wait_t_pref = 0
                stored_wait_t_pref = wait_t_pref
                try_another_gap = 1
                oppVeh_ind = 0
                while try_another_gap == 1:
                    for yield_2lane in lanes2yield:
                        # Check the remaining opposite vehicles to consider after the last one which is known that
                        # after it it is possible to cross for certain lane to yield (avoid recalculation)
                        diff_len = len(edg_obj.lane_discharging_vehs_phase[yield_2lane]) \
                                   - len(edg_obj.lane_discharging_vehs_phase[yield_2lane]
                                         [edg_obj.yield2lane_last_oppVeh[in_laneID][yield_2lane]:])
                        for vehInd, (oppVehNum, oppVeh_lane_ind, oppVeh_movgID) \
                                in enumerate(edg_obj.lane_discharging_vehs_phase[yield_2lane]
                                             [edg_obj.yield2lane_last_oppVeh[in_laneID][yield_2lane]:]):
                            # Adjust opposite vehicle index
                            oppVeh_ind = diff_len + vehInd
                            if (oppVehNum, oppVeh_lane_ind) in edg_obj.veh_already_disharged:
                                # if opposite vehicle already discharged, just get its time of departure (if it crosses)
                                pass
                            else:
                                # vehicle not discharge yet, discharge it
                                self.defVehMovement(jct_obj, movg_obj, edg_obj, jctID, oppVehNum, oppVeh_lane_ind,
                                                    oppVeh_movgID,  phase_green_movgs, lanes_order, all_toNeigh_lanes,
                                                    tc_it)

                            try:
                                if self.connection[oppVehNum][oppVeh_lane_ind] in conns2yield:
                                    # oppVehNum uses connection that vehicle (veh_num) must yield to
                                    # Adjust critical gap based on preferential traffic mean speed
                                    # (base values are for 50 km/h so any lower/higher speeds require different gaps)
                                    if edg_obj.conn_direction[self.connection[veh_num][lane_ind]] == "r":
                                        # If left or right turn movements
                                        if self.phase_end_speed[veh_num][lane_ind] == 0:
                                            # If vehicle at subject connection must stop, change 1 second per 40 km/h
                                            yield_veh_speed_adj = ((self.phase_end_speed[oppVehNum][oppVeh_lane_ind]
                                                                    - 13.89) * 1) / 11.11
                                        else:
                                            yield_veh_speed_adj = ((self.phase_end_speed[oppVehNum][oppVeh_lane_ind]
                                                                    - 13.89) * 0.5) / 11.11
                                    else:
                                        # If left turn or through movements
                                        if self.phase_end_speed[veh_num][lane_ind] == 0:
                                            # If vehicle at subject connection must stop, change 1.5 second per 40 km/h
                                            yield_veh_speed_adj = ((self.phase_end_speed[oppVehNum][oppVeh_lane_ind]
                                                                    - 13.89) * 1.5) / 11.11
                                        else:
                                            yield_veh_speed_adj = ((self.phase_end_speed[oppVehNum][oppVeh_lane_ind]
                                                                    - 13.89) * 1) / 11.11

                                    if firstVeh_bool == 0 \
                                    and (self.len_crossed_t[veh_ahead_num][veh_ahead_lane_ind]
                                         + self.min_gap_t[veh_num] >= ref_beg_cross):
                                        # If not first vehicle and the time of veh. ahead cross + gap time is after
                                        # the time vehicle is able to begin to cross, then it is following veh. ahead
                                        # the required gap time to yield is the follow up time of vehicle's type
                                        # for this connection
                                        ref_gap_t = \
                                        edg_obj.conn_follow_up[self.connection[veh_num][lane_ind]]["std"] \
                                        [self.vType[veh_num][0]]
                                    else:
                                        # If first vehicle or not following vehicle ahead (veh. ahead departs much earlier)
                                        # then vehicle is not following veh. ahead
                                        # the required gap time to yield is the average critical gap time of vehicle's type
                                        # for this connection
                                        ref_gap_t = \
                                        edg_obj.conn_ave_cr_gap[self.connection[veh_num][lane_ind]]["std"] \
                                        [self.vType[veh_num][0]] + ego_veh_speed_adj + yield_veh_speed_adj

                                    if self.crossing_bool[oppVehNum][oppVeh_lane_ind] in (0, None):
                                        if self.phase_end_t[oppVehNum][oppVeh_lane_ind]  \
                                        < self.phase_end_t[veh_num][lane_ind]:
                                            # the oppVehNum will have to yield to this veh_num, so this veh_num will
                                            # cross first,
                                            break
                                        else:
                                            len_crossed_t = self.phase_end_t[oppVehNum][oppVeh_lane_ind] \
                                                            + (self.phase_end_speed[oppVehNum][oppVeh_lane_ind]
                                                               * self.len[oppVehNum])
                                            if ref_beg_cross < len_crossed_t < (ref_beg_cross + ref_gap_t) \
                                            or ref_beg_cross < stored_wait_t_pref:
                                                # Vehicle veh_num will yield to oppVehNum
                                                ref_beg_cross = len_crossed_t
                                            else:
                                                # Vehicle veh_num will not yield to oppVehNum,
                                                # finish checking arrival of pref. vehicles for this lane
                                                break
                                    else:
                                        if ref_beg_cross < self.len_crossed_t[oppVehNum][oppVeh_lane_ind] \
                                                         < (ref_beg_cross + ref_gap_t) \
                                        or ref_beg_cross < stored_wait_t_pref:
                                            # Vehicle veh_num will yield to oppVehNum
                                            ref_beg_cross = self.len_crossed_t[oppVehNum][oppVeh_lane_ind]
                                        else:
                                            # Vehicle veh_num will not yield to oppVehNum,
                                            # finish checking arrival of pref. vehicles for this lane
                                            break
                                else:
                                    # If oppVehNum don't use connection that veh_num needs to yield
                                    # (go to the next vehicle)
                                    raise IndexError
                            except IndexError:
                                DontNeed2Yield = 1
                        else:
                            edg_obj.yield2lane_last_oppVeh[in_laneID][yield_2lane] = oppVeh_ind
                            # store the highest waiting time amongst all lanes to yield
                            wait_t_pref = round(max(wait_t_pref, ref_beg_cross - self.phase_end_t[veh_num][lane_ind]), 2)

                    if wait_t_pref > stored_wait_t_pref and len(lanes2yield) > 1:
                        # if trying to find a suitable gap again for all lanes to yield the waiting time is longer
                        # than the last, it will have to continue waiting for a suitable gap on all lanes
                        stored_wait_t_pref = wait_t_pref
                        try_another_gap = 1
                    else:
                        try_another_gap = 0

                    if self.phase_end_t[veh_num][lane_ind] + wait_t_pref > endGreen_t \
                    or self.phase_end_t[veh_num][lane_ind] >= ACTUAL_T + (MAX_T_MAX_CYCLE * jct_obj.t_max[jctID]):
                        # If reached maximum number of gap trials or the possible time to depart is after green time,
                        # stop finding the delay time due preferential traffic because it may cross during amber
                        # at a later point, if there is no amber time
                        break
            else:
                # If not known vehicles departure during the cycle,
                # Then consider the opposite (preferential) flow of the same phase on previous cycles
                if endGreen_t == float("inf"):
                    # If unsignalized junction
                    if firstVeh_bool == 0:
                        ref_hdwy_t = self.phase_end_t[veh_ahead_num][veh_ahead_lane_ind]
                    else:
                        ref_hdwy_t = self.phase_end_t[veh_num][lane_ind]

                    suitable_end = self.phase_end_t[veh_num][lane_ind] + (MAX_T_MAX_CYCLE * jct_obj.t_max[jctID])
                else:
                    # If signalized junction
                    if firstVeh_bool == 0:
                        ref_hdwy_t = self.phase_end_t[veh_ahead_num][veh_ahead_lane_ind]
                    else:
                        ref_hdwy_t = beginGreen_t

                    suitable_end = endGreen_t + amber_t

                wait_t_pref = edg_obj.estDelayConnCap(jct_obj, movg_obj, self,
                                                      jctID, jct_obj.t_max_cycle[jctID],
                                                      phase_num, suitable_end,
                                                      self.phase_end_t[veh_num][lane_ind],
                                                      ref_hdwy_t,
                                                      conn=self.connection[veh_num][lane_ind])
                # Vehicle may wait only until end of phase
                wait_t_pref = round(min(wait_t_pref, suitable_end - self.phase_end_t[veh_num][lane_ind]), 2)

        if veh_will_cross == 1 and wait_t_pref > 0:
            stored_phase_end_t = self.phase_end_t[veh_num][lane_ind]
            stored_phase_end_speed = self.phase_end_speed[veh_num][lane_ind]
            # Update time and speed if vehicle needs to yield to preferential traffic
            if self.phase_end_t[veh_num][lane_ind] + wait_t_pref < endGreen_t \
            and self.phase_end_t[veh_num][lane_ind] < ACTUAL_T + (MAX_T_MAX_CYCLE * jct_obj.t_max[jctID]):
                if self.phase_end_speed[veh_num][lane_ind] > self.thresh_stopped:
                    (_,
                     self.phase_end_speed[veh_num][lane_ind],
                     _) = self.getTimeSpeedDistOfDecelMov(0, wait_t_pref,
                                                          self.phase_end_speed[veh_num][lane_ind],
                                                          self.decel_ave[veh_num]
                                                          * edg_obj.edge_decel_adj[in_edgeID],
                                                          self.vType_vals["reaction_t"][self.vType[veh_num][0]],
                                                          0, 0)
                # Store departure time considering the possibility of yielding to preferential traffic
                self.phase_end_t[veh_num][lane_ind] = self.phase_end_t[veh_num][lane_ind] + wait_t_pref

                self.detailed_tts[veh_num][lane_ind].append((self.phase_end_t[veh_num][lane_ind], wait_t_pref,
                                                             "delay due preferential traffic"))
            else:
                # This happens when self.phase_end_t[veh_num][lane_ind] + wait_t_pref >= endGreen_t
                if self.phase_end_t[veh_num][lane_ind] < ACTUAL_T + (MAX_T_MAX_CYCLE * jct_obj.t_max[jctID]):
                    if firstVeh_bool == 1 and amber_t > 0:
                        # First vehicle will pass anyway during amber time
                        pass
                    else:
                        # Not first vehicle, check if the vehicle ahead will pass already during amber
                        if amber_t == 0 or self.phase_end_t[veh_ahead_num][veh_ahead_lane_ind] >= endGreen_t:
                            veh_will_cross = 0

                    if self.phase_end_speed[veh_num][lane_ind] > self.thresh_stopped:
                        (_,
                         self.phase_end_speed[veh_num][lane_ind],
                         _) = self.getTimeSpeedDistOfDecelMov(0,
                                                              endGreen_t - self.phase_end_t[veh_num][lane_ind],
                                                              self.phase_end_speed[veh_num][lane_ind],
                                                              self.decel_ave[veh_num]
                                                              * edg_obj.edge_decel_adj[in_edgeID],
                                                              self.vType_vals["reaction_t"][self.vType[veh_num][0]],
                                                              0, 0)
                    if veh_will_cross == 1:

                        temp_stored = self.phase_end_t[veh_num][lane_ind]

                        self.phase_end_t[veh_num][lane_ind] = max(endGreen_t, self.phase_end_t[veh_num][lane_ind])
                        self.phase_end_t[veh_num][lane_ind] = min(endGreen_t + amber_t,
                                                                  self.phase_end_t[veh_num][lane_ind])

                        self.detailed_tts[veh_num][lane_ind].append((self.phase_end_t[veh_num][lane_ind],
                                                                     self.phase_end_t[veh_num][lane_ind] - temp_stored,
                                                                     "delay due preferential traffic"))
                else:
                    veh_will_cross = 0


            # Make "checkpoint" for next algorithm update if suitable
            if stored_phase_end_t < ACTUAL_T + LEN_RANGE \
            and self.phase_end_t[veh_num][lane_ind] >= ACTUAL_T + LEN_RANGE and veh_in_already == 0:
                veh_in_already = self.estRefNextArrRange(edg_obj,
                                                         in_laneID,
                                                         veh_num,
                                                         lane_ind,
                                                         stored_phase_end_t,
                                                         0,
                                                         stored_phase_end_speed,
                                                         self.decel_ave[veh_num] * edg_obj.edge_decel_adj[in_edgeID],
                                                         self.vType_vals["reaction_t"][self.vType[veh_num][0]],
                                                         veh_in_already,
                                                         min_dist=0)

        if veh_will_cross == 1:
            # If vehicle will cross during the planning horizon (t_max_cycle)
            # Estimate the time for the length of the vehicle to cross
            (_,
             self.len_crossed_t[veh_num][lane_ind],
             _) = self.getTimeSpeedDistOfAccConstMov(self.phase_end_t[veh_num][lane_ind],
                                                     float("inf"), self.len[veh_num],
                                                     veh_max_speed,
                                                     self.phase_end_speed[veh_num][lane_ind],
                                                     veh_acc)
            # Estimate the arrival speed and time at the next edge/lane
            (out_lane_arr_speed,
             out_lane_arr_t,
             _) = self.getTimeSpeedDistOfAccConstMov(self.phase_end_t[veh_num][lane_ind], float("inf"),
                                                     edg_obj.conn_length[self.connection[veh_num][lane_ind]],
                                                     veh_max_speed,
                                                     self.phase_end_speed[veh_num][lane_ind],
                                                     veh_acc)

            # Estimate the travel time
            self.travel_time[veh_num][lane_ind] = np.ceil(out_lane_arr_t - self.in_lane_arr_t[veh_num][lane_ind])
        else:
            # If vehicle will not cross during the planning horizon (t_max_cycle)
            out_lane_arr_speed = None
            out_lane_arr_t = None


        end34 = time.time()
        timings[34][1] += end34 - start34
        end31 = time.time()
        timings[31][1] += end31 - start31

        return veh_will_cross, out_lane_arr_t, out_lane_arr_speed


# Install custom exception hook that logs error info on program crash
sys.excepthook = log_uncaught_exceptions 


## NETWORK INFORMATION

# Try if there was a database with DB_FILE name and load the objects, otherwise create a new DB and parse the files
if os.path.isfile(DB_FILE):

    LOGGER.info("Reading network database file `{:s}` ...".format(DB_FILE))
    with open(DB_FILE, "rb") as input:
        msgs_obj_data = pickle.load(input)
        jct_obj_data = pickle.load(input)
        movg_obj_data = pickle.load(input)
        edg_obj_data = pickle.load(input)
        veh_obj_data = pickle.load(input)

    msgs_objs = []
    for msgs_zipped_data in msgs_obj_data:
        msgs_obj_name = msgs_zipped_data[0]
        exec(msgs_obj_name + " = msgs_zipped_data[1]")
        exec("msgs_objs.append(" + msgs_obj_name + ")")
    del msgs_zipped_data
    del msgs_obj_data
    del msgs_obj_name

    jct_objs = []
    for jct_zipped_data in jct_obj_data:
        jct_obj_name = jct_zipped_data[0]
        exec(jct_obj_name + " = jct_zipped_data[1]")
        exec("jct_objs.append(" + jct_obj_name + ")")
    del jct_zipped_data
    del jct_obj_data
    del jct_obj_name

    movg_objs = []
    for movg_zipped_data in movg_obj_data:
        movg_obj_name = movg_zipped_data[0]
        exec(movg_obj_name + " = movg_zipped_data[1]")
        exec("movg_objs.append(" + movg_obj_name + ")")
    del movg_zipped_data
    del movg_obj_data
    del movg_obj_name

    edg_objs = []
    for edg_zipped_data in edg_obj_data:
        edg_obj_name = edg_zipped_data[0]
        exec(edg_obj_name + " = edg_zipped_data[1]")
        exec("edg_objs.append(" + edg_obj_name + ")")
    del edg_zipped_data
    del edg_obj_data
    del edg_obj_name

    veh_objs = []
    for veh_zipped_data in veh_obj_data:
        veh_obj_name = veh_zipped_data[0]
        exec(veh_obj_name + " = veh_zipped_data[1]")
        exec("veh_objs.append(" + veh_obj_name + ")")
    del veh_zipped_data
    del veh_obj_data
    del veh_obj_name
    LOGGER.info("... reading finished.")

else:
    # GETTING LOCAL LEVEL ROUTING INPUTS USING SUMO DATA (OR ANY XML WITH SAME TAGS)
    # If no database, define and parse all files and save the inputs for quicker initialization later
    LOGGER.info("Database file not found, creating it.")
    logger.info('Not found database file, creating it')

    # Read Network File
    tree_network = ElementTree.parse(os.path.abspath(os.path.join(NET_FILE)))

    # Read Additional File (For Movement Groups)
    tree_mov_groups = ElementTree.parse(os.path.abspath(os.path.join(MOV_GROUPS_FILE)))

    # Read Vehicle Composition File
    tree_vtypes = ElementTree.parse(os.path.abspath(os.path.join(VTYPES_FILE)))

    # Define the SUMO XML junction, edge, connection, mov. groups and veh. distributions children used for the area LLR
    net_jct_children = list(tree_network.iter('junction'))
    net_jct_children = {jct.get("id"): jct for jct in net_jct_children}
    net_edg_children = list(tree_network.iter('edge'))
    net_internal_edg_children = [edge for edge in net_edg_children if edge.get('function') == "internal"]
    net_edg_children = [edge for edge in net_edg_children if edge.get('function') != "internal"]
    net_conn_children = list(tree_network.iter('connection'))
    # don't get internal junction connections
    net_conn_children = [conn for conn in net_conn_children
                         if conn.get('via') != None] # avoid connections from an internal lane int_lane to out_lane
                                                     # this still takes any in_lane to int_lane and int_lane to int_lane
    tls_signal_children = list(tree_mov_groups.iter('tlLogic'))
    tls_signal_children = {signal.get('id'): signal for signal in tls_signal_children}
    vTypeDistribution_children = list(tree_vtypes.iter('vTypeDistribution'))
    # vehComposition_children = []
    # for vTypeDistribution_child in vTypeDistribution_children:
    vTypeDistribution_child = vTypeDistribution_children[0] # assuming only one distribution
    vehComposition_children = list(vTypeDistribution_child.findall("vType"))
    # vehComposition_children.extend(vehComposition_children)

    # Automatically add all/selected junctions parameters for each traffic controller for a traffic network
    if len(TCS_REF_JCTS) != 0:
        # Initialize for adding selected junctions grouped in the same traffic controller according to TCS_REF_JCTS
        given_tcs_ref = 1
        tcs_jcts = [[] for tc_ind in range(0, len(TCS_REF_JCTS))]
        all_junctions_tc = dict()
        for tc_ind, tc_inputs in enumerate(TCS_REF_JCTS):
            for jctID in tc_inputs[0][1]:
                all_junctions_tc.update({jctID: tc_ind})
    else:
        # Initialize for adding whole network junctions, making each traffic light controlled junction
        # a traffic controller, and other junctions are grouped by their closest distance to a traffic light
        given_tcs_ref = 0
        tcs_jcts = []
        TCS_REF_JCTS = []
        all_junctions_tc = {jctID: None for jctID in net_jct_children}

    tcs_jct_attrs = [[]]
    non_deadEnd_jcts = []
    jcts_param = dict()
    for jctID in all_junctions_tc:
        jct_child = net_jct_children[jctID]
        jct_type = jct_child.get("type")
        if jct_type != "dead_end" and jct_type != "internal":
            # Exclude junctions at the border of the network (usually where the flow of vehicles are generated),
            # and also internal junctions (junctions inside other junctions)
            non_deadEnd_jcts.append(jctID)
            jctx = float(jct_child.get("x"))
            jcty = float(jct_child.get("y"))
            neighJcts = {edge_child.get("from") for edge_child in net_edg_children
                         if edge_child.get("to") == jctID}
            neighJcts = neighJcts.union({edge_child.get("to") for edge_child in net_edg_children
                                         if edge_child.get("from") == jctID})
            jcts_param.update({jctID: [jct_type, neighJcts, jctx, jcty]})

            if given_tcs_ref == 0 and jct_type == "traffic_light":
                tcs_jcts.append([])
                # Set traffic light controlled junction as reference junction and as a new traffic controller
                TCS_REF_JCTS.append((("ref.", (jctID)), (jctx, jcty)))

    if ORIGIN_COORD != ():
        # Define the order of TCs in case not given TCS_REF_JCTS
        # Order based on reference junctions distance to ORIGIN_COORD
        origin_x = ORIGIN_COORD[0]
        origin_y = ORIGIN_COORD[1]
        origin_dists_tcs = sorted([(tc_ind, math.sqrt((origin_x - tc_ref_x) ** 2 + (origin_y - tc_ref_y) ** 2))
                                   for tc_ind, (_, (tc_ref_x, tc_ref_y)) in enumerate(TCS_REF_JCTS)],
                                  key=lambda sort_it: (sort_it[1]))
        old_TCS_REF_JCTS = TCS_REF_JCTS[:]
        TCS_REF_JCTS = []
        for tc_ind, _ in origin_dists_tcs:
            # Modify TCS_REF_JCTS order/sequence of TCs (traffic controllers) based on origin_dists_tcs
            TCS_REF_JCTS.append(old_TCS_REF_JCTS[tc_ind])

    copied_jcts_param = jcts_param.copy()
    for jctID in copied_jcts_param:
        jctx = jcts_param[jctID][2]
        jcty = jcts_param[jctID][3]
        try:
            # Define the tc_ind if junction is set to be modelled by certain TC
            jct_tc_ind = [tc_ind for tc_ind, ((_, tc_jcts), _) in enumerate(TCS_REF_JCTS) if jctID in tc_jcts][0]
            if TCS_REF_JCTS[jct_tc_ind][0][0] == "ref.":
                jct_dist_origin = math.sqrt((jctx - TCS_REF_JCTS[jct_tc_ind][1][0]) ** 2
                                            + (jcty - TCS_REF_JCTS[jct_tc_ind][1][1]) ** 2)
            else:
                jct_dist_origin = TCS_REF_JCTS[jct_tc_ind][0][1].index(jctID)
        except IndexError:
            # Define the tc_ind based on the distance to each tc reference point (the closest one)
            jct_dists_tcs = sorted([(tc_ind, math.sqrt((jctx - tc_ref_x) ** 2 + (jcty - tc_ref_y) ** 2))
                                    for tc_ind, (_, (tc_ref_x, tc_ref_y)) in enumerate(TCS_REF_JCTS)],
                                   key=lambda sort_it: (sort_it[1]))
            # Choose the closest one
            jct_tc_ind, jct_dist_origin = jct_dists_tcs[0]

        if given_tcs_ref == 0:
            all_junctions_tc.update({jctID: jct_tc_ind})
        tcs_jcts[jct_tc_ind].append((jctID, jct_dist_origin))

    connected_jcts = []
    for tc_ind in range(0, len(TCS_REF_JCTS)):
        # As the TC reference points are supposed to be where it is expected the first junction of the TC following the
        # order (sequence of jcts. to be modelled), then order junctions based on their distances from the tcs. ref. point
        tcs_jcts[tc_ind] = [jct[0] for jct in sorted(tcs_jcts[tc_ind], key=lambda sort_it: sort_it[1])]
        if CENTRALIZED_SYS_BOOL or tc_ind == 0:
            # If only one traffic controller or first one, not needed to create a new tcs_jct_attrs list
            pass
        else:
            tcs_jct_attrs.append([])
        for jctID in tcs_jcts[tc_ind]:
            jct_type = jcts_param[jctID][0]
            neighJcts = jcts_param[jctID][1]
            if CENTRALIZED_SYS_BOOL:
                jct_neighJcts_other_TC = ""
            else:
                # Add neighbouring junctions (junctions modelled by different traffic controllers
                # but they are connected by same edges as outgoing from one and incoming for other)
                jct_neighJcts_other_TC = [neighJct for neighJct in neighJcts
                                          if neighJct in non_deadEnd_jcts and neighJct not in tcs_jcts[tc_ind]]
                if jct_neighJcts_other_TC == []:
                    jct_neighJcts_other_TC = ""
            if jctID in JCTS_WH_RSU:
                # If junction should have Road Side Unit (RSU), add "Coop." (from Cooperative) in its types
                jct_types = ["Coop."]
            else:
                jct_types = []
            if jct_type == "traffic_light":
                # tcs_jct_attrs = [[ Ego Traffic Controller list of junctions
                                #  (ego junction ID,
                                #   ego junction types (allowed Coop., if applies, with either Mod. C. or Mod. Non-C.),
                                #   ego junction's traffic light ID,
                                #   ego junctions's neighbouring junctions modelled by another Traffic Controller),
                                #  (inputs for another ego junction...)
                                #  ]
                                #  Then another list for another ego TC...
                                # ]
                if jctID in SIGNALIZED_JCTS_NON_C:
                    jct_types.append("Mod. Non-C.")
                else:
                    jct_types.append("Mod. C.")
                    connected_jcts.append(jctID)

                # Get junction traffic light ID (tlID)
                global tlID
                for conn in net_conn_children:
                    # Old way to check if connection is within jctID
                    # if ("_").join(conn.get('via').split("_")[:-2])[1:] == jctID and conn.get('tl') != None:
                    # New way to check if connection is within jctID
                    if jctID in conn.get('via') and conn.get('tl') != None:
                        tlID = conn.get('tl')
                        break

                tcs_jct_attrs[-1].append((jctID, jct_types, tlID, jct_neighJcts_other_TC))
            else:
                if jctID in UNSIGNALIZED_JCTS_C:
                    jct_types.append("Mod. C.")
                    connected_jcts.append(jctID)
                else:
                    jct_types.append("Mod. Non-C.")
                tcs_jct_attrs[-1].append((jctID, jct_types, "", jct_neighJcts_other_TC))

    # Define the traffic network objects
    msgs_objs = []

    # Define class objects of the traffic controllers for area A
    subnetA_mailbox = LdmDataMsgs("subnetA_mailbox")
    msgs_objs.append(subnetA_mailbox)
    subnet = subnetA_mailbox

    # Define TC objects
    jct_objs = []
    movg_objs = []
    edg_objs = []
    veh_objs = []
    num_tcs = len(tcs_jct_attrs)
    subnet_traffic_controllers = range(0, num_tcs)
    for tc in subnet_traffic_controllers:
        jct_obj_name = "tc" + str(tc) + "_jct"
        exec (jct_obj_name + " = Junctions(jct_obj_name)")
        jct_objs.append(eval(jct_obj_name))
        movg_obj_name = "tc" + str(tc) + "_movg"
        exec (movg_obj_name + " = MovementGroups(movg_obj_name)")
        movg_objs.append(eval(movg_obj_name))
        edg_obj_name = "tc" + str(tc) + "_edg"
        exec (edg_obj_name + " = Edges(edg_obj_name)")
        edg_objs.append(eval(edg_obj_name))
        veh_obj_name = "tc" + str(tc) + "_veh"
        exec (veh_obj_name + " = Vehicles(veh_obj_name)")
        veh_objs.append(eval(veh_obj_name))

    # Define args for files to be called
    edges_args = ["python", "create_file_net_selected_objs.py", "-n", NET_FILE]
    all_jcts = ""
    # Call functions to add the junctions
    for tc_ind, jct_attrs in enumerate(tcs_jct_attrs):
        tc_jcts = ""
        for jct_attr in jct_attrs:
            tc_jcts += jct_attr[0] + " "
            all_jcts += jct_attr[0] + " "
            if jct_attr[2] != "" and jct_attr[3] != "":
                exec ("tc" + str(tc_ind) + "_jct" + ".addJunction(subnet, \"" + str(jct_attr[0]) + "\",\""
                                                                              + str(jct_attr[1]) + "\",\""
                                                                              + str(jct_attr[2]) + "\","
                                                              + "neigh_jcts=" + str(jct_attr[3]) + ")")
            elif jct_attr[2] != "":
                exec ("tc" + str(tc_ind) + "_jct" + ".addJunction(subnet, \"" + str(jct_attr[0]) + "\",\""
                                                                              + str(jct_attr[1]) + "\",\""
                                                                              + str(jct_attr[2]) + "\")")
            elif jct_attr[3] != "":
                exec ("tc" + str(tc_ind) + "_jct" + ".addJunction(subnet, \"" + str(jct_attr[0]) + "\",\""
                                                                              + str(jct_attr[1]) + "\","
                                                              + "neigh_jcts=" + str(jct_attr[3]) + ")")
            else:
                exec ("tc" + str(tc_ind) + "_jct" + ".addJunction(subnet, \"" + str(jct_attr[0]) + "\",\""
                                                                              + str(jct_attr[1]) + "\")")
        else:
            tc_jcts = tc_jcts[:-1]

        if os.path.isfile(PATH_OUTPUT + "selected_tc" + str(tc_ind) + "_edges."
                          + SCE_NAME.split(".")[0] + ".png") == False:
            # Create TXT with Incoming Edges of the TC
            tc_edges_args = edges_args + ["--incoming"] + \
                            ["-r", PATH_OUTPUT + "tc" + str(tc_ind) + "_edges."
                             + SCE_NAME.split(".")[0] + ".txt"] + ["-j", tc_jcts]
            process = subprocess.Popen(tc_edges_args, env=MY_ENV)
            process.wait()

            # Plot Edge Selection per TC
            process = subprocess.Popen(["python", "plot_net_selected_edgesLanes.py", "-n", NET_FILE,
                                        "-i", PATH_OUTPUT + "tc" + str(tc_ind) + "_edges."
                                        + SCE_NAME.split(".")[0] + ".txt", "-b", "--no-ticks",
                                        "--xlim", NET_XLIM, "--ylim", NET_YLIM,
                                        "--title", "TC" + str(tc_ind) + " Modelled Edges",
                                        "-o", PATH_OUTPUT + "selected_tc" + str(tc_ind) + "_edges."
                                        + SCE_NAME.split(".")[0] + ".png"], env=MY_ENV)
            # -b = If set, the figure will not be shown
            process.wait()

    else:
        # Delete the last space character " " after adding all junctions in the string
        all_jcts = all_jcts[:-1]

    if os.path.isfile(PATH_OUTPUT + "jcts_TLS." + SCE_NAME.split(".")[0] + ".png") == False:
        # Plot Junctions with Traffic Light
        process = subprocess.Popen(["python", "plot_net_selected_junctions.py", "-n", NET_FILE,
                                    "-b", "--no-ticks",
                                    "--xlim", NET_XLIM, "--ylim", NET_YLIM,
                                    "--title", "Location of Traffic Lights",
                                    "-o", PATH_OUTPUT + "jcts_TLS." + SCE_NAME.split(".")[0] + ".png"], env=MY_ENV)
        # -b = If set, the figure will not be shown
        process.wait()

    if os.path.isfile(PATH_OUTPUT + "jcts_C." + SCE_NAME.split(".")[0] + ".png") == False:
        # Plot Connected Junctions
        jctsC = ""
        for jct in connected_jcts:
            jctsC += jct + " "
        else:
            jctsC = jctsC[:-1]
        process = subprocess.Popen(["python", "plot_net_selected_junctions.py", "-n", NET_FILE,
                                    "-b", "--no-ticks", "-j", jctsC,
                                    "--xlim", NET_XLIM, "--ylim", NET_YLIM,
                                    "--title", "Connected Junctions (with Queue Length Est.)",
                                    "-o", PATH_OUTPUT + "jcts_C." + SCE_NAME.split(".")[0] + ".png"], env=MY_ENV)
        # -b = If set, the figure will not be shown
        process.wait()

    if os.path.isfile(PATH_OUTPUT + "jcts_RSU." + SCE_NAME.split(".")[0] + ".png") == False:
        # Plot Junctions with RSU
        jctsRSU = ""
        for jct in JCTS_WH_RSU:
            jctsRSU += jct + " "
        else:
            jctsRSU = jctsRSU[:-1]
        process = subprocess.Popen(["python", "plot_net_selected_junctions.py", "-n", NET_FILE,
                                    "-j", jctsRSU, "-b", "--no-ticks",
                                    "--xlim", NET_XLIM, "--ylim", NET_YLIM,
                                    "--title", "Junctions with RSU",
                                    "-o", PATH_OUTPUT + "jcts_RSU." + SCE_NAME.split(".")[0] + ".png"], env=MY_ENV)
        # -b = If set, the figure will not be shown
        process.wait()

    if EDGE2CLOSE != [] and END_T > BEGIN_LANE_CLOSURE and BEGIN_T < END_LANE_CLOSURE:
        if os.path.isfile(PATH_OUTPUT + "closed_lanes." + SCE_NAME.split(".")[0] + ".png") == False:
            # Create TXT with Closed Lanes
            closed_edges_args = edges_args + ["-r", PATH_OUTPUT + "closed_lanes."
                                              + SCE_NAME.split(".")[0] + ".txt"] + ["-l", ",".join(LANES2CLOSE)]
            process = subprocess.Popen(closed_edges_args, env=MY_ENV)
            process.wait()

            # Plot Closed Lanes
            process = subprocess.Popen(["python", "plot_net_selected_edgesLanes.py", "-n", NET_FILE, "-b", "--no-ticks",
                                        "-i", PATH_OUTPUT + "closed_lanes." + SCE_NAME.split(".")[0] + ".txt", "-b",
                                        "--no-ticks", "--xlim", NET_XLIM, "--ylim", NET_YLIM,
                                        "--title", "Closed Lanes",
                                        "-o", PATH_OUTPUT + "closed_lanes." + SCE_NAME.split(".")[0] + ".png"],
                                       env=MY_ENV)
            # -b = If set, the figure will not be shown
            process.wait()

            # Create another file with a zoom around the closed lanes
            for zoom_ind in range(0, len(NET_XLIM_CLOSED)):
                process = subprocess.Popen(["python", "plot_net_selected_edgesLanes.py", "-n", NET_FILE, "-b",
                                            "--no-ticks",
                                            "-i", PATH_OUTPUT + "closed_lanes."
                                            + SCE_NAME.split(".")[0] + ".txt", "-b", "--no-ticks",
                                            "--selected-width=5", "--edge-width=4",
                                            "--xlim", NET_XLIM_CLOSED[zoom_ind], "--ylim", NET_YLIM_CLOSED[zoom_ind],
                                            "--title", "Closed Lanes Location " + str(zoom_ind),
                                            "-o", PATH_OUTPUT + "closed_lanes_zoom" + str(zoom_ind) + "."
                                            + SCE_NAME.split(".")[0] + ".png"], env=MY_ENV)
                # -b = If set, the figure will not be shown
                process.wait()

    if os.path.isfile(PATH_OUTPUT + "probes_OD." + SCE_NAME.split(".")[0] + ".png") == False:
        # Plot Empty Network (For Probe OD Pairs)
        process = subprocess.Popen(["python", "plot_net_selected_edgesLanes.py", "-n", NET_FILE, "-b", "--no-ticks",
                                    "--xlim", NET_XLIM, "--ylim", NET_YLIM,
                                    "--title", "Origin-Destination Pair of Probe Vehicles",
                                    "-o", PATH_OUTPUT + "probes_OD." + SCE_NAME.split(".")[0] + ".png"], env=MY_ENV)
        # -b = If set, the figure will not be shown
        process.wait()

    if os.path.isfile(PATH_OUTPUT + "allInEdges_modelled_area." + SCE_NAME.split(".")[0] + ".xlsx") == False:
        # Create XLSX with Incoming Edges modeled by LLR
        inc_edges_args = edges_args + ["--incoming"] \
                         + ["-r", PATH_OUTPUT + "allInEdges_modelled_area." + SCE_NAME.split(".")[0] + ".xlsx"] \
                         + ["-j", all_jcts]
        process = subprocess.Popen(inc_edges_args, env=MY_ENV)
        process.wait()

    if os.path.isfile(PATH_OUTPUT + "allInOutEdges_modelled_area." + SCE_NAME.split(".")[0] + ".xlsx") == False:
        # Create XLSX with Incoming Edges modeled by LLR
        inc_edges_args = edges_args + ["--incoming"] + ["--outgoing"] \
                         + ["-r", PATH_OUTPUT + "allInOutEdges_modelled_area." + SCE_NAME.split(".")[0] + ".xlsx"] \
                         + ["-j", all_jcts]
        process = subprocess.Popen(inc_edges_args, env=MY_ENV)
        process.wait()

    if os.path.isfile(PATH_OUTPUT + "allConn_modelled_area." + SCE_NAME.split(".")[0] + ".xlsx") == False:
        # Create XLSX with Connection Edges modeled by LLR
        conn_edges_args = edges_args + ["--connections"] \
                          + ["-r", PATH_OUTPUT + "allConn_modelled_area." + SCE_NAME.split(".")[0] + ".xlsx"] \
                          + ["-j", all_jcts]
        process = subprocess.Popen(conn_edges_args, env=MY_ENV)
        process.wait()

    if os.path.isfile(PATH_OUTPUT + "allEdges_modelled_area." + SCE_NAME.split(".")[0] + ".xlsx") == False:
        # Create XLSX with Modelled Edges modeled by LLR
        allEdges_modelled_area_args = edges_args + ["--incoming"] + ["--outgoing"] + ["--connections"] \
                                      + ["-r", PATH_OUTPUT + "allEdges_modelled_area." + SCE_NAME.split(".")[0]
                                         + ".xlsx"] \
                                      + ["-j", all_jcts]
        process = subprocess.Popen(allEdges_modelled_area_args, env=MY_ENV)
        process.wait()

    for tc_ind in range(0, len(jct_objs)):
        # Get Vehicle Composition
        if vehComposition_children == []:
            # If no vehicle composition, use only the standard value for a passenger car
            vehComposition_children = [None]
            veh_objs[tc_ind].vType_vals["probability"][0] = 1

        # If probabilities of vehicle composition don't sum 100%, distribute the difference among the vehicle types
        global sum_probs, prob_correction
        sum_probs = 0
        for vehComposition_child in vehComposition_children:
            sum_probs += float(vehComposition_child.get("probability"))

        prob_correction = 1 - sum_probs
        prob_correction = prob_correction / len(vehComposition_children)

        # Get attributes of the vehicle composition
        # When no information is available, use default values from veh_objs[tc_ind].vType_vals[attribute][0]
        for vehComp_ind, vehComposition_child in enumerate(vehComposition_children):
            if vehComposition_child.get("id") == None:
                veh_objs[tc_ind].vType_vals["ids"].append("CUSTOM_" + str(vehComp_ind))
            else:
                veh_objs[tc_ind].vType_vals["ids"].append(vehComposition_child.get("id"))
            if vehComposition_child.get("probability") == None:
                veh_objs[tc_ind].vType_vals["probability"].append(1)
            else:
                veh_objs[tc_ind].vType_vals["probability"].append(float(vehComposition_child.get("probability"))
                                                                  + prob_correction)
            if vehComposition_child.get("accel") == None:
                veh_objs[tc_ind].vType_vals["acc"].append(veh_objs[tc_ind].vType_vals["acc"][0])
            else:
                veh_objs[tc_ind].vType_vals["acc"].append(float(vehComposition_child.get("accel")))
            if vehComposition_child.get("decel") == None:
                veh_objs[tc_ind].vType_vals["decel"].append(veh_objs[tc_ind].vType_vals["decel"][0])
            else:
                veh_objs[tc_ind].vType_vals["decel"].append(float(vehComposition_child.get("decel")))
            if vehComposition_child.get("tau") == None:
                veh_objs[tc_ind].vType_vals["min_gap_t"].append(veh_objs[tc_ind].vType_vals["min_gap_t"][0])
            else:
                veh_objs[tc_ind].vType_vals["min_gap_t"].append(float(vehComposition_child.get("tau")))
            if vehComposition_child.get("length") == None:
                veh_objs[tc_ind].vType_vals["length"].append(veh_objs[tc_ind].vType_vals["length"][0])
            else:
                veh_objs[tc_ind].vType_vals["length"].append(float(vehComposition_child.get("length")))
            if vehComposition_child.get("minGap") == None:
                veh_objs[tc_ind].vType_vals["min_gap_d"].append(veh_objs[tc_ind].vType_vals["min_gap_d"][0])
            else:
                veh_objs[tc_ind].vType_vals["min_gap_d"].append(float(vehComposition_child.get("minGap")))
            if vehComposition_child.get("actionStepLength") == None:
                veh_objs[tc_ind].vType_vals["reaction_t"].append(STD_REACTION_T)
            else:
                veh_objs[tc_ind].vType_vals["reaction_t"].append(float(vehComposition_child.get("actionStepLength")))

        for jct_ind,jctID in enumerate(jct_objs[tc_ind].ids):
            # Define edges from junctions not modelled by the traffic controller
            edges_from_unknown_jcts = [edge.get('id') for edge in net_edg_children if
                                       all_junctions_tc.has_key(edge.get('from')) == False and edge.get('to') == jctID]
            jct_edg_children = [edge for edge in net_edg_children
                                if edge.get('from') == jctID or edge.get('to') == jctID]
            if TC_IT_MAX == 1:
                # Add Lanes that connect junctions according to the order of insertion of modelled junctions by the TC
                next_jctIDs = jct_objs[tc_ind].ids[jct_ind + 1:]
                previous_jctIDs = jct_objs[tc_ind].ids[:jct_ind]
            else:
                # Add Lanes that connect junctions according to the order of insertion of modelled junctions by all TCs
                next_jctIDs = []
                for tc_obj in range(tc_ind, len(jct_objs)):
                    if tc_obj == tc_ind:
                        next_jctIDs.extend(jct_objs[tc_obj].ids[jct_ind + 1:])
                    else:
                        next_jctIDs.extend(jct_objs[tc_obj].ids[:])
                previous_jctIDs = []
                for tc_obj in range(0, tc_ind + 1):
                    if tc_obj == tc_ind:
                        previous_jctIDs.extend(jct_objs[tc_obj].ids[:jct_ind])
                    else:
                        previous_jctIDs.extend(jct_objs[tc_obj].ids[:])

                edge_next_jcts_children = [edge for edge in jct_edg_children
                                           if (edge.get('from') == jctID and edge.get('to') in next_jctIDs)
                                           or (edge.get('from') in previous_jctIDs and edge.get('to') == jctID)]

                edge_prev_jcts_children = [edge for edge in jct_edg_children
                                           if (edge.get('from') == jctID and edge.get('to') in previous_jctIDs)
                                           or (edge.get('from') in next_jctIDs and edge.get('to') == jctID)]

                edge_from_jct_children = [edge for edge in jct_edg_children
                                          if edge.get('from') == jctID
                                          and edge.get('to') not in previous_jctIDs+next_jctIDs]

                edge_to_jct_children = [edge for edge in jct_edg_children
                                        if edge.get('to') == jctID
                                        and edge.get('from') not in previous_jctIDs+next_jctIDs]

            # Define the edges based on the order of insertion of the junctions and TC objects
            forwardOrder = []
            reverseOrder = []
            lane_fromJct_children = [edge_child.findall("lane") for edge_child in edge_from_jct_children]
            for lane_children in lane_fromJct_children:
                forwardOrder.extend([lane_child.get("id") for lane_child in lane_children])
                reverseOrder.extend([lane_child.get("id") for lane_child in lane_children])

            lane_toJct_children = [edge_child.findall("lane") for edge_child in edge_to_jct_children]
            for lane_children in lane_toJct_children:
                forwardOrder.extend([lane_child.get("id") for lane_child in lane_children])
                reverseOrder.extend([lane_child.get("id") for lane_child in lane_children])

            lane_order_children = [edge_child.findall("lane") for edge_child in edge_next_jcts_children]
            for lane_children in lane_order_children:
                forwardOrder.extend([lane_child.get("id") for lane_child in lane_children])

            lane_reverseOrder_children = [edge_child.findall("lane") for edge_child in edge_prev_jcts_children]
            for lane_children in lane_reverseOrder_children:
                reverseOrder.extend([lane_child.get("id") for lane_child in lane_children])
            jct_objs[tc_ind].lanesForwardOrder.update({jctID: forwardOrder})
            jct_objs[tc_ind].lanesReverseOrder.update({jctID: reverseOrder})

            # Add common lanes with neighbouring border junctions
            for neigh_jct in jct_objs[tc_ind].neigh_jcts[jctID]:
                commonTo_edge_children = [edge for edge in jct_edg_children
                                          if edge.get('from') == jctID and edge.get('to') == neigh_jct]
                commonFrom_edge_children = [edge for edge in jct_edg_children
                                            if edge.get('from') == neigh_jct and edge.get('to') == jctID]
                commonTo_edg_lanes_children = [edge_child.findall("lane") for edge_child in commonTo_edge_children]
                commonFrom_edg_lanes_children = [edge_child.findall("lane") for edge_child in commonFrom_edge_children]
                commonTo_laneIDs = []
                commonFrom_laneIDs = []
                for edge_lane_children in commonTo_edg_lanes_children:
                    commonTo_laneIDs.extend([lane_child.get("id") for lane_child in edge_lane_children])
                for edge_lane_children in commonFrom_edg_lanes_children:
                    commonFrom_laneIDs.extend([lane_child.get("id") for lane_child in edge_lane_children])
                jct_objs[tc_ind].commonToNeigh_laneIDs[jctID].update({neigh_jct: commonTo_laneIDs[:]})
                jct_objs[tc_ind].commonFromNeigh_laneIDs[jctID].update({neigh_jct: commonFrom_laneIDs[:]})

            # Add Lanes
            for edge_child in jct_edg_children:
                edgeID = edge_child.get("id")
                edge_type = []
                laneIDs = []
                lanes_length = []
                lanes_max_speed = []
                if edge_child.get('to') == jctID:
                    edge_type.append("incoming")
                if edge_child.get('from') == jctID:
                    edge_type.append("outgoing")
                for lane_child in list(edge_child.findall("lane")):
                    laneIDs.append(lane_child.get("id"))
                    if edgeID not in edg_objs[tc_ind].edge_ids:
                        lanes_length.append(float(lane_child.get("length")))
                        lanes_max_speed.append(float(lane_child.get("speed")))

                if edgeID not in edg_objs[tc_ind].edge_ids:
                    edg_objs[tc_ind].addEdgeLanes(jct_objs[tc_ind], edgeID, laneIDs, edge_type, lanes_length,
                                                  lanes_max_speed=lanes_max_speed)
                else:
                    edg_objs[tc_ind].addEdgeLanes(jct_objs[tc_ind], edgeID, laneIDs, edge_type)

                if "outgoing" in edge_type:
                    tcs_already_added = []
                    for neighID in jct_objs[tc_ind].neigh_jcts[jctID]:
                        neigh_jct_tc_ind = all_junctions_tc[neighID]
                        if neigh_jct_tc_ind not in tcs_already_added:
                            tcs_already_added.append(neigh_jct_tc_ind)
                            # If edge is in common with a neighbouring junction, also add edge information into
                            # traffic controller of neighbouring junction
                            edg_objs[neigh_jct_tc_ind].addEdgeLanes(jct_objs[neigh_jct_tc_ind], edgeID, laneIDs, add_neigh_tc_bool=1)

    for tc_ind in range(0, len(jct_objs)):
        for jctID in jct_objs[tc_ind].ids:
            # Add Connections
            all_commom_lanes_tc = dict()
            for neighID in jct_objs[tc_ind].neigh_jcts[jctID]:
                for lane in jct_objs[tc_ind].commonFromNeigh_laneIDs[jctID][neighID]:
                    all_commom_lanes_tc.update({lane: all_junctions_tc[neighID]})

            jct_conn_children = {conn.get('via'): conn for conn in net_conn_children
                                 # net_conn_children is filtered already for connection from edges to be modelled by LLR
                                 # if conn.get('via') != None
                                 if jctID in conn.get('via')}
                                 # if ("_").join(conn.get('via').split("_")[:-2])[1:] == jctID}
            jct_conn_children_not_internal = {conn.get('via'): conn for conn in net_conn_children
                                              # if conn.get('via') != None
                                              # if ("_").join(conn.get('via').split("_")[:-2])[1:] == jctID
                                              if jctID in conn.get('via')
                                              and conn.get('from') in edg_objs[tc_ind].edge_ids
                                              and conn.get('to') in edg_objs[tc_ind].edge_ids}

            jct_internal_edg_children = {edg_conn.get('id'): edg_conn for edg_conn in net_internal_edg_children
                                         if jctID in edg_conn.get('id')}
                                         # if ("_").join(edg_conn.get('id').split("_")[:-1])[1:] == jctID}

            for connID in jct_conn_children_not_internal:
                conn_child = jct_conn_children_not_internal[connID]
                edge_connID = ("_").join(connID.split("_")[:-1])
                in_laneID = conn_child.get('from') + "_" + conn_child.get('fromLane')
                out_laneID = conn_child.get('to') + "_" + conn_child.get('toLane')
                conn_dir = conn_child.get("dir").lower()

                via_connID = connID
                edg_conn_children = [[via_connID, jct_internal_edg_children[edge_connID]]]
                while 1:
                    for internal_connID in jct_conn_children:
                        if jct_conn_children[internal_connID].get("from") == ("_").join(via_connID.split("_")[:-1]):
                            edge_connID = ("_").join(internal_connID.split("_")[:-1])
                            edg_conn_children.append([internal_connID, jct_internal_edg_children[edge_connID]])
                            via_connID = internal_connID
                            # Stop searching internal connections
                            break
                    else:
                        # If looked for all connections, it means it was the last internal connection to out_lane
                        break

                # The connection length is the sum of each subsequent internal connections until outgoing lane
                conn_length = []
                for via_connID,edg_conn_child in edg_conn_children:
                    for edg_conn_lane in list(edg_conn_child.findall("lane")):
                        if edg_conn_lane.get("id") == via_connID:
                            conn_length.append(float(edg_conn_lane.get("length")))
                            break

                edg_objs[tc_ind].addConnection(veh_objs[tc_ind], connID, in_laneID, out_laneID, jct_objs[tc_ind],
                                               conn_dir, jctID, sum(conn_length))

                if all_commom_lanes_tc.has_key(in_laneID):
                    # If incoming lane is in common with a neighbouring junction, also add connection information into
                    # traffic controller of neighbouring junction
                    edg_objs[all_commom_lanes_tc[in_laneID]].addConnection(veh_objs[tc_ind],
                                                                           connID, in_laneID, out_laneID,
                                                                           add_neigh_tc_bool=1)

            # For yielding/priority SUMO uses the internal lanes, which are either:
            # a) the connection/via name when direct connection from in_lane to out_lane
            # b) internal lane int_lane where vehicle waits when indirect conn. from in_lane to int_lane to out_lane
            # Assign the connection name from the via to the internal lane when there is indirect connection
            tcs_common_conns = dict()
            for tc in range(0, len(jct_objs)):
                tcs_common_conns.update({tc: []})
            for neighID in jct_objs[tc_ind].neigh_jcts[jctID]:
                for lane in jct_objs[tc_ind].commonFromNeigh_laneIDs[jctID][neighID]:
                    for conn in edg_objs[tc_ind].inLaneID_to_connIDs[lane]:
                        tcs_common_conns[all_junctions_tc[neighID]].append(conn)

            # Child of the junction jctID
            jct_child = net_jct_children[jctID]
            jct_type = jct_child.get("type")
            # Child of each internal junction within jctID (when vehicles yield and wait in the middle of the junction)
            jct_internal_children = [net_jct_children[jct] for jct in net_jct_children
                                     if jctID in net_jct_children[jct].get('id')]
                                     # if ("_").join(net_jct_children[jct].get('id').split("_")[:-2])[1:] == jctID]
            all_conns = jct_child.get("intLanes").split(" ")  # all connections are internal lanes of the junction
            copied_all_conns = all_conns[:]
            for connID in jct_objs[tc_ind].connIDs[jctID]:
                # Check if connID correspond to an internal lane of intLanes from SUMO
                if connID in copied_all_conns:
                    # if yes, it is a direct connection
                    pass
                else:
                    # if not, it is an indirect connection
                    # Get the internal lane that has its incoming lane the via name of the connection from in_lane
                    # to out_lane
                    conn2assign = connID
                    it = 0
                    while conn2assign not in copied_all_conns:
                        it += 1
                        conn2assign = [intJct_child.get("id") for intJct_child in jct_internal_children if
                                       conn2assign in intJct_child.get("incLanes").split(" ")][0]

                    conn_ind = copied_all_conns.index(conn2assign)
                    # Assign the via name to the indirect connection
                    all_conns[conn_ind] = connID

            # Add Movement Groups from Signal Groups (connections in SUMO that have same green pattern)
            # and Yield to Connections for Traffic Light Groups
            if "traffic_light" in jct_type:
                tlID = jct_objs[tc_ind].tlID[jctID]
                signal_child = tls_signal_children[tlID]
                phases_children = list(signal_child.findall("phase"))
                linkIndexs_pattern = []
                cycle_t = []

                # Get linkIndexs_pattern
                for phase_num, phase_child in enumerate(phases_children):
                    cycle_t.append(float(phase_child.get("duration")))
                    links_state = phase_child.get("state")
                    for linkIndex, linkState in enumerate(links_state):
                        try:
                            if linkState == "G":
                                linkIndexs_pattern[linkIndex][0].append(phase_num)
                            elif linkState == "g":
                                linkIndexs_pattern[linkIndex][1].append(phase_num)
                        except IndexError:
                            num_links = len(links_state)
                            linkIndexs_pattern = [[[],  # [linkIndexs with protected green per phase]
                                                   []]  # [linkIndexs with permissive green per phase]
                                                  for _ in range(0, num_links)]
                            if linkState == "G":
                                linkIndexs_pattern[linkIndex][0].append(phase_num)
                            elif linkState == "g":
                                linkIndexs_pattern[linkIndex][1].append(phase_num)

                # Define the pattern of each movement group
                cycle_t = sum(cycle_t)
                if STD_T_MAX > cycle_t:
                    t_max = np.ceil(STD_T_MAX / float(cycle_t)) * cycle_t
                else:
                    t_max = cycle_t
                jct_objs[tc_ind].t_max.update({jctID: t_max})
                mov_groups_links = []
                mov_groups_pattern = []
                for linkIndex, linkPattern in enumerate(linkIndexs_pattern):
                    try:
                        mov_group_num = mov_groups_pattern.index(linkPattern)
                        mov_groups_links[mov_group_num].append(linkIndex)
                    except ValueError:
                        mov_groups_links.append([linkIndex])
                        mov_groups_pattern.append(linkPattern)

                # Assign connections based on link indexes
                for movg_num, movg_links in enumerate(mov_groups_links):
                    linkIndexs = movg_links[:]
                    connIDs = [[] for _ in range(0, len(linkIndexs))]
                    for connID in jct_conn_children:
                        conn_child = jct_conn_children[connID]
                        try:
                            linkIndex = int(conn_child.get("linkIndex"))
                            if linkIndex in linkIndexs:
                                link_ind = linkIndexs.index(linkIndex)
                                connIDs[link_ind] = connID
                            else:
                                try:
                                    # Check if remaining connections to assign their linkIndex
                                    _ = connIDs.index([])
                                except ValueError:
                                    # Found all connections
                                    break
                        except TypeError:
                            # Internal Connection, it doesn't have linkIndex
                            internal_conn = 1

                    while 1:
                        try:
                            # Remove the missing connections when they are located on other junction with same tlID
                            missing_conn = connIDs.index([])
                            del linkIndexs[missing_conn]
                            del connIDs[missing_conn]
                        except ValueError:
                            break

                    if connIDs != []:
                        # Add movement group information
                        movg_objs[tc_ind].addMovGroup(edg_objs[tc_ind], jctID, movg_num, connIDs, jct_objs[tc_ind],
                                                      tlID, linkIndexs)

                        for neigh_tc_ind in tcs_common_conns.keys():
                            movg_common_conns = set(connIDs).intersection(set(tcs_common_conns[neigh_tc_ind]))
                            if movg_common_conns != set([]):
                                # Add movement group information into a traffic controller modelling a neighbouring
                                # junction
                                movg_objs[neigh_tc_ind].addMovGroup(edg_objs[neigh_tc_ind], jctID, movg_num,
                                                                    movg_common_conns, add_neigh_tc_bool=1)

                for conn_ind, conn_child in enumerate(list(jct_child.findall("request"))):
                    # Get which connections yield to each connection
                    connID = all_conns[conn_ind]
                    response = conn_child.get("response")[::-1] # the conn_index is inverted
                    yield2conns = [all_conns[conn_index] for conn_index, conn_val in enumerate(response) if
                                   conn_val == "1"]

                    # Add Yield to Movement Group of Signilised Control
                    edg_objs[tc_ind].addConnYield2Conns(movg_objs[tc_ind], connID, yield2conns)

            # Add Movement Groups from Priority Groups (connections in SUMO that have same priority pattern)
            # and Yield to Connections for Priority Groups
            else:
            # elif jct_child.get("type") == "priority":
                yield2conn_patterns = [[]  # [foeConnIndexs with permissive green per connection]
                                       for _ in range(0, len(all_conns))]
                for conn_ind, conn_child in enumerate(list(jct_child.findall("request"))):
                    response = conn_child.get("response")[::-1] # [::-1] inverts the response due SUMO gives it inverted
                    for foeConnIndex, conn_response in enumerate(response):
                        if conn_response == "1":
                            # If connection has to yield to another connection
                            foeConnID = all_conns[foeConnIndex]
                            yield2conn_patterns[conn_ind].append(foeConnID)

                mov_groups_conns = []
                mov_groups_pattern = []
                # Define movement groups by the pattern of yielding to same connections
                for conn_ind, connPattern in enumerate(yield2conn_patterns):
                    connID = all_conns[conn_ind]  # correspondent via name connection to the internal lane
                    try:
                        mov_group_num = mov_groups_pattern.index(connPattern)
                        mov_groups_conns[mov_group_num].append(connID)
                    except ValueError:
                        mov_groups_conns.append([connID])
                        mov_groups_pattern.append(connPattern)

                for movg_num, movg_conns in enumerate(mov_groups_conns):
                    # Add Movement Group of Unsignilized Junction
                    movg_objs[tc_ind].addMovGroup(edg_objs[tc_ind], jctID, movg_num, movg_conns, jct_objs[tc_ind])

                    for neigh_tc_ind in tcs_common_conns.keys():
                        movg_common_conns = set(movg_conns).intersection(set(tcs_common_conns[neigh_tc_ind]))
                        if movg_common_conns != set([]):
                            # Add movement group information into a traffic controller modelling a neighbouring
                            # junction
                            movg_objs[neigh_tc_ind].addMovGroup(edg_objs[neigh_tc_ind],
                                                                jctID, movg_num, movg_common_conns, add_neigh_tc_bool=1)

                for movg_num, movg_conns in enumerate(mov_groups_conns):
                    # Set Static Time of Movement Group for Priority Groups
                    movgID = str(jctID) + "_MG" + str(movg_num)
                    if mov_groups_pattern[movg_num] == []:
                        # 6 for protected green
                        movg_objs[tc_ind].setUnsignalizedJctMovGroupPlans(jct_objs[tc_ind], jctID, movgID,6)
                    else:
                        # 5 for permissive green
                        movg_objs[tc_ind].setUnsignalizedJctMovGroupPlans(jct_objs[tc_ind], jctID, movgID,5)

                    for connID in movg_conns:
                        # Add Yield to Movement Group of Unsignilised Control
                        edg_objs[tc_ind].addConnYield2Conns(movg_objs[tc_ind], connID, mov_groups_pattern[movg_num])

                        # Set if vehicle is obligatory to stop
                        if jct_type == "allway_stop" \
                        or (jct_type == "priority_stop" and mov_groups_pattern[movg_num] != []):
                            in_laneID = edg_objs[tc_ind].connID_to_inLaneID[connID]
                            edg_objs[tc_ind].lane_must_stop[in_laneID] = 1

    # Define the SUMO XML detector children used for the area with local level routing (if options.use_detectors_data)
    # NOTICE THAT END AND BEGIN REPRESENT THE POSITION OF THE DETECTOR ON THE LANE ON THE FLOW
    # Read Detectors File
    if INF_DETECTORS_FILE != None:
        # Get detectors data from files
        tree_detectors_end = ElementTree.parse(os.path.abspath(os.path.join(END_DETECTORS_FILE)))
        tree_detectors_begin = ElementTree.parse(os.path.abspath(os.path.join(BEGIN_DETECTORS_FILE)))
        tree_detectors_info = ElementTree.parse(os.path.abspath(os.path.join(INF_DETECTORS_FILE)))
        # Get all lanes
        all_lanes = set()
        for edg_obj in edg_objs:
            for laneID in edg_obj.lane_ids:
                all_lanes.add(laneID)
        # Get raw data
        detectors_children = list(tree_detectors_info.iter('instantInductionLoop'))
        detectors_children = [detector for detector in detectors_children if detector.get('lane') in all_lanes]
        all_detectors = [detector.get("id") for detector in detectors_children]
        # Define intervals for aggregation of the data
        intervals_edge_attrs = [BEGIN_T + STEP_LENGTH + (interv * ((END_T - BEGIN_T) / NUM_INTERVS))
                                for interv in range(0, NUM_INTERVS)]
        detections_end_children = []
        detections_begin_children = []
        for interv, beg_interv in enumerate(intervals_edge_attrs):
            # Get each vehicle detection per desired interval of edge attributes
            det_children = list(tree_detectors_end.iter('instantOut'))
            detections_end_children.append([detection for detection in det_children
                                            if detection.get('id') in all_detectors
                                            and float(detection.get('time')) >= beg_interv
                                            and float(detection.get('time')) < \
                                            beg_interv + (END_T - BEGIN_T) / NUM_INTERVS
                                            and detection.get("state") == "leave"])
            det_children = list(tree_detectors_begin.iter('instantOut'))
            detections_begin_children.append([detection for detection in det_children
                                              if detection.get('id') in all_detectors
                                              and float(detection.get('time')) >= beg_interv
                                              and float(detection.get('time')) < \
                                              beg_interv + (END_T - BEGIN_T) / NUM_INTERVS
                                              and detection.get("state") == "leave"])
    else:
        # Read turn definitions and flows when not using detectors data
        tree_turndefs = ElementTree.parse(os.path.abspath(os.path.join(TURN_DEFS_FILE)))
        tree_flows = ElementTree.parse(os.path.abspath(os.path.join(FLOWS_FILE)))

        turndefs_children = list(tree_turndefs.iter('interval'))
        turndefs_children = [interv_child for interv_child in turndefs_children
                             if float(interv_child.get('end')) > BEGIN_T
                             and float(interv_child.get('begin')) < END_T]

        flows_children = list(tree_flows.iter('interval'))
        flows_children = [interv_child for interv_child in flows_children
                          if float(interv_child.get('end')) > BEGIN_T
                          and float(interv_child.get('begin')) < END_T]

    for tc_ind in range(0, len(jct_objs)):
        for edgeID in edg_objs[tc_ind].edge_ids:
            parallel_laneIDs = edg_objs[tc_ind].edgeID_to_laneIDs[edgeID]
            num_lanes = len(edg_objs[tc_ind].edgeID_to_laneIDs[edgeID])
            # Initialize vehicles starting within the edge and probability of vehicle to finish at the edge
            edg_objs[tc_ind].last_edg_prob.update({edgeID: dict()})
            edg_objs[tc_ind].last_edg_prob[edgeID].update({"interv": None})
            if edg_objs[tc_ind].inLaneID_to_jctID.has_key(parallel_laneIDs[0]):
                # If lanes are incoming lanes
                for laneID in parallel_laneIDs:
                    edg_objs[tc_ind].lane_num_beg_vehs.update({laneID: dict()})
                    edg_objs[tc_ind].lane_num_beg_vehs[laneID].update({"interv": None})

            if INF_DETECTORS_FILE != None:
                lanes_detectorIDs = [detector.get("id") for detector in detectors_children if
                                     detector.get("lane") in parallel_laneIDs]
                lanes_end_detectorIDs = [det_id for det_id in lanes_detectorIDs if
                                         det_id.split("_")[1][0:3] == "end"]
                lanes_begin_detectorIDs = [det_id for det_id in lanes_detectorIDs if
                                           det_id.split("_")[1][0:5] == "begin"]
                for interv,beg_interv in enumerate(intervals_edge_attrs):
                    # Get vehicles at the end and begin of the lane detectors within each interval
                    vehs_end_detector = set([detection.get("vehID") for detection in detections_end_children[interv]
                                             if detection.get("id") in lanes_end_detectorIDs])
                    vehs_begin_detector = set([detection.get("vehID") for detection in detections_begin_children[interv]
                                               if detection.get("id") in lanes_begin_detectorIDs])

                    # Vehicles that finished their trip on the edge are those that passed begin detector but not
                    # the end one
                    finished_vehicles = float(len(vehs_begin_detector.difference(vehs_end_detector)))
                    if len(vehs_begin_detector) != 0:
                        edg_objs[tc_ind].last_edg_prob[edgeID].update({beg_interv:
                                                                       finished_vehicles / len(vehs_begin_detector)})
                    else:
                        edg_objs[tc_ind].last_edg_prob[edgeID].update({beg_interv: 0})

                    if edg_objs[tc_ind].inLaneID_to_jctID.has_key(parallel_laneIDs[0]):
                        # Vehicles that start their trip on the edge are those that passed end detector but
                        # not the begin one
                        new_vehicles_lane = float(len(vehs_end_detector.difference(vehs_begin_detector))) / num_lanes
                        flow = new_vehicles_lane / ((END_T - BEGIN_T) / NUM_INTERVS)  # veh/s
                        num_veh_arr_range = flow * LEN_RANGE  # veh per arrival range
                        for laneID in parallel_laneIDs:
                            edg_objs[tc_ind].lane_num_beg_vehs[laneID].update({beg_interv:
                                                                              (num_veh_arr_range, ["random"], [1])})
            else:
                # When using only flows file and not detectors data, it is not known the number of vehs. starting
                # and finishing on lane, but edges going to dead end junctions are sink edges (end of network)
                tojctID = [edge.get('to') for edge in net_edg_children if edge.get('id') == edgeID][0]
                if net_jct_children[tojctID].get("type") == "dead_end":
                    edg_objs[tc_ind].last_edg_prob[edgeID].update({BEGIN_T + STEP_LENGTH: 1})
                else:
                    edg_objs[tc_ind].last_edg_prob[edgeID].update({BEGIN_T + STEP_LENGTH: 0})
                if edg_objs[tc_ind].inLaneID_to_jctID.has_key(parallel_laneIDs[0]):
                    for laneID in parallel_laneIDs:
                        edg_objs[tc_ind].lane_num_beg_vehs[laneID].update({BEGIN_T + STEP_LENGTH: (0, ["random"], [1])})

            # Estimate the next edges probability and their turning radius
            parallel_laneIDs = edg_objs[tc_ind].edgeID_to_laneIDs[edgeID]
            edg_objs[tc_ind].next_edg_prob.update({edgeID: dict()})
            edg_objs[tc_ind].next_edg_radius.update({edgeID: dict()})
            edg_objs[tc_ind].next_edg_prob[edgeID].update({"interv": None})
            try:
                jctID = edg_objs[tc_ind].inLaneID_to_jctID[parallel_laneIDs[0]]
                # If lane is modelled by traffic controller tc_ind
                calc_turn_probs = 1
                calc_turn_radius = 1
            except KeyError:
                # If lane is not modelled by traffic controller tc_ind, but goes to a lane in common with a neighbouring
                # junction, then get the turning probabilities
                jctID = edg_objs[tc_ind].outLaneID_to_jctID[parallel_laneIDs[0]]
                calc_turn_probs = 0
                calc_turn_radius = 0
                common_lanes_neigh = []
                common_edgs_neigh = set()
                for neighID in jct_objs[tc_ind].neigh_jcts[jctID]:
                    try:
                        common_lanes_neigh.extend(jct_objs[tc_ind].commonToNeigh_laneIDs[jctID][neighID])
                    except KeyError:
                        skip = 1
                for laneID in common_lanes_neigh:
                    common_edgs_neigh.add(edg_objs[tc_ind].laneID_to_edgeID[laneID])
                all_common_edges = list(common_edgs_neigh)
                if edgeID in all_common_edges:
                    calc_turn_probs = 1

            if calc_turn_probs == 1:
                possible_out_edges = set([])
                possible_conns = []
                for laneID in parallel_laneIDs:
                    possible_conns.extend(edg_objs[tc_ind].inLaneID_to_connIDs[laneID])
                for conn in possible_conns:
                    possible_out_edges.add(edg_objs[tc_ind].laneID_to_edgeID[edg_objs[tc_ind].connID_to_outLaneID[conn]])

                if INF_DETECTORS_FILE != None:
                    # If using detectors data, get number of vehicles going from one edge to another
                    lanes_detectorIDs = [detector.get("id") for detector in detectors_children if
                                         detector.get("lane") in parallel_laneIDs]
                    lanes_end_detectorIDs = [det_id for det_id in lanes_detectorIDs
                                             if det_id.split("_")[1][0:3] == "end"]

                    for interv, beg_interv in enumerate(intervals_edge_attrs):
                        # (Begin and End mean position of the detector on the lane through the direction of flow)
                        vehs_end_detector = {detection.get("vehID"): detection.get("time")
                                             for detection in detections_end_children[interv]
                                             if detection.get("id") in lanes_end_detectorIDs}
                        count_vehs2out_edge = dict()

                        edg_objs[tc_ind].next_edg_prob[edgeID].update({beg_interv: dict()})
                        for out_edgeID in possible_out_edges:
                            # Check if there is a connection from in_edgeID to out_edgeID
                            parallel_out_laneIDs = edg_objs[tc_ind].edgeID_to_laneIDs[out_edgeID]
                            outLanes_detectorIDs = [detector.get("id") for detector in detectors_children if
                                                    detector.get("lane") in parallel_out_laneIDs]
                            outLanes_begin_detectorIDs = [det_id for det_id in outLanes_detectorIDs if
                                                          det_id.split("_")[1][0:5] == "begin"]
                            vehs_begin_detector = [detection.get("vehID")
                                                   for detection in detections_begin_children[interv]
                                                   if detection.get("id") in outLanes_begin_detectorIDs
                                                   and vehs_end_detector.has_key(detection.get("vehID"))
                                                   and float(detection.get("time")) <= float(
                                                   vehs_end_detector[detection.get("vehID")]) + 30]
                            # after 30 seconds assumes the vehicle is coming from another edge
                            count_vehs2out_edge.update({out_edgeID: float(len(vehs_begin_detector))})

                        sum_all_vehs = sum(count_vehs2out_edge.values())
                        if sum_all_vehs == 0:
                            # If not detected any vehicle, assumes equal turning rates according to the number of
                            # possible outgoing edges from the incoming edge
                            num_out_edges = len(count_vehs2out_edge.keys())
                            for out_edgeID in count_vehs2out_edge.keys():
                                edg_objs[tc_ind].next_edg_prob[edgeID][beg_interv].update({out_edgeID:
                                                                                           1.0 / num_out_edges})
                        else:
                            for out_edgeID in count_vehs2out_edge.keys():
                                edg_objs[tc_ind].next_edg_prob[edgeID][beg_interv].update({out_edgeID:
                                                                                           count_vehs2out_edge[out_edgeID]
                                                                                           / sum_all_vehs})
                else:
                    for interv_child in turndefs_children:
                        # The turning rates are already defined, just collect them
                        beg_interv = round(max(float(interv_child.get('begin')) + STEP_LENGTH, BEGIN_T + STEP_LENGTH),
                                           1)
                        try:
                            ing_egde_children = list([ing_egde_child
                                                      for ing_egde_child in interv_child.findall("fromEdge")
                                                      if ing_egde_child.get("id") == edgeID][0].findall("toEdge"))
                        except IndexError:
                            noToEdges = 1

                        edg_objs[tc_ind].next_edg_prob[edgeID].update({beg_interv: dict()})
                        for out_edgeID in possible_out_edges:
                            try:
                                prob_egrEdge = float([out_edge_child.get("probability")
                                                      for out_edge_child in ing_egde_children
                                                      if out_edge_child.get("id") == out_edgeID][0]) / 100
                            except IndexError:
                                # No probability on turn definition file
                                prob_egrEdge = 0

                            edg_objs[tc_ind].next_edg_prob[edgeID][beg_interv].update({out_edgeID: prob_egrEdge})

                if calc_turn_radius == 1:
                    # Define Turning Radius
                    for out_edgeID in possible_out_edges:
                        conns_radius_curve = []
                        for in_laneID in edg_objs[tc_ind].edgeID_to_laneIDs[edgeID]:
                            for possible_conn in edg_objs[tc_ind].inLaneID_to_connIDs[in_laneID]:
                                if edg_objs[tc_ind].connID_to_outLaneID[possible_conn] \
                                in edg_objs[tc_ind].edgeID_to_laneIDs[out_edgeID]:
                                    in_laneID = edg_objs[tc_ind].connID_to_inLaneID[possible_conn]
                                    out_laneID = edg_objs[tc_ind].connID_to_outLaneID[possible_conn]
                                    inOut_edgs_children = {edge.get('id'): edge for edge in net_edg_children if
                                                           edge.get('id') == edgeID or edge.get('id') == out_edgeID}
                                    # Get coordinates of the last segment of the incoming edge
                                    inLane_child = [lane_child
                                                    for lane_child in inOut_edgs_children[edgeID].findall("lane")
                                                    if lane_child.get("id") == in_laneID][0]
                                    inEdge_beg_x, inEdge_beg_y = inLane_child.get("shape").split(" ")[-2].split(",")
                                    inEdge_end_x, inEdge_end_y = inLane_child.get("shape").split(" ")[-1].split(",")

                                    inEdge_beg_x = float(inEdge_beg_x)
                                    inEdge_beg_y = float(inEdge_beg_y)
                                    inEdge_end_x = float(inEdge_end_x)
                                    inEdge_end_y = float(inEdge_end_y)

                                    # Get coordinates of the first segment of the outgoing edge
                                    outLane_child = [lane_child
                                                     for lane_child in inOut_edgs_children[out_edgeID].findall("lane")
                                                     if lane_child.get("id") == out_laneID][0]
                                    outEdge_beg_x, outEdge_beg_y = outLane_child.get("shape").split(" ")[0].split(",")

                                    outEdge_beg_x = float(outEdge_beg_x)
                                    outEdge_beg_y = float(outEdge_beg_y)

                                    # Getting the line of the vehicle trajectory approaching the junction
                                    A_inLine = inEdge_beg_y - inEdge_end_y
                                    B_inLine = inEdge_end_x - inEdge_beg_x
                                    C_inLine = inEdge_beg_x * inEdge_end_y - inEdge_end_x * inEdge_beg_y

                                    # Getting the perpendicular line, Ax + By + C = 0, to the line vehicle trajectory
                                    # approaching
                                    # the junction
                                    A_perp_inLine = -1 / A_inLine
                                    perp_inLine_y = inEdge_end_y - A_perp_inLine
                                    perp_inLine_x = inEdge_end_x + 1
                                    B_perp_inLine = perp_inLine_x - inEdge_end_x
                                    C_perp_inLine = inEdge_end_x * perp_inLine_y - perp_inLine_x * inEdge_end_y

                                    # Check if outEdge begin coordinates are within the line of vehicle approaching the
                                    # junction
                                    # Done by checking the distance from point to the line
                                    # Using standard linear equation Ax + By + C = 0
                                    adj_side = abs(A_inLine * outEdge_beg_x + B_inLine * outEdge_beg_y + C_inLine) / \
                                               float(math.sqrt(A_inLine ** 2 + B_inLine ** 2))

                                    opp_side = abs(A_perp_inLine * outEdge_beg_x + B_perp_inLine * outEdge_beg_y
                                                   + C_perp_inLine) / float(math.sqrt(A_perp_inLine ** 2
                                                                                      + B_perp_inLine ** 2))

                                    global radius_curve

                                    if (adj_side <= 2 and opp_side >= 3):
                                        # If no curve, the radius is infinity
                                        conns_radius_curve.append(float("inf"))
                                    else:
                                        # calculate the distance from outEdge begin coordinates to the perpendicular
                                        # (perp.) line
                                        # which is the opposite side for finding the central angle
                                        dist_beg_end = math.sqrt((outEdge_beg_y - inEdge_end_y) ** 2
                                                                 + (outEdge_beg_x - inEdge_end_x) ** 2)

                                        central_angle = 2 * math.acos(opp_side / dist_beg_end)
                                        conns_radius_curve.append(edg_objs[tc_ind].conn_length[possible_conn]
                                                                  / central_angle)
                                        # radius_curve_alt_2 = opp_side / math.sin(central_angle)
                                        # radius_curve_alt_3 = dist_beg_end / (2 * math.sin(central_angle / 2))

                        edg_objs[tc_ind].next_edg_radius[edgeID].update({out_edgeID: max(conns_radius_curve)})

        # Add Fixed Inflow Profile
        for jctID in jct_objs[tc_ind].ids:
            # Get edges from junctions not modelled by traffic controller tc_ind
            edges_from_unknown_jcts = [edge.get('id') for edge in net_edg_children
                                       if all_junctions_tc.has_key(edge.get('from')) == False
                                       and edge.get('to') == jctID]
            common_lanes_neigh = []
            common_edgs_neigh = set()
            # Get all lanes in common with neighbouring junctions
            for neighID in jct_objs[tc_ind].neigh_jcts[jctID]:
                try:
                    common_lanes_neigh.extend(jct_objs[tc_ind].commonFromNeigh_laneIDs[jctID][neighID])
                except KeyError:
                    skip = 1
            for laneID in common_lanes_neigh:
                common_edgs_neigh.add(edg_objs[tc_ind].laneID_to_edgeID[laneID])
            all_edges_inflow = edges_from_unknown_jcts + list(common_edgs_neigh)
            if all_edges_inflow != []:
                for edgeID in all_edges_inflow:
                    num_lanes = len(edg_objs[tc_ind].edgeID_to_laneIDs[edgeID])
                    for laneID in edg_objs[tc_ind].edgeID_to_laneIDs[edgeID]:
                        edg_objs[tc_ind].lane_inflow_profile.update({laneID: [dict(), ()]})
                        edg_objs[tc_ind].lane_inflow_profile[laneID][0].update({"interv": None})

                    if INF_DETECTORS_FILE != None:
                        for interv, beg_interv in enumerate(intervals_edge_attrs):
                            for laneID in edg_objs[tc_ind].edgeID_to_laneIDs[edgeID]:
                                # Get number of vehicles arriving on the lane begin detector within the interval
                                lane_detectorIDs = [detector.get("id") for detector in detectors_children
                                                    if detector.get("lane") == laneID]
                                # But take just the begin lane detector children
                                lane_detections_begin_children = [detection
                                                                  for detection in detections_begin_children[interv]
                                                                  if detection.get("id") in lane_detectorIDs]
                                if lane_detections_begin_children == []:
                                    # If no detected vehicle
                                    num_veh_arr_range = 0
                                    flow = 0
                                    mean_speed = 0
                                    mean_hdwy = 0
                                    std_dev_hdwy = 0
                                else:
                                    num_veh = float(len(lane_detections_begin_children)) # veh in one simulation time
                                    flow = num_veh / ((END_T - BEGIN_T) / NUM_INTERVS)  # veh/s
                                    num_veh_arr_range = flow * LEN_RANGE # veh per arrival range
                                    speed_all = [float(lane_detection.get("speed"))
                                                 for lane_detection in lane_detections_begin_children]
                                    mean_speed = np.mean(speed_all)
                                    arr_t_all = [float(lane_detection.get("time"))
                                                 for lane_detection in lane_detections_begin_children]
                                    hdwy_all = [0] # first vehicle has headway equal to zero
                                    for arr_ind, arr_t in enumerate(arr_t_all[1:]):
                                        hdwy_all.append(arr_t - arr_t_all[arr_ind])
                                    mean_hdwy =  np.mean(hdwy_all)
                                    std_dev_hdwy = np.std(hdwy_all)

                                if flow >= LOW_FLOW and flow < MID_FLOW:
                                    # Medium Flow of Vehicles, create the flow profile using Pearson Type III Distribution
                                    edg_objs[tc_ind].lane_inflow_profile[laneID][0].update({beg_interv:
                                                                                            (round(num_veh_arr_range,2),
                                                                                            round(mean_speed,2),
                                                                                            round(mean_hdwy,2),
                                                                                            round(std_dev_hdwy,2))})
                                else:
                                    # Low Flow of Vehicles, create the flow profile using Negative Exponential Distribution
                                    # or
                                    # High Flow of Vehicles, create the flow profile using Normal Distribution
                                    edg_objs[tc_ind].lane_inflow_profile[laneID][0].update({beg_interv:
                                                                                            (round(num_veh_arr_range,2),
                                                                                            round(mean_speed,2),
                                                                                            round(mean_hdwy,2))})
                    else:
                        # Use intervals_edge_attrs when want to have edge attributes for intervals of simulation time
                        # for interv, beg_interv in enumerate(intervals_edge_attrs):
                        for interv_child in flows_children:
                            vehsPerHour = []
                            beg_interv = round(max(float(interv_child.get('begin'))
                                                   + STEP_LENGTH, BEGIN_T + STEP_LENGTH), 1)
                            end_interv = round(min(float(interv_child.get('end'))
                                                   + STEP_LENGTH, END_T + STEP_LENGTH), 1)
                            interv_dur = end_interv - beg_interv
                            real_dur = float(interv_child.get('end')) - float(interv_child.get('begin'))
                            interv_flows_children = list(interv_child.findall("flow"))
                            for interv_flows_child in interv_flows_children:
                                if interv_flows_child.get("from") == edgeID:
                                    try:
                                        vehsPerHour.append(float(interv_flows_child.get("vehsPerHour")))
                                    except AttributeError:
                                        try:
                                            vehsPerHour.append(float(interv_flows_child.get("number")) \
                                                                     * (interv_dur / real_dur) / (interv_dur / 3600.0))
                                        except AttributeError:
                                            vehsPerHour.append((interv_dur / float(interv_flows_child.get("period")))
                                                               * 3600)

                            for laneID in edg_objs[tc_ind].edgeID_to_laneIDs[edgeID]:
                                if vehsPerHour == []:
                                    # If no vehicles
                                    num_veh_arr_range = 0
                                    flow = 0
                                    mean_speed = 0
                                    mean_hdwy = 0
                                    std_dev_hdwy = 0
                                else:
                                    lane_vehsPerHour = sum(vehsPerHour) / num_lanes
                                    flow = lane_vehsPerHour / 3600.0 # in veh/s
                                    num_veh_arr_range = flow * LEN_RANGE
                                    mean_speed = edg_objs[tc_ind].lane_max_speed[laneID]
                                    mean_hdwy = 1 / flow
                                    std_dev_hdwy = 0

                                if flow >= LOW_FLOW and flow < MID_FLOW:
                                    # Medium Flow of Vehicles, create the flow profile using Pearson Type III Distribution
                                    edg_objs[tc_ind].lane_inflow_profile[laneID][0].update({beg_interv:
                                                                                            (round(num_veh_arr_range,2),
                                                                                            round(mean_speed,2),
                                                                                            round(mean_hdwy,2),
                                                                                            round(std_dev_hdwy,2))})
                                else:
                                    # Low Flow of Vehicles, create the flow profile using Negative Exponential Distribution
                                    # or
                                    # High Flow of Vehicles, create the flow profile using Normal Distribution
                                    edg_objs[tc_ind].lane_inflow_profile[laneID][0].update({beg_interv:
                                                                                            (round(num_veh_arr_range,2),
                                                                                            round(mean_speed,2),
                                                                                            round(mean_hdwy,2))})

    for tc_ind in range(0, len(jct_objs)):
        for jctID in jct_objs[tc_ind].ids:
            # Update the critical gap times and follow up times according to the connection context
            jct_edg_priorities = {edge.get('id'): float(edge.get('priority')) for edge in net_edg_children
                                  if edge.get('to') == jctID}
            jct_edg_from_jct = {edge.get('id'): edge.get('from') for edge in net_edg_children
                                  if edge.get('to') == jctID}
            # Group artificial parallel edges based on their common junction they start from,
            # and define the same stream priority with the edge with highest priority of the stream
            # and also the total number of lanes
            lanes_to_yield = set([])
            for connID in jct_objs[tc_ind].connIDs[jctID]:
                for yield_conn in edg_objs[tc_ind].connID_to_yield2connIDs[connID]:
                    lanes_to_yield.add(edg_objs[tc_ind].connID_to_inLaneID[yield_conn])
            edgIDs_from_same_jct = dict()
            for edgeID in jct_edg_priorities:
                try:
                    edgIDs_from_same_jct[jct_edg_from_jct[edgeID]][0].append(edgeID)
                    edgIDs_from_same_jct[jct_edg_from_jct[edgeID]][1] = max(
                        edgIDs_from_same_jct[jct_edg_from_jct[edgeID]][1], jct_edg_priorities[edgeID])
                    priority_lanes = [lane for lane in edg_objs[tc_ind].edgeID_to_laneIDs[edgeID]
                                      if lane in lanes_to_yield]
                    edgIDs_from_same_jct[jct_edg_from_jct[edgeID]][2] += len(priority_lanes)
                except KeyError:
                    priority_lanes = [lane for lane in edg_objs[tc_ind].edgeID_to_laneIDs[edgeID]
                                      if lane in lanes_to_yield]
                    edgIDs_from_same_jct.update({jct_edg_from_jct[edgeID]:
                                                     [[edgeID], jct_edg_priorities[edgeID],
                                                      len(priority_lanes)]})

            # Define the context for each connection and adjust critical and follow up times
            num_streams = len(edgIDs_from_same_jct.keys())
            if num_streams > 2:
                sum_num_priority_lanes = sum([num_lanes for edges, priority, num_lanes
                                              in sorted(edgIDs_from_same_jct.values(),
                                                        key=lambda sort_it: (sort_it[1]))[::-1][:2]])
            else:
                sum_num_priority_lanes = sorted(edgIDs_from_same_jct.values(),
                                                key=lambda sort_it: (sort_it[1]))[::-1][0][2]

            for connID in jct_objs[tc_ind].connIDs[jctID]:
                if sum_num_priority_lanes <= 2:
                    context = "2-lanes/"
                else:
                    context = "4-lanes/"
                in_laneID = edg_objs[tc_ind].connID_to_inLaneID[connID]
                out_laneID = edg_objs[tc_ind].connID_to_outLaneID[connID]
                in_edgID = edg_objs[tc_ind].laneID_to_edgeID[in_laneID]
                out_edgID = edg_objs[tc_ind].laneID_to_edgeID[out_laneID]
                yield_edgIDs = set([])
                # Define all edges that this connection must yield to
                for yield_conn in edg_objs[tc_ind].connID_to_yield2connIDs[connID]:
                    yield_edgID = edg_objs[tc_ind].laneID_to_edgeID[edg_objs[tc_ind].connID_to_inLaneID[yield_conn]]
                    if yield_edgID != in_edgID:
                        yield_edgIDs.add(yield_edgID)

                if len(yield_edgIDs) > 0:
                    if edg_objs[tc_ind].conn_direction[connID] == "s":
                        context += "through/"
                    elif edg_objs[tc_ind].conn_direction[connID] == "l":
                        context += "left/"
                    elif edg_objs[tc_ind].conn_direction[connID] == "r":
                        context += "right/"
                    in_edgID_priority = jct_edg_priorities[in_edgID]

                    num_yield_streams = 0
                    for from_jctID in edgIDs_from_same_jct:
                        stream_info = edgIDs_from_same_jct[from_jctID]
                        stream_edges = stream_info[0]
                        if 1 in [1 for stream_edge in stream_edges if stream_edge in yield_edgIDs]:
                            yield_stream_priority = stream_info[1]
                            if yield_stream_priority >= in_edgID_priority:
                                num_yield_streams += 1
                    else:
                        if jct_objs[tc_ind].tlID[jctID] != None and "left" in context:
                            # If traffic light controlled, consider all streams as major if turning left
                            context += "major"
                        else:
                            if num_yield_streams == 0 or (num_yield_streams == 1 and "left" in context):
                                context += "major"
                            else:
                                context += "minor"

                    if num_yield_streams > 0:
                        geometry_adj = 0
                        if num_streams < 4 and "left" in context:
                            # If 3-leg intersection and turning left, reduce critical gap
                            geometry_adj += 0.7
                        if edg_objs[tc_ind].next_edg_radius[in_edgID][out_edgID] > 15.25:
                            geometry_adj += 0.5

                        adjusted_cr_gap = edg_objs[tc_ind].base_ave_cr_gap[context]
                        adjusted_follow_up = edg_objs[tc_ind].base_follow_up[context[8:]]
                        standard_cr_gap = edg_objs[tc_ind].std_cr_gap
                        standard_follow_up = edg_objs[tc_ind].std_follow_up
                        adjusted_cr_gap = max(adjusted_cr_gap - geometry_adj, 0)
                        standard_cr_gap = max(adjusted_cr_gap - geometry_adj, 0)
                        for veh_ind, vehType in enumerate(veh_objs[tc_ind].vType_vals["ids"]):
                            if "HDV" in vehType:
                                if "2-lanes" in context:
                                    edg_objs[tc_ind].conn_ave_cr_gap[connID]["adj"][veh_ind] = adjusted_cr_gap + 1
                                    edg_objs[tc_ind].conn_follow_up[connID]["adj"][veh_ind] = adjusted_follow_up + 0.9
                                    edg_objs[tc_ind].conn_ave_cr_gap[connID]["std"][veh_ind] = standard_cr_gap + 1
                                    edg_objs[tc_ind].conn_follow_up[connID]["std"][veh_ind] = standard_follow_up + 0.9
                                else:
                                    edg_objs[tc_ind].conn_ave_cr_gap[connID]["adj"][veh_ind] = adjusted_cr_gap + 2
                                    edg_objs[tc_ind].conn_follow_up[connID]["adj"][veh_ind] = adjusted_follow_up + 1
                                    edg_objs[tc_ind].conn_ave_cr_gap[connID]["std"][veh_ind] = standard_cr_gap + 2
                                    edg_objs[tc_ind].conn_follow_up[connID]["std"][veh_ind] = standard_follow_up + 1
                            else:
                                edg_objs[tc_ind].conn_ave_cr_gap[connID]["adj"][veh_ind] = adjusted_cr_gap
                                edg_objs[tc_ind].conn_follow_up[connID]["adj"][veh_ind] = adjusted_follow_up
                                edg_objs[tc_ind].conn_ave_cr_gap[connID]["std"][veh_ind] = standard_cr_gap
                                edg_objs[tc_ind].conn_follow_up[connID]["std"][veh_ind] = standard_follow_up
                    else:
                        for veh_ind, vehType in enumerate(veh_objs[tc_ind].vType_vals["ids"]):
                            # Assuming 2/3 of average needed gap when yielding to the same stream
                            edg_objs[tc_ind].conn_ave_cr_gap[connID]["adj"][veh_ind] = edg_objs[tc_ind].std_cr_gap * 2 / 3
                            edg_objs[tc_ind].conn_follow_up[connID]["adj"][veh_ind] = edg_objs[tc_ind].std_follow_up * 2 / 3
                            edg_objs[tc_ind].conn_ave_cr_gap[connID]["std"][veh_ind] = edg_objs[tc_ind].std_cr_gap * 2 / 3
                            edg_objs[tc_ind].conn_follow_up[connID]["std"][veh_ind] = edg_objs[tc_ind].std_follow_up * 2 / 3
                else:
                    for veh_ind, vehType in enumerate(veh_objs[tc_ind].vType_vals["ids"]):
                        # Assuming 2/3 of average needed gap when yielding to the same edge
                        edg_objs[tc_ind].conn_ave_cr_gap[connID]["adj"][veh_ind] = edg_objs[tc_ind].std_cr_gap * 2 / 3
                        edg_objs[tc_ind].conn_follow_up[connID]["adj"][veh_ind] = edg_objs[tc_ind].std_follow_up * 2 / 3
                        edg_objs[tc_ind].conn_ave_cr_gap[connID]["std"][veh_ind] = edg_objs[tc_ind].std_cr_gap * 2 / 3
                        edg_objs[tc_ind].conn_follow_up[connID]["std"][veh_ind] = edg_objs[tc_ind].std_follow_up * 2 / 3

    # Create a new database and save the inputs defined by parsing SUMO configuration files
    msgs_obj_names = []
    msgs_obj_values = []
    for msgs_obj in msgs_objs:
        msgs_obj_names.append(msgs_obj.name)
        msgs_obj_values.append(msgs_obj)
    msgs_obj_data = zip(msgs_obj_names, msgs_obj_values)

    jct_obj_names = []
    jct_obj_values = []
    for jct_obj in jct_objs:
        jct_obj_names.append(jct_obj.name)
        jct_obj_values.append(jct_obj)
    jct_obj_data = zip(jct_obj_names, jct_obj_values)

    movg_obj_names = []
    movg_obj_values = []
    for movg_obj in movg_objs:
        movg_obj_names.append(movg_obj.name)
        movg_obj_values.append(movg_obj)
    movg_obj_data = zip(movg_obj_names, movg_obj_values)

    edg_obj_names = []
    edg_obj_values = []
    for edg_obj in edg_objs:
        edg_obj_names.append(edg_obj.name)
        edg_obj_values.append(edg_obj)
    edg_obj_data = zip(edg_obj_names, edg_obj_values)

    veh_obj_names = []
    veh_obj_values = []
    for veh_obj in veh_objs:
        veh_obj_names.append(veh_obj.name)
        veh_obj_values.append(veh_obj)
    veh_obj_data = zip(veh_obj_names,veh_obj_values)


    with open(DB_FILE, "wb") as output:
        pickle.dump(msgs_obj_data, output)
        pickle.dump(jct_obj_data, output)
        pickle.dump(movg_obj_data, output)
        pickle.dump(edg_obj_data, output)
        pickle.dump(veh_obj_data, output)

    LOGGER.info("Database created, exitting. Restart this script.")
    logger.info('Created databased, restart this script.')
    os._exit(1)


## BEGIN SIMULATION

# Subscribe to get data from all vehicles within modelled edges/lanes
allConn_modelled_area = set(getExcelTargetList(ALL_CONN_AREA_FILE))
allInEdges_modelled_area = set([])
for jct_obj in jct_objs:
    for jctID in jct_obj.ids:
        if INTERFACE == "traci" and "Coop." in jct_obj.jctType[jctID]:
            # Subscribe to get all vehicles inside the communication range of each RSU
            traci.junction.subscribeContext(jctID, DOMAIN, jct_obj.comm_range[jctID])
        for edgeID in jct_obj.in_edgeIDs[jctID]:
            # allInEdges_modelled_area used for defining the edges to have real travel times being measured
            allInEdges_modelled_area.add(edgeID)

# Initializing for adding error of queue length estimation
vehs_last_range = dict()

# Initializing for counting numer of vehicles
last_step_edge_vehs = dict()

# Initialize dictionary with time to reroute for each vehicle equipped with rerouteing device (CAV)
vehs_time2route = dict()

# Get lanes and length of the artificial edges and also the edges sending vehicles to them (called prior edges),
# in addition, define vehs_last_range for error of queue length estimation and last_step_edge_vehs for counting
priorAndArtificial_edgesLanes = dict()
edge_endLane_vehs = dict()
edge_defined_vehs = dict()
for priorEdgeID in PRIOR_ARTIFICIAL_EDGES.keys():
    edge_part = priorEdgeID + "_"
    priorAndArtificial_edgesLanes.update({priorEdgeID: [edge_part + str(lane_num)
                                                       for lane_num in range(0, traci.edge.getLaneNumber(priorEdgeID))]})
    # Init. list of vehicles to be set as their next edges (artificial one) already defined
    edge_endLane_vehs.update({priorEdgeID: set([])})
    edge_defined_vehs.update({priorEdgeID: set([])})
    for artEdgeID in PRIOR_ARTIFICIAL_EDGES[priorEdgeID]:
        edge_part = artEdgeID + "_"
        priorAndArtificial_edgesLanes.update({artEdgeID: [edge_part + str(lane_num)
                                                       for lane_num in range(0, traci.edge.getLaneNumber(artEdgeID))]})
all_InOut_edges = dict()
prior_edges_len = dict()
for tc_ind,edg_obj in enumerate(edg_objs):
    all_in_lanes = set([])
    all_InOut_lanes = set([])
    for edgeID in edg_obj.edge_ids:
        # Define all incoming and outgoing modelled edges
        all_InOut_edges.update({edgeID: tc_ind})
    all_in_lanes.update(edg_obj.lane_inflow_profile.keys())
    all_InOut_lanes.update(edg_obj.lane_ids)
    for laneID in all_in_lanes:
        edgeID = edg_obj.laneID_to_edgeID[laneID]
        vehs_last_range.update({edgeID: set([])})
        last_step_edge_vehs.update({edgeID: set([])})
        if priorAndArtificial_edgesLanes.has_key(edgeID):
            prior_edges_len.update({edgeID: edg_obj.lane_length[laneID]})
        # Subscribe to get data from modelled lanes
        traci.lane.subscribe(laneID, LANE_METRICS + [LANE_OCCUPANCY, VEHIDS])
    only_out_lanes = all_InOut_lanes.difference(all_in_lanes)
    for laneID in only_out_lanes:
        # Subscribe to get data from modelled lanes
        traci.lane.subscribe(laneID, [VEHIDS])

# Subscribe to get data from modelled edges
for edgeID in allConn_modelled_area:
     traci.edge.subscribe(edgeID, [VEHIDS])

# Initialize Dictionaries
for ind_trafficNet in range(0, len(msgs_objs)):
    msgs_objs[ind_trafficNet].setMsgsDicts()
    for tc_ind in range(0, len(jct_objs)):
        edg_objs[tc_ind].setEdgesDicts(jct_objs[tc_ind], movg_objs[tc_ind])
        movg_objs[tc_ind].setMovGroupsDicts(jct_objs[tc_ind], edg_objs[tc_ind])

# Initialize Simulation Values
ACTUAL_T = round(float(traci.simulation.getTime()), 1)

# Initialize Algorithm Values
time2updte = 0.0
jct_order = 1

# Initializing for counting number of vehicles
lasttStep_area_error_updDeps = []
LLR_nexttStep_area_updDeps = []
SUMO_lasttStep_area_updDeps = []
LLR_lasttStep_area_updDeps = []

lasttStep_area_error_updGen = []
LLR_nexttStep_area_updGen = []
SUMO_lasttStep_area_updGen = []
LLR_lasttStep_area_updGen = []

# Initializing for storing CAVs routed by the LLR algorithm
cavs_area = set([])
noncavs_area = set([])
# all_vehs_area = set([]) # only when not measuring travel time accuracy
if BEGIN_LANE_CLOSURE > BEGIN_T + STEP_LENGTH:
    # When lane closure happens not at the beginning of the simulation
    global routedCAVsLaneClosure
    routedCAVsLaneClosure = set([])
if INTERFACE in ("libsumo", "traci") and TTS_PER_VEH:
    # Only when needed to traci.setAdaptTravelTime for each vehicle
    global CAVsAlreadyBroadcastedTo

# Initializing for measuring accuracy of travel time prediction
tts_accur_vals = {edgeID: {"measured": dict(),
                           "estimation": {beginTime: [] for beginTime in frange(BEGIN_T + WARM_UP_T + STEP_LENGTH,
                                                                                END_T + STEP_LENGTH + STEP_LENGTH,
                                                                                LEN_RANGE)},
                           "RMSE": dict()}
                  for edgeID in allInEdges_modelled_area}
allVehs_Edge_ArrTime = dict()

# Initializing for measuring accuracy of departures modelling
num_ranges = int(STD_T_MAX / LEN_RANGE)
rges = [0 for _ in range(0, num_ranges)]
deps_accur_vals = {edgeID: {"measured": dict(),
                            "estimation": {beginTime: rges for beginTime in frange(BEGIN_T + WARM_UP_T + STEP_LENGTH,
                                                                                   END_T + STEP_LENGTH + STEP_LENGTH,
                                                                                   LEN_RANGE)},
                            "RMSE": dict()}
                   for edgeID in allInEdges_modelled_area}

# Set travel times for all connections of modelled junctions as zero
# This is because the LLR express time of edges between junctions while SUMO also consider travel times of junctions
# internal edges (connections). LLR includes the connections travel time on their incoming edges travel time
# This avoids SUMO routing taking LLR travel times for edges and hourly travel times for connections.
for conn_edgeID in allConn_modelled_area:
    # Set traffic network travel times to for every vehicle
    traci.edge.adaptTraveltime(conn_edgeID,         # SUMO edgeID
                               0)                   # traveltime
                               # BEGIN_T,             # beginTime
                               # END_T + STEP_LENGTH) # endTime

# try:
# Constant Loop to Communicate with SUMO/TraCI
LOGGER.info('All set up, starting simulation loop')
while ACTUAL_T < END_T + STEP_LENGTH:
    if INTERFACE == "libsumo":
        # Do simulation step using Libsumo
        stored_ACTUAL_T = ACTUAL_T
        # If Libsumo open a thread to avoid the simulation getting stuck
        Thread(target=traci.simulationStep()).start()
        time.sleep(0.02)
        # Returns the current simulation time in seconds
        ACTUAL_T = round(float(traci.simulation.getTime()), 1)
        if ACTUAL_T == stored_ACTUAL_T:
            LOGGER.debug('SUMO stopped')
            logger.info('SUMO stopped')
            Thread(target=traci.simulationStep()).start()
            time.sleep(1)
            ACTUAL_T = round(float(traci.simulation.getTime()), 1)
            if ACTUAL_T == stored_ACTUAL_T:
                raise SystemExit
            else:
                LOGGER.debug('SUMO is back')
                logger.info('SUMO is back')

        sys.stdout.write('\r' + "Step " + str(ACTUAL_T))
    elif INTERFACE == "traci":
        # Do simulation step using Traci
        traci.simulationStep()

    if INTERFACE in ("libsumo", "traci"):
        # Returns the current simulation time in seconds
        ACTUAL_T = round(float(traci.simulation.getTime()), 1)
        LOGGER.debug("SIMULATION TIME "+str(ACTUAL_T))
    else:
        LOGGER.debug("Get current time in seconds via other interface")

    # For timing statistics of functions
    master_start = time.time()
    for f_ind,f_val in enumerate(timings):
        timings[f_ind][1] = 0

    # Initializing for counting number of all vehicles
    tStep_area_CAVs = set()
    tStep_area_nonCAVs = set()
    tStep_area_queues = []
    tStep_area_queuesError = []
    tStep_area_finishing = []
    tStep_area_alreadyOnLane = []
    tStep_area_CAVreplacedAlready = []
    tStep_area_AlreadyNotGenCAVBalance = []
    tStep_area_QueueBalance = []
    tStep_area_newArrivals = []
    tStep_area_previousArrivals = []
    tStep_area_beforeIT1 = []
    tStep_area_afterIT1 = []
    tStep_area_beforeIT2 = []
    tStep_area_afterIT2 = []

    # For each traffic network
    for ind_trafficNet in range(0, len(msgs_objs)):
        for tc_it in range(1, TC_IT_MAX + 1):
            # As the simulation is for all traffic controllers and different traffic networks,
            # it includes TC_seq and ind_trafficNet but the script is for each traffic controller,
            # so only jct_order is needed from getTCorder function.
            # It is possible to maintain them and avoid changes in the code.
            TC_seq, jct_order = getTCorder(jct_objs, tc_it, time2updte, jct_order)

            # For simulation purposes, get SUMO CAV data
            if tc_it == 1:
                handleSUMOsimulationData(time2updte)
                if INTERFACE == "libsumo":
                    # Get CAV data via Libsumo
                    getCAVdataViaLibsumo(TC_seq)
                elif INTERFACE == "traci":
                    # Get CAV data via Traci
                    getCAVdataViaTraci(TC_seq)

            # For each traffic controller in the traffic network ind_trafficNet
            # The final script is for one TC, the for loop below is for the simulation of each TC
            for tc_ind in TC_seq:
                # Define the interval for the time-dependent edge/lane attributes
                edg_objs[tc_ind].setEdgesAttrbInterval(ACTUAL_T)

                if tc_it == 1:
                    # For counting number of all vehicles
                    tStep_area_alreadyAccounted = []

                    # For accounting closing lane on LLR system
                    if ACTUAL_T == BEGIN_LANE_CLOSURE:
                        # Define the closed lanes and lanes and edges that no vehicle will
                        # take because they only lead to closed lanes
                        edg_objs[tc_ind].setClosedLanes(LANES2CLOSE, DIST2CLOSE)

                    # For acounting opening lane on LLR system
                    elif ACTUAL_T == END_LANE_CLOSURE:
                        # Define the edges that vehicles will take again because it reopened lanes
                        edg_objs[tc_ind].clearClosedLanes(LANES2CLOSE)

                    # Get CAVs in the range of the RSUs of each cooperative junction modelled by the TC
                    jct_objs[tc_ind].getCAVinRange(edg_objs[tc_ind], veh_objs[tc_ind])

                    # Get CAVs' data for sampling queue length in non-connected junctions
                    MonitorQueueLenUsingCAVs(jct_objs[tc_ind], edg_objs[tc_ind])

                # If it is time to update the junction Local Level Routing Algorithm
                if time2updte <= 0:
                    if tc_it == 1:

                        # Reset CAVs already routed between algorithm updates (only for simulation purposes)
                        if INTERFACE in ("libsumo", "traci") and TTS_PER_VEH:
                            # Only when needed to traci.setAdaptTravelTime for each vehicle
                            CAVsAlreadyBroadcastedTo = set([])

                        # Get data from LDM messages
                        msgs_objs[ind_trafficNet].rcvLdmDataMsgs(jct_objs[tc_ind], edg_objs[tc_ind], time2updte)

                        # Make sure the queue prediction ranges combine with the ACTUAL_T
                        edg_objs[tc_ind].adjQueuePred(jct_objs[tc_ind])

                        # Get SPaT
                        movg_objs[tc_ind].getSignalizedJctMovGroupPlans(jct_objs[tc_ind])

                        # Get CAVs around RSUs (from subscription of Cooperative Junctions) and generate CAVs on lane
                        veh_objs[tc_ind].genCAVOnLane(jct_objs[tc_ind], movg_objs[tc_ind], edg_objs[tc_ind])

                        # Get the expected vehicles to be on lane at ACTUAL_T
                        edg_objs[tc_ind].getVehAlreadyOnLane(jct_objs[tc_ind], veh_objs[tc_ind])

                        # Generate vehicles queueing as well as those beginning their trip on each lane
                        veh_objs[tc_ind].genStartingVehOnLane(jct_objs[tc_ind], movg_objs[tc_ind], edg_objs[tc_ind])

                        # For adding error of queue length estimation
                        for edgeID in all_InOut_edges:
                            # Restart number of vehicles on the lane within last range
                            vehs_last_range.update({edgeID: set([])})

                    if tc_it == TC_IT_MAX and tc_ind == TC_seq[-1]:
                        # Last iteration and last traffic controller, decrease the time2updte
                        time2updte = LEN_RANGE - STEP_LENGTH

                    # Get the lanes that are in the order of junctions to be explored
                    lanes_order = jct_objs[tc_ind].getLanesOrder(edg_objs[tc_ind], jct_order)

                    # Estimate vehicle arrivals
                    edg_objs[tc_ind].estArrivals(jct_objs[tc_ind], movg_objs[tc_ind], veh_objs[tc_ind], lanes_order)

                    # Estimate departures
                    edg_objs[tc_ind].estDepartures(jct_objs[tc_ind], movg_objs[tc_ind], veh_objs[tc_ind], tc_it,
                                                   jct_order, lanes_order)

                    # Estimate edge travel times
                    edg_objs[tc_ind].estEdgeTravelTime(jct_objs[tc_ind], veh_objs[tc_ind], movg_objs[tc_ind], tc_it)

                    # Reset data for vehicles on lanes
                    veh_objs[tc_ind].ResetVehsOnLane(jct_objs[tc_ind], edg_objs[tc_ind], tc_it, jct_order)

                    # Send dynamic inflow profile, edge travel times and queue prediction to other TCs
                    msgs_objs[ind_trafficNet].sendLdmDataMsgs(jct_objs[tc_ind], edg_objs[tc_ind], tc_it, jct_order,
                                                              lanes_order)

                    if tc_it == TC_IT_MAX and tc_ind == TC_seq[-1]:
                        # Broadcast last updated travel times for all vehicles either in the whole network or just
                        # those within comm. range of RSUs in last iteration and last traffic controller to model
                        BroadcastAreaTravelTime(TC_seq)

                        # Print travel times of alternative routes from south east to north west
                        if DO_PRINT and NOGUI == False and ACTUAL_T > BEGIN_T + STEP_LENGTH:
                            global sum_tt_vals
                            LOGGER.debug("\n")
                            LOGGER.debug("Time Step:", ACTUAL_T)
                            num_ranges = int(STD_T_MAX / LEN_RANGE)
                            zero_ranges = [0 for _ in range(0, num_ranges)]
                            max_ranges = [{"None": 0} for _ in range(0, num_ranges)]
                            sum_tt_vals = {"south counter-clock": (zero_ranges[:], max_ranges[:]),
                                           "left south counter-clock": (zero_ranges[:], max_ranges[:]),
                                           "straight south counter-clock": (zero_ranges[:], max_ranges[:]),
                                           "down west counter-clock": (zero_ranges[:], max_ranges[:]),
                                           "up west counter-clock": (zero_ranges[:], max_ranges[:]),
                                           "left up west counter-clock": (zero_ranges[:], max_ranges[:]),
                                           "straight up west counter-clock": (zero_ranges[:], max_ranges[:]),
                                           "down east counter-clock": (zero_ranges[:], max_ranges[:]),
                                           "left down east counter-clock": (zero_ranges[:], max_ranges[:]),
                                           "straight down east counter-clock": (zero_ranges[:], max_ranges[:]),
                                           "up east counter-clock": (zero_ranges[:], max_ranges[:]),
                                           "left up east counter-clock": (zero_ranges[:], max_ranges[:]),
                                           "straight up east counter-clock": (zero_ranges[:], max_ranges[:]),
                                           "center EW": (zero_ranges[:], max_ranges[:]),
                                           "right center EW": (zero_ranges[:], max_ranges[:]),
                                           "left center EW": (zero_ranges[:], max_ranges[:]),
                                           "north counter-clock": (zero_ranges[:], max_ranges[:]),

                                           "south clockwise": (zero_ranges[:], max_ranges[:]),
                                           "straight south clockwise": (zero_ranges[:], max_ranges[:]),
                                           "right south clockwise": (zero_ranges[:], max_ranges[:]),
                                           "down west clockwise": (zero_ranges[:], max_ranges[:]),
                                           "up west clockwise": (zero_ranges[:], max_ranges[:]),
                                           "left up west clockwise": (zero_ranges[:], max_ranges[:]),
                                           "right up west clockwise": (zero_ranges[:], max_ranges[:]),
                                           "down east clockwise": (zero_ranges[:], max_ranges[:]),
                                           "right down east clockwise": (zero_ranges[:], max_ranges[:]),
                                           "left down east clockwise": (zero_ranges[:], max_ranges[:]),
                                           "up east clockwise": (zero_ranges[:], max_ranges[:]),
                                           "straight up east clockwise": (zero_ranges[:], max_ranges[:]),
                                           "right up east clockwise": (zero_ranges[:], max_ranges[:]),
                                           "center WE": (zero_ranges[:], max_ranges[:]),
                                           "left center WE": (zero_ranges[:], max_ranges[:]),
                                           "right center WE": (zero_ranges[:], max_ranges[:]),
                                           "north clockwise": (zero_ranges[:], max_ranges[:]),
                                           "left north clockwise": (zero_ranges[:], max_ranges[:]),
                                           "right north clockwise": (zero_ranges[:], max_ranges[:]),

                                           "Curve north up east": (zero_ranges[:], max_ranges[:]),
                                           "SE to NW via north": (zero_ranges[:], max_ranges[:]),
                                           "SE to NW via center": (zero_ranges[:], max_ranges[:]),
                                           "SE to NW via south": (zero_ranges[:], max_ranges[:]),
                                           "SW to NE via north": (zero_ranges[:], max_ranges[:]),
                                           "SW to NE via center": (zero_ranges[:], max_ranges[:]),
                                           "SW to NE via south": (zero_ranges[:], max_ranges[:]),
                                           "NE to SW via north": (zero_ranges[:], max_ranges[:]),
                                           "NE to SW via center": (zero_ranges[:], max_ranges[:]),
                                           "NE to SW via south": (zero_ranges[:], max_ranges[:]),
                                           "NW to SE via north": (zero_ranges[:], max_ranges[:]),
                                           "NW to SE via center": (zero_ranges[:], max_ranges[:]),
                                           "NW to SE via south": (zero_ranges[:], max_ranges[:])}

                            try:
                                len(edges2show)
                            except NameError:
                                des_routes = ("Curve north up east",
                                               "SE to NW via north",
                                               "SE to NW via center",
                                               "SE to NW via south",
                                               "SW to NE via north",
                                               "SW to NE via center",
                                               "SW to NE via south",
                                               "NE to SW via north",
                                               "NE to SW via center",
                                               "NE to SW via south",
                                               "NW to SE via north",
                                               "NW to SE via center",
                                               "NW to SE via south")
                                if "mid" in SCE_NAME:
                                    # If scenario with closed lanes, use a differente edge
                                    non_closed_edge = ("gneE45", "left up east counter-clock")
                                else:
                                    non_closed_edge = ("gneE46", "left up east counter-clock")

                                edges2show = (('fl_135706746#2', "first south counter-clock"),
                                              ("fl_135706746#7", "second south counter-clock"),
                                              ("fl_135706746#10.18", "third south counter-clock"),
                                              ("gneE244", "left south counter-clock"),
                                              ("gneE243", "straight south counter-clock"),

                                              ("gneE10", "first down west counter-clock"),
                                              ("gneE10.433", "second down west counter-clock"),
                                              ("dz_gneE33", "first up west counter-clock"),
                                              ("dz_-4641287#0", "second up west counter-clock"),
                                              ("gneE52", "third up west counter-clock"),
                                              ("dz_gneE44", "fourth up west counter-clock"),
                                              ("gneE17", "fifth up west counter-clock"),
                                              ("gneE18", "sixth up west counter-clock"),
                                              ("gneE55", "straight up west counter-clock"),
                                              ("gneE54", "left up west counter-clock"),

                                              ("gneE293", "first down east counter-clock"),
                                              ("fl_80779632#4", "second down east counter-clock"),
                                              ("gneE148", "straight down east counter-clock"),
                                              ("gneE149", "left down east counter-clock"),
                                              ("dz_330512997", "first up east counter-clock"),
                                              ("dz_50178031", "second up east counter-clock"),
                                              ("dz_50178032#1", "third up east counter-clock"),
                                              ("dz_50178032#1.132", "fourth up east counter-clock"),
                                              ("gneE47", "straight up east counter-clock"),
                                              non_closed_edge,

                                              ("fl_4641299#11", "first center EW"),
                                              ("gneE239", "right center EW"),
                                              ("gneE69", "left center EW"),

                                              ("dz_gneE5", "first north counter-clock"),
                                              ("dz_-172345911#13", "second north counter-clock"),
                                              ("dz_-172345911#2", "third north counter-clock"),
                                              ("dz_128493058#0", "fourth north counter-clock"),



                                              ('fl_483444577#0', "first south clockwise"),
                                              ("fl_261867261#4", "second south clockwise"),
                                              ("fl_483444576#1", "third south clockwise"),
                                              ("fl_483444573#0", "fourth south clockwise"),
                                              ("gneE153", "straight south clockwise"),
                                              ("gneE152", "right south clockwise"),

                                              ("gneE11", "first down west clockwise"),
                                              ("gneE11.74", "second down west clockwise"),
                                              ("gneE203", "third down west clockwise"),
                                              ("gneE210", "fourth down west clockwise"),
                                              ("gneE235.10", "first up west clockwise"),
                                              ("dz_4641564#0", "second up west clockwise"),
                                              ("gneE51", "first left up west clockwise"),
                                              ("dz_gneE14", "second left up west clockwise"),
                                              ("dz_4641287#0", "third left up west clockwise"),
                                              ("gneE31", "fourth left up west clockwise"),
                                              ("gneE53", "first right up west clockwise"),
                                              ("dz_8102967#0", "second right up west clockwise"),
                                              ("gneE35", "third right up west clockwise"),

                                              ("fl_330512996#6.19", "first down east clockwise"),
                                              ("gneE241.49", "left down east clockwise"),
                                              ("gneE240.50", "right down east clockwise"),
                                              ("dz_142147143#0.19", "first up east clockwise"),
                                              ("dz_142147143#0.19.220", "second up east clockwise"),
                                              ("dz_142147143#9", "third up east clockwise"),
                                              ("dz_142147141", "fourth up east clockwise"),
                                              ("gneE143", "right up east clockwise"),
                                              ("gneE144", "straight up east clockwise"),

                                              ("fl_4641298#3", "first center WE"),
                                              ("gneE147", "left center WE"),
                                              ("gneE145", "right center WE"),

                                              ("dz_172345911#2", "first north clockwise"),
                                              ("dz_172345911#13", "second north clockwise"),
                                              ("gneE37", "left north clockwise"),
                                              ("gneE41", "right north clockwise"))

                            for in_edgID, edgName in edges2show:
                                # LOGGER.debug(edgName,in_edgID)
                                for arr_rge in range(0, num_ranges):
                                    # LOGGER.debug(arr_rge, edg_objs[tc_ind].traffic_net_tt[in_edgID][1][arr_rge])
                                    splited = edgName.split(" ")
                                    if "up" in splited or "down" in splited:
                                        # splited = [direction or position, up or down, region, clockwise or not]
                                        segment = " ".join(splited[-4:])
                                        if sum_tt_vals.has_key(segment) == False:
                                            # First word is not direction, but position (first, second)
                                            segment = " ".join(splited[-3:])
                                        sum_tt_vals[segment][0][arr_rge] \
                                            += edg_objs[tc_ind].traffic_net_tt[in_edgID][1][arr_rge]
                                    else:
                                        # splited = [direction or position, region, clockwise or not]
                                        segment = " ".join(splited[-3:])
                                        if sum_tt_vals.has_key(segment) == False:
                                            segment = " ".join(splited[-2:])
                                        sum_tt_vals[segment][0][arr_rge] \
                                            += edg_objs[tc_ind].traffic_net_tt[in_edgID][1][arr_rge]

                                    if edg_objs[tc_ind].traffic_net_tt[in_edgID][1][arr_rge] \
                                            > sum_tt_vals[segment][1][arr_rge].values()[0]:
                                        sum_tt_vals[segment][1][arr_rge] = {in_edgID: edg_objs[tc_ind].
                                                                            traffic_net_tt[in_edgID][1][arr_rge]}

                            for route in des_routes:
                                for arr_rge in range(0, num_ranges):
                                    if route == "Curve north up east":
                                        segments = ["dz_135646528#1", "dz_142147138#0"]

                                        for segment in segments:
                                            sum_tt_vals[route][0][arr_rge] += \
                                              edg_objs[tc_ind].traffic_net_tt[segment][1][arr_rge]
                                            if edg_objs[tc_ind].traffic_net_tt[segment][1][arr_rge] \
                                            > sum_tt_vals[route][1][arr_rge].values()[0]:
                                                sum_tt_vals[route][1][arr_rge] = {segment: edg_objs[tc_ind].
                                                                                  traffic_net_tt[segment][1][arr_rge]}
                                    else:
                                        if route == "SE to NW via north":
                                            segments = ["down east counter-clock","straight down east counter-clock",
                                                        "up east counter-clock", "left up east counter-clock",
                                                        "north counter-clock"]

                                        elif route == "SE to NW via center":
                                            segments = ["down east counter-clock", "left down east counter-clock",
                                                        "center EW", "right center EW", "up west clockwise",
                                                        "left up west clockwise"]

                                        elif route == "SE to NW via south":
                                            segments = ["south clockwise","right south clockwise",
                                                        "down west clockwise","up west clockwise",
                                                        "left up west clockwise"]

                                        elif route == "SW to NE via north":
                                            segments = ["down west clockwise", "up west clockwise",
                                                        "right up west clockwise", "north clockwise",
                                                        "left north clockwise"]

                                        elif route == "SW to NE via center":
                                            segments = ["down west clockwise", "center WE", "left center WE",
                                                        "up east counter-clock", "straight up east counter-clock"]

                                        elif route == "SW to NE via south":
                                            segments = ["south counter-clock", "left south counter-clock",
                                                        "down east counter-clock", "straight down east counter-clock",
                                                        "up east counter-clock", "straight up east counter-clock"]

                                        elif route == "NE to SW via north":
                                            segments = ["north counter-clock", "up west counter-clock",
                                                        "straight up west counter-clock", "down west counter-clock"]

                                        elif route == "NE to SW via center":
                                            segments = ["up east clockwise", "right up east clockwise",
                                                        "center EW", "left center EW", "down west counter-clock"]

                                        elif route == "NE to SW via south":
                                            segments = ["up east clockwise", "straight up east clockwise",
                                                        "down east clockwise", "left down east clockwise",
                                                        "south clockwise", "straight south clockwise"]

                                        elif route == "NW to SE via north":
                                            segments = ["north clockwise", "right north clockwise", "up east clockwise",
                                                        "straight up east clockwise", "down east clockwise",
                                                        "left down east clockwise"]

                                        elif route == "NW to SE via center":
                                            segments = ["up west counter-clock", "left up west counter-clock",
                                                        "center WE", "right center WE", "down east clockwise",
                                                        "left down east clockwise"]

                                        elif route == "NW to SE via south":
                                            segments = ["up west counter-clock", "straight up west counter-clock",
                                                        "down west counter-clock", "south counter-clock",
                                                        "straight south counter-clock"]

                                        for segment in segments:
                                            try:
                                                sum_tt_vals[route][0][arr_rge] += sum_tt_vals[segment][0][arr_rge]
                                                if sum_tt_vals[segment][1][arr_rge].values()[0] \
                                                > sum_tt_vals[route][1][arr_rge].values()[0]:
                                                    sum_tt_vals[route][1][arr_rge] = \
                                                        sum_tt_vals[segment][1][arr_rge].copy()
                                            except IndexError:
                                                skip = 1

                            last_OD = des_routes[0].split("via")[0]
                            for route in des_routes:
                                OD = route.split("via")[0]
                                if OD != last_OD:
                                    LOGGER.debug("\n")
                                LOGGER.debug(route, sum_tt_vals[route][0])
                                LOGGER.debug("- Longest per range", sum_tt_vals[route][1])
                                last_OD = OD
                else:
                    if tc_it == TC_IT_MAX and tc_ind == TC_seq[-1]:
                            time2updte -= STEP_LENGTH

            if INTERFACE in ("libsumo", "traci") and TTS_PER_VEH and tc_it == TC_IT_MAX and time2updte > 0:
                # Broadcast last updated travel times when setting TTs to each vehicle within comm. range of RSUs.
                # When needed traci.setAdaptTravelTime for each vehicle,
                # it must broadcast every time step as new vehicles
                # may arrive within comm. range between algorithhm updates
                BroadcastAreaTravelTime(TC_seq)


    # For counting number of all vehicles
    if NOGUI == False and time2updte == LEN_RANGE - STEP_LENGTH:
    # if time2updte == LEN_RANGE - STEP_LENGTH:
        sum_all_ids = sum([len([veh for veh, _ in enumerate(veh_objs[obj_ind].ids_all)
                                if veh_objs[obj_ind].vClass[veh] != "Art. Probe"])
                           for obj_ind in range(0, len(jct_objs))])

        sum_already_CAV_balance = 0
        all_already_CAV_balance = []
        for obj_ind in range(0, len(jct_objs)):
            for edgeID in edg_objs[obj_ind].CAVsAlreadyOnLaneBalance:
                if edg_objs[obj_ind].CAVsAlreadyOnLaneBalance[edgeID] > 0:
                    all_already_CAV_balance.append((edgeID, edg_objs[obj_ind].CAVsAlreadyOnLaneBalance[edgeID]))
                    sum_already_CAV_balance += edg_objs[obj_ind].CAVsAlreadyOnLaneBalance[edgeID]
        # LOGGER.debug("\nCAV/Already Balance\n")
        # for edgeID,balance in sorted(all_already_CAV_balance, key=lambda sort_it: sort_it[1])[::-1]:
        #     LOGGER.debug(edgeID,balance)

        # LOGGER.debug("\nRemaining Already on Lane\n")
        sorted_tStep_area_alreadyOnLane = sorted(tStep_area_alreadyOnLane, key=lambda sort_it: sort_it[1])[::-1]
        tStep_area_alreadyOnLane = []
        for laneID, num_vehs in sorted_tStep_area_alreadyOnLane:
            if num_vehs > 0:
                tStep_area_alreadyOnLane.append(num_vehs)
        #         LOGGER.debug(laneID, num_vehs)

        sum_all_already_on_lane_next = 0
        for obj_ind in range(0, len(jct_objs)):
            for laneID in edg_objs[obj_ind].already_on_lane:
                for veh, _ in edg_objs[obj_ind].already_on_lane[laneID]:
                    if "CAV" not in veh_objs[obj_ind].vClass[veh]:
                        sum_all_already_on_lane_next += 1

        if ACTUAL_T > BEGIN_T + STEP_LENGTH:
            lasttStep_area_error_updDeps = sum(LLR_lasttStep_area_updDeps) - sum(SUMO_lasttStep_area_updDeps)
            lasttStep_area_error_updGen = sum(LLR_lasttStep_area_updGen) - sum(SUMO_lasttStep_area_updGen)
        else:
            lasttStep_area_error_updDeps = 0
            lasttStep_area_error_updGen = 0
        
        if DO_PRINT:
            LOGGER.debug("\nTime Step:", ACTUAL_T, "\n",
                  "- Already on lane replaced by CAVs:", len(tStep_area_CAVreplacedAlready), "\n"
                  "- Already on lane not generated due CAVs balance:", len(tStep_area_AlreadyNotGenCAVBalance), "\n"
                  "- Remaining already on Lane (non-CAVs):", sum(tStep_area_alreadyOnLane), "\n"
                  "- Already on lane/CAVs balance:", sum_already_CAV_balance, "\n"
                  "- Queuing Generated/Deleted balance:", sum(tStep_area_QueueBalance), "\n"
                  "- All (SUMO) CAVs:", len(tStep_area_CAVs), "\n"
                  "- All (SUMO) non-CAVs:", len(tStep_area_nonCAVs), "\n"
                  "- All (SUMO) queuing:", sum(tStep_area_queues), "\n"
                  "- Applied error queue estimation:", sum(tStep_area_queuesError), "\n"
                  "- Planning Horizon Forward arrivals:", sum(tStep_area_newArrivals), "\n",
                  "- Planning Horizon Reverse arrivals:", sum(tStep_area_previousArrivals), "\n"
                  "- Until next update departures:", sum(LLR_nexttStep_area_updDeps), "\n",
                  "- (SUMO) departures from last update:", sum(SUMO_lasttStep_area_updDeps), "\n",
                  "- Error departures from last update:", lasttStep_area_error_updDeps, "\n",
                  "- Until next update vehicle generation:", sum(LLR_nexttStep_area_updGen), "\n",
                  "- (SUMO) vehicle generation from last update:", sum(SUMO_lasttStep_area_updGen), "\n",
                  "- Error vehicle generation from last update:", lasttStep_area_error_updGen, "\n",
                  "- Finished within net:", len(tStep_area_finishing), "\n"
                  "- tc_it 1 - All veh num before:", sum(tStep_area_beforeIT1), "\n",
                  "- tc_it 1 - All veh num after:", sum(tStep_area_afterIT1), "\n"
                  "- tc_it 2 - All veh num before:", sum(tStep_area_beforeIT2), "\n",
                  "- tc_it 2 - All veh num after:", sum(tStep_area_afterIT2), "\n"
                  "- Num. vehs for next update:", sum_all_ids, "\n"
                  "- (of which) already on lane (non-CAVs):", sum_all_already_on_lane_next)

        LLR_lasttStep_area_updDeps = LLR_nexttStep_area_updDeps[:]
        LLR_nexttStep_area_updDeps = []
        SUMO_lasttStep_area_updDeps = []

        LLR_lasttStep_area_updGen = LLR_nexttStep_area_updGen[:]
        LLR_nexttStep_area_updGen = []
        SUMO_lasttStep_area_updGen = []

    # For timing statistics of functions
    master_end = time.time()
    timings[-1][1] = master_end - master_start
    if DO_PRINT and NOGUI == False and timings[-1][1] > STEP_LENGTH:
        # If total time longer than the step length
        # For Checking the Runtime of each Function
        for f_ind, f_val in enumerate(timings):
            try:
                timings[f_ind][2] = timings[f_ind][1] / timings[-1][1]
            except (ZeroDivisionError, FloatingPointError):
                timings[f_ind][2] = 0

        longest_func_name = 0
        for f_name, _, _ in timings:
            if len(f_name) > longest_func_name:
                longest_func_name = len(f_name)

        LOGGER.debug("\n")
        LOGGER.debug("Time step:", ACTUAL_T)
        headers = ["Function Name", "Sum time (seconds)", "% Total"]
        for it, header in enumerate(headers):
            if it == len(headers) - 1:
                LOGGER.debug(header)  # last table column headings
            else:
                text = header + ' ' * (longest_func_name - len(header)) + "\t"
                sys.stdout.write(text) # table column headings
        for it, header in enumerate(headers):
            if it == len(headers) - 1:
                LOGGER.debug("-" * len(header)) # last table column separator
            else:
                text = "-" * longest_func_name + "\t"
                sys.stdout.write(text) # table column separator

        sorted_timings = sorted([(f_name, sum_time, perc_total)
                                 for (f_name, sum_time, perc_total) in timings], key=lambda sort_it: sort_it[2])
        for f_ind, f_val in enumerate(sorted_timings):
            if sorted_timings[f_ind][1] > 0:
                if NOGUI == False:
                    LOGGER.debug(sorted_timings[f_ind][0], ' ' * (longest_func_name - len(sorted_timings[f_ind][0])), "\t",
                          "{:4.2f}".format(sorted_timings[f_ind][1]), ' '
                          * (longest_func_name - 3 - len("{:4.2f}".format(sorted_timings[f_ind][1]))), '\t',
                          "{:.2%}".format(sorted_timings[f_ind][2]))  # generate values for columns
                else:
                    # In case desired to write the function time in the log file
                    logger.info((sorted_timings[f_ind][0] + ' ' * (longest_func_name - len(sorted_timings[f_ind][0]))
                                 + "{:4.2f}".format(sorted_timings[f_ind][1]) + ' '
                                 * (longest_func_name - 3 - len("{:4.2f}".format(sorted_timings[f_ind][1]))) +
                                 "{:.2%}".format(sorted_timings[f_ind][2])))  # generate values for columns

LOGGER.info('creating workbook for CAVs area')
# Create workbook with and a new sheet for the CAVs Area
if os.path.isfile(CAVS_AREA_FILE) == True:
    os.remove(CAVS_AREA_FILE)
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "CAVs Modelled Area"
for cav_ind, cav in enumerate(cavs_area):
    sheet.cell(row=cav_ind+1, column=1).value = cav
wb.save(CAVS_AREA_FILE)

# Create workbook with and a new sheet for the nonCAVs Area
LOGGER.info('creating workbook for nonCAVs area')
if os.path.isfile(NONCAVS_AREA_FILE) == True:
    os.remove(NONCAVS_AREA_FILE)
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "nonCAVs Modelled Area"
for cav_ind,cav in enumerate(noncavs_area):
    sheet.cell(row=cav_ind+1, column=1).value = cav
wb.save(NONCAVS_AREA_FILE)

# Create workbook with and a new sheet for all vehicles in the modelled area
LOGGER.info('creating workbook for all vehicles')
if os.path.isfile(ALL_VEHS_AREA_FILE) == True:
    os.remove(ALL_VEHS_AREA_FILE)
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "All Vehs. Modelled Area"
for cav_ind,cav in enumerate(allVehs_Edge_ArrTime.keys()):
# for cav_ind, cav in enumerate(all_vehs_area):
    sheet.cell(row=cav_ind+1, column=1).value = cav
wb.save(ALL_VEHS_AREA_FILE)

# For measuring accuracy of travel time prediction
LOGGER.info('processing data for travel time prediction accuracy')
for edgeID in tts_accur_vals:
    copied_tts_accur_vals = tts_accur_vals[edgeID]["estimation"].copy()
    for beginTime in copied_tts_accur_vals:
        if tts_accur_vals[edgeID]["estimation"][beginTime] == []:
            del tts_accur_vals[edgeID]["estimation"][beginTime]
    for beginTime in tts_accur_vals[edgeID]["measured"]:
        if tts_accur_vals[edgeID]["estimation"].has_key(beginTime):
            measured_tt = np.mean(tts_accur_vals[edgeID]["measured"][beginTime])
            tts_accur_vals[edgeID]["RMSE"].update({beginTime: []})
            # tts_accur_vals[edgeID]["RMSE"].update({beginTime: []})
            for est_tt in tts_accur_vals[edgeID]["estimation"][beginTime]:
                tts_accur_vals[edgeID]["RMSE"][beginTime].append((est_tt - measured_tt) ** 2)
            tts_accur_vals[edgeID]["RMSE"][beginTime] = round(
                math.sqrt(np.sum(tts_accur_vals[edgeID]["RMSE"][beginTime])
                          / len(tts_accur_vals[edgeID]["RMSE"][beginTime])), 1)

# For measuring accuracy of departures modelling
LOGGER.info('preprocessing data for evauation of depatrues model')
for edgeID in deps_accur_vals:
    copied_deps_accur_vals = deps_accur_vals[edgeID]["estimation"].copy()
    for beginTime in deps_accur_vals[edgeID]["measured"]:
        measured_deps = deps_accur_vals[edgeID]["measured"][beginTime]
        deps_accur_vals[edgeID]["RMSE"].update({beginTime: []})
        for est_deps in deps_accur_vals[edgeID]["estimation"][beginTime]:
            deps_accur_vals[edgeID]["RMSE"][beginTime].append((est_deps - measured_deps) ** 2)
        deps_accur_vals[edgeID]["RMSE"][beginTime] = round(
            math.sqrt(np.sum(deps_accur_vals[edgeID]["RMSE"][beginTime])
                      / len(deps_accur_vals[edgeID]["RMSE"][beginTime])), 1)

# Create XML file with measured, estimated and ratio of std. deviations to measured travel time for each edge.
# Create the file structure
LOGGER.info('creating XML log with measured and estimated travel times')
mean_data = ElementTree.Element('meandata')
items = ElementTree.SubElement(mean_data, 'interval')
for edgeID in tts_accur_vals:
    item = ElementTree.SubElement(items, 'edge')
    item.set('id', str(edgeID))
    all_vals_measured = []
    all_vals_estimated = []
    all_vals_RMSE = []
    for beginTime in tts_accur_vals[edgeID]["measured"]:
        all_vals_measured.append(np.mean(tts_accur_vals[edgeID]["measured"][beginTime]))
        try:
            all_vals_estimated.append(np.mean(tts_accur_vals[edgeID]["estimation"][beginTime]))
            all_vals_RMSE.append(tts_accur_vals[edgeID]["RMSE"][beginTime])
        except KeyError:
            noEstimation = 1

    if all_vals_measured != []:
        item.set('measuredTravelTime', str(round(np.mean(all_vals_measured), 1)))
    if all_vals_estimated != []:
        item.set('estimatedTravelTime', str(round(np.mean(all_vals_estimated), 1)))
    if all_vals_RMSE != []:
        item.set('RMSE', str(round(np.mean(all_vals_RMSE), 1)))

# Save the new XML file with the results
LOGGER.info('saving the XML log')
if os.path.isfile(TTS_ACCUR_AGGREGATED) == True:
    os.remove(TTS_ACCUR_AGGREGATED)
xml_string = ElementTree.tostring(mean_data)
parsed_xml_string = xml.dom.minidom.parseString(xml_string)
python_edge_traffic = parsed_xml_string.toprettyxml()
with open(TTS_ACCUR_AGGREGATED, "w") as f:
    f.write(python_edge_traffic)

# Create workbook with the measured, and estimated travel times as well as RMSE
LOGGER.info('creating workbook with the travel time data')
if os.path.isfile(TTS_ACCUR_FULL) == True:
    os.remove(TTS_ACCUR_FULL)
wb = openpyxl.Workbook()
mapping_time = dict()
# Create a new sheet for the function
sheet = wb.active
sheet.title = "Measured TTs"
sheet.cell(row=1, column=1).value = "EdgeID"
for time_ind, beginTime in enumerate(frange(BEGIN_T + WARM_UP_T + STEP_LENGTH,
                                            END_T + STEP_LENGTH + STEP_LENGTH,
                                            LEN_RANGE)):
    mapping_time.update({beginTime: time_ind})
    sheet.cell(row=1, column=time_ind + 2).value = beginTime
for edge_ind,edgeID in enumerate(tts_accur_vals):
    sheet.cell(row=edge_ind + 2, column=1).value = edgeID
    if tts_accur_vals[edgeID]["measured"] != {}:
        for beginTime in tts_accur_vals[edgeID]["measured"]:
            time_ind = mapping_time[beginTime]
            sheet.cell(row=edge_ind + 2, column=time_ind + 2).value = str(
                tts_accur_vals[edgeID]["measured"][beginTime])

LOGGER.info('creating sheet `Estimated TTs`')
sheet = wb.create_sheet("Estimated TTs")
sheet.cell(row=1, column=1).value = "EdgeID"
for time_ind, beginTime in enumerate(frange(BEGIN_T + WARM_UP_T + STEP_LENGTH,
                                            END_T + STEP_LENGTH + STEP_LENGTH,
                                            LEN_RANGE)):
    sheet.cell(row=1, column=time_ind + 2).value = beginTime
for edge_ind, edgeID in enumerate(tts_accur_vals):
    sheet.cell(row=edge_ind + 2, column=1).value = edgeID
    for beginTime in tts_accur_vals[edgeID]["estimation"]:
        time_ind = mapping_time[beginTime]
        sheet.cell(row=edge_ind + 2, column=time_ind + 2).value = str(
            tts_accur_vals[edgeID]["estimation"][beginTime])

# Replace all values for its average for next calculations
for edgeID in tts_accur_vals:
    for beginTime in tts_accur_vals[edgeID]["measured"]:
        tts_accur_vals[edgeID]["measured"][beginTime] = \
            round(np.mean(tts_accur_vals[edgeID]["measured"][beginTime]), 1)

for edgeID in tts_accur_vals:
    for beginTime in tts_accur_vals[edgeID]["estimation"]:
        tts_accur_vals[edgeID]["estimation"][beginTime] = \
            round(np.mean(tts_accur_vals[edgeID]["estimation"][beginTime]), 1)

LOGGER.info('creating sheet `Mean Measured TTs`')
sheet = wb.create_sheet("Mean Measured TTs")
sheet.cell(row=1, column=1).value = "EdgeID"
for time_ind, beginTime in enumerate(frange(BEGIN_T + WARM_UP_T + STEP_LENGTH,
                                            END_T + STEP_LENGTH + STEP_LENGTH,
                                            LEN_RANGE)):
    sheet.cell(row=1, column=time_ind + 2).value = beginTime
for edge_ind,edgeID in enumerate(tts_accur_vals):
    sheet.cell(row=edge_ind + 2, column=1).value = edgeID
    if tts_accur_vals[edgeID]["measured"] != {}:
        for beginTime in tts_accur_vals[edgeID]["measured"]:
            time_ind = mapping_time[beginTime]
            sheet.cell(row=edge_ind + 2, column=time_ind + 2).value = tts_accur_vals[edgeID]["measured"][beginTime]

LOGGER.info('creating sheet `Mean Estimated TTs`')
sheet = wb.create_sheet("Mean Estimated TTs")
sheet.cell(row=1, column=1).value = "EdgeID"
for time_ind, beginTime in enumerate(frange(BEGIN_T + WARM_UP_T + STEP_LENGTH,
                                            END_T + STEP_LENGTH + STEP_LENGTH,
                                            LEN_RANGE)):
    sheet.cell(row=1, column=time_ind + 2).value = beginTime
for edge_ind,edgeID in enumerate(tts_accur_vals):
    sheet.cell(row=edge_ind + 2, column=1).value = edgeID
    for beginTime in tts_accur_vals[edgeID]["estimation"]:
        time_ind = mapping_time[beginTime]
        sheet.cell(row=edge_ind + 2, column=time_ind + 2).value = tts_accur_vals[edgeID]["estimation"][beginTime]

LOGGER.info('creating sheet `RMSE`')
sheet = wb.create_sheet("RMSE")
sheet.cell(row=1, column=1).value = "EdgeID"
for time_ind, beginTime in enumerate(frange(BEGIN_T + WARM_UP_T + STEP_LENGTH,
                                            END_T + STEP_LENGTH + STEP_LENGTH,
                                            LEN_RANGE)):
    sheet.cell(row=1, column=time_ind + 2).value = beginTime
for edge_ind, edgeID in enumerate(tts_accur_vals):
    sheet.cell(row=edge_ind + 2, column=1).value = edgeID
    for beginTime in tts_accur_vals[edgeID]["RMSE"]:
        time_ind = mapping_time[beginTime]
        sheet.cell(row=edge_ind + 2, column=time_ind + 2).value = tts_accur_vals[edgeID]["RMSE"][beginTime]

LOGGER.info('creating sheet `Normalized RMSE`')
sheet = wb.create_sheet("Normalized RMSE")
sheet.cell(row=1, column=1).value = "EdgeID"
for time_ind, beginTime in enumerate(frange(BEGIN_T + WARM_UP_T + STEP_LENGTH,
                                            END_T + STEP_LENGTH + STEP_LENGTH,
                                            LEN_RANGE)):
    sheet.cell(row=1, column=time_ind + 2).value = beginTime
for edge_ind,edgeID in enumerate(tts_accur_vals):
    sheet.cell(row=edge_ind + 2, column=1).value = edgeID
    if tts_accur_vals[edgeID]["RMSE"] != {}:
        for beginTime in tts_accur_vals[edgeID]["RMSE"]:
            time_ind = mapping_time[beginTime]
            sheet.cell(row=edge_ind + 2,
                       column=time_ind + 2).value = "{:.2%}".format(tts_accur_vals[edgeID]["RMSE"][beginTime]
                                                                    / tts_accur_vals[edgeID]["measured"][beginTime])

LOGGER.info('creating sheet `Stats per Edge`')
sheet = wb.create_sheet("Stats per Edge")
sheet.cell(row=1, column=1).value = "EdgeID"
sheet.cell(row=1, column=2).value = "Mean Estimated"
sheet.cell(row=1, column=3).value = "Mean Measured"
sheet.cell(row=1, column=4).value = "Mean RMSE"
sheet.cell(row=1, column=5).value = "Mean Norm. RMSE"
for edge_ind, edgeID in enumerate(tts_accur_vals):
    all_Estimated = []
    all_Measured = []
    all_RMSE = []
    sheet.cell(row=edge_ind + 2, column=1).value = edgeID
    if tts_accur_vals[edgeID]["estimation"] != {}:
        for beginTime in tts_accur_vals[edgeID]["estimation"]:
            all_Estimated.append(tts_accur_vals[edgeID]["estimation"][beginTime])
        sheet.cell(row=edge_ind + 2, column=2).value = round(np.mean(all_Estimated), 1)
    if tts_accur_vals[edgeID]["measured"] != {}:
        for beginTime in tts_accur_vals[edgeID]["measured"]:
            all_Measured.append(tts_accur_vals[edgeID]["measured"][beginTime])
        sheet.cell(row=edge_ind + 2, column=3).value = round(np.mean(all_Measured), 1)
    if tts_accur_vals[edgeID]["RMSE"] != {}:
        for beginTime in tts_accur_vals[edgeID]["RMSE"]:
            all_RMSE.append(tts_accur_vals[edgeID]["RMSE"][beginTime])
        sheet.cell(row=edge_ind + 2, column=4).value = round(np.mean(all_RMSE), 1)
        sheet.cell(row=edge_ind + 2, column=5).value = "{:.2%}".format(np.mean(all_RMSE) / np.mean(all_Measured))

LOGGER.info('creating sheet `Stats per Time`')
sheet = wb.create_sheet("Stats per Time")
sheet.cell(row=1, column=1).value = "Time"
sheet.cell(row=1, column=2).value = "Mean Estimated"
sheet.cell(row=1, column=3).value = "Mean Measured"
sheet.cell(row=1, column=4).value = "Mean RMSE"
sheet.cell(row=1, column=5).value = "Mean Norm. RMSE"
for time_ind, beginTime in enumerate(frange(BEGIN_T + WARM_UP_T + STEP_LENGTH,
                                            END_T + STEP_LENGTH + STEP_LENGTH,
                                            LEN_RANGE)):
    sheet.cell(row=time_ind + 2, column=1).value = beginTime
    all_RMSE = []
    all_Measured = []
    all_Estimated = []
    for edge_ind, edgeID in enumerate(tts_accur_vals):
        try:
            all_Estimated.append(tts_accur_vals[edgeID]["estimation"][beginTime])
        except KeyError:
            skip = 1
        try:
            all_Measured.append(tts_accur_vals[edgeID]["measured"][beginTime])
        except KeyError:
            skip = 1
        try:
            all_RMSE.append(tts_accur_vals[edgeID]["RMSE"][beginTime])
        except KeyError:
            skip = 1
    if all_Estimated != []:
        sheet.cell(row=time_ind + 2, column=2).value = round(np.mean(all_Estimated), 1)
    if all_Measured != []:
        sheet.cell(row=time_ind + 2, column=3).value = round(np.mean(all_Measured), 1)
    if all_RMSE != []:
        sheet.cell(row=time_ind + 2, column=4).value = round(np.mean(all_RMSE), 1)
        sheet.cell(row=time_ind + 2, column=5).value = "{:.2%}".format(np.mean(all_RMSE) / np.mean(all_Measured))

LOGGER.info('creating sheet `Stats total`')
sheet = wb.create_sheet("Stats Total")
sheet.cell(row=1, column=1).value = "Sum Edge Mean Estimated"
sheet.cell(row=1, column=2).value = "Sum Edge Mean Measured"
sheet.cell(row=1, column=3).value = "Sum RMSE"
sheet.cell(row=1, column=4).value = "Norm. Sum RMSE"
all_RMSE = []
all_Measured = []
all_Estimated = []
for edge_ind, edgeID in enumerate(tts_accur_vals):
    if tts_accur_vals[edgeID]["estimation"] != {}:
        edge_mean = []
        for time_ind, beginTime in enumerate(tts_accur_vals[edgeID]["estimation"]):
            edge_mean.append(tts_accur_vals[edgeID]["estimation"][beginTime])
        all_Estimated.append(np.mean(edge_mean))
    if tts_accur_vals[edgeID]["measured"] != {}:
        edge_mean = []
        for time_ind, beginTime in enumerate(tts_accur_vals[edgeID]["measured"]):
            edge_mean.append(tts_accur_vals[edgeID]["measured"][beginTime])
        all_Measured.append(np.mean(edge_mean))
    if tts_accur_vals[edgeID]["RMSE"] != {}:
        edge_mean = []
        for time_ind, beginTime in enumerate(tts_accur_vals[edgeID]["RMSE"]):
            edge_mean.append(tts_accur_vals[edgeID]["RMSE"][beginTime])
        all_RMSE.append(np.mean(edge_mean))

if all_Estimated != []:
    sheet.cell(row=2, column=1).value = round(sum(all_Estimated), 1)
if all_Measured != []:
    sheet.cell(row=2, column=2).value = round(sum(all_Measured), 1)
if all_RMSE != []:
    sheet.cell(row=2, column=3).value = round(sum(all_RMSE), 1)
    sheet.cell(row=2, column=4).value = "{:.2%}".format(sum(all_RMSE) / sum(all_Measured))

# Finally, save the file and give it a name
wb.save(TTS_ACCUR_FULL)


# Create workbook with the measured, and estimated departures as well as RMSE
LOGGER.info('creating workbook for measured and esimated departures')
if os.path.isfile(DEPS_ACCUR_FULL) == True:
    os.remove(DEPS_ACCUR_FULL)
wb = openpyxl.Workbook()
mapping_time = dict()
# Create a new sheet for the function
sheet = wb.active
sheet.title = "Measured Departures"
sheet.cell(row=1, column=1).value = "EdgeID"
for time_ind, beginTime in enumerate(frange(BEGIN_T + WARM_UP_T + STEP_LENGTH,
                                            END_T + STEP_LENGTH + STEP_LENGTH,
                                            LEN_RANGE)):
    mapping_time.update({beginTime: time_ind})
    sheet.cell(row=1, column=time_ind + 2).value = beginTime
for edge_ind,edgeID in enumerate(deps_accur_vals):
    sheet.cell(row=edge_ind + 2, column=1).value = edgeID
    if deps_accur_vals[edgeID]["measured"] != {}:
        for beginTime in deps_accur_vals[edgeID]["measured"]:
            time_ind = mapping_time[beginTime]
            sheet.cell(row=edge_ind + 2, column=time_ind + 2).value = str(deps_accur_vals[edgeID]["measured"][beginTime])

LOGGER.info('creating sheet `Estimated Departures`')
sheet = wb.create_sheet("Estimated Departures")
sheet.cell(row=1, column=1).value = "EdgeID"
for time_ind, beginTime in enumerate(frange(BEGIN_T + WARM_UP_T + STEP_LENGTH,
                                            END_T + STEP_LENGTH + STEP_LENGTH,
                                            LEN_RANGE)):
    sheet.cell(row=1, column=time_ind + 2).value = beginTime
for edge_ind, edgeID in enumerate(deps_accur_vals):
    sheet.cell(row=edge_ind + 2, column=1).value = edgeID
    for beginTime in deps_accur_vals[edgeID]["estimation"]:
        time_ind = mapping_time[beginTime]
        sheet.cell(row=edge_ind + 2, column=time_ind + 2).value = str(
            deps_accur_vals[edgeID]["estimation"][beginTime])

# Replace all values for its average for next calculations
for edgeID in deps_accur_vals:
    for beginTime in deps_accur_vals[edgeID]["estimation"]:
        deps_accur_vals[edgeID]["estimation"][beginTime] = \
            round(np.mean(deps_accur_vals[edgeID]["estimation"][beginTime]), 1)

LOGGER.info('creating sheet `Mean Measured Departures`')
sheet = wb.create_sheet("Mean Measured Departures")
sheet.cell(row=1, column=1).value = "EdgeID"
for time_ind, beginTime in enumerate(frange(BEGIN_T + WARM_UP_T + STEP_LENGTH,
                                            END_T + STEP_LENGTH + STEP_LENGTH,
                                            LEN_RANGE)):
    sheet.cell(row=1, column=time_ind + 2).value = beginTime
for edge_ind,edgeID in enumerate(deps_accur_vals):
    sheet.cell(row=edge_ind + 2, column=1).value = edgeID
    if deps_accur_vals[edgeID]["measured"] != {}:
        for beginTime in deps_accur_vals[edgeID]["measured"]:
            time_ind = mapping_time[beginTime]
            sheet.cell(row=edge_ind + 2, column=time_ind + 2).value = deps_accur_vals[edgeID]["measured"][beginTime]

LOGGER.info('creating sheet `Mean Estimated Departures`')
sheet = wb.create_sheet("Mean Estimated Departures")
sheet.cell(row=1, column=1).value = "EdgeID"
for time_ind, beginTime in enumerate(frange(BEGIN_T + WARM_UP_T + STEP_LENGTH,
                                            END_T + STEP_LENGTH + STEP_LENGTH,
                                            LEN_RANGE)):
    sheet.cell(row=1, column=time_ind + 2).value = beginTime
for edge_ind,edgeID in enumerate(deps_accur_vals):
    sheet.cell(row=edge_ind + 2, column=1).value = edgeID
    for beginTime in deps_accur_vals[edgeID]["estimation"]:
        time_ind = mapping_time[beginTime]
        sheet.cell(row=edge_ind + 2, column=time_ind + 2).value = deps_accur_vals[edgeID]["estimation"][beginTime]

LOGGER.info('creating sheet `RMSE`')
sheet = wb.create_sheet("RMSE")
sheet.cell(row=1, column=1).value = "EdgeID"
for time_ind, beginTime in enumerate(frange(BEGIN_T + WARM_UP_T + STEP_LENGTH,
                                            END_T + STEP_LENGTH + STEP_LENGTH,
                                            LEN_RANGE)):
    sheet.cell(row=1, column=time_ind + 2).value = beginTime
for edge_ind, edgeID in enumerate(deps_accur_vals):
    sheet.cell(row=edge_ind + 2, column=1).value = edgeID
    for beginTime in deps_accur_vals[edgeID]["RMSE"]:
        time_ind = mapping_time[beginTime]
        sheet.cell(row=edge_ind + 2, column=time_ind + 2).value = deps_accur_vals[edgeID]["RMSE"][beginTime]

LOGGER.info('creating sheet `Normalized RMSE`')
sheet = wb.create_sheet("Normalized RMSE")
sheet.cell(row=1, column=1).value = "EdgeID"
for time_ind, beginTime in enumerate(frange(BEGIN_T + WARM_UP_T + STEP_LENGTH,
                                            END_T + STEP_LENGTH + STEP_LENGTH,
                                            LEN_RANGE)):
    sheet.cell(row=1, column=time_ind + 2).value = beginTime
for edge_ind,edgeID in enumerate(deps_accur_vals):
    sheet.cell(row=edge_ind + 2, column=1).value = edgeID
    if deps_accur_vals[edgeID]["RMSE"] != {}:
        for beginTime in deps_accur_vals[edgeID]["RMSE"]:
            time_ind = mapping_time[beginTime]
            sheet.cell(row=edge_ind + 2,
                       column=time_ind + 2).value = "{:.2%}".format(deps_accur_vals[edgeID]["RMSE"][beginTime]
                                                                    / deps_accur_vals[edgeID]["measured"][beginTime])

LOGGER.info('creating sheet `Stats per Edge`')
sheet = wb.create_sheet("Stats per Edge")
sheet.cell(row=1, column=1).value = "EdgeID"
sheet.cell(row=1, column=2).value = "Mean Estimated"
sheet.cell(row=1, column=3).value = "Mean Measured"
sheet.cell(row=1, column=4).value = "Mean RMSE"
sheet.cell(row=1, column=5).value = "Mean Norm. RMSE"
for edge_ind, edgeID in enumerate(deps_accur_vals):
    all_Estimated = []
    all_Measured = []
    all_RMSE = []
    sheet.cell(row=edge_ind + 2, column=1).value = edgeID
    if deps_accur_vals[edgeID]["estimation"] != {}:
        for beginTime in deps_accur_vals[edgeID]["estimation"]:
            all_Estimated.append(deps_accur_vals[edgeID]["estimation"][beginTime])
        sheet.cell(row=edge_ind + 2, column=2).value = round(np.mean(all_Estimated), 1)
    if deps_accur_vals[edgeID]["measured"] != {}:
        for beginTime in deps_accur_vals[edgeID]["measured"]:
            all_Measured.append(deps_accur_vals[edgeID]["measured"][beginTime])
        sheet.cell(row=edge_ind + 2, column=3).value = round(np.mean(all_Measured), 1)
    if deps_accur_vals[edgeID]["RMSE"] != {}:
        for beginTime in deps_accur_vals[edgeID]["RMSE"]:
            all_RMSE.append(deps_accur_vals[edgeID]["RMSE"][beginTime])
        sheet.cell(row=edge_ind + 2, column=4).value = round(np.mean(all_RMSE), 1)
        sheet.cell(row=edge_ind + 2, column=5).value = "{:.2%}".format(np.mean(all_RMSE) / np.mean(all_Measured))

LOGGER.info('creating sheet `Stats per Time`')
sheet = wb.create_sheet("Stats per Time")
sheet.cell(row=1, column=1).value = "Time"
sheet.cell(row=1, column=2).value = "Mean Estimated"
sheet.cell(row=1, column=3).value = "Mean Measured"
sheet.cell(row=1, column=4).value = "Mean RMSE"
sheet.cell(row=1, column=5).value = "Mean Norm. RMSE"
for time_ind, beginTime in enumerate(frange(BEGIN_T + WARM_UP_T + STEP_LENGTH,
                                            END_T + STEP_LENGTH + STEP_LENGTH,
                                            LEN_RANGE)):
    sheet.cell(row=time_ind + 2, column=1).value = beginTime
    all_RMSE = []
    all_Measured = []
    all_Estimated = []
    for edge_ind, edgeID in enumerate(deps_accur_vals):
        try:
            all_Estimated.append(deps_accur_vals[edgeID]["estimation"][beginTime])
        except KeyError:
            skip = 1
        try:
            all_Measured.append(deps_accur_vals[edgeID]["measured"][beginTime])
        except KeyError:
            skip = 1
        try:
            all_RMSE.append(deps_accur_vals[edgeID]["RMSE"][beginTime])
        except KeyError:
            skip = 1
    if all_Estimated != []:
        sheet.cell(row=time_ind + 2, column=2).value = round(np.mean(all_Estimated), 1)
    if all_Measured != []:
        sheet.cell(row=time_ind + 2, column=3).value = round(np.mean(all_Measured), 1)
    if all_RMSE != []:
        sheet.cell(row=time_ind + 2, column=4).value = round(np.mean(all_RMSE), 1)
        sheet.cell(row=time_ind + 2, column=5).value = "{:.2%}".format(np.mean(all_RMSE) / np.mean(all_Measured))

# Finally, save the file and give it a name
wb.save(DEPS_ACCUR_FULL)

LOGGER.info('calling traci.close()')
traci.close()

logger.info('End Simulation')
LOGGER.info('simulation finished')

sys.stdout.flush()

# except (SystemExit, KeyboardInterrupt):
#     traci.close()
#     raise
# except Exception, e:
#     LOGGER.debug("\n There was an error, check the log file \n")
#     logger.error('Crash', exc_info=True)
#     sys.exit(1)