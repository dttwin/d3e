# coding=utf-8

## Read xml files after each replication of certain script and output aggregated results into excel using confidence
## intervals and plot a Box Plot for each desired measure. It is possible to use excel files containing IDs that filter
## the results. It uses functions for a specific file or a standard function that gets values from the xml file. It
## also compares the performance of different scenarios.


## IMPORT MODULES

from __future__ import absolute_import
from __future__ import print_function

import os
import sys
import subprocess
from xml.etree import ElementTree
import numpy as np
import openpyxl
import math
import matplotlib.pyplot as plt
import shlex
import warnings
warnings.filterwarnings("ignore")

## GLOBALS

global broke_filter

## INPUTS

path_outputs = "D:/Andre_Back_Up/a_Projects_Files/Local_Level_Routing/Prague_Zizkov/Inputs_Outputs/Necessary_Outputs/"
path_inputs = "D:/Andre_Back_Up/a_Projects_Files/Local_Level_Routing/Prague_Zizkov/Inputs_Outputs/"
path_scenarios = os.path.dirname(os.path.abspath(__file__))

excel_file = 'experiment_results.mid.xlsx'

begin_t_consider = 43200 + 1800 # Initial time to consider results, begin time of simulation plus the warm_up time

scenarios = [# ((scenario name, inputs file),...)
             # 15 scenarios
             # 22 scenarios including centralized system
             ("LLR 100PR", "inputs.mid.LLR.100PR.txt"),
             # ("LLR 100PR C", "inputs.mid.LLR.100PR.C.txt"),
             ("HTT 100PR", "inputs.mid.HTT.100PR.txt"),
             ("LLR 85PR", "inputs.mid.LLR.85PR.txt"),
             # ("LLR 85PR C", "inputs.mid.LLR.85PR.C.txt"),
             ("HTT 85PR", "inputs.mid.HTT.85PR.txt"),
             ("LLR 70PR", "inputs.mid.LLR.70PR.txt"),
             # ("LLR 70PR C", "inputs.mid.LLR.70PR.C.txt"),
             ("HTT 70PR", "inputs.mid.HTT.70PR.txt"),
             ("LLR 50PR", "inputs.mid.LLR.50PR.txt"),
             # ("LLR 50PR C", "inputs.mid.LLR.50PR.C.txt"),
             ("HTT 50PR", "inputs.mid.HTT.50PR.txt"),
             ("LLR 30PR", "inputs.mid.LLR.30PR.txt"),
             # ("LLR 30PR C", "inputs.mid.LLR.30PR.C.txt"),
             ("HTT 30PR", "inputs.mid.HTT.30PR.txt"),
             ("LLR 20PR", "inputs.mid.LLR.20PR.txt"),
             # ("LLR 20PR C", "inputs.mid.LLR.20PR.C.txt"),
             ("HTT 20PR", "inputs.mid.HTT.20PR.txt"),
             ("LLR 10PR", "inputs.mid.LLR.10PR.txt"),
             # ("LLR 10PR C", "inputs.mid.LLR.10PR.C.txt"),
             ("HTT 10PR", "inputs.mid.HTT.10PR.txt"),
             ("TURNS", "inputs.mid.TURNS.txt")
             ]

# Define the number of replications for each scenario. Same index of scenario
# Scenarios that are comparing to each other either must have same number of replications,
# or they can have different but the one with more will use the last replication of the one with less replications
number_replications = [10,10,10,10,10,10,10,10,10,10,10,10,10,10,10]

# Define critical_value_t_student for each scenario. Same index of scenario
# Critical Value T-Student:
    # alpha = usually 0.25 for one tail or 0.5 when two tail
    # degrees of freedom = number_replications - 1
critical_value_t_student = [2.2622] * len(scenarios)

# Define wich scenario to compare with. Same index of scenario
base_scenario_ind = [1,None,3,None,5,None,7,None,9,None,11,None,13,None,None]

# In case of a file to build (generate) the travel demand (routes).
# (travel demand generator file, SUMO router). Use Same index of scenario.
# Use none if skip to build a file (the scenario uses the last generated routes of a previous scenario)
# Useful for same generated routes for proposed case and base case with same flow volumes and PR
# In case same routes file for all scenarios. Ex.: "gen_high_travel_demand.txt"
gen_routes_file = "gen_routes.mid.txt"
# In case of a specific routes file per scenario.
# Ex: ["gen_high_travel_demand.TURNS.txt",
     # "gen_high_travel_demand.10PR.txt",
     # None,
     # "gen_high_travel_demand.20PR.txt"]
# gen_routes_files =

fltr_classif_objs = [# List with dictionary, key is the attribute to filter using dictionary key its values
                         # Use {"any attribute: [["all"], ["All"]]} to indicate no filtering/classification
                         # Use in the dict. key for the attribute and then slash </> and a number in case of repeting
                         # the same attribute
                         # Use in the key "elemenent1,elemenent2,attribute:" when attribute under an tree of elements
                         # For the values, use [[attr. values],[attr. names]]
                         # Ex.: [{"id": [["LLR_routed_vehs"],["CAVs in Area"]], "rerouteNo": [["1", "2"],["One", "Two"]]}]
                         # This will filter the results of ids in LLR_routed_vehs that have reroutedNo == 1 or  == 2
                         # Or if used as classification it will create three results one with the id and other two with
                         # reroute number
                         # and the aggregation will have the name CAVs in Area, One and Two.
                         # Ex.: [{"emissions,electricity_abs": [["0"],["No Consumption"]]}]
                         # Ir will display the results from the emissions tree element for all veicles that have
                         # no eletricity comsumption
                         # Regarding using {"id": [["all"], ["All"]]} for classification,
                         # it will show the results for all filtered values, under the aggregation name All.
                          {"id": [["all"], ["All"]]},
                          {"id": [["CAVs_modelled_area"],["CAVs in LLR area"]]},
                          # {"id": ["nonCAVs_modelled_area","nonCAVs in LLR area"]},
                          {"id": [["allVehs_modelled_area"],["All vehs. in LLR area"]]},
                          {"id": [["allEdges_modelled_area"], ["All Edges in LLR area"]]},
                          {"devices": [["routing"], ["CAVs whole net."]]},
                          {"type": [["probe1","probe2","probe3","probe4","probe5"],
                                    ["Probe 1","Probe 2","Probe 3","Probe 4","Probe 5"]],
                           "id": [["CAVs_modelled_area"],["CAVs in LLR area"]]},
                          {"id": [["CAVs_modelled_area","nonCAVs_modelled_area","allVehs_modelled_area"],
                                  ["CAVs in LLR area","nonCAVs in LLR area","all vehs. in LLR area"]],
                           "vType/1": [["probe1","probe2","probe3","probe4","probe5"],
                                       ["probe 1","probe 2","probe 3","probe 4","probe 5"]],
                           "vType/2": [["probe"], ["probes"]]}
                          ]

fltr_classif_objs_files = [# Target Objects to read after after the scenario runs (read its outputs)
                        # If file to be read is not an output of any scenario, put on the first scenario index.
                        # This file will be read at the same time of the output of the first scenario
                        # requires:
                        # key of each filter/classification object index and the index of the its list
                        # name of the file with its file format for the key of fltr_classif_objs
                        # 1 if when running this scenario the file should be read (but every replication)
                        # 0 if when running this scenario the file should be read (but only first replication)
                        # Otherwise use None
                        # Example:
                        # ({"fltr_classif_objs key": list with fltr_classif_objs indexes and its list indexes}
                        # "filename",
                        # bool if append scenario name on the filename,
                        # "file_extension in this format")
                        ({"id": [[1,0],[5,0],[6,0]]}, "CAVs_modelled_area", 1, ".xlsx",
                         [1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1]),
                        ({"id": [[6,1]]}, "nonCAVs_modelled_area", 1, ".xlsx",
                         [1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1]),
                        ({"id": [[2,0],[6,2]]}, "allVehs_modelled_area", 1, ".xlsx",
                         [1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1]),
                        ({"id": [[3,0]]}, "allEdges_modelled_area.mid", 0, ".xlsx",
                        [0,None,None,None,None,None,None,None,None,None,None,None,None,None,None])
                        ]

measures = {# dictionary with key: tuple in which 1 if the key and 2 is the tuple.
            # 1) output function,output file name, tree root name
            # 1.a) Name of tree output format, i.e. the functions of this script. Options: Std, Vehroute, Timestep
            # 1.b) SUMO output file name. It will read files changing <inputs> from scenario file to <file_output_name>
            # and xml instead of txt.
                # Ex: if inputs.mid.TURNS.txt > tripinfo.mid.TURNS.xml
                # Some options: vehroute, edge_traffic, edge_emission,
            # 1.c) Root name of the tree from SUMO. Usually tripinfo, vehicle or edge/lane
            # 2) ((Measures inputs),scenarios that uses the measures, and (in case of VehRoute function) index to
            # the scenario to compare
                # where measures inputs is 2.a, scenarios that uses is 2.b, and scenario to compare is 2.c
                # 2.a # measures inputs =
                    #("python function,name of this function for figure and instance naming",
                       # Ex of functions: np.nanmean, np.nansum, len (for counting vehs), np.nanmax,...,
                       # Name is compulsory and it must be in one word but can contain underscore <_>
                    #"tree of elements until the desired attribute,attribute,name of this attribute for figure and
                    # instance naming",
                       # attribute is the attributes to get the measure from XML output. Ex: id, traveltime, etc.
                       # Tree of elements may contain several elements, use comma between each
                       # Name is optional and it must be in one word but can contain underscore <_>, even if no name
                       # the comma must appear anyway
                    #"measure name",
                    #attributes to filter the results,
                    #attributes to classify the results)
                # 2.b # scenarios that use at least one of these measures. Ex.: [0] + (len(scenarios) - 2) * [1] + [0],
                # 2.c # index to the scenario should be compared with. Ex.: [None, 2, None, None])),
                 "Vehroute,vehroute,vehicle": ((
                                      # Warm-up Time here based on arrivals before end of warm-up time
                                      ("np.nanmean,mean", "percDiffEdg,", "Avg. Different Edges [%]",
                                       fltr_classif_objs[1], fltr_classif_objs[5], (len(scenarios) - 1) * [1] + [0]),
                                      ("np.nansum,sum", "percVehDiffRoute,", "Avg. Vehicles Diff. Route [%]",
                                       fltr_classif_objs[1],fltr_classif_objs[5], (len(scenarios) - 1) * [1] + [0])),
                                      (# scenarios that use at least one of these measures
                                       (len(scenarios) - 1) * [1] + [0],
                                       # index to the scenario should be compared with
                                       [1, None, 3, None, 5, None, 7, None, 9, None, 11, None, 13, None, None])),
                 "Std,tripinfo,tripinfo": ((
                                   # Warm-up Time here based on arrivals before end of warm-up time
                                   ("np.nanmean,mean", "duration,", "Mean Trip Duration [s]",
                                    fltr_classif_objs[2], fltr_classif_objs[6], len(scenarios) * [1]),
                                   ("np.nanmean,mean", "routeLength,", "Mean Trip Length [m]",
                                    fltr_classif_objs[2], fltr_classif_objs[6], len(scenarios) * [1]),
                                   ("np.nanmean,mean", "div__routeLength__duration,", "Mean Trip Speed [m]",
                                    fltr_classif_objs[2], fltr_classif_objs[6], len(scenarios) * [1]),
                                   ("np.nanmean,mean", "waitingTime,", "Mean Trip Waiting Time [s]",
                                    fltr_classif_objs[2], fltr_classif_objs[6], len(scenarios) * [1]),
                                   ("np.nanmean,mean", "rerouteNo,", "Avg. Rerouting Times [#]",
                                    fltr_classif_objs[2], fltr_classif_objs[6], len(scenarios) * [1]),
                                   ("np.nanmean,mean", "emissions,CO2_abs,", "Mean Trip Carbon Dioxide (CO2) [mg]",
                                    fltr_classif_objs[2], fltr_classif_objs[6], len(scenarios) * [1]),
                                   ("np.nanmean,mean", "emissions,CO_abs,", "Mean Trip Carbon Monoxide (CO) [mg]",
                                    fltr_classif_objs[2], fltr_classif_objs[6], len(scenarios) * [1]),
                                   ("np.nanmean,mean", "emissions,HC_abs,", "Mean Trip Hydrocarbon (HC) [mg]",
                                    fltr_classif_objs[2], fltr_classif_objs[6], len(scenarios) * [1]),
                                   ("np.nanmean,mean", "emissions,NOx_abs,", "Mean Trip Nitrogen Oxides (NOx) [mg]",
                                    fltr_classif_objs[2], fltr_classif_objs[6], len(scenarios) * [1]),
                                   ("np.nanmean,mean", "emissions,PMx_abs,", "Mean Trip Particulate Matter (PMx) [mg]",
                                    fltr_classif_objs[2], fltr_classif_objs[6], len(scenarios) * [1]),
                                   ("np.nanmean,mean", "emissions,fuel_abs,", "Mean Trip Fuel Consumption [ml]",
                                    fltr_classif_objs[2], fltr_classif_objs[6], len(scenarios) * [1]),
                                   ("len,number", "id,LLRrouted", "LLR CAVs in Managed Area [#]",
                                    fltr_classif_objs[1], fltr_classif_objs[1], (len(scenarios) - 1) * [1] + [0]),
                                   ("len,number", "id,CAVs", "CAVs in Whole Net. [#]",
                                    fltr_classif_objs[4], fltr_classif_objs[4], (len(scenarios) - 1) * [1] + [0])),
                                    len(scenarios) * [1]),
                 "Std,python_edge_traffic,edge": ((
                               # Warm-up Time included already in the travel time reliability report file
                               ("np.nansum,sum", "measuredTravelTime,", "Sum Mean Measured Travel Times [s]",
                                fltr_classif_objs[3], fltr_classif_objs[3], len(scenarios) * [1]),
                               ("np.nansum,sum", "estimatedTravelTime,", "Sum Mean Estimated Travel Times [s]",
                                fltr_classif_objs[3], fltr_classif_objs[3], (len(scenarios) - 1) * [1] + [0]),
                               ("np.nansum,sum", "RMSE,", "Sum Mean RMSE [s]",
                                fltr_classif_objs[3], fltr_classif_objs[3], (len(scenarios) - 1) * [1] + [0]),
                               ("np.nansum,sum", "norm__RMSE__measuredTravelTime,", "Normalized Mean RMSE [%]",
                                fltr_classif_objs[3], fltr_classif_objs[3], (len(scenarios) - 1) * [1] + [0])),
                                len(scenarios) * [1]),
                 "Std,edge_traffic,edge": ((
                               # Warm-up Time included already in the addition file
                               ("np.nansum,sum", "waitingTime,", "Sum Waited Time [s]",
                                fltr_classif_objs[3], fltr_classif_objs[3], len(scenarios) * [1]),
                               # ("np.nansum,sum", "traveltime,", "Sum Mean Travel Time [s]",
                               # fltr_classif_objs[3], fltr_classif_objs[0], len(scenarios) * [1]),
                               ("np.nanmean,mean", "speed,", "Mean Speed [m/s]",
                                fltr_classif_objs[3], fltr_classif_objs[3], len(scenarios) * [1]),
                               ("np.nanmean,mean", "density,", "Mean Density [veh/km]",
                                fltr_classif_objs[3], fltr_classif_objs[3], len(scenarios) * [1]),
                               ("np.nanmean,mean", "occupancy,", "Mean Occupancy [%]",
                                fltr_classif_objs[3], fltr_classif_objs[3], len(scenarios) * [1]),
                               ("np.nanmax,max", "occupancy,", "Max Mean Occupancy [%]",
                                fltr_classif_objs[3], fltr_classif_objs[3], len(scenarios) * [1])),
                                len(scenarios) * [1]),
                 "Std,edge_emission,edge": ((
                               # Warm-up Time included already in the addition file
                               ("np.nansum,sum", "CO2_perVeh,", "Sum Carbon Dioxide (CO2) per Veh. [mg]",
                                fltr_classif_objs[3], fltr_classif_objs[3], len(scenarios) * [1]),
                               ("np.nansum,sum", "CO_perVeh,", "Sum Carbon Monoxide (CO) per Veh. [mg]",
                                fltr_classif_objs[3], fltr_classif_objs[3], len(scenarios) * [1]),
                               ("np.nansum,sum", "HC_perVeh,", "Sum Hydrocarbon (HC) per Veh. [mg]",
                                fltr_classif_objs[3], fltr_classif_objs[3], len(scenarios) * [1]),
                               ("np.nansum,sum", "NOx_perVeh,", "Sum Nitrogen Oxides (NOx) per Veh. [mg]",
                                fltr_classif_objs[3], fltr_classif_objs[3], len(scenarios) * [1]),
                               ("np.nansum,sum", "PMx_perVeh,", "Sum Particulate Matter (PMx) per Veh. [mg]",
                                fltr_classif_objs[3], fltr_classif_objs[3], len(scenarios) * [1]),
                               ("np.nansum,sum", "fuel_perVeh,", "Sum Fuel per Veh. [ml]",
                                fltr_classif_objs[3], fltr_classif_objs[3], len(scenarios) * [1])),
                                len(scenarios) * [1])
                                }

## FUNCTIONS

def getExcelTargetList(filename):
    """Get Values of Each Row of the First Collum of an Excel file"""

    book = openpyxl.load_workbook(filename, data_only=True)
    sheet = book.active
    rows = sheet.rows
    file_input_vals = []
    for row in rows:
        for cell in row:
            file_input_vals.append(cell.value)

    return file_input_vals


def getVehrouteTreeValues(replication, sce, func_name, tree_results, tree_inputs):
    """Get last route of vehicles from SUMO which fulfill the filter conditions
    and define vehicle's aggregation (classification)"""

    global vehRoutes, measure_aggr_vehs
    input_measures = tree_inputs[0]
    sce_that_uses = tree_inputs[1][0]
    root_name = func_name.split(",")[2]
    if "vehRoutes" not in globals():
        # Initialize dictionaries
        vehRoutes = {sce_ind: {measure_ind: dict()
                               for measure_ind,(_,_,_,_,_,sce_uses_measure) in enumerate(input_measures)
                               if sce_uses_measure[sce_ind] == 1}
                     for sce_ind in range(0, len(sce_that_uses))
                     if sce_that_uses[sce_ind] == 1}

        measure_aggr_vehs = {sce_ind: {measure_ind: {aggr_ind: []
                                                     for aggr_ind,_ in enumerate(aggregations[func_name][measure_ind])}
                                       for measure_ind,(_,_,_,_,_,sce_uses_measure) in enumerate(input_measures)
                                       if sce_uses_measure[sce_ind] == 1}
                             for sce_ind in range(0, len(sce_that_uses))
                             if sce_that_uses[sce_ind] == 1}

    # Get last route
    root_children = list(tree_results.iter(root_name))
    for root_child in root_children:
        route_children = list(root_child.findall("route"))
        if route_children == []:
            routeDistribution_child = root_child.findall("routeDistribution")[0]
            route_children = routeDistribution_child.findall("route")
        last_route = set(route_children[-1].get("edges").split())
        
        # If vehicle departure is after the begin time + warm up time, check if vehicle fulfils the filter conditions
        # and define the aggregations (classification) the vehicle belongs
        if float(root_child.get("depart")) >= begin_t_consider:
            for measure_ind,measure_inputs in enumerate(input_measures):
                sce_uses_measure = measure_inputs[5]
                if sce_uses_measure[sce] == 1:
                    filter_objs = measure_inputs[3]
                    broke_filter = 0
                    for filter_key in filter_objs.keys():
                        # For each filter condition
                        filter_key_list = filter_key.split(",")
                        filter_attrb = filter_key_list[-1].split("/")[0]
                        filter_info = filter_objs[filter_attrb]  # include second part with name
                        filter_vals = filter_info[0]
                        filter_child = root_child.copy()
                        filter_tree_elems = filter_key_list[0:-1]
                        for elem in filter_tree_elems:
                            filter_child = list(filter_child.findall(elem))[0]
                        filter_child_val = filter_child.get(filter_attrb)
                        for filter_val in filter_vals:
                            if filter_child_val == filter_val or filter_val == "all":
                                pass # fulfiled the filter condition
                            elif (type(filter_val) == list or type(filter_val) == tuple) \
                            and filter_child_val in filter_val:
                                pass # fulfiled the filter condition
                            elif type(filter_child_val) == str and type(filter_val) == str\
                            and filter_val in filter_child_val:
                                pass # fulfiled the filter condition
                            else:
                                broke_filter = 1
                                break # didn't fulfil the filter condition
                        if broke_filter == 1:
                            break # don't continue to next filter as last one was not fulfiled
                    else:
                        # If vehicle fulfils all filter conditions, define aggregations it belongs
                        aggrs_veh_belongs = []
                        classif_child = root_child.copy()
                        classif_objs = measure_inputs[4]
                        aggr_ind = 0
                        for classif_key in classif_objs.keys():
                            # For each possible aggregation
                            classif_key_list = classif_key.split(",")
                            classif_attrb = classif_key_list[-1].split("/")[0]
                            classif_tree_elems = classif_key_list[0:-1]
                            for elem in classif_tree_elems:
                                classif_child = list(classif_child.findall(elem))[0]
                            classif_child_val = classif_child.get(classif_attrb)
                            aggr_vals = classif_objs[classif_key][0]
                            for aggr in aggr_vals:
                                if classif_child_val == aggr or aggr == "all":
                                    pass # fulfiled the aggregation/classification condition
                                elif (type(aggr) == list or type(aggr) == tuple) and classif_child_val in aggr:
                                    pass # fulfiled the aggregation/classification condition
                                elif type(classif_child_val) == str and type(aggr) == str and aggr in classif_child_val:
                                    pass # fulfiled the aggregation/classification condition
                                else:
                                    aggr_ind += 1
                                    continue # didn't fulfil the aggregation/classification condition
                                aggrs_veh_belongs.append(aggr_ind)
                                measure_aggr_vehs[sce][measure_ind][aggr_ind].append(root_child.get("id"))
                                aggr_ind += 1

                        vehRoutes[sce][measure_ind].update({root_child.get("id"): (aggrs_veh_belongs, last_route)})


def getStdTreeValues(replication, sce, func_name, tree_results, tree_inputs):
    """Get the desired values of certain xml entry that fulfill the filter conditions
        and define vehicle's aggregation (classification)"""

    input_measures = tree_inputs[0]
    root_name = func_name.split(",")[2]

    # Initialize list of values for each desired measure/metric
    measure_aggr_list.update({func_name: dict()})
    for measure_ind,measure_inputs in enumerate(input_measures):
        sce_uses_measure = measure_inputs[5]
        if sce_uses_measure[sce] == 1:
            (python_func, python_func_name) = measure_inputs[0].split(",")
            measure_key = measure_inputs[1]
            measure_key_list = measure_key.split(",")
            measure_attrb = measure_key_list[-2]
            measure_attrb_name = measure_key_list[-1]
            if measure_attrb_name == '':
                measure_attrb_name = measure_attrb
            if measure_aggr_list[func_name].has_key(measure_attrb_name) == False:
                measure_aggr_list[func_name].update({measure_attrb_name: dict()})
            measure_aggr_list[func_name][measure_attrb_name].update({python_func_name: [[]
                                                                     for _ in range(0, len(aggregations[func_name]
                                                                                           [measure_ind]))]})

    root_children = list(tree_results.iter(root_name))
    for child_ind,root_child in enumerate(root_children):
        # If output is the vehicle's tripInfo from SUMO check if vehicle's departure is after
        # the begin time + warm up time or if not, ignore the departure time
        if (func_name == "tripinfo" and float(root_child.get("depart")) >= begin_t_consider) or func_name != "tripinfo":
            # Check if vehicle fulfils the filter conditions
            # and define the aggregations (classification) the vehicle belongs
            for measure_ind,measure_inputs in enumerate(input_measures):
                sce_uses_measure = measure_inputs[5]
                if sce_uses_measure[sce] == 1:
                    filter_objs = measure_inputs[3]
                    broke_filter = 0
                    for filter_key in filter_objs.keys():
                        filter_key_list = filter_key.split(",")
                        filter_attrb = filter_key_list[-1].split("/")[0]
                        filter_info = filter_objs[filter_attrb]  # include second part with name
                        filter_vals = filter_info[0]
                        filter_child = root_child.copy()
                        filter_tree_elems = filter_key_list[0:-1]
                        for elem in filter_tree_elems:
                            filter_child = list(filter_child.findall(elem))[0]
                        filter_child_val = filter_child.get(filter_attrb)
                        for filter_val in filter_vals:
                            if filter_child_val == filter_val or filter_val == "all":
                                pass # fulfiled the filter condition
                            elif (type(filter_val) == list or type(filter_val) == tuple) \
                            and filter_child_val in filter_val:
                                pass # fulfiled the filter condition
                            elif type(filter_child_val) == str and type(filter_val) == str \
                            and filter_val in filter_child_val:
                                pass # fulfiled the filter condition
                            else:
                                broke_filter = 1
                                break # didn't fulfil the filter condition
                        if broke_filter == 1:
                            break # don't continue to next filter as last one was not fulfiled
                    else:
                        # If passed all filters, define which aggregation/classification it belongs to
                        classif_child = root_child.copy()
                        classif_objs = measure_inputs[4]
                        aggr_ind = 0
                        for classif_key in classif_objs.keys():
                            classif_key_list = classif_key.split(",")
                            classif_attrb = classif_key_list[-1].split("/")[0]
                            classif_tree_elems = classif_key_list[0:-1]
                            for elem in classif_tree_elems:
                                classif_child = list(classif_child.findall(elem))[0]
                            classif_child_val = classif_child.get(classif_attrb)
                            (python_func, python_func_name) = measure_inputs[0].split(",")
                            measure_child = root_child.copy()
                            measure_key = measure_inputs[1]
                            measure_key_list = measure_key.split(",")
                            measure_attrb = measure_key_list[-2]
                            measure_attrb_name = measure_key_list[-1]
                            measure_tree_elems = measure_key_list[0:-2]
                            if measure_attrb_name == '':
                                measure_attrb_name = measure_attrb
                            for elem in measure_tree_elems:
                                measure_child = list(measure_child.findall(elem))[0]

                            if "div__" in measure_attrb:
                                # If div__ in the measure name, divide each value directly
                                # (division)
                                nominator, denominator = measure_attrb.split("__")[1:]
                                try:
                                    nominator_val = float(measure_child.get(nominator))
                                    denominator_val = float(measure_child.get(denominator))
                                    measure_val = round(nominator_val / denominator_val, 1)
                                except TypeError:
                                    noData = 1
                            elif "norm__" in measure_attrb:
                                # If norm__ in the measure name, store each value separately for dividing later
                                # (normalization)
                                nominator,denominator = measure_attrb.split("__")[1:]
                                try:
                                    nominator_val = float(measure_child.get(nominator))
                                    denominator_val = float(measure_child.get(denominator))
                                    measure_val = (nominator_val, denominator_val)
                                except TypeError:
                                    noData = 1
                            else:
                                # Normal value to get data
                                measure_val = measure_child.get(measure_attrb)

                            aggr_vals = classif_objs[classif_key][0]
                            for aggr in aggr_vals:
                                if classif_child_val == aggr or aggr == "all":
                                    pass # fulfiled the aggregation/classification condition
                                elif (type(aggr) == list or type(aggr) == tuple) and classif_child_val in aggr:
                                    pass # fulfiled the aggregation/classification condition
                                elif type(classif_child_val) == str and type(aggr) == str and aggr in classif_child_val:
                                    pass # fulfiled the aggregation/classification condition
                                else:
                                    aggr_ind += 1
                                    continue # didn't fulfil the aggregation/classification condition
                                if measure_val != None and (python_func == "len" or "norm__" in measure_attrb):
                                    # When not a number append the raw value in the list of values of the desired measure
                                    measure_aggr_list[func_name][measure_attrb_name][python_func_name] \
                                                     [aggr_ind].append(measure_val)
                                else:
                                    try:
                                        # When a number append the float value in the list of values of the desired measure
                                        float_measure_val = float(measure_val)
                                        measure_aggr_list[func_name][measure_attrb_name][python_func_name] \
                                                         [aggr_ind].append(float_measure_val)
                                    except TypeError:
                                        noData = 1
                                aggr_ind += 1

    for measure_ind,measure_inputs in enumerate(input_measures):
        sce_uses_measure = measure_inputs[5]
        if sce_uses_measure[sce] == 1:
            (python_func, python_func_name) = measure_inputs[0].split(",")
            measure_key = measure_inputs[1]
            measure_key_list = measure_key.split(",")
            measure_attrb = measure_key_list[-2]
            measure_attrb_name = measure_key_list[-1]
            if measure_attrb_name == '':
                measure_attrb_name = measure_attrb
            for aggr_ind, _ in enumerate(aggregations[func_name][measure_ind]):
                # Store the result of the python function for the scenario for each aggregation
                try:
                    if "norm__" in measure_attrb:
                        # In case of normalization unzip the nominators and denominators of each values of the list
                        norm_list = list(zip(*measure_aggr_list[func_name][measure_attrb_name][python_func_name]
                                                               [aggr_ind]))
                        try:
                            # Divide the result of the python function
                            if python_func == "np.nansum":
                                results_sce_replication[func_name][measure_attrb_name][python_func_name] \
                                                       [aggr_ind][replication, sce, 0] = \
                                    round(np.nansum(norm_list[0]) / np.nansum(norm_list[1]), 3) * 100
                            elif python_func == "np.nanmean":
                                results_sce_replication[func_name][measure_attrb_name][python_func_name] \
                                                       [aggr_ind][replication, sce, 0] = \
                                    round(np.nanmean(norm_list[0]) / np.nanmean(norm_list[1]), 3) * 100
                            elif python_func == "len":
                                results_sce_replication[func_name][measure_attrb_name][python_func_name] \
                                                       [aggr_ind][replication, sce, 0] = \
                                    round(len(norm_list[0]) / len(norm_list[1]), 3) * 100
                            elif python_func == "np.nanmax":
                                results_sce_replication[func_name][measure_attrb_name][python_func_name] \
                                                       [aggr_ind][replication, sce, 0] = \
                                    round(np.nanmax(norm_list[0]) / np.nanmax(norm_list[1]), 3) * 100
                            elif python_func == "np.nanmin":
                                results_sce_replication[func_name][measure_attrb_name][python_func_name] \
                                                       [aggr_ind][replication, sce, 0] = \
                                    round(np.nanmin(norm_list[0]) / np.nanmin(norm_list[1]), 3) * 100
                            else:
                                warnings.warn("Not expected Python function \"" + python_func + "\" please check")
                                raise sys.exit()

                        except IndexError:
                            # When empty list use np.nan
                            results_sce_replication[func_name][measure_attrb_name][python_func_name] \
                                                   [aggr_ind][replication, sce, 0] = np.nan
                    else:
                        if measure_aggr_list[func_name][measure_attrb_name][python_func_name][aggr_ind] == []:
                            # When empty list use np.nan
                            results_sce_replication[func_name][measure_attrb_name][python_func_name] \
                                                   [aggr_ind][replication, sce, 0] = np.nan
                        else:
                            # Divide the result of the python function
                            if python_func == "np.nansum":
                                results_sce_replication[func_name][measure_attrb_name][python_func_name] \
                                                       [aggr_ind][replication, sce, 0] = \
                                    np.nansum(measure_aggr_list[func_name][measure_attrb_name][python_func_name]
                                                               [aggr_ind])
                            elif python_func == "np.nanmean":
                                results_sce_replication[func_name][measure_attrb_name][python_func_name] \
                                                       [aggr_ind][replication, sce, 0] = \
                                    np.nanmean(measure_aggr_list[func_name][measure_attrb_name][python_func_name]
                                               [aggr_ind])
                            elif python_func == "len":
                                results_sce_replication[func_name][measure_attrb_name][python_func_name] \
                                                       [aggr_ind][replication, sce, 0] = \
                                    len(measure_aggr_list[func_name][measure_attrb_name][python_func_name][aggr_ind])
                            elif python_func == "np.nanmax":
                                results_sce_replication[func_name][measure_attrb_name][python_func_name] \
                                                       [aggr_ind][replication, sce, 0] = \
                                    np.nanmax(measure_aggr_list[func_name][measure_attrb_name][python_func_name]
                                              [aggr_ind])
                            elif python_func == "np.nanmin":
                                results_sce_replication[func_name][measure_attrb_name][python_func_name] \
                                                       [aggr_ind][replication, sce, 0] = \
                                    np.nanmin(measure_aggr_list[func_name][measure_attrb_name][python_func_name]
                                              [aggr_ind])
                            else:
                                warnings.warn("Not expected Python function \"" + python_func + "\" please check")
                                raise sys.exit()

                except (ValueError, TypeError):
                    noData = 1


def compareVehroute(replication, sce, func_name, tree_inputs):
    """Compare the last route of vehicles from the function getVehrouteTreeValues"""

    input_measures = tree_inputs[0]
    sce_to_comp = tree_inputs[1][1]

    # Initialize list of values for each desired measure/metric
    measure_aggr_list.update({func_name: dict()})
    for measure_ind, measure_inputs in enumerate(input_measures):
        sce_uses_measure = measure_inputs[5]
        if sce_uses_measure[sce] == 1 and sce_to_comp[sce] != None:
            (python_func, python_func_name) = measure_inputs[0].split(",")
            measure_key = measure_inputs[1]
            measure_key_list = measure_key.split(",")
            measure_attrb = measure_key_list[-2]
            measure_attrb_name = measure_key_list[-1]
            if measure_attrb_name == '':
                measure_attrb_name = measure_attrb
            if measure_aggr_list[func_name].has_key(measure_attrb_name) == False:
                measure_aggr_list[func_name].update({measure_attrb_name: dict()})
            measure_aggr_list[func_name][measure_attrb_name].update({python_func_name: [[]
                                                                     for _ in range(0, len(aggregations[func_name]
                                                                                           [measure_ind]))]})

    for measure_ind,measure_inputs in enumerate(input_measures):
        sce_uses_measure = measure_inputs[5]
        if sce_uses_measure[sce] == 1 and sce_to_comp[sce] != None:
            (python_func, python_func_name) = measure_inputs[0].split(",")
            measure_key = measure_inputs[1]
            measure_key_list = measure_key.split(",")
            measure_attrb = measure_key_list[-2]
            measure_attrb_name = measure_key_list[-1]
            if measure_attrb_name == '':
                measure_attrb_name = measure_attrb
            for vehID in vehRoutes[sce][measure_ind]:
                aggrs_veh_belongs = vehRoutes[sce][measure_ind][vehID][0]
                try:
                    diff_edges = vehRoutes[sce_to_comp[sce]][measure_ind][vehID][1].difference(
                                                                               vehRoutes[sce][measure_ind][vehID][1])
                    if measure_attrb == "percDiffEdg":
                        # If desired measure is percentage of different edges
                        result = (float(len(diff_edges)) / len(vehRoutes[sce_to_comp[sce]][measure_ind][vehID][1])) * 100
                        for aggr_ind in aggrs_veh_belongs:
                            measure_aggr_list[func_name][measure_attrb_name][python_func_name][aggr_ind].append(result)

                    elif measure_attrb == "percVehDiffRoute":
                        # If desired measure is percentage of vehicles with different route
                        if len(diff_edges) > 0:
                            for aggr_ind in aggrs_veh_belongs:
                                result = (1.0 / len(measure_aggr_vehs[sce][measure_ind][aggr_ind])) * 100
                                measure_aggr_list[func_name][measure_attrb_name][python_func_name] \
                                                 [aggr_ind].append(result)
                        else:
                            for aggr_ind in aggrs_veh_belongs:
                                measure_aggr_list[func_name][measure_attrb_name][python_func_name][aggr_ind].append(0)
                except KeyError:
                    # Use np.nan if not found route of same vehicle in one of the comparing scenarios
                    for aggr_ind in aggrs_veh_belongs:
                        measure_aggr_list[func_name][measure_attrb_name][python_func_name][aggr_ind].append(np.nan)

    for measure_ind, measure_inputs in enumerate(input_measures):
        # Store the result of the python function for the scenario for each aggregation
        sce_uses_measure = measure_inputs[5]
        if sce_uses_measure[sce] == 1 and sce_to_comp[sce] != None:
            (python_func, python_func_name) = measure_inputs[0].split(",")
            measure_key = measure_inputs[1]
            measure_key_list = measure_key.split(",")
            measure_attrb = measure_key_list[-2]
            measure_attrb_name = measure_key_list[-1]
            if measure_attrb_name == '':
                measure_attrb_name = measure_attrb
            for aggr_ind, _ in enumerate(aggregations[func_name][measure_ind]):
                try:
                    if measure_aggr_list[func_name][measure_attrb_name][python_func_name][aggr_ind] == []:
                        # When empty list use np.nan
                        results_sce_replication[func_name][measure_attrb_name][python_func_name] \
                                               [aggr_ind][replication, sce, 0] = np.nan
                    else:
                        # Divide the result of the python function
                        if python_func == "np.nansum":
                            results_sce_replication[func_name][measure_attrb_name][python_func_name] \
                                                   [aggr_ind][replication, sce, 0] = \
                                np.nansum(measure_aggr_list[func_name][measure_attrb_name][python_func_name][aggr_ind])
                        elif python_func == "np.nanmean":
                            results_sce_replication[func_name][measure_attrb_name][python_func_name] \
                                                   [aggr_ind][replication, sce, 0] = \
                                np.nanmean(measure_aggr_list[func_name][measure_attrb_name][python_func_name][aggr_ind])
                        elif python_func == "len":
                            results_sce_replication[func_name][measure_attrb_name][python_func_name] \
                                                   [aggr_ind][replication, sce, 0] = \
                                len(measure_aggr_list[func_name][measure_attrb_name][python_func_name][aggr_ind])
                        elif python_func == "np.nanmax":
                            results_sce_replication[func_name][measure_attrb_name][python_func_name] \
                                                   [aggr_ind][replication, sce, 0] = \
                                np.nanmax(measure_aggr_list[func_name][measure_attrb_name][python_func_name][aggr_ind])
                        elif python_func == "np.nanmin":
                            results_sce_replication[func_name][measure_attrb_name][python_func_name] \
                                                   [aggr_ind][replication, sce, 0] = \
                                np.nanmin(measure_aggr_list[func_name][measure_attrb_name][python_func_name][aggr_ind])
                        else:
                            warnings.warn("Not expected Python function \"" + python_func + "\" please check")
                            raise sys.exit()

                except (ValueError, TypeError):
                    noData = 1


def compareScenarios(replication, sce, func_name, tree_inputs):
    """Compare the values of each meeasure for each aggregation between two scenarios"""

    input_measures = tree_inputs[0]
    for measure_ind, measure_inputs in enumerate(input_measures):
        sce_uses_measure = measure_inputs[5]
        if sce_uses_measure[sce] == 1:
            (python_func, python_func_name) = measure_inputs[0].split(",")
            measure_key = measure_inputs[1]
            measure_key_list = measure_key.split(",")
            measure_attrb = measure_key_list[-2]
            measure_attrb_name = measure_key_list[-1]
            if measure_attrb_name == '':
                measure_attrb_name = measure_attrb
            for aggr_ind, _ in enumerate(aggregations[func_name][measure_ind]):
                # Check if base scenario measure value is zero
                isBase_zero = results_sce_replication[func_name][measure_attrb_name][python_func_name] \
                                                     [aggr_ind][replication, base_scenario_ind[sce], 0] == 0
                # Check if comparing scenario measure value is zero
                isComparing_zero = results_sce_replication[func_name][measure_attrb_name][python_func_name] \
                                                          [aggr_ind][replication, sce, 0] == 0
                # If base or comparing is zero use different formulas to avoid wrong percentage difference comparison
                if isBase_zero:
                    if isComparing_zero:
                        results_sce_replication[func_name][measure_attrb_name][python_func_name] \
                                               [aggr_ind][replication, sce, 1] = 0
                    else:
                        isComparing_positive = results_sce_replication[func_name][measure_attrb_name] \
                                                                      [python_func_name][aggr_ind] \
                                                                      [replication, sce, 0] > 0
                        if isComparing_positive:
                            results_sce_replication[func_name][measure_attrb_name][python_func_name] \
                                                   [aggr_ind][replication, sce, 1] = 1
                        else:
                            results_sce_replication[func_name][measure_attrb_name][python_func_name] \
                                                   [aggr_ind][replication, sce, 1] = -1
                else:
                    results_sce_replication[func_name][measure_attrb_name][python_func_name] \
                                           [aggr_ind][replication, sce, 1] = float(
                    results_sce_replication[func_name][measure_attrb_name][python_func_name]
                                           [aggr_ind][replication, sce, 0] -
                    results_sce_replication[func_name][measure_attrb_name][python_func_name]
                                           [aggr_ind][replication, base_scenario_ind[sce], 0]) / \
                    results_sce_replication[func_name][measure_attrb_name][python_func_name] \
                                           [aggr_ind][replication, base_scenario_ind[sce], 0]


## INITIALIZE FILES

# Row is the replication
# Column is the scenario
# Values are the tuple (average, deviation)
num_scenarios = len(scenarios)
aggregations = dict()
results_sce_replication = dict()
for func_name,func_inputs in measures.items():
    input_measures = func_inputs[0]
    results_sce_replication.update({func_name: dict()})
    aggregations.update({func_name: []})
    for measure_ind,measure_inputs in enumerate(input_measures):
        (python_func, python_func_name) = measure_inputs[0].split(",")
        aggregations[func_name].append([])
        measure_key = measure_inputs[1]
        measure_key_list = measure_key.split(",")
        measure_attrb = measure_key_list[-2]
        measure_attrb_name = measure_key_list[-1]
        if measure_attrb_name == '':
            measure_attrb_name = measure_attrb
        classif_objs = measure_inputs[4]
        for classif_key in classif_objs.keys():
            aggr_vals,aggr_names = classif_objs[classif_key]
            for aggr_name in aggr_names:
                aggregations[func_name][measure_ind].append(aggr_name)
        # Initialize one matrix for each measure and classification to store the values
        # of each scenario in each replication
        matrix_sce_replication = np.full([max(number_replications), num_scenarios, 2], np.nan)
        if results_sce_replication[func_name].has_key(measure_attrb_name) == False:
            results_sce_replication[func_name].update({measure_attrb_name: dict()})
        results_sce_replication[func_name][measure_attrb_name].update({python_func_name: [matrix_sce_replication.copy()
                                                                       for _ in range(0, len(aggregations[func_name]
                                                                                             [measure_ind]))]})

# Initialize the list with aggregate results for each scenario
conf_int_sce_measure = [[] for sce in scenarios]

# Initialize lists with each measure values for every aggregation
measure_aggr_list = dict()

# Get Enviroment Variables
my_env = os.environ.copy()
my_env["PYTHONPATH"] = "/usr/sbin:/sbin:" + my_env["PYTHONPATH"]

# Build the travel/traffic demand that is the same for all scenarios and all replications
global route_args
try:
    path_gen_routes = os.path.join(path_inputs, gen_routes_file)
    with open(path_gen_routes, 'r') as f:
        genRoutes_inputs = f.read()
    route_args = shlex.split(genRoutes_inputs)
except NameError:
    skip = 1

for replication in range(0, max(number_replications)):

    # Build the travel/traffic demands that will be equal to all scenarios on same replication
    try:
        process = subprocess.Popen(route_args, env=my_env)
        print("\n Generating routes for replication", replication + 1)
        process.wait()
    except (NameError,IndexError):
        skip = 1

    for sce,(sce_name, sce_inputs_file) in enumerate(scenarios):
        if replication < number_replications[sce]:
            try:
                # Build the travel/traffic demands that will be specific for the scenario
                if gen_routes_files[sce] != None:
                    path_gen_routes = os.path.join(path_inputs, gen_routes_files[sce])
                    with open(path_gen_routes, 'r') as f:
                        genRoutes_inputs = f.read()
                    route_args = shlex.split(genRoutes_inputs)
                    process = subprocess.Popen(route_args, env=my_env)
                    print("\n Generating routes of scenario", (sce + 1), "for replication", replication + 1)
                    process.wait()
            except NameError:
                skip = 1

            print("\n Replication", (replication + 1), "for scenario", (sce + 1), "of", num_scenarios)
            path_scenario = os.path.join(path_scenarios,sce_inputs_file)
            # Get scenario inputs text file to be used
            with open(path_scenario, 'r') as f:
                sce_inputs = f.read()
            sce_args = shlex.split(sce_inputs)
            sce_ran_ok = 0

            while sce_ran_ok == 0:
                # Run the Scenario File until it doesn't return an error
                process = subprocess.Popen(sce_args, env=my_env)
                if process.wait() == 0:
                    # Check if scenario ran without error
                    sce_ran_ok = 1

            # Parse the Results of the Output XML Files to input for target objects (filter and classification/aggregation)
            print("Getting results of nonXML files")
            for trg_dict, filename, bool_sce_name, file_formt, sces_to_use in fltr_classif_objs_files:
                if sces_to_use[sce] != None \
                and ((replication == 0 and sces_to_use[sce] >= 0) or sces_to_use[sce] == 1):
                    if bool_sce_name == 1:
                        filename = filename + sce_inputs_file.replace("inputs", "").replace(".txt", file_formt)
                    else:
                        filename = filename + file_formt
                    path_filename = os.path.join(path_outputs, filename)
                    if file_formt == ".xlsx":
                        for trg_key in trg_dict.keys():
                            last_trg_ind = None
                            last_trg_list_ind = None
                            for trg_objs_ind,trg_obj_list_ind in trg_dict[trg_key]:
                                try:
                                    fltr_classif_objs[trg_objs_ind][trg_key][0][trg_obj_list_ind] = \
                                        fltr_classif_objs[last_trg_ind][trg_key][0][last_trg_list_ind]
                                except TypeError:
                                    fltr_classif_objs[trg_objs_ind][trg_key][0][trg_obj_list_ind] = \
                                        getExcelTargetList(path_filename)
                                    last_trg_ind = trg_objs_ind
                                    last_trg_list_ind = trg_obj_list_ind

            # Parse the Results of the Output XML Files for scenarios results
            for func_name in measures.keys():
                (chosen_func,output_name,_) = func_name.split(",")
                if chosen_func == "Vehroute":
                    sce_that_uses = measures[func_name][1][0]
                else:
                    sce_that_uses = measures[func_name][1]
                if sce_that_uses[sce] == 1:
                    out_file = sce_inputs_file.replace("inputs", output_name).replace(".txt", ".xml")
                    path_results = os.path.join(path_outputs, out_file)
                    tree_results = ElementTree.parse(path_results)
                    print("Parsing results of XML file: " + output_name)
                    if chosen_func == "Std":
                        getStdTreeValues(replication, sce, func_name, tree_results, measures[func_name])
                    elif chosen_func == "Vehroute":
                        getVehrouteTreeValues(replication, sce, func_name, tree_results, measures[func_name])
                    else:
                        warnings.warn("Not expected XML parsing function \"" + chosen_func + "\" please check")
                        raise sys.exit()
        else:
            # Use the last replication values in case the replication number is above the number of replications for
            # the scenario
            for func_name, func_inputs in measures.items():
                input_measures = func_inputs[0]
                for measure_ind, measure_inputs in enumerate(input_measures):
                    sce_uses_measure = measure_inputs[5]
                    if sce_uses_measure[sce] == 1:
                        (python_func, python_func_name) = measure_inputs[0].split(",")
                        measure_key = measure_inputs[1]
                        measure_key_list = measure_key.split(",")
                        measure_attrb = measure_key_list[-2]
                        measure_attrb_name = measure_key_list[-1]
                        if measure_attrb_name == '':
                            measure_attrb_name = measure_attrb
                        for aggr_ind, _ in enumerate(aggregations[func_name][measure_ind]):
                            results_sce_replication[func_name][measure_attrb_name][python_func_name] \
                                                   [aggr_ind][replication, sce, :] = \
                            results_sce_replication[func_name][measure_attrb_name][python_func_name] \
                                                   [aggr_ind][replication - 1, sce, :]

    print("\n Completed all simulations of replication", (replication + 1))
    # Compare Routes (if applied)
    print("Comparing routes")
    for sce, _ in enumerate(scenarios):
        for func_name in measures.keys():
            chosen_func = func_name.split(",")[0]
            if chosen_func == "Vehroute":
                sce_that_uses = measures[func_name][1][0]
                sce_to_comp = measures[func_name][1][1]
                if sce_that_uses[sce] == 1 and sce_to_comp[sce] != None:
                    compareVehroute(replication, sce, func_name, measures[func_name])

    # Compare Scenarios (with base cases)
    print("Comparing scenarios")
    for sce, _ in enumerate(scenarios):
        if base_scenario_ind[sce] != None:
            for func_name in measures.keys():
                compareScenarios(replication, sce, func_name, measures[func_name])

    # Output Results into Excel File (First Part - Writing Replication Results)
    print("Outputting replication results into Excel file")
    for func_name,func_inputs in measures.items():
        tab_name = ",".join(func_name.split(",")[:2])
        input_measures = func_inputs[0]
        sce_that_uses = []
        if func_name.split(",")[0] == "Vehroute":
            # Define if a vehicle uses or not certain measure/metric if it compares the route
            # or if it compares against other scenarios
            for sce_val in func_inputs[1][1]:
                sce_that_uses.append((lambda val: 1 if val != None else 0)(sce_val))
        elif func_name.split(",")[0] == "Std":
            # Define if a vehicle uses or not certain measure/metric if it compares against other scenarios
            sce_that_uses = func_inputs[1]

        # Define the new scenarios index with only the ones that use the measure
        new_sce_ind = [0] + [reduce((lambda x, y: x + y), sce_that_uses[0:sce + 1])
                             for sce,_ in enumerate(sce_that_uses[:-1])]
        try:
            # Read the workbook and create a new sheet for the measure
            wb = openpyxl.load_workbook(excel_file)
            try:
                # Get sheet for the function
                sheet = wb[tab_name]
            except KeyError:
                # Create a new sheet for the function
                sheet = wb.create_sheet(tab_name)
        except IOError:
            # Create workbook
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = tab_name

        colunm_cells_sce = 0
        for measure_ind, _ in enumerate(input_measures):
            # one for average and one for % difference
            colunm_cells_sce += len(aggregations[func_name][measure_ind]) * 2
        # one for space between scenarios
        colunm_cells_sce += 1

        for sce, (sce_name, _) in enumerate(scenarios):
            new_sce = new_sce_ind[sce]
            if sce_that_uses[sce] == 1 and replication < number_replications[sce]:
                # Add scenario name in the first row
                sce_begin_column = colunm_cells_sce * new_sce + 1
                if replication == 0:
                    cell = sheet.cell(row=1, column=sce_begin_column)
                    cell.value = sce_name
                    sheet.merge_cells(start_row=1, start_column=sce_begin_column,
                                      end_row=1, end_column=sce_begin_column + (colunm_cells_sce - 2))
                    cell.alignment = openpyxl.styles.Alignment(horizontal='center')
                aggr_begin_column = sce_begin_column
                for measure_ind, measure_inputs in enumerate(input_measures):
                    (python_func, python_func_name) = measure_inputs[0].split(",")
                    sce_uses_measure = measure_inputs[5]
                    measure_key = measure_inputs[1]
                    measure_key_list = measure_key.split(",")
                    measure_attrb = measure_key_list[-2]
                    measure_attrb_name = measure_key_list[-1]
                    if measure_attrb_name == '':
                        measure_attrb_name = measure_attrb
                    measure_name = measure_inputs[2]
                    # Add measure name in the second row
                    num_aggrs = len(aggregations[func_name][measure_ind])
                    if replication == 0:
                        cell = sheet.cell(row=2, column=aggr_begin_column)
                        cell.value = measure_name
                        sheet.merge_cells(start_row=2, start_column=aggr_begin_column,
                                          end_row=2, end_column=aggr_begin_column + num_aggrs * 2 - 1)
                        cell.alignment = openpyxl.styles.Alignment(horizontal='center')
                    for aggr_ind, aggr_name in enumerate(aggregations[func_name][measure_ind]):
                        # Add aggregation name in the third row
                        if replication == 0:
                            cell = sheet.cell(row=3, column=aggr_begin_column)
                            cell.value = aggr_name
                            sheet.merge_cells(start_row=3, start_column=aggr_begin_column,
                                              end_row=3, end_column=aggr_begin_column + 1)
                            cell.alignment = openpyxl.styles.Alignment(horizontal='center')
                            # Add value name in the fourth row
                            cell = sheet.cell(row=4, column=aggr_begin_column)
                            cell.value = 'Average'
                            cell.alignment = openpyxl.styles.Alignment(horizontal='center')
                            cell = sheet.cell(row=4, column=aggr_begin_column + 1)
                            cell.value = '% Difference'
                            cell.alignment = openpyxl.styles.Alignment(horizontal='center')
                        cell = sheet.cell(row=replication + 5, column=aggr_begin_column)
                        if sce_uses_measure[sce] == 1:
                            # Put into excel the average value of measure values for the aggregation in the replication
                            ave = results_sce_replication[func_name][measure_attrb_name][python_func_name] \
                                                         [aggr_ind][replication, sce, 0]
                            if math.isnan(ave):
                                # If no value for the measure, put a dash
                                cell.value = "-"
                            else:
                                cell.value = "{:,.2f}".format(ave)
                        else:
                            # If scenario doesn't use the measure/metric, put a dash
                            cell.value = "-"
                        cell.alignment = openpyxl.styles.Alignment(horizontal='center')
                        if base_scenario_ind[sce] == None or func_name.split(",")[0] == "Vehroute" \
                        or sce_uses_measure[sce] == 0:
                            cell = sheet.cell(row=replication + 5, column=aggr_begin_column + 1)
                            # If scenario doesn't use the measure/metric, put a dash
                            cell.value = "-"
                            cell.alignment = openpyxl.styles.Alignment(horizontal='center')
                        else:
                            # Put into excel the percentage difference value of the measure value for the aggregation
                            # in the replication.
                            # This is the difference between the comparing scenario and against base scenario
                            perc = results_sce_replication[func_name][measure_attrb_name][python_func_name] \
                                                          [aggr_ind][replication, sce, 1]
                            cell = sheet.cell(row=replication + 5, column=aggr_begin_column + 1)
                            if math.isnan(perc):
                                # If no value for the measure, put a dash
                                cell.value = "-"
                            else:
                                cell.value = "{:.2%}".format(perc)
                            cell.alignment = openpyxl.styles.Alignment(horizontal='center')
                        aggr_begin_column += 2

        # Finally, save the file and give it a name
        wb.save(excel_file)

# Output Results into Excel File (Second part - Estimating Confidence Intervals)
print("\n Completed all replications")
print("Outputting aggregated results of all replications into Excel file")
for func_name,func_inputs in measures.items():
    tab_name = ",".join(func_name.split(",")[:2])
    input_measures = func_inputs[0]
    sce_that_uses = []
    max_num_replications = max(number_replications)
    if func_name.split(",")[0] == "Vehroute":
        # Define if a vehicle uses or not certain measure/metric if it compares the route
        # or if it compares against other scenarios
        for sce_val in func_inputs[1][1]:
            # Define if a vehicle uses or not certain measure/metric if it compares against other scenarios
            sce_that_uses.append((lambda val: 1 if val != None else 0)(sce_val))
    elif func_name.split(",")[0] == "Std":
        sce_that_uses = func_inputs[1]
    # Define the new scenarios index with only the ones that use the measure
    new_sce_ind = [0] + [reduce((lambda x, y: x + y), sce_that_uses[0:sce + 1])
                         for sce, _ in enumerate(sce_that_uses[:-1])]

    # Read the workbook and create a new sheet for the measure
    wb = openpyxl.load_workbook(excel_file)
    # Get sheet for the function
    sheet = wb[tab_name]

    colunm_cells_sce = 0
    for measure_ind, _ in enumerate(input_measures):
        # one for average and one for % difference
        colunm_cells_sce += len(aggregations[func_name][measure_ind]) * 2
    # one for space between scenarios
    colunm_cells_sce += 1

    for sce, _ in enumerate(scenarios):
        new_sce = new_sce_ind[sce]
        if sce_that_uses[sce] == 1:
            aggr_begin_column = colunm_cells_sce * new_sce + 1
            for measure_ind, measure_inputs in enumerate(input_measures):
                (python_func, python_func_name) = measure_inputs[0].split(",")
                sce_uses_measure = measure_inputs[5]
                measure_key = measure_inputs[1]
                measure_key_list = measure_key.split(",")
                measure_attrb = measure_key_list[-2]
                measure_attrb_name = measure_key_list[-1]
                if measure_attrb_name == '':
                    measure_attrb_name = measure_attrb
                for aggr_ind, _ in enumerate(aggregations[func_name][measure_ind]):
                    # Estimate Confidence Intervals
                    if sce_uses_measure[sce] == 1:
                        # Calculate the average of the average values of the measure for all replications
                        # of each aggregation
                        ave_mean = np.nanmean(results_sce_replication[func_name][measure_attrb_name]
                                                                     [python_func_name][aggr_ind][:, sce, 0])
                        # Calculate the average of the standard deviation values of the measure for all replications
                        # of each aggregation
                        ave_std = np.nanstd(results_sce_replication[func_name][measure_attrb_name]
                                                                   [python_func_name][aggr_ind][:, sce, 0])
                        if base_scenario_ind[sce] == None or func_name.split(",")[0] == "Vehroute" \
                        or sce_uses_measure[sce] == 0:
                            pass # scenario doesn't use the measure,
                                 # or it is the base case when comparing route and/or scenarios
                        else:
                            perc_mean = np.nanmean(results_sce_replication[func_name][measure_attrb_name]
                                                                          [python_func_name][aggr_ind][:, sce, 1])
                            perc_std = np.nanstd(results_sce_replication[func_name][measure_attrb_name]
                                                                        [python_func_name][aggr_ind][:, sce, 1])
                    else:
                        # If no value, store np.nan
                        perc_mean = np.nan
                        ave_mean =  np.nan

                    cell = sheet.cell(row=max_num_replications + 5 + 1, column=aggr_begin_column)
                    if sce_uses_measure[sce] == 1:
                        if math.isnan(ave_mean):
                            # If no value, use dash
                            cell.value = "-"
                        else:
                            # Calculate confidence interval of the average values
                            variation_ave = critical_value_t_student[sce] * math.sqrt(ave_std ** 2 /
                                                                                      float(number_replications[sce]))
                            conf_int_ave = "{:,.2f} ± {:,.2f}".format(ave_mean, variation_ave)
                            cell.value = "{:,.2f}".format(ave_mean)
                    else:
                        # If don't use measure, use dash
                        cell.value = "-"
                    cell.alignment = openpyxl.styles.Alignment(horizontal='center')

                    cell = sheet.cell(row=max_num_replications + 5 + 2, column=aggr_begin_column)
                    if math.isnan(ave_mean) or sce_uses_measure[sce] == 0:
                        # If no value or don't use measure, use dash
                        cell.value = "-"
                    else:
                        cell.value = "{:,.2f}".format(ave_std)
                    cell.alignment = openpyxl.styles.Alignment(horizontal='center')

                    cell = sheet.cell(row=max_num_replications + 5 + 3, column=aggr_begin_column)
                    if math.isnan(ave_mean) or sce_uses_measure[sce] == 0:
                        # If no value or don't use measure, use dash
                        cell.value = "-"
                    else:
                        cell.value = conf_int_ave
                    cell.alignment = openpyxl.styles.Alignment(horizontal='center')

                    if base_scenario_ind[sce] == None or func_name.split(",")[0] == "Vehroute" \
                    or math.isnan(perc_mean) or sce_uses_measure[sce] == 0:
                        # If no value or don't use measure, use dash for the percentage different mean
                        cell = sheet.cell(row=max_num_replications + 5 + 1, column=aggr_begin_column + 1)
                        cell.value = "-"
                        cell.alignment = openpyxl.styles.Alignment(horizontal='center')
                        cell = sheet.cell(row=max_num_replications + 5 + 2, column=aggr_begin_column + 1)
                        cell.value = "-"
                        cell.alignment = openpyxl.styles.Alignment(horizontal='center')
                        cell = sheet.cell(row=max_num_replications + 5 + 3, column=aggr_begin_column + 1)
                        cell.value = "-"
                        cell.alignment = openpyxl.styles.Alignment(horizontal='center')
                        if math.isnan(ave_mean):
                            # If no value, use dash
                            conf_int_sce_measure[sce].append(("-", "-"))
                        else:
                            conf_int_sce_measure[sce].append((conf_int_ave, "-"))
                    else:
                        # Calculate confidence interval of the percentage different between comparing scenarios
                        variation_perc = critical_value_t_student[sce] * math.sqrt(perc_std ** 2 /
                                                                                   float(number_replications[sce]))
                        conf_int_perc = "{:.2%} ± {:.2%}".format(perc_mean, variation_perc)

                        # Put calulated values on the cells
                        cell = sheet.cell(row=max_num_replications + 5 + 1, column=aggr_begin_column + 1)
                        cell.value = "{:.2%}".format(perc_mean)
                        cell.alignment = openpyxl.styles.Alignment(horizontal='center')
                        cell = sheet.cell(row=max_num_replications + 5 + 2, column=aggr_begin_column + 1)
                        cell.value = "{:.2%}".format(perc_std)
                        cell.alignment = openpyxl.styles.Alignment(horizontal='center')
                        cell = sheet.cell(row=max_num_replications + 5 + 3, column=aggr_begin_column + 1)
                        cell.value = conf_int_perc
                        cell.alignment = openpyxl.styles.Alignment(horizontal='center')
                        conf_int_sce_measure[sce].append((conf_int_ave,conf_int_perc))
                    aggr_begin_column += 2
        else:
            for measure_ind, _ in enumerate(input_measures):
                for _ in aggregations[func_name][measure_ind]:
                    conf_int_sce_measure[sce].append(("-", "-"))

    # Finally, save the file and give it a name
    wb.save(excel_file)

# Read the workbook and create a new sheet for the aggregated values
wb = openpyxl.load_workbook(excel_file)
sheet = wb.create_sheet('Aggregated')
cell = sheet.cell(row=3, column=1)
cell.value="Scenario"
cell.alignment = openpyxl.styles.Alignment(horizontal='center')
for sce, (sce_name, _) in enumerate(scenarios):
    cell = sheet.cell(row=sce + 4, column=1)
    cell.value = sce_name
    cell.alignment = openpyxl.styles.Alignment(horizontal='center')
last_col = 2
measure_aggr = 0
for func_name,func_inputs in measures.items():
    input_measures = func_inputs[0]
    for measure_ind, measure_inputs in enumerate(input_measures):
        (python_func, python_func_name) = measure_inputs[0].split(",")
        measure_key = measure_inputs[1]
        measure_key_list = measure_key.split(",")
        measure_attrb = measure_key_list[-2]
        measure_attrb_name = measure_key_list[-1]
        if measure_attrb_name == '':
            measure_attrb_name = measure_attrb
        measure_name = measure_inputs[2]
        num_aggrs = len(aggregations[func_name][measure_ind])
        cell = sheet.cell(row=1, column=last_col)
        cell.value = measure_name
        sheet.merge_cells(start_row=1, start_column=last_col,
                          end_row=1, end_column=last_col + num_aggrs * 2 - 1)
        cell.alignment = openpyxl.styles.Alignment(horizontal='center')
        for aggr_name in aggregations[func_name][measure_ind]:
            cell = sheet.cell(row=2, column=last_col)
            cell.value = aggr_name
            sheet.merge_cells(start_row=2, start_column=last_col,
                              end_row=2, end_column=last_col + 1)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center')
            cell = sheet.cell(row=3, column=last_col)
            cell.value = "Average"
            cell.alignment = openpyxl.styles.Alignment(horizontal='center')
            cell = sheet.cell(row=3, column=last_col + 1)
            cell.value = "% Difference"
            cell.alignment = openpyxl.styles.Alignment(horizontal='center')
            for sce, _ in enumerate(scenarios):
                cell = sheet.cell(row=sce + 4, column=last_col)
                cell.value = conf_int_sce_measure[sce][measure_aggr][0]
                cell.alignment = openpyxl.styles.Alignment(horizontal='center')
                cell = sheet.cell(row=sce + 4, column=last_col + 1)
                cell.value = conf_int_sce_measure[sce][measure_aggr][1]
                cell.alignment = openpyxl.styles.Alignment(horizontal='center')
            measure_aggr += 1
            last_col += 2

# Finally, save the file and give it a name
wb.save(excel_file)

# Plot the measured actions
print("Creating figures with plots")
for func_name,func_inputs in measures.items():
    output_name = func_name.split(",")[1]
    sce_that_uses = []
    input_measures = func_inputs[0]
    if func_name.split(",")[0] == "Vehroute":
        # Define if a vehicle uses or not certain measure/metric if it compares the route
        # or if it compares against other scenarios
        for sce_val in func_inputs[1][1]:
            sce_that_uses.append((lambda val: 1 if val != None else 0)(sce_val))
    elif func_name.split(",")[0] == "Std":
        # Define if a vehicle uses or not certain measure/metric if it compares against other scenarios
        sce_that_uses = func_inputs[1]

    scenarios_names = [sce_name for sce_name,_ in scenarios]
    for measure_ind, measure_inputs in enumerate(input_measures):
        (python_func, python_func_name) = measure_inputs[0].split(",")
        measure_key = measure_inputs[1]
        measure_key_list = measure_key.split(",")
        measure_attrb = measure_key_list[-2]
        measure_attrb_name = measure_key_list[-1]
        if measure_attrb_name == '':
            measure_attrb_name = measure_attrb
        measure_name = measure_inputs[2]
        num_aggrs = len(aggregations[func_name][measure_ind])
        for aggr_ind, aggr_name in enumerate(aggregations[func_name][measure_ind]):
            # Create a figure instance
            fig = plt.figure()
            # fig = plt.figure(figsize=(3, 5))
            # Create an axes instance
            ax = fig.add_subplot(111)
            # Create the boxplot
            flierprops = dict(marker='o', markerfacecolor='black', markersize=2, linestyle='none')
            meanpointprops = dict(marker='D', markeredgecolor='blue', markerfacecolor='blue', markersize=3)
            # Exclude from the data those with np.isnan
            data = results_sce_replication[func_name][measure_attrb_name][python_func_name][aggr_ind][:, :, 0]
            filtered_values = []
            filtered_sce_names = []
            for sce, sce_name in enumerate(scenarios_names):
                valid_vals = []
                for sce_vals in data[:, sce]:
                    if np.isnan(sce_vals) == False:
                        valid_vals.append(sce_vals)
                if valid_vals != []:
                    filtered_sce_names.append(sce_name)
                    filtered_values.append(valid_vals)

            # Create box plot
            bp = plt.boxplot(filtered_values, flierprops=flierprops, vert=False, widths=0.8,
                             showmeans=True, showfliers=True, meanprops=meanpointprops)
            ## Custom x-axis labels
            ax.set_yticklabels(filtered_sce_names)
            if num_aggrs == 1:
                ax.set_xlabel(measure_name)
            else:
                ax.set_xlabel(measure_name + " of " + aggr_name)
            # Legend
            mean_line = plt.Line2D([], [], color='blue', linestyle="None", marker="D", label='Mean')
            median_line = plt.Line2D([], [], color='orange', linestyle="-", label='Median')
            outlier_line = plt.Line2D([], [], color='black', linestyle="None", marker="o", label='Outlier')
            plt.legend(handles=[mean_line, median_line, outlier_line])
            # Save the figure
            fig_name = "plots_results." + excel_file.split(".")[1] + " - " + output_name + " - " + measure_attrb_name \
                       + " - " + python_func_name + " - " + "aggr" + str(aggr_ind) + ".png"
            fig.savefig(fig_name, bbox_inches='tight')