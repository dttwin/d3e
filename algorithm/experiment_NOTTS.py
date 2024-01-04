## Get traffic measures via SUMO and real travel times of vehicles when vehicles are not rerouted.
## The decision of artificial edge is based on turning rates and lane utilization.

## IMPORT MODULES

from __future__ import absolute_import
from __future__ import print_function

import os
import sys
import optparse
import math
import itertools
import numpy as np
import openpyxl
from xml.etree import ElementTree
import xml.dom.minidom
import logging
import time
from threading import Thread
from sumolib.net import readNet  # noqa
import csv

timestamp = time.strftime("%Y%m%dT%H%M%S")
log_file_name = "notts_zizkov_" + timestamp + ".log"

# Create a logger object
LOGGER = logging.getLogger('notts_zizkov')
LOGGER.setLevel(logging.DEBUG)
# Create file handler which logs even debug messages
fh = logging.FileHandler(log_file_name, mode='w')
fh.setLevel(logging.DEBUG)
# Create console handler with a higher log level
ch = logging.StreamHandler()
ch.setLevel(logging.INFO)
# create formatter and add it to the handlers
formatter = logging.Formatter('%(levelname)-8s %(message)s')
ch.setFormatter(formatter)
formatter = logging.Formatter('%(asctime)s - %(levelname)-8s %(message)s')
fh.setFormatter(formatter)
# add the handlers to logger
LOGGER.addHandler(ch)
LOGGER.addHandler(fh)

np.seterr(all='raise')

# Import python modules from the $SUMO_HOME/tools directory
try:
    # sys.path.append(os.path.join(os.path.dirname(__file__), '..', '..', '..', '..', "tools"))
    sys.path.append(
        os.path.join(os.environ.get("SUMO_HOME", os.path.join(os.path.dirname(__file__), "..", "..", "..")), "tools"))
    import sumolib  # noqaAnd
except ImportError:
    sys.exit(
        "missing declare environment variable 'SUMO_HOME' as the root directory of SUMO installation "
        "(it should contain folders 'bin', 'tools' and 'docs')")


## INPUTS THAT CAN BE CHANGED BY OPTION PARSER

def get_options():
    """Initialing Script from Command Line"""
    optParser = optparse.OptionParser()

    # SUMO GUI Option
    optParser.add_option("--nogui", action="store_true", default=False, help="run the commandline version of sumo")

    # Python Interface Connection
    optParser.add_option("--interface", action="store", type="string", default="traci",
                         help="Define which connection interface to choose, if SUMO either traci or libsumo")

    # Experiment Scenario Name
    optParser.add_option("--scenario_name",
                         default="mid.TURNS",
                         action="store", type="string",
                         help="Name of the Simulation Scenario")

    # SUMO Simulation Paths Inputs and Outputs
    optParser.add_option("--simul_filename",
                         default="zizkov.sumocfg",
                         action="store", type="string",
                         help="SUMO simulation file")
    optParser.add_option("--net_filename",
                         default="zizkov_flora_closed.net.xml",
                         action="store", type="string",
                         help="SUMO network file")
    optParser.add_option("--fixed_route_files",
                         default="routes.mid.rou.xml,routes.probes.rou.xml",
                         action="store", type="string",
                         help="SUMO fixed (same for all scenarios) vehicle route files, use comma as separator")
    optParser.add_option("--dyn_route_files",
                         default="",
                         action="store", type="string",
                         help="SUMO dynamic (appending sce. name) vehicle route files, use comma as separator")
    optParser.add_option("--fixed_additional_files",
                         default="tll_static_closed.xml",
                         action="store", type="string",
                         help="SUMO fixed (same for all scenarios) additional files, use comma as separator")
    optParser.add_option("--dyn_additional_files",
                         default="additional.add.xml",
                         action="store", type="string",
                         help="SUMO dynamic (appending sce. name) additional files, use comma as separator")
    optParser.add_option("--path_inputs",
                         default="./Inputs_Outputs/",
                         action="store", type="string",
                         help="path of SUMO input files")
    optParser.add_option("--path_outputs",
                         default="./Inputs_Outputs/Necessary_Outputs/",
                         action="store", type="string",
                         help="path of SUMO output files")

    # Input for Estimating Accuracy of TTs Prediction
    optParser.add_option("--warm_up_t", default=1800, action="store", type="int",
                         help="Warm up time for the travel time reliability report")

    # SUMO Simulation Time Interval
    optParser.add_option("--begin", default="43200", action="store", type="string",
                         help="Begin Step Simulation")
    optParser.add_option("--end", default="55800", action="store", type="string",
                         help="End Step Simulation")
    optParser.add_option("--step_length", default="1", action="store", type="string",
                         help="Simulation Step Length in seconds")

    # SUMO Lane Closure Input for Local Level Routing and Choosing Artificial Edge
    optParser.add_option("--interval_lane_closure", nargs=2, default=(43201, 55801), action="store", type="float",
                         help="begin and end time of lane closure")
    optParser.add_option("--lane_to_close", default="dz_50178032#1.132_1,gneE46_0", action="store",
                         type="string", help="network id of lane to close. Use comma as separator")
    optParser.add_option("--closed_art_edge", default="gneE46", action="store",
                         type="string", help="network id of artificial edge that will close. Use comma as separator")
    optParser.add_option("--dist_to_close", default=-1, action="store", type="float",
                         help="distance to stop line the lane should be closed, if -1 it uses lane length")

    options, args = optParser.parse_args()
    return options

# Main Entry Point for SUMO and to Parse Variables
if __name__ == "__main__":
    options = get_options()

    net = readNet(options.path_inputs + options.net_filename)
    
    # this script has been called from the command line. It will start sumo as a server, then connect and run
    try:
        if options.interface == "libsumo":
            import libsumo as traci
            # As GUI is not working yet with libsumo, use only without GUI
            sumoBinary = sumolib.checkBinary('sumo')
            INTERFACE = "libsumo"
        elif options.interface == "traci":
            raise ImportError
        else:
            # Import here module that connects the Python script with CAVs data interface
            pass
    except ImportError:
        import traci

        INTERFACE = "traci"
        if options.nogui:
            sumoBinary = sumolib.checkBinary('sumo')
        else:
            sumoBinary = sumolib.checkBinary('sumo-gui')

    # Define SUMO additional files
    add_filenames = options.fixed_additional_files.split(",")
    if options.dyn_additional_files != "":
        add_filenames.extend([dynf.split(".")[0] + "." + options.scenario_name + "." + ".".join(dynf.split(".")[1:])
                              for dynf in options.dyn_additional_files.split(",")])
    additional_files = options.path_inputs + options.path_inputs.join([afile + "," for afile in add_filenames][:-1]
                                                                      + [add_filenames[-1]])

    # Define SUMO route files
    route_filenames = options.fixed_route_files.split(",")
    if options.dyn_route_files != "":
        route_filenames.extend(
            [dynf.split(".")[0] + "." + options.scenario_name + "." + ".".join(dynf.split(".")[1:])
             for dynf in options.dyn_route_files.split(",")])
    route_files = options.path_inputs + options.path_inputs.join([rfile + "," for rfile in route_filenames][:-1]
                                                                 + [route_filenames[-1]])

    # this is the normal way of using traci. sumo is started as a
    # subprocess and then the python script connects and runs
    params = [sumoBinary, "-c", options.simul_filename,
                 "--net", options.path_inputs + options.net_filename,
                 "--route-files", route_files,
                 "--additional-files", additional_files,
                 "--error-log", options.path_outputs.replace("Necessary_Outputs", "Unecessary_Outputs") + "SUMOlog."
                 + options.scenario_name + ".txt",
                 "--vehroute-output",options.path_outputs + "vehroute." + options.scenario_name + ".xml",
                 "--tripinfo-output",options.path_outputs + "tripinfo." + options.scenario_name + ".xml",
                 "--begin",options.begin,
                 "--end",options.end,
                 "--emission-output", "notts_emission.out.xml",
                 "--gui-settings-file", "zizkov_alt.view.xml",
                 "--window-size", "900,1400",
                 "--window-pos", "0,0",
                 "--step-length", options.step_length,
                 "--device.emissions.probability", "1"]
    traci.start(params)
    LOGGER.info("Started SUMO as: " + " ".join(params))

options = get_options()

# TraCI Constants
LANE_VEH_NUM = traci.constants.LAST_STEP_VEHICLE_NUMBER
LANE_QUEUE_NUM = traci.constants.LAST_STEP_VEHICLE_HALTING_NUMBER
LANE_OCCUPANCY = traci.constants.LAST_STEP_OCCUPANCY
VEHIDS = traci.constants.LAST_STEP_VEHICLE_ID_LIST
LANE_METRICS = [LANE_QUEUE_NUM, LANE_VEH_NUM]

# The prior edges are edges before artificial edges,
# artificial edges are edges representing one lane of a multi-lane incoming edge at certain junction.
# The index of artificial edge is the lane number of the prior edge
NET_FILE = options.path_inputs + options.net_filename
PRIOR_ARTIFICIAL_EDGES = {"gneE247": ("gneE248","gneE249"),
                          "fl_156262933#4": ("gneE246","gneE245"),
                          "fl_330512996#6.19": ("gneE240.50","gneE241.49"),
                          "gneE57": ("gneE60","gneE59"),
                          "fl_80779632#4": ("gneE148","gneE149"),
                          "dz_142147141": ("gneE143","gneE144"),
                          "dz_50178032#1.132": ("gneE47","gneE45")}

PRIOR_COMMON_AFTER_ART_EDGES = {"gneE247": "gneE293",
                                "fl_156262933#4": "fl_483444577#0",
                                "fl_330512996#6.19": "fl_448565742.11",
                                "gneE57": "fl_4641298#3",
                                "fl_80779632#4": "dz_330512997",
                                "dz_142147141": "fl_330512996#6.19",
                                "dz_50178032#1.132": "dz_142147149#0"}

# Scenario Name
SCE_NAME = options.scenario_name

# SUMO Simulation Time Interval
BEGIN_T = float(options.begin)
END_T = float(options.end)
STEP_LENGTH = float(options.step_length)
ACTUAL_T = BEGIN_T

# Road Closure
BEGIN_LANE_CLOSURE = options.interval_lane_closure[0]
END_LANE_CLOSURE = options.interval_lane_closure[1]
LANES2CLOSE = options.lane_to_close.split(",")
EDGE2CLOSE = options.closed_art_edge.split(",")
DIST2CLOSE = options.dist_to_close

# CAVs that entered LLR area
CAVS_AREA_FILE = options.path_outputs + "CAVs_modelled_area." + SCE_NAME + ".xlsx"

# nonCAVs that entered LLR area
NONCAVS_AREA_FILE = options.path_outputs + "nonCAVs_modelled_area." + SCE_NAME + ".xlsx"

# All vehicles that entered LLR area
ALL_VEHS_AREA_FILE = options.path_outputs + "allVehs_modelled_area." + SCE_NAME + ".xlsx"

# Edges to the Traffic Network, for the report of accuracy of TTs
ALL_INEDGES_AREA_FILE = options.path_outputs + "allInEdges_modelled_area." + SCE_NAME.split(".")[0] + ".xlsx"
ALL_INOUTEDGES_AREA_FILE = options.path_outputs + "allInOutEdges_modelled_area." + SCE_NAME.split(".")[0] + ".xlsx"

# File with Measured, Estimated and Std. of Travel Times
TTS_ACCUR_AGGREGATED = options.path_outputs + "python_edge_traffic." + SCE_NAME + ".xml"
WARM_UP_T = round(float(options.warm_up_t), 1)


## LOGGING
logging.basicConfig(filename=options.path_outputs.replace("Necessary_Outputs/","Unecessary_Outputs/") + 'PYTHONlog.' +
                             options.scenario_name + '.log',level=logging.INFO,format='%(asctime)s %(message)s')
logger = logging.getLogger(__name__)
logger.info('Start Simulation')

def add_marker(vehID, polygonID):
    vx,vy = traci.vehicle.getPosition(vehID)
    r = 50.0
    np = 16
    circle_shape = list()
    for i in range(np):
        phi = i * 2 * math.pi / np
        p = (vx + r*math.sin(phi), vy + r*math.cos(phi))
        circle_shape.append(p)
    traci.polygon.add(polygonID, shape=circle_shape, color=(0,0,255,255), fill=True, polygonType="circle", layer=10)
    # traci.polygon.setFilled(polygonID, True)
    traci.polygon.addDynamics(polygonID, vehID, rotate=False)


def get_fcd_row(vehID):
    accel = traci.vehicle.getAcceleration(vehID)
    wt = traci.vehicle.getWaitingTime(vehID)
    fuel = traci.vehicle.getFuelConsumption(vehID)
    e_co = traci.vehicle.getCOEmission(vehID)
    e_co2 = traci.vehicle.getCO2Emission(vehID)
    e_nox = traci.vehicle.getNOxEmission(vehID)
    return ACTUAL_T, vehID, wt, accel, fuel, e_co, e_co2, e_nox


## MAIN SCRIPT FUNCTIONS
def frange(start, end=None, inc=None):
    """A range function, that does accept float increments"""

    if end == None:
        end = start + 0.0
        start = 0.0

    if inc == None:
        inc = 1.0

    L = []
    while 1:
        next = start + len(L) * inc
        if inc > 0 and next >= end:
            break
        elif inc < 0 and next <= end:
            break
        L.append(next)

    return L

def getExcelTargetList(filename):
    """Get Values of Each Row of the First Collum of an Excel file"""

    book = openpyxl.load_workbook(filename, data_only=True)
    sheet = book.active
    rows = sheet.rows
    file_input_vals = []
    for row in rows:
        for cell in row:
            file_input_vals.append(cell.value)

    return file_input_vals


def defPriorArtEdgeVehiclesNextEdge(edgeID, edge_vehs, lanes_vehs):
    """Define the next artificial edge of vehicles on prior artificial edges"""

    global lowest_art_lanes_measure, lowest_prior_lanes_measure

    num_lanes = len(priorAndArtificial_edgesLanes[edgeID])
    CAVs_route = dict()
    # Select artificial edges only that are not closed
    aval_art_edges = [art_edge for art_edge in PRIOR_ARTIFICIAL_EDGES[edgeID]
                      if art_edge not in EDGE2CLOSE]
    # Update vehicles that already had defined their next edges (artificial one)
    edge_endLane_vehs[edgeID].intersection_update(edge_vehs)
    edge_defined_vehs[edgeID].intersection_update(edge_vehs)

    # For the closest vehicle to stop line on each lane, check if going to next edge on the next time step.
    closest_stopline_cavs = set([])
    for laneID in lanes_vehs:
        try:
            # Assuming that only 1 vehicle per lane can decide next edge per time step
            # Assuming edge_vehs_list are ordered by their distance to stop line
            first_vehID = lanes_vehs[laneID][-1]
            try:
                # Get vehicle next edges
                veh_route_index = traci.vehicle.getRouteIndex(first_vehID)
                veh_route = list(traci.vehicle.getRoute(first_vehID)[veh_route_index:])
                if veh_route[2] == PRIOR_COMMON_AFTER_ART_EDGES[edgeID]:
                    veh_speed = traci.vehicle.getSpeed(first_vehID)
                    veh_displacement = traci.vehicle.getLanePosition(first_vehID)
                    # Define distance to stop line after next time step
                    dist_nextSteps = veh_displacement + veh_speed
                    if dist_nextSteps >= prior_edges_len[edgeID]:
                        closest_stopline_cavs.add(first_vehID)
                        CAVs_route.update({first_vehID: veh_route[:]})
                else:
                    edge_defined_vehs[edgeID].add(first_vehID)
            except IndexError:
                # Vehicle finishing trip
                skip = 1
        except IndexError:
            NoVehicle = 1
    if len(closest_stopline_cavs) > 0:
        # If so, get artificial edge metric and define next edge
        # Among all lanes of artificial edges, choose edge with following priority:
        # 1) shortest queue
        # 2) lowest number of vehicles
        sum_lane = 0
        for lane_metric in LANE_METRICS:
            lowest_art_lanes_measure = (None, float("inf"))
            for art_edgeID in PRIOR_ARTIFICIAL_EDGES[edgeID]:
                if art_edgeID not in EDGE2CLOSE or (ACTUAL_T < BEGIN_LANE_CLOSURE or ACTUAL_T > END_LANE_CLOSURE):
                    # If artificial edge is not closed
                    for lane_num, laneID in enumerate(priorAndArtificial_edgesLanes[art_edgeID]):
                        # For each lane of artificial edge,
                        # choose lane with either shortest queue or number of vehicles
                        # and take its respective artificial edge
                        lane_measure = traci.lane.getSubscriptionResults(laneID)[lane_metric]
                        sum_lane += lane_measure
                        if lane_measure < lowest_art_lanes_measure[1]:
                            # Update lowest metric
                            lowest_art_lanes_measure = (art_edgeID, lane_measure)
                        elif lane_num == (num_lanes - 1) and lane_measure == lowest_art_lanes_measure[1]:
                            # All queue lanes are the same, then maintain same prior lane or change
                            lowest_art_lanes_measure = None
            if sum_lane > 0:
                # If sum of lane metric is zero, use next measure
                break

    # For other vehicles, define current lane only for those not defined before
    edge_vehs.difference_update(edge_defined_vehs[edgeID])
    edge_vehs_to_monitor = set([])
    # but if lanes of current edge are too much different, make for all vehicles aiming to balance the load
    if num_lanes > 1:
        # If only more than 1 possible next artificial edge choose lane with following priority:
        # 1) shortest queue
        # 2) lowest number of vehicles
        sum_lane = 0
        for lane_metric in LANE_METRICS:
            lowest_prior_lanes_measure = (None, float("inf"))
            for lane_num, laneID in enumerate(priorAndArtificial_edgesLanes[edgeID]):
                # For each lane of prior edge, choose lane with lowest metric and its respective artificial edge
                try:
                    art_edgeID = PRIOR_ARTIFICIAL_EDGES[edgeID][lane_num]
                except IndexError:
                    art_edgeID = aval_art_edges[-1]
                if PRIOR_ARTIFICIAL_EDGES[edgeID][lane_num] not in EDGE2CLOSE \
                or (ACTUAL_T < BEGIN_LANE_CLOSURE or ACTUAL_T > END_LANE_CLOSURE):
                    lane_measure = traci.lane.getSubscriptionResults(laneID)[lane_metric]
                    if lane_metric == LANE_QUEUE_NUM:
                        # Due SUMO bug of new vehicle being inserted with speed zero, decrease it from the lane
                        # queue number
                        try:
                            last_vehID = lanes_vehs[laneID][0]
                            if traci.vehicle.getSpeed(last_vehID) <= 0.1:
                                lane_measure -= 1
                        except IndexError:
                            NoVehs = 1
                    sum_lane += lane_measure
                    if lane_measure < lowest_prior_lanes_measure[1]:
                        if lane_metric == LANE_QUEUE_NUM \
                        and 5 < lowest_prior_lanes_measure[1] < float("inf") \
                        and lane_measure < 0.6 * lowest_prior_lanes_measure[1]:
                            # If shortest lane is half of the longest queue and bigger than 5 vehicles,
                            # reset all vehicles to distribute better
                            edge_defined_vehs[edgeID] = set([])
                        # Update lowest queue
                        lowest_prior_lanes_measure = (art_edgeID, lane_measure)
                    elif lane_num == (num_lanes - 1) and lane_measure == lowest_prior_lanes_measure[1]:
                        # All queue lanes are the same, then maintain same prior lane
                        lowest_prior_lanes_measure = None

            if sum_lane > 0:
                # If sum of lane metric is zero, use next measure
                break
    else:
        # If only 1 possible next artificial edge (only one lane)
        lowest_prior_lanes_measure = None

    for vehID in edge_vehs:
        if lowest_prior_lanes_measure != None \
        and traci.vehicle.getParameter(vehID, "has.rerouting.device") == "true":
            try:
                # Get vehicle next edges
                veh_displacement = traci.vehicle.getLanePosition(vehID)
                veh_len = traci.vehicle.getLength(vehID)
                if veh_displacement >= veh_len + 0.2:
                    # Due SUMO bug that inserts vehicle with speed zero, skip vehicles being inserted right now,
                    # consider vehicle only on next time step
                    veh_route_index = traci.vehicle.getRouteIndex(vehID)
                    veh_route = list(traci.vehicle.getRoute(vehID)[veh_route_index:])
                    if veh_route[2] == PRIOR_COMMON_AFTER_ART_EDGES[edgeID]:
                        edge_vehs_to_monitor.add(vehID)
                        CAVs_route.update({vehID: veh_route[:]})
                    else:
                        edge_defined_vehs[edgeID].add(vehID)
            except IndexError:
                # Vehicle finishing trip
                skip = 1
        else:
            edge_defined_vehs[edgeID].add(vehID)

    # For vehicles that hasn't defined yet current lanes or next edges running on the prior artificial edges
    decide_current_lane = set([])
    decide_next_lane = set([])
    maintain_next_lane = set([])
    for vehID in closest_stopline_cavs:
        if vehID not in edge_endLane_vehs[edgeID]:
            # Set vehicle as already defined next artificial edge
            edge_endLane_vehs[edgeID].add(vehID)
            if lowest_art_lanes_measure != None:
                # Decide next lane (respective artificial edge),
                # based on the shortest queue or lowest number of vehicles of next lane
                decide_next_lane.add(vehID)
            else:
                # Maintain same current lane, to define respective artificial edge
                # of current lane of prior edge
                maintain_next_lane.add(vehID)
        else:
            # Maintain same current lane, to define respective artificial edge of current lane of prior edge
            maintain_next_lane.add(vehID)

    for vehID in edge_vehs_to_monitor:
        edge_defined_vehs[edgeID].add(vehID)
        veh_speed = traci.vehicle.getSpeed(vehID)
        if veh_speed > 0.1 and lowest_prior_lanes_measure != None:
            # Decide current lane, to define respective artificial edge
            # of the shortest queue of current prior edge
            decide_current_lane.add(vehID)
        # elif veh_speed == 0:
        #     # Maintain same current lane, to define respective artificial edge
        #     # of current lane of prior edge
        #     maintain_next_lane.add(vehID)

    # Input into SUMO suitable next artificial edge
    for vehID in maintain_next_lane:
        veh_route = CAVs_route[vehID][:]
        lane_num = traci.vehicle.getLaneIndex(vehID)
        if PRIOR_ARTIFICIAL_EDGES[edgeID][lane_num] not in EDGE2CLOSE \
        or (ACTUAL_T < BEGIN_LANE_CLOSURE or ACTUAL_T > END_LANE_CLOSURE):
            # The next artificial edge is based on the current lane of the prior edge
            veh_route[1] = PRIOR_ARTIFICIAL_EDGES[edgeID][lane_num]
        else:
            # If more prior lanes than artificial edges
            veh_route[1] = aval_art_edges[-1]
        LOGGER.debug('maintain_next_lane({:s}): setting route for vehicle {:s} to {:s}'.format(edgeID, vehID, str(veh_route)))
        traci.vehicle.setRoute(vehID, veh_route)
    for vehID in decide_current_lane:
        veh_route = CAVs_route[vehID][:]
        if veh_route[1] != lowest_prior_lanes_measure[0]:
            veh_route = CAVs_route[vehID][:]
            # The artificial edge is based on the shortest queue of current lane of the prior edge
            veh_route[1] = lowest_prior_lanes_measure[0]
            LOGGER.debug('decide_current_lane({:s}): setting route for vehicle {:s} to {:s}'.format(edgeID, vehID, str(veh_route)))
            traci.vehicle.setRoute(vehID, veh_route)
    for vehID in decide_next_lane:
        veh_route = CAVs_route[vehID][:]
        if lowest_art_lanes_measure != None and veh_route[1] != lowest_art_lanes_measure[0]:
            # The next artificial edge is based on the lowest metric of all artificial edges
            veh_route[1] = lowest_art_lanes_measure[0]
            LOGGER.debug('decide_next_lane({:s}): setting route for vehicle {:s} to {:s}'.format(edgeID, vehID, str(veh_route)))
            traci.vehicle.setRoute(vehID, veh_route)


# Initializing for measuring accuracy of travel time prediction
allInEdges_modelled_area = getExcelTargetList(ALL_INEDGES_AREA_FILE)
all_InOut_edges = getExcelTargetList(ALL_INOUTEDGES_AREA_FILE)
tts_accur_vals = {edgeID: [] for edgeID in allInEdges_modelled_area}
allVehs_Edge_ArrTime = dict()

# Get lanes and length of the artificial edges and also the edges sending vehicles to them (called prior edges),
# in addition, define vehs_last_range for error of queue length estimation and last_step_edge_vehs for counting
prior_edges_len = dict()
priorAndArtificial_edgesLanes = dict()
edge_endLane_vehs = dict()
edge_defined_vehs = dict()
for priorEdgeID in PRIOR_ARTIFICIAL_EDGES.keys():
    edge_part = priorEdgeID + "_"
    priorAndArtificial_edgesLanes.update({priorEdgeID: [edge_part + str(lane_num)
                                                       for lane_num in range(0, traci.edge.getLaneNumber(priorEdgeID))]})
    # Init. list of vehicles to be set as their next edges (artificial one) already defined
    edge_endLane_vehs.update({priorEdgeID: set([])})
    edge_defined_vehs.update({priorEdgeID: set([])})
    for artEdgeID in PRIOR_ARTIFICIAL_EDGES[priorEdgeID]:
        edge_part = artEdgeID + "_"
        priorAndArtificial_edgesLanes.update({artEdgeID: [edge_part + str(lane_num)
                                                       for lane_num in range(0, traci.edge.getLaneNumber(artEdgeID))]})

net = sumolib.net.readNet(NET_FILE)
edgeID_to_laneIDs = dict()
for edgeID in all_InOut_edges:
    edge_child = net.getEdge(edgeID)
    if priorAndArtificial_edgesLanes.has_key(edgeID):
        prior_edges_len.update({edgeID: edge_child.getLength()})
    edgeID_to_laneIDs.update({edgeID: [lane_child.getID() for lane_child in edge_child.getLanes()]})
    for laneID in edgeID_to_laneIDs[edgeID]:
        # Subscribe to get data from modelled lanes
        if edgeID not in allInEdges_modelled_area:
            traci.lane.subscribe(laneID, [VEHIDS])
        else:
            traci.lane.subscribe(laneID, LANE_METRICS + [LANE_OCCUPANCY, VEHIDS])

# Initializing for list of vehicles within the modelled area
cavs_area = set([])
noncavs_area = set([])
# all_vehs_area = set([]) # only when not measuring travel time accuracy

TRACKED_VEHID = None
polygonID="marker"
fcd_data = []
routes_vehicles = []
entrance_time = {}
have_vehicle_69 = False
veh_69_route = []

# Constant Loop to Communicate with SUMO/TraCI
# try:
while ACTUAL_T < END_T + STEP_LENGTH:
    if INTERFACE == "libsumo":
        # If Libsumo open a thread to avoid the simulation get stuck
        stored_ACTUAL_T = ACTUAL_T
        Thread(target=traci.simulationStep()).start()
        time.sleep(0.02)
        # Simulation data
        ACTUAL_T = round(float(traci.simulation.getTime()), 1)  # Returns the current simulation time in seconds
        if ACTUAL_T == stored_ACTUAL_T:
            LOGGER.debug("SUMO stopped")
            Thread(target=traci.simulationStep()).start()
            time.sleep(1)
            ACTUAL_T = round(float(traci.simulation.getTime()), 1)
            if ACTUAL_T == stored_ACTUAL_T:
                raise SystemExit
            else:
                LOGGER.debug("SUMO is back")

        sys.stdout.write('\r' + "Step " + str(ACTUAL_T))
    else:
        # If Traci
        traci.simulationStep()
        # Simulation Data
        ACTUAL_T = traci.simulation.getTime()  # Returns the current simulation time in seconds

    if ACTUAL_T % 900 == 0:
        LOGGER.debug("saving vehicle route data at time {:f}".format(ACTUAL_T))
        with open('notts_routes_{:.0f}.csv'.format(ACTUAL_T), 'wb') as csvfile:
            r_writer = csv.writer(csvfile, delimiter=',')
            for row in routes_vehicles:
                r_writer.writerow(row)

    arrived = traci.simulation.getArrivedIDList()
    for vehID in arrived:
        # Check 6.9
        if vehID == '6.9':
            have_vehicle_69 = False
        routes_vehicles.append([ACTUAL_T, vehID, "left_simulation", ACTUAL_T - entrance_time[vehID]])
        if vehID == TRACKED_VEHID:
            traci.polygon.remove("route")
            with open('fcd_notts_{:s}.csv'.format(vehID), 'wb') as csvfile:
                fcd_writer = csv.writer(csvfile, delimiter=',')
                for row in fcd_data:
                    fcd_writer.writerow(row)
            # forget the vehicle
            TRACKED_VEHID = None
            LOGGER.debug("removing highlight of {:s}".format(vehID))

    if TRACKED_VEHID is not None:
        row = get_fcd_row(TRACKED_VEHID)
        fcd_data.append(row)
        
    # Mark selected vehicles
    departed = traci.simulation.getDepartedIDList()
    for vehID in departed:
        # Check 6.9
        if vehID == '6.9':
            have_vehicle_69 = True
        # Store vehicle route
        route = traci.vehicle.getRoute(vehID) 
        routes_vehicles.append([ACTUAL_T, vehID, "entered_simulation"] + list(route))
        entrance_time[vehID] = ACTUAL_T
        if 'probe' in vehID: 
            if TRACKED_VEHID is None:
                # rerouted vehicle
                LOGGER.debug("vehicle {:s} marked at time {:f}".format(vehID, ACTUAL_T))
                # traci.vehicle.highlight(vehID, size=50)
                add_marker(vehID, polygonID)
                TRACKED_VEHID = vehID
                route = traci.vehicle.getRoute(vehID)
                lanes = []
                for e in route:
                    lanes.append(net.getEdge(e).getLane(0))
                route_shape = list(itertools.chain(*list(l.getShape() for l in lanes)))
                LOGGER.debug(route_shape)
                traci.polygon.add("route", shape=route_shape, color=(255,0,0,255), lineWidth=10, polygonType="route")
                row = get_fcd_row(TRACKED_VEHID)
                fcd_data.append(row)
            else:
                LOGGER.debug('cannot mark vehicle {:s} at time {:f} as {:s} already highlited'.format(vehID, ACTUAL_T, TRACKED_VEHID))
        
    # Check vehicle 6.9 route
    if have_vehicle_69:
        t69 = traci.vehicle.getRoute('6.9')
        if t69 and not veh_69_route:
            veh_69_route = t69
            LOGGER.info('vehicle `6.9` entered simulation at time {:f} with route {:s}'.format(ACTUAL_T, str(list(veh_69_route))))
        elif t69 != veh_69_route:
            veh_69_route = t69
            LOGGER.info('vehicle `6.9` at lane `{:s}` at time {:f} new route {:s}'.format(traci.vehicle.getLaneID('6.9'), ACTUAL_T, str(list(veh_69_route))))
    
    
    for edgeID in all_InOut_edges:
        # For all modelled edges (incoming and outgoing edges of junctions)
        lanes_vehs = dict()
        edge_vehs = set([])
        for laneID in edgeID_to_laneIDs[edgeID]:
            # Get list of vehicles running on the modelled lanes
            lanes_vehs.update({laneID: traci.lane.getSubscriptionResults(laneID)[VEHIDS]})
            edge_vehs.update(lanes_vehs[laneID])

        if len(edge_vehs) > 0:
            for vehID in edge_vehs:
                # all_vehs_area.add(vehID) # only when travel time accuracy not estimated
                if traci.vehicle.getParameter(vehID, "has.rerouting.device") == "true":
                    # Add vehicle to the set of CAVs detected in the modelled area
                    cavs_area.add(vehID)
                else:
                    # Add vehicle to the set of nonCAVs detected in the modelled area
                    noncavs_area.add(vehID)
                try:
                    if allVehs_Edge_ArrTime[vehID][0] != edgeID:
                        # If vehicle is in another edge than the last one detected
                        if ACTUAL_T >= BEGIN_T + WARM_UP_T + STEP_LENGTH and ACTUAL_T <= END_T + STEP_LENGTH:
                            # Add measured travel time to the list of travel times of edgeID
                            tts_accur_vals[allVehs_Edge_ArrTime[vehID][0]].append(
                                ACTUAL_T - allVehs_Edge_ArrTime[vehID][1])
                        # Update new edge to monitor
                        allVehs_Edge_ArrTime.update({vehID: (edgeID, ACTUAL_T)})
                except KeyError:
                    # If vehicle not detected before, start monitoring vehicle's edge
                    allVehs_Edge_ArrTime.update({vehID: (edgeID, ACTUAL_T)})

            # Algorithm to spread vehicles evenly to next edge (only artificial ones) based on queue length and/or number
            if PRIOR_ARTIFICIAL_EDGES.has_key(edgeID):
                defPriorArtEdgeVehiclesNextEdge(edgeID, edge_vehs, lanes_vehs)


# Create workbook with and a new sheet for the CAVs Area
if os.path.isfile(CAVS_AREA_FILE) == True:
    os.remove(CAVS_AREA_FILE)
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "CAVs Modelled Area"
for cav_ind, cav in enumerate(cavs_area):
    sheet.cell(row=cav_ind + 1, column=1).value = cav
wb.save(CAVS_AREA_FILE)

# Create workbook with and a new sheet for the nonCAVs Area
if os.path.isfile(NONCAVS_AREA_FILE) == True:
    os.remove(NONCAVS_AREA_FILE)
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "nonCAVs Modelled Area"
for cav_ind, cav in enumerate(noncavs_area):
    sheet.cell(row=cav_ind + 1, column=1).value = cav
wb.save(NONCAVS_AREA_FILE)

# Create workbook with and a new sheet for all vehicles in the modelled area
if os.path.isfile(ALL_VEHS_AREA_FILE) == True:
    os.remove(ALL_VEHS_AREA_FILE)
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "All Vehs. Modelled Area"
for cav_ind, cav in enumerate(allVehs_Edge_ArrTime.keys()):
    sheet.cell(row=cav_ind + 1, column=1).value = cav
wb.save(ALL_VEHS_AREA_FILE)

# Create XML file with measured travel time for each edge
# Create the file structure
mean_data = ElementTree.Element('meandata')
items = ElementTree.SubElement(mean_data, 'interval')
for edgeID in tts_accur_vals:
    item = ElementTree.SubElement(items, 'edge')
    item.set('id', str(edgeID))
    if tts_accur_vals[edgeID] != []:
        item.set('measuredTravelTime', str(round(np.mean(tts_accur_vals[edgeID]), 1)))

# Create a new XML file with the results
if os.path.isfile(TTS_ACCUR_AGGREGATED) == True:
    os.remove(TTS_ACCUR_AGGREGATED)
xml_string = ElementTree.tostring(mean_data)
parsed_xml_string = xml.dom.minidom.parseString(xml_string)
python_edge_traffic = parsed_xml_string.toprettyxml()
with open(TTS_ACCUR_AGGREGATED, "w") as f:
    f.write(python_edge_traffic)

# Close Traci/Libsumo connection
traci.close()

logger.info('End Simulation')

sys.stdout.write('\n')
sys.stdout.flush()

# except (SystemExit, KeyboardInterrupt):
#     # Close Traci/Libsumo connection
#     traci.close()
#     raise
# except Exception, e:
#     LOGGER.debug("\n There was an error, check the log file \n")
#     logger.error('Crash', exc_info=True)
#     sys.exit(1)