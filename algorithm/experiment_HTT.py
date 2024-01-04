## Get traffic measures via SUMO, real travel times of vehicles and the accuracy of prediction when vehicles are
## rerouted every routing period using information of time-dependant travel times inputed by option --path_weight_files.
## It also uses an algorithm that chooses which artificial edge a vehicle uses based on queue length and/or number
## of vehicles, instead of travel times.
## Otherwise all vehicles would choose only one of the artificial edges that has lowest travel time.

## IMPORT MODULES

from __future__ import absolute_import
from __future__ import print_function

import os
import sys
import optparse
import numpy as np
import math
import openpyxl
from xml.etree import ElementTree
import xml.dom.minidom
import logging
import time
from threading import Thread

np.seterr(all='raise')

# Import python modules from the $SUMO_HOME/tools directory
try:
    sys.path.append(os.path.join(os.path.dirname(__file__), '..', '..', '..', '..', "tools"))
    sys.path.append(
        os.path.join(os.environ.get("SUMO_HOME", os.path.join(os.path.dirname(__file__), "..", "..", "..")), "tools"))
    import sumolib
except ImportError:
    sys.exit(
        "missing declare environment variable 'SUMO_HOME' as the root directory of SUMO installation "
        "(it should contain folders 'bin', 'tools' and 'docs')")



## INPUTS THAT CAN BE CHANGED BY OPTION PARSER

def get_options():
    """Initialing Script from Command Line"""
    optParser = optparse.OptionParser()

    # SUMO GUI Option and Connection
    optParser.add_option("--nogui", action="store_true", default=False, help="run the commandline version of sumo")

    # Python Interface Connection
    optParser.add_option("--interface", action="store", type="string", default="traci",
                         help="Define which connection interface to choose, if SUMO either traci or libsumo")

    # Experiment Scenario Name
    optParser.add_option("--scenario_name",
                         default="mid.HTT.70PR",
                         action="store", type="string",
                         help="Name of the Simulation Scenario")

    # SUMO Simulation Paths Inputs and Outputs
    optParser.add_option("--simul_filename",
                         default="zizkov.sumocfg",
                         action="store", type="string",
                         help="SUMO simulation file")
    optParser.add_option("--net_filename",
                         default="zizkov_flora_closed.net.xml",
                         action="store", type="string",
                         help="SUMO network file")
    optParser.add_option("--fixed_route_files",
                         default="routes.mid.rou.xml,routes.probes.rou.xml",
                         action="store", type="string",
                         help="SUMO fixed (same for all scenarios) vehicle route files, use comma as separator")
    optParser.add_option("--dyn_route_files",
                         default="",
                         action="store", type="string",
                         help="SUMO dynamic (appending sce. name) vehicle route files, use comma as separator")
    optParser.add_option("--fixed_additional_files",
                         default="tll_static_closed.xml",
                         action="store", type="string",
                         help="SUMO fixed (same for all scenarios) additional files, use comma as separator")
    optParser.add_option("--dyn_additional_files",
                         default="additional.add.xml",
                         action="store", type="string",
                         help="SUMO dynamic (appending sce. name) additional files, use comma as separator")
    optParser.add_option("--path_inputs",
                         default="D:/Andre_Back_Up/a_Projects_Files/Local_Level_Routing/Prague_Zizkov/Inputs_Outputs/",
                         action="store", type="string",
                         help="path of SUMO input files")
    optParser.add_option("--path_outputs",
                         default="D:/Andre_Back_Up/a_Projects_Files/Local_Level_Routing/Prague_Zizkov/"
                                 "Inputs_Outputs/Necessary_Outputs/",
                         action="store", type="string",
                         help="path of SUMO output files")

    # Input for Estimating Accuracy of TTs Prediction
    optParser.add_option("--std_len_range", default=20, action="store", type="int",
                         help="time interval of TTs comparison between HTT and real values, in seconds (integer)")
    optParser.add_option("--warm_up_t", default=1800, action="store", type="int",
                         help="Warm up time for the travel time reliability report")

    # SUMO Simulation Time Interval
    optParser.add_option("--begin", default="43200", action="store", type="string",
                         help="Begin Step Simulation")
    optParser.add_option("--end", default="55800", action="store", type="string",
                         help="End Step Simulation")
    optParser.add_option("--step_length", default="1", action="store", type="string",
                         help="Simulation Step Length in seconds")

    # SUMO Routing Parameters
    optParser.add_option("--routing_probability", default="0.7", action="store", type="string",
                         help="Probability of a vehicle with routing device")
    optParser.add_option("--routing_period", default="120", action="store", type="string",
                         help="Interval between routing queries")
    optParser.add_option("--routing_pre_period", default="60", action="store", type="string",
                         help="Interval between routing queries before trip")

    # SUMO Edge Closure Input for Artificial Edge Selection
    optParser.add_option("--interval_edge_closure", nargs=2, default=(43201, 55801), action="store", type="float",
                         help="begin and end time of lane closure")
    optParser.add_option("--closed_art_edge", default="gneE46", action="store",
                         type="string", help="network id of artificial edge that will close. Use comma as separator")

    # Path Weight Files (Travel Times)
    optParser.add_option("--path_weight_files",
                         default="D:/Andre_Back_Up/a_Projects_Files/Local_Level_Routing/"
                                 "Prague_Zizkov/Inputs_Outputs/Scenario_Dependent_Travel_Times/",
                         action="store", type="string",
                         help="path of file with Travel Times")

    options, args = optParser.parse_args()
    return options

# Main Entry Point for SUMO and to Parse Variables
if __name__ == "__main__":
    options = get_options()

    # this script has been called from the command line. It will start sumo as a server, then connect and run
    try:
        if options.interface == "libsumo":
            import libsumo as traci
            # As GUI is not working yet with libsumo, use only without GUI
            sumoBinary = sumolib.checkBinary('sumo')
            INTERFACE = "libsumo"
        elif options.interface == "traci":
            raise ImportError
        else:
            # Import here module that connects the Python script with CAVs data interface
            pass

    except ImportError:
        import traci
        INTERFACE = "traci"
        if options.nogui:
            sumoBinary = sumolib.checkBinary('sumo')
        else:
            sumoBinary = sumolib.checkBinary('sumo-gui')

    # Define SUMO additional files
    add_filenames = options.fixed_additional_files.split(",")
    if options.dyn_additional_files != "":
        add_filenames.extend([dynf.split(".")[0] + "." + options.scenario_name + "." + ".".join(dynf.split(".")[1:])
                              for dynf in options.dyn_additional_files.split(",")])
    additional_files = options.path_inputs + options.path_inputs.join([afile + "," for afile in add_filenames][:-1]
                                                                      + [add_filenames[-1]])

    # Define SUMO route files
    route_filenames = options.fixed_route_files.split(",")
    if options.dyn_route_files != "":
        route_filenames.extend(
            [dynf.split(".")[0] + "." + options.scenario_name + "." + ".".join(dynf.split(".")[1:])
             for dynf in options.dyn_route_files.split(",")])
    route_files = options.path_inputs + options.path_inputs.join([rfile + "," for rfile in route_filenames][:-1]
                                                                 + [route_filenames[-1]])

    # this is the normal way of using traci. sumo is started as a
    # subprocess and then the python script connects and runs
    traci.start([sumoBinary, "-c", options.simul_filename,
                 "--net", options.path_inputs + options.net_filename,
                 "--route-files", route_files,
                 "--additional-files", additional_files,
                 "--vehroute-output",options.path_outputs + "vehroute." + options.scenario_name + ".xml",
                 "--tripinfo-output",options.path_outputs + "tripinfo." + options.scenario_name + ".xml",
                 "--weight-files",options.path_weight_files + "travel_times." + options.scenario_name.replace("HTT.","")
                 + "." + options.routing_period + "P" + ".xml",
                 "--begin",options.begin,
                 "--end",options.end,
                 "--step-length", options.step_length,
                 "--device.emissions.probability", "1",
                 "--device.rerouting.period",options.routing_period,
                 "--device.rerouting.pre-period",options.routing_pre_period,
                 "--device.rerouting.init-with-loaded-weights","true"])

options = get_options()

# TraCI Constants
LANE_QUEUE_NUM = traci.constants.LAST_STEP_VEHICLE_HALTING_NUMBER
LANE_VEH_NUM = traci.constants.LAST_STEP_VEHICLE_NUMBER
EDGE_VEHIDS = traci.constants.LAST_STEP_VEHICLE_ID_LIST
LANE_METRICS = [LANE_QUEUE_NUM, LANE_VEH_NUM]

# Prior edges are edges before artificial edges,
# artificial edges are edges representing one lane of a multi-lane incoming edge at certain junction.
# The index of artificial edge is the lane number of the prior edge
PRIOR_ARTIFICIAL_EDGES = {"gneE247": ("gneE248","gneE249"),
                          "fl_156262933#4": ("gneE246","gneE245"),
                          "fl_330512996#6.19": ("gneE240.50","gneE241.49"),
                          "gneE57": ("gneE60","gneE59"),
                          "fl_80779632#4": ("gneE148","gneE149"),
                          "dz_142147141": ("gneE143","gneE144"),
                          "dz_50178032#1.132": ("gneE47","gneE45")}

PRIOR_COMMON_AFTER_ART_EDGES = {"gneE247": "gneE293",
                                "fl_156262933#4": "fl_483444577#0",
                                "fl_330512996#6.19": "fl_448565742.11",
                                "gneE57": "fl_4641298#3",
                                "fl_80779632#4": "dz_330512997",
                                "dz_142147141": "fl_330512996#6.19",
                                "dz_50178032#1.132": "dz_142147149#0"}

# SUMO Connection
NOGUI = options.nogui

# Scenario Name
SCE_NAME = options.scenario_name

# Net File
NET_FILE = options.path_inputs + options.net_filename

# Road Closure
BEGIN_EDGE_CLOSURE = options.interval_edge_closure[0]
END_EDGE_CLOSURE = options.interval_edge_closure[1]
EDGE2CLOSE = options.closed_art_edge.split(",")

# Routing Probability for Setting Vehicles to Have Routing Device if not in List
ROUTING_PROB = float(options.routing_probability)

# SUMO Simulation Time Interval
BEGIN_T = round(float(options.begin),1)
END_T = round(float(options.end),1)
STEP_LENGTH = round(float(options.step_length),1)
VEH_LIST_FILE = options.path_outputs + "tripinfo." + options.scenario_name.replace("HTT.","LLR.") + ".xml"
ACTUAL_T = BEGIN_T
LEN_RANGE = round(float(options.std_len_range), 1) # time interval of arrival ranges in seconds (integer)

# CAVs that entered LLR area
CAVS_AREA_FILE = options.path_outputs + "CAVs_modelled_area." + SCE_NAME + ".xlsx"

# nonCAVs that entered LLR area
NONCAVS_AREA_FILE = options.path_outputs + "nonCAVs_modelled_area." + SCE_NAME + ".xlsx"

# All vehicles that entered LLR area
ALL_VEHS_AREA_FILE = options.path_outputs + "allVehs_modelled_area." + SCE_NAME + ".xlsx"

# All Incoming Edges of the LLR area, for the report of accuracy of TTs
ALL_INEDGES_AREA_FILE = options.path_outputs + "allInEdges_modelled_area." + SCE_NAME.split(".")[0] + ".xlsx"
ALL_INOUTEDGES_AREA_FILE = options.path_outputs + "allInOutEdges_modelled_area." + SCE_NAME.split(".")[0] + ".xlsx"

# File with Measured, Estimated and Std. of Travel Times
TTS_ACCUR_FULL = options.path_outputs + "TTs_accur." + SCE_NAME + ".xlsx"
TTS_ACCUR_AGGREGATED = options.path_outputs + "python_edge_traffic." + SCE_NAME + ".xml"
WARM_UP_T = round(float(options.warm_up_t), 1)


## LOGGING

logging.basicConfig(filename=options.path_outputs.replace("Necessary_Outputs/","Unecessary_Outputs/") + 'PYTHONlog.' +
                             options.scenario_name + '.log',level=logging.INFO,format='%(asctime)s %(message)s')
logger = logging.getLogger(__name__)
logger.info('Start Simulation')


## MAIN SCRIPT FUNCTIONS

def frange(start, end=None, inc=None):
    """A range function, that does accept float increments"""

    if end == None:
        end = start + 0.0
        start = 0.0

    if inc == None:
        inc = 1.0

    L = []
    while 1:
        next = start + len(L) * inc
        if inc > 0 and next >= end:
            break
        elif inc < 0 and next <= end:
            break
        L.append(next)

    return L

def getExcelTargetList(filename):
    """Get Values of Each Row of the First Collum of an Excel file"""

    book = openpyxl.load_workbook(filename, data_only=True)
    sheet = book.active
    rows = sheet.rows
    file_input_vals = []
    for row in rows:
        for cell in row:
            file_input_vals.append(cell.value)

    return file_input_vals

def measureVehicleRealTravelTime(vehID, edgeID, time2updte):
    """Store which edge vehicle is at each time step and measure travel time once changes edge"""

    # This configuration makes that vehicles detected on a internal lanes will not
    # be acounted as left the edge, only when arrived at the next one.
    try:
        if allVehs_Edge_ArrTime[vehID][0] != edgeID:
            # If vehicle is in another edge than the last one detected
            time_after_updte = LEN_RANGE - allVehs_Edge_ArrTime[vehID][2]
            # range is between [ACTUAL_T, ACTUAL_T + LEN_RANGE],
            # when time_after_updte = 0, means a new range
            # and time_after_updte is always lower and equal than LEN_RANGE
            # used to define which range the travel time is suitable
            if time_after_updte == LEN_RANGE:
                time_after_updte = 0

            beginTime = allVehs_Edge_ArrTime[vehID][1] - time_after_updte
            if beginTime >= BEGIN_T + WARM_UP_T + STEP_LENGTH and beginTime <= END_T + STEP_LENGTH:
                try:
                    # Store measured travel time for the range
                    tts_accur_vals[allVehs_Edge_ArrTime[vehID][0]]["measured"][beginTime].append(
                        ACTUAL_T - allVehs_Edge_ArrTime[vehID][1])
                except KeyError:
                    # If not found values for the range, create a key for it and store measured
                    # travel time
                    tts_accur_vals[allVehs_Edge_ArrTime[vehID][0]]["measured"].update({
                        beginTime: [ACTUAL_T - allVehs_Edge_ArrTime[vehID][1]]})

            # Update new edge to monitor
            allVehs_Edge_ArrTime.update({vehID: (edgeID, ACTUAL_T, time2updte)})

    except KeyError:
        # If vehicle not detected before, start monitoring vehicle's edge
        # time2updte is necessary to define which range the measurement is suitable
        # as this function runs at every simulation time step and not every algorithm update
        allVehs_Edge_ArrTime.update({vehID: (edgeID, ACTUAL_T, time2updte)})


# Initializing for measuring accuracy of travel time prediction
allInEdges_modelled_area = getExcelTargetList(ALL_INEDGES_AREA_FILE)
all_InOut_edges = getExcelTargetList(ALL_INOUTEDGES_AREA_FILE)
tts_accur_vals = {edgeID: {"measured": dict(),
                           "estimation": {beginTime: [] for beginTime in frange(BEGIN_T + WARM_UP_T + STEP_LENGTH,
                                                                                END_T + STEP_LENGTH + STEP_LENGTH,
                                                                                LEN_RANGE)},
                           "RMSE": dict()}
                  for edgeID in allInEdges_modelled_area}
allVehs_Edge_ArrTime = dict()
del allInEdges_modelled_area

# Initializing for list of vehicles within the modelled area
cavs_area = set([])
noncavs_area = set([])
all_vehs_area = set([])
rem_cavs_list = set([])
rem_noncavs_list = set([])

# Get CAVs of last simulated scenario to ensure same vehicles as CAV
tree_all_vehs = ElementTree.parse(VEH_LIST_FILE)
trips = list(tree_all_vehs.iter('tripinfo'))
for t in trips:
    if "routing" in t.get('devices'):
        rem_cavs_list.add(t.get('id'))
    else:
        rem_noncavs_list.add(t.get('id'))

# Subscribe to get data from modelled edges
for edgeID in all_InOut_edges:
    traci.edge.subscribe(edgeID, [EDGE_VEHIDS])

# Get lanes and length of the artificial edges and also the edges sending vehicles to them (called prior edges)
net = sumolib.net.readNet(NET_FILE)
priorAndArtificial_edgesLanes = dict()
prior_edges_len = dict()
edge_endLane_vehs = dict()
for priorEdgeID in PRIOR_ARTIFICIAL_EDGES.keys():
    prior_edges_len.update({priorEdgeID: net.getEdge(priorEdgeID).getLength()})
    edge_part = priorEdgeID + "_"
    priorAndArtificial_edgesLanes.update({priorEdgeID: [edge_part + str(lane_num)
                                                       for lane_num in range(0, traci.edge.getLaneNumber(priorEdgeID))]})
    # Init. list of vehicles to be set as their next edges (artificial one) already defined
    edge_endLane_vehs.update({priorEdgeID: set([])})
    for artEdgeID in PRIOR_ARTIFICIAL_EDGES[priorEdgeID]:
        edge_part = artEdgeID + "_"
        priorAndArtificial_edgesLanes.update({artEdgeID: [edge_part + str(lane_num)
                                                       for lane_num in range(0, traci.edge.getLaneNumber(artEdgeID))]})

# Subscribe to get data from modelled lanes
for laneIDs in priorAndArtificial_edgesLanes.values():
    for laneID in laneIDs:
        traci.lane.subscribe(laneID, LANE_METRICS)

del net

# Constant Loop to Communicate with SUMO/TraCI
time2updte = 0.0
try:
    while ACTUAL_T < END_T + STEP_LENGTH:
        if INTERFACE == "libsumo":
            # If Libsumo open a thread to avoid the simulation get stuck
            stored_ACTUAL_T = ACTUAL_T
            Thread(target=traci.simulationStep()).start()
            time.sleep(0.02)
            # Simulation Data
            ACTUAL_T = round(float(traci.simulation.getTime()), 1) # Returns the current simulation time in seconds
            if ACTUAL_T == stored_ACTUAL_T:
                print("SUMO stopped")
                Thread(target=traci.simulationStep()).start()
                time.sleep(1)
                ACTUAL_T = round(float(traci.simulation.getTime()), 1)
                if ACTUAL_T == stored_ACTUAL_T:
                    raise SystemExit
                else:
                    print("SUMO is back")

            sys.stdout.write('\r' + "Step " + str(ACTUAL_T))
        else:
            # If Traci
            traci.simulationStep()
            # Simulation Data
            ACTUAL_T = round(float(traci.simulation.getTime()), 1)  # Returns the current simulation time in seconds

        # Get list of vehicles from SUMO
        net_departed_vehs = set(traci.simulation.getDepartedIDList())
        net_new_CAVs = rem_cavs_list.intersection(net_departed_vehs)
        net_new_nonCAVs = rem_noncavs_list.intersection(net_departed_vehs)

        # Vehicles not simulated in the previous scenario
        net_to_define = net_departed_vehs.difference(net_new_CAVs.union(net_new_nonCAVs))
        for veh2define in net_to_define:
            # If vehicle was not simulated in the previous scenario, define it CAV or not based on penetration rate
            if np.random.binomial(1, ROUTING_PROB):
                # Define as CAV
                net_new_CAVs.add(veh2define)
            else:
                # Define as nonCAV
                net_new_nonCAVs.add(veh2define)

        for cavID in net_new_CAVs:
            try:
                # Update list of remaining CAVs to be simulated
                rem_cavs_list.remove(cavID)
            except KeyError:
                skip = 1

            # Set vehicle as CAV
            traci.vehicle.setParameter(cavID, "has.rerouting.device", "true")
            if NOGUI == False and "probe" not in cavID:
                # Change colour of CAV to blue (default is white)
                traci.vehicle.setColor(cavID, (0,0,255))

        # Update list of remaining nonCAVs to be simulated
        for nonCavID in net_new_nonCAVs:
            try:
                rem_noncavs_list.remove(nonCavID)
            except KeyError:
                skip = 1

        # Algorithm to spread vehicles evenly to next edge (only artificial ones) based on queue length and/or number
        global lowest_art_lanes_measure,lowest_prior_lanes_queue
        # For all modelled edges (incoming and outgoing edges of junctions)
        for edgeID in all_InOut_edges:
            # Get list of vehicles running on the modelled edges
            edge_vehs = set(traci.edge.getSubscriptionResults(edgeID)[EDGE_VEHIDS])
            if edge_vehs != set([]):
                try:
                    # Define vehicles that already defined their next edges (artificial one)
                    edge_endLane_vehs[edgeID].intersection_update(edge_vehs)
                    # Select artificial edges only that are not closed
                    aval_art_edges = [art_edge for art_edge in PRIOR_ARTIFICIAL_EDGES[edgeID]
                                      if art_edge not in EDGE2CLOSE]
                    num_lanes = len(priorAndArtificial_edgesLanes[edgeID])
                    if num_lanes > 1:
                        # If only more than 1 possible next artificial edge
                        lowest_prior_lanes_queue = (None, float("inf"))
                        for lane_num,laneID in enumerate(priorAndArtificial_edgesLanes[edgeID]):
                            # For each lane of prior edge, choose lane with shortest queue
                            # and its respective artificial edge
                            try:
                                art_edgeID = PRIOR_ARTIFICIAL_EDGES[edgeID][lane_num]
                            except IndexError:
                                art_edgeID = aval_art_edges[-1]
                            if PRIOR_ARTIFICIAL_EDGES[edgeID][lane_num] not in EDGE2CLOSE \
                            or (ACTUAL_T < BEGIN_EDGE_CLOSURE or ACTUAL_T > END_EDGE_CLOSURE):
                                lane_queue = traci.lane.getSubscriptionResults(laneID)[LANE_QUEUE_NUM]
                                if lane_queue < lowest_prior_lanes_queue[1]:
                                    # Update lowest queue
                                    lowest_prior_lanes_queue = (art_edgeID, lane_queue)
                                elif lane_num == (num_lanes - 1) and lane_queue == lowest_prior_lanes_queue[1]:
                                    # All queue lanes are the same, then maintain same prior lane or change
                                    lowest_prior_lanes_queue = None
                    else:
                        # If only 1 possible next artificial edge
                        lowest_prior_lanes_queue = None

                    # Among all lanes of artificial edges, choose edge with following priority:
                    # 1) shortest queue
                    # 2) lowest number of vehicles
                    sum_lane = 0
                    for lane_metric in LANE_METRICS:
                        lowest_art_lanes_measure = (None, float("inf"))
                        for art_edgeID in PRIOR_ARTIFICIAL_EDGES[edgeID]:
                            if art_edgeID not in EDGE2CLOSE \
                            or (ACTUAL_T < BEGIN_EDGE_CLOSURE or ACTUAL_T > END_EDGE_CLOSURE):
                                # If artificial edge is not closed
                                for lane_num, laneID in enumerate(priorAndArtificial_edgesLanes[art_edgeID]):
                                    # For each lane of artificial edge,
                                    # choose lane with either shortest queue or number of vehicles
                                    # and take its respective artificial edge
                                    lane_measure = traci.lane.getSubscriptionResults(laneID)[lane_metric]
                                    sum_lane += lane_measure
                                    if lane_measure < lowest_art_lanes_measure[1]:
                                        # Update lowest metric
                                        lowest_art_lanes_measure = (art_edgeID, lane_measure)
                                    elif lane_num == (num_lanes - 1) and lane_measure == lowest_art_lanes_measure[1]:
                                        # All queue lanes are the same, then maintain same prior lane or change
                                        lowest_art_lanes_measure = None
                        if sum_lane > 0:
                            # If sum of lane metric is zero, use next measure
                            break
                except KeyError:
                    # If edge not a prior edge
                    lowest_prior_lanes_queue = None
                    lowest_art_lanes_measure = None

                # For all vehicles running on the modelled edges
                for vehID in edge_vehs:
                    # all_vehs_area.add(vehID) # only when travel time accuracy not estimated
                    # Measure vehicle edge travel time for measuring accuracy of travel time prediction
                    measureVehicleRealTravelTime(vehID, edgeID, time2updte)
                    if traci.vehicle.getParameter(vehID, "has.rerouting.device") == "true":
                        # Add vehicle to the set of CAVs detected in the modelled area
                        cavs_area.add(vehID)

                        # Initialize boolean values for deciding next artificial edge
                        decide_current_lane = 0
                        decide_next_lane = 0
                        maintain_current_lane = 0
                        try:
                            if PRIOR_ARTIFICIAL_EDGES.has_key(edgeID):
                                # Get vehicle next edges
                                veh_route_index = traci.vehicle.getRouteIndex(vehID)
                                veh_route = list(traci.vehicle.getRoute(vehID)[veh_route_index:])
                                if vehID not in edge_endLane_vehs[edgeID]:
                                    # If not defined yet vehicle's next artificial edge
                                    if veh_route[2] == PRIOR_COMMON_AFTER_ART_EDGES[edgeID]:
                                        veh_accel = traci.vehicle.getAcceleration(vehID)
                                        veh_speed = traci.vehicle.getSpeed(vehID)
                                        veh_dist = traci.vehicle.getLanePosition(vehID)
                                        # Define distance to stop line after next time step
                                        dist_nextSteps = veh_dist + veh_speed
                                        if (veh_accel < -traci.vehicle.getDecel(vehID) / 2.5
                                        and lowest_prior_lanes_queue != None
                                        and dist_nextSteps < prior_edges_len[edgeID]):
                                            # If vehicle braking more than the maximum deceleration / 2.5
                                            # Decide current lane, to define respective artificial edge
                                            # of the shortest queue of current prior edge
                                            decide_current_lane = 1
                                        elif veh_speed == 0:
                                            # Maintain same current lane, to define respective artificial edge
                                            # of current lane of prior edge
                                            maintain_current_lane = 1
                                        elif dist_nextSteps >= prior_edges_len[edgeID]:
                                            # Set vehicle as already defined next artificial edge
                                            edge_endLane_vehs[edgeID].add(vehID)
                                            if (lowest_prior_lanes_queue == None or lowest_prior_lanes_queue[1] == 0) \
                                            and lowest_art_lanes_measure != None:
                                                # Decide next lane (respective artificial edge),
                                                # based on the shortest queue or lowest number of vehicles of next lane
                                                decide_next_lane = 1
                                            else:
                                                # Maintain same current lane, to define respective artificial edge
                                                # of current lane of prior edge
                                                maintain_current_lane = 1
                                else:
                                    # Maintain same current lane, to define respective artificial edge
                                    # of current lane of prior edge
                                    maintain_current_lane = 1
                        except IndexError:
                            # Vehicle finishing trip
                            skip = 1

                        if decide_current_lane == 1 or decide_next_lane == 1 or maintain_current_lane == 1:
                            if maintain_current_lane == 1:
                                # When decide_current_lane == 1
                                lane_num = traci.vehicle.getLaneIndex(vehID)
                                if PRIOR_ARTIFICIAL_EDGES[edgeID][lane_num] not in EDGE2CLOSE \
                                or (ACTUAL_T < BEGIN_EDGE_CLOSURE or ACTUAL_T > END_EDGE_CLOSURE):
                                    # The next artificial edge is based on the current lane of the prior edge
                                    veh_route[1] = PRIOR_ARTIFICIAL_EDGES[edgeID][lane_num]
                                else:
                                    # If more prior lanes than artificial edges
                                    veh_route[1] = aval_art_edges[-1]
                                traci.vehicle.setRoute(vehID, veh_route)
                            elif decide_current_lane == 1 and veh_route[1] != lowest_prior_lanes_queue[0]:
                                # When decide_current_lane == 1
                                # The artificial edge is based on the shortest queue of current lane of the prior edge
                                veh_route[1] = lowest_prior_lanes_queue[0]
                                traci.vehicle.setRoute(vehID, veh_route)
                            elif lowest_art_lanes_measure != None and veh_route[1] != lowest_art_lanes_measure[0]:
                                # When decide_next_lane == 1
                                # The next artificial edge is based on the lowest metric of all artificial edges
                                veh_route[1] = lowest_art_lanes_measure[0]
                                traci.vehicle.setRoute(vehID, veh_route)
                    else:
                        # Add vehicle to the set of nonCAVs detected in the modelled area
                        noncavs_area.add(vehID)

        # Store estimated travel times for measuring accuracy of travel time prediction
        if time2updte == 0:
            if ACTUAL_T >= BEGIN_T + WARM_UP_T + STEP_LENGTH and ACTUAL_T <= END_T + STEP_LENGTH:
                for edgeID in tts_accur_vals:
                    traveltime = traci.edge.getAdaptedTraveltime(edgeID, ACTUAL_T)
                    # Account Hourly Travel Times from OBUs
                    tts_accur_vals[edgeID]["estimation"][ACTUAL_T].append(traveltime)
                time2updte = LEN_RANGE - STEP_LENGTH
        else:
            time2updte -= STEP_LENGTH

    # Create workbook with and a new sheet for the CAVs Area
    if os.path.isfile(CAVS_AREA_FILE) == True:
        os.remove(CAVS_AREA_FILE)
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "CAVs modelled Area"
    for cav_ind, cav in enumerate(cavs_area):
        sheet.cell(row=cav_ind + 1, column=1).value = cav
    wb.save(CAVS_AREA_FILE)

    # Create workbook with and a new sheet for the nonCAVs Area
    if os.path.isfile(NONCAVS_AREA_FILE) == True:
        os.remove(NONCAVS_AREA_FILE)
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "nonCAVs modelled Area"
    for cav_ind, cav in enumerate(noncavs_area):
        sheet.cell(row=cav_ind + 1, column=1).value = cav
    wb.save(NONCAVS_AREA_FILE)

    # Create workbook with and a new sheet for all vehicles in the modelled area
    if os.path.isfile(ALL_VEHS_AREA_FILE) == True:
        os.remove(ALL_VEHS_AREA_FILE)
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "All Vehs. modelled Area"
    for cav_ind, cav in enumerate(allVehs_Edge_ArrTime.keys()):
        sheet.cell(row=cav_ind + 1, column=1).value = cav
    wb.save(ALL_VEHS_AREA_FILE)

    # For measuring accuracy of travel time prediction
    for edgeID in tts_accur_vals:
        for beginTime in tts_accur_vals[edgeID]["measured"]:
            measured_tt = np.mean(tts_accur_vals[edgeID]["measured"][beginTime])
            tts_accur_vals[edgeID]["RMSE"].update({beginTime: []})
            for est_tt in tts_accur_vals[edgeID]["estimation"][beginTime]:
                tts_accur_vals[edgeID]["RMSE"][beginTime].append((est_tt - measured_tt) ** 2)
            tts_accur_vals[edgeID]["RMSE"][beginTime] = round(math.sqrt(np.sum(tts_accur_vals[edgeID]["RMSE"][beginTime])
                                                                    / len(tts_accur_vals[edgeID]["RMSE"][beginTime])), 1)

    # Create XML file with measured, estimated and ratio of std. deviations to measured travel time for each edge.
    # Create the file structure
    mean_data = ElementTree.Element('meandata')
    items = ElementTree.SubElement(mean_data, 'interval')
    for edgeID in tts_accur_vals:
        item = ElementTree.SubElement(items, 'edge')
        item.set('id', str(edgeID))
        all_vals_measured = []
        all_vals_estimated = []
        all_vals_RMSE = []
        for beginTime in tts_accur_vals[edgeID]["measured"]:
            all_vals_measured.append(np.mean(tts_accur_vals[edgeID]["measured"][beginTime]))
            all_vals_estimated.append(np.mean(tts_accur_vals[edgeID]["estimation"][beginTime]))
            all_vals_RMSE.append(tts_accur_vals[edgeID]["RMSE"][beginTime])
        if all_vals_measured != []:
            item.set('measuredTravelTime', str(round(np.mean(all_vals_measured), 1)))
        if all_vals_estimated != []:
            item.set('estimatedTravelTime', str(round(np.mean(all_vals_estimated), 1)))
        if all_vals_RMSE != []:
            item.set('RMSE', str(round(np.mean(all_vals_RMSE), 1)))

    # Save the new XML file with the results
    if os.path.isfile(TTS_ACCUR_AGGREGATED) == True:
        os.remove(TTS_ACCUR_AGGREGATED)
    xml_string = ElementTree.tostring(mean_data)
    parsed_xml_string = xml.dom.minidom.parseString(xml_string)
    python_edge_traffic = parsed_xml_string.toprettyxml()
    with open(TTS_ACCUR_AGGREGATED, "w") as f:
        f.write(python_edge_traffic)

    # Create workbook with the measured, and estimated travel times as well as
    if os.path.isfile(TTS_ACCUR_FULL) == True:
        os.remove(TTS_ACCUR_FULL)
    wb = openpyxl.Workbook()
    mapping_time = dict()
    # Create a new sheet for the function
    sheet = wb.active
    sheet.title = "Measured TTs"
    sheet.cell(row=1, column=1).value = "EdgeID"
    for time_ind, beginTime in enumerate(frange(BEGIN_T + WARM_UP_T + STEP_LENGTH,
                                                END_T + STEP_LENGTH + STEP_LENGTH,
                                                LEN_RANGE)):
        mapping_time.update({beginTime: time_ind})
        sheet.cell(row=1, column=time_ind + 2).value = beginTime
    for edge_ind,edgeID in enumerate(tts_accur_vals):
        sheet.cell(row=edge_ind + 2, column=1).value = edgeID
        if tts_accur_vals[edgeID]["measured"] != {}:
            for beginTime in tts_accur_vals[edgeID]["measured"]:
                time_ind = mapping_time[beginTime]
                sheet.cell(row=edge_ind + 2, column=time_ind + 2).value = str(
                    tts_accur_vals[edgeID]["measured"][beginTime])

    sheet = wb.create_sheet("Estimated TTs")
    sheet.cell(row=1, column=1).value = "EdgeID"
    for time_ind, beginTime in enumerate(frange(BEGIN_T + WARM_UP_T + STEP_LENGTH,
                                                END_T + STEP_LENGTH + STEP_LENGTH,
                                                LEN_RANGE)):
        sheet.cell(row=1, column=time_ind + 2).value = beginTime
    for edge_ind, edgeID in enumerate(tts_accur_vals):
        sheet.cell(row=edge_ind + 2, column=1).value = edgeID
        for beginTime in tts_accur_vals[edgeID]["estimation"]:
            time_ind = mapping_time[beginTime]
            sheet.cell(row=edge_ind + 2, column=time_ind + 2).value = str(
                tts_accur_vals[edgeID]["estimation"][beginTime])

    # Replace all values for its average for next calculations
    for edgeID in tts_accur_vals:
        for beginTime in tts_accur_vals[edgeID]["measured"]:
            tts_accur_vals[edgeID]["measured"][beginTime] = round(
                np.mean(tts_accur_vals[edgeID]["measured"][beginTime]), 1)

    for edgeID in tts_accur_vals:
        for beginTime in tts_accur_vals[edgeID]["estimation"]:
            tts_accur_vals[edgeID]["estimation"][beginTime] = round(
                np.mean(tts_accur_vals[edgeID]["estimation"][beginTime]), 1)

    sheet = wb.create_sheet("Mean Measured TTs")
    sheet.cell(row=1, column=1).value = "EdgeID"
    for time_ind, beginTime in enumerate(frange(BEGIN_T + WARM_UP_T + STEP_LENGTH,
                                                END_T + STEP_LENGTH + STEP_LENGTH,
                                                LEN_RANGE)):
        sheet.cell(row=1, column=time_ind + 2).value = beginTime
    for edge_ind,edgeID in enumerate(tts_accur_vals):
        sheet.cell(row=edge_ind + 2, column=1).value = edgeID
        if tts_accur_vals[edgeID]["measured"] != {}:
            for beginTime in tts_accur_vals[edgeID]["measured"]:
                time_ind = mapping_time[beginTime]
                sheet.cell(row=edge_ind + 2, column=time_ind + 2).value = tts_accur_vals[edgeID]["measured"][beginTime]

    sheet = wb.create_sheet("Mean Estimated TTs")
    sheet.cell(row=1, column=1).value = "EdgeID"
    for time_ind, beginTime in enumerate(frange(BEGIN_T + WARM_UP_T + STEP_LENGTH,
                                                END_T + STEP_LENGTH + STEP_LENGTH,
                                                LEN_RANGE)):
        sheet.cell(row=1, column=time_ind + 2).value = beginTime
    for edge_ind,edgeID in enumerate(tts_accur_vals):
        sheet.cell(row=edge_ind + 2, column=1).value = edgeID
        for beginTime in tts_accur_vals[edgeID]["estimation"]:
            time_ind = mapping_time[beginTime]
            sheet.cell(row=edge_ind + 2, column=time_ind + 2).value = tts_accur_vals[edgeID]["estimation"][beginTime]

    sheet = wb.create_sheet("RMSE")
    sheet.cell(row=1, column=1).value = "EdgeID"
    for time_ind, beginTime in enumerate(frange(BEGIN_T + WARM_UP_T + STEP_LENGTH,
                                                END_T + STEP_LENGTH + STEP_LENGTH,
                                                LEN_RANGE)):
        sheet.cell(row=1, column=time_ind + 2).value = beginTime
    for edge_ind, edgeID in enumerate(tts_accur_vals):
        sheet.cell(row=edge_ind + 2, column=1).value = edgeID
        for beginTime in tts_accur_vals[edgeID]["RMSE"]:
            time_ind = mapping_time[beginTime]
            sheet.cell(row=edge_ind + 2, column=time_ind + 2).value = tts_accur_vals[edgeID]["RMSE"][beginTime]

    sheet = wb.create_sheet("Normalized RMSE")
    sheet.cell(row=1, column=1).value = "EdgeID"
    for time_ind, beginTime in enumerate(frange(BEGIN_T + WARM_UP_T + STEP_LENGTH,
                                                END_T + STEP_LENGTH + STEP_LENGTH,
                                                LEN_RANGE)):
        sheet.cell(row=1, column=time_ind + 2).value = beginTime
    for edge_ind,edgeID in enumerate(tts_accur_vals):
        sheet.cell(row=edge_ind + 2, column=1).value = edgeID
        if tts_accur_vals[edgeID]["RMSE"] != {}:
            for beginTime in tts_accur_vals[edgeID]["RMSE"]:
                time_ind = mapping_time[beginTime]
                sheet.cell(row=edge_ind + 2,
                           column=time_ind + 2).value = "{:.2%}".format(tts_accur_vals[edgeID]["RMSE"][beginTime]
                                                                        / tts_accur_vals[edgeID]["measured"][beginTime])

    sheet = wb.create_sheet("Stats per Edge")
    sheet.cell(row=1, column=1).value = "EdgeID"
    sheet.cell(row=1, column=2).value = "Mean Estimated"
    sheet.cell(row=1, column=3).value = "Mean Measured"
    sheet.cell(row=1, column=4).value = "Mean RMSE"
    sheet.cell(row=1, column=5).value = "Mean Norm. RMSE"
    for edge_ind, edgeID in enumerate(tts_accur_vals):
        all_Estimated = []
        all_Measured = []
        all_RMSE = []
        sheet.cell(row=edge_ind + 2, column=1).value = edgeID
        if tts_accur_vals[edgeID]["estimation"] != {}:
            for beginTime in tts_accur_vals[edgeID]["estimation"]:
                all_Estimated.append(tts_accur_vals[edgeID]["estimation"][beginTime])
            sheet.cell(row=edge_ind + 2, column=2).value = round(np.mean(all_Estimated), 1)
        if tts_accur_vals[edgeID]["measured"] != {}:
            for beginTime in tts_accur_vals[edgeID]["measured"]:
                all_Measured.append(tts_accur_vals[edgeID]["measured"][beginTime])
            sheet.cell(row=edge_ind + 2, column=3).value = round(np.mean(all_Measured), 1)
        if tts_accur_vals[edgeID]["RMSE"] != {}:
            for beginTime in tts_accur_vals[edgeID]["RMSE"]:
                all_RMSE.append(tts_accur_vals[edgeID]["RMSE"][beginTime])
            sheet.cell(row=edge_ind + 2, column=4).value = round(np.mean(all_RMSE), 1)
            sheet.cell(row=edge_ind + 2, column=5).value = "{:.2%}".format(np.mean(all_RMSE) / np.mean(all_Measured))

    sheet = wb.create_sheet("Stats per Time")
    sheet.cell(row=1, column=1).value = "Time"
    sheet.cell(row=1, column=2).value = "Mean Estimated"
    sheet.cell(row=1, column=3).value = "Mean Measured"
    sheet.cell(row=1, column=4).value = "Mean RMSE"
    sheet.cell(row=1, column=5).value = "Mean Norm. RMSE"
    for time_ind, beginTime in enumerate(frange(BEGIN_T + WARM_UP_T + STEP_LENGTH,
                                                END_T + STEP_LENGTH + STEP_LENGTH,
                                                LEN_RANGE)):
        sheet.cell(row=time_ind + 2, column=1).value = beginTime
        all_RMSE = []
        all_Measured = []
        all_Estimated = []
        for edge_ind, edgeID in enumerate(tts_accur_vals):
            try:
                all_Estimated.append(tts_accur_vals[edgeID]["estimation"][beginTime])
            except KeyError:
                skip = 1
            try:
                all_Measured.append(tts_accur_vals[edgeID]["measured"][beginTime])
            except KeyError:
                skip = 1
            try:
                all_RMSE.append(tts_accur_vals[edgeID]["RMSE"][beginTime])
            except KeyError:
                skip = 1
        if all_Estimated != []:
            sheet.cell(row=time_ind + 2, column=2).value = round(np.mean(all_Estimated), 1)
        if all_Measured != []:
            sheet.cell(row=time_ind + 2, column=3).value = round(np.mean(all_Measured), 1)
        if all_RMSE != []:
            sheet.cell(row=time_ind + 2, column=4).value = round(np.mean(all_RMSE), 1)
            sheet.cell(row=time_ind + 2, column=5).value = "{:.2%}".format(np.mean(all_RMSE) / np.mean(all_Measured))

    sheet = wb.create_sheet("Stats Total")
    sheet.cell(row=1, column=1).value = "Sum Edge Mean Estimated"
    sheet.cell(row=1, column=2).value = "Sum Edge Mean Measured"
    sheet.cell(row=1, column=3).value = "Sum RMSE"
    sheet.cell(row=1, column=4).value = "Norm. Sum RMSE"
    all_RMSE = []
    all_Measured = []
    all_Estimated = []
    for edge_ind, edgeID in enumerate(tts_accur_vals):
        if tts_accur_vals[edgeID]["estimation"] != {}:
            edge_mean = []
            for time_ind, beginTime in enumerate(tts_accur_vals[edgeID]["estimation"]):
                edge_mean.append(tts_accur_vals[edgeID]["estimation"][beginTime])
            all_Estimated.append(np.mean(edge_mean))
        if tts_accur_vals[edgeID]["measured"] != {}:
            edge_mean = []
            for time_ind, beginTime in enumerate(tts_accur_vals[edgeID]["measured"]):
                edge_mean.append(tts_accur_vals[edgeID]["measured"][beginTime])
            all_Measured.append(np.mean(edge_mean))
        if tts_accur_vals[edgeID]["RMSE"] != {}:
            edge_mean = []
            for time_ind, beginTime in enumerate(tts_accur_vals[edgeID]["RMSE"]):
                edge_mean.append(tts_accur_vals[edgeID]["RMSE"][beginTime])
            all_RMSE.append(np.mean(edge_mean))

    if all_Estimated != []:
        sheet.cell(row=2, column=1).value = round(sum(all_Estimated), 1)
    if all_Measured != []:
        sheet.cell(row=2, column=2).value = round(sum(all_Measured), 1)
    if all_RMSE != []:
        sheet.cell(row=2, column=3).value = round(sum(all_RMSE), 1)
        sheet.cell(row=2, column=4).value = "{:.2%}".format(sum(all_RMSE) / sum(all_Measured))

    # Finally, save the file and give it a name
    wb.save(TTS_ACCUR_FULL)

    # Close Traci/Libsumo connection
    traci.close()

    logger.info('End Simulation')

    sys.stdout.write('\n')
    sys.stdout.flush()

except (SystemExit, KeyboardInterrupt):
    # Close Traci/Libsumo connection
    traci.close()
    raise
except Exception, e:
    print("\n There was an error, check the log file \n")
    logger.error('Crash', exc_info=True)
    sys.exit(1)